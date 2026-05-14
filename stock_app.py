#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""個人股票管理工具 - 買賣記錄 / 庫存樹狀圖(finviz風格) / 個股分析"""

import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os, threading
import warnings

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import load_workbook

from tkcalendar import Calendar

import matplotlib, matplotlib.ticker
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.font_manager as fm
from matplotlib.patches import Patch
from matplotlib.lines import Line2D

import squarify
import yfinance as yf
import math

# 全域靜音 yfinance 所有子 logger（含 yfinance.base 等），避免下架/查無股票時噴 WARNING
import logging as _logging
for _yf_logger_name in ('yfinance', 'yfinance.base', 'yfinance.utils',
                         'yfinance.scrapers', 'yfinance.cache'):
    _logging.getLogger(_yf_logger_name).setLevel(_logging.CRITICAL)
import warnings as _warnings
_warnings.filterwarnings('ignore', module='yfinance')

# ── PyInstaller exe SSL 憑證修正 ────────────────────────────────────────────
# certifi CA bundle 需要明確指定路徑，否則 requests/yfinance 在 exe 環境無法驗證 SSL
import sys as _sys
try:
    import certifi as _certifi
    import os as _os
    _ca = _certifi.where()
    _os.environ.setdefault('SSL_CERT_FILE',      _ca)
    _os.environ.setdefault('REQUESTS_CA_BUNDLE', _ca)
except Exception:
    pass
import json as _json
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

try:
    from curl_cffi import requests as _cffi_req
    _CFFI_OK = True
except ImportError:
    _CFFI_OK = False

def _net_log(msg: str):
    """寫入網路除錯日誌（僅在 PyInstaller 打包環境下）。"""
    try:
        import sys as _sys
        if getattr(_sys, 'frozen', False):
            import os as _os, datetime as _dt
            log_path = _os.path.join(_os.path.dirname(_sys.executable), 'net_debug.log')
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write(f"[{_dt.datetime.now():%H:%M:%S}] {msg}\n")
    except Exception:
        pass

def _cffi_get_json(url, params=None, headers=None, timeout=15):
    """curl_cffi-based JSON GET that handles corporate DNS/firewall issues."""
    _hdrs = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/122.0.0.0',
        'Accept': 'application/json',
    }
    if headers:
        _hdrs.update(headers)
    if _CFFI_OK:
        try:
            r = _cffi_req.get(url, params=params, headers=_hdrs, timeout=timeout, verify=False)
            return _json.loads(r.content.decode('utf-8'))
        except Exception as _e:
            _net_log(f'curl_cffi FAIL [{url[:60]}]: {_e}')
            # curl_cffi 執行時失敗，降級為 urllib
    import urllib.request, urllib.parse, ssl
    try:
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        if params:
            url = url + '?' + urllib.parse.urlencode(params)
        req = urllib.request.Request(url, headers=_hdrs)
        with urllib.request.urlopen(req, context=ctx, timeout=timeout) as resp:
            return _json.loads(resp.read().decode('utf-8'))
    except Exception as _e2:
        _net_log(f'urllib FAIL [{url[:60]}]: {_e2}')
        raise

def _cffi_get_text(url: str, headers: dict | None = None, timeout: int = 15) -> str:
    """curl_cffi-based HTML GET returning raw UTF-8 text."""
    _hdrs = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/122.0.0.0 Safari/537.36',
        **(headers or {}),
    }
    if _CFFI_OK:
        try:
            r = _cffi_req.get(url, headers=_hdrs, timeout=timeout,
                              impersonate='chrome110', verify=False)
            return r.content.decode('utf-8', errors='replace')
        except Exception as _e:
            _net_log(f'curl_cffi text FAIL [{url[:60]}]: {_e}')
    import urllib.request, ssl
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    with urllib.request.urlopen(
            urllib.request.Request(url, headers=_hdrs),
            context=ctx, timeout=timeout) as r:
        return r.read().decode('utf-8', errors='replace')


warnings.filterwarnings('ignore')

# ─── 路徑與欄位 ──────────────────────────────────────────────────────────────
if getattr(_sys, 'frozen', False):
    BASE_DIR = os.path.dirname(_sys.executable)  # 執行檔所在目錄
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, '股票紀錄.xlsx')
SHEET_NAME = '交易記錄'

COLUMNS = ['日期', '股票代號', '股票名稱', '分類', '買賣',
           '數量(股)', '價格(元)', '手續費(元)', '交易稅(元)', '淨金額(元)']
CATEGORIES = ['定期定額', '波段操作', '爸爸合資', '獲利賣出']

# ─── 股票分割紀錄 ─────────────────────────────────────────────────────────────
# 格式：股票代號 → [(分割生效日, 換股比例), ...]
# 換股比例 = 分割後每1舊股換得新股數（例如1拆22 → ratio=22）
# 生效日當天及之後的交易已是新單位，不需換算；生效日之前的交易數量 × ratio
from datetime import date as _date
STOCK_SPLITS: dict[str, list[tuple[_date, int]]] = {
    '00631L': [(_date(2026, 3, 31), 22)],   # 元大台灣50正2  1拆22，首交易日 3/31
}

ETF_LIST = [
    # ── 市值型 ──────────────────────────────────────────────────────────────
    ('0050',   '元大台灣50'),
    ('006208', '富邦台灣50'),
    ('0051',   '元大中型100'),
    ('00733',  '富邦臺灣中小'),
    ('00850',  '元大MSCI台灣'),
    # ── 主動型 ──────────────────────────────────────────────────────────────
    ('00981A', '統一台股增長'),
    ('00992A', '中信優先順位'),
    ('00994A', '台新臺灣動能'),
    ('00987A', '野村優息'),
    ('00985A', '第一金優選'),
    ('00991A', '群益多因子'),
    ('00995A', '統一台灣動力'),
    # ── 高股息 / 配息 ────────────────────────────────────────────────────────
    ('0056',   '元大高股息'),
    ('00878',  '國泰永續高股息'),
    ('00919',  '群益台灣精選高息'),
    ('00900',  '富邦特選高股息30'),
    ('00929',  '復華台灣科技優息'),
    ('00934',  '中信成長高股息'),
    ('00940',  '元大台灣價值高息'),
    ('00713',  '元大台灣高息低波'),
    ('00915',  '凱基優選高股息30'),
    ('00918',  '大華優利高填息30'),
    ('00939',  '統一台灣高息動能'),
    ('00943',  '兆豐永續高息等權'),
    ('00948',  '街口臺灣高息ETF'),
    ('00960',  '復華臺灣科技中小'),
    # ── 科技 / 半導體 ────────────────────────────────────────────────────────
    ('0052',   '富邦科技'),
    ('0053',   '元大電子'),
    ('00891',  '中信關鍵半導體'),
    ('00892',  '富邦台灣半導體'),
    ('00881',  '國泰台灣5G+'),
    ('00830',  '國泰費城半導體'),
    ('00912',  '中信台灣智慧50'),
    # ── ESG / 永續 ───────────────────────────────────────────────────────────
    ('00692',  '富邦公司治理'),
    ('00757',  '統一MSCI台灣ESG'),
    ('00896',  '中信綠能及電動車'),
    # ── 槓桿 / 反向 ──────────────────────────────────────────────────────────
    ('00631L', '元大台灣50正2'),
    ('00632R', '元大台灣50反1'),
    # ── 海外 ────────────────────────────────────────────────────────────────
    ('00646',  '元大S&P500'),
    ('00864B', '中信美國公債20年'),
    ('00679B', '元大美債20年'),
    ('00772B', '中信高評級公司債'),
]

SECTOR_ZH = {
    'technology':          '科技業',
    'financialServices':   '金融保險業',
    'communication':       '通訊服務',
    'industrials':         '工業',
    'healthcare':          '醫療保健',
    'consumerCyclical':    '非必需消費',
    'consumerDefensive':   '必需消費',
    'energy':              '能源',
    'basicMaterials':      '原物料',
    'realestate':          '不動產',
    'utilities':           '公用事業',
    'cash':                '現金',
    'stock':               '股票',
    'bond':                '債券',
    'other':               '其他',
}

# ─── 台股總覽：概念股分組 ────────────────────────────────────────────────────
# 每個概念組列出主要成分股代號（第一個命中的概念組為該股的主分組）
_MKT_CONCEPT_GROUPS: dict[str, list[str]] = {
    # ── AI / 高效能運算（現代主題）───────────────────────────────────────────
    'AI 伺服器':      ['2382', '2356', '6669', '3231', '2376', '2377', '2357', '2324',
                       '4953', '3189', '6116', '3017', '6274', '3706', '2301'],
    'AI 晶片半導體':  ['2330', '2454', '2303', '3034', '2379', '6415', '3661', '2338',
                       '6770', '3231', '3443', '6533'],
    '先進封裝 CoWoS': ['3711', '2325', '2449', '8150', '6147', '2330', '2303', '5469',
                       '3014', '6274'],
    'HBM 記憶體':     ['4919', '3570', '3260', '2408'],
    'NVIDIA 供應鏈':  ['2330', '2454', '6669', '2382', '3231', '2376', '6116', '3189'],

    # ── 半導體製造與設計 ──────────────────────────────────────────────────────
    'IC 設計':        ['3034', '2379', '3661', '6488', '2337', '3536', '6515',
                       '3443', '6533', '5274', '4958', '3005', '3051'],
    'IC 製造封測':    ['2330', '2303', '2454', '3711', '2325', '6146', '2449', '5469',
                       '8150', '6147', '3014'],
    '矽智財 IP':      ['3661', '6533', '6488', '5274'],
    '第三代半導體':   ['3016', '8044', '6770', '5222'],

    # ── 電子零組件 ────────────────────────────────────────────────────────────
    'PCB 電路板':     ['2382', '2383', '3037', '2049', '3149', '5243', '6188', '3376',
                       '6274', '3044', '4974', '6269'],
    '被動元件':       ['2327', '2312', '2316', '2492', '3489', '2453'],
    '連接器':         ['2392', '3081', '6243', '2367', '5285'],
    '散熱模組':       ['3017', '6116', '3622', '1563', '2354', '3324'],
    '電源供應器':     ['6509', '3515', '1516', '6415', '3504', '1519'],
    'USB Type C':     ['2392', '3081', '5285', '6243', '2367', '3044'],

    # ── 網通與儲存 ────────────────────────────────────────────────────────────
    '網通設備':       ['3045', '2345', '3094', '3706', '6277', '3044', '2498', '4406',
                       '6456', '2454'],
    '資料儲存':       ['2387', '8299', '3005', '2393', '6245'],

    # ── 電腦周邊與消費電子 ────────────────────────────────────────────────────
    '主機板顯示卡':   ['2376', '2377', '2324', '2357'],
    '筆電代工':       ['2382', '2356', '2301', '2317', '4938'],
    '手機供應鏈':     ['2317', '4938', '2498', '3008', '6409', '6515', '3231', '2360'],
    '車用電子':       ['2308', '3665', '6284', '1590', '3703', '2395', '6239', '2049',
                       '6116', '3034', '4958'],
    'AMOLED':         ['3481', '5269', '6409', '2354'],

    # ── 面板與光電 ────────────────────────────────────────────────────────────
    '面板光電':       ['2408', '3481', '6409', '3008', '2489', '5269', '3019'],
    'LED 照明':       ['2641', '4974', '3698', '3016'],

    # ── 綠能與電動化 ──────────────────────────────────────────────────────────
    '離岸風電':       ['1605', '1590', '2637', '3682', '6510', '1584'],
    '太陽能':         ['3576', '3563', '6510', '3317', '6282', '5483'],
    '電動車儲能':     ['2308', '3665', '1590', '6121', '1513', '3518', '1597', '1598',
                       '6279', '1536', '3703'],
    '水資源':         ['1521', '1504', '2911', '5907', '1508', '9914'],

    # ── 5G / 通訊 ────────────────────────────────────────────────────────────
    '5G 通訊':        ['3045', '4904', '2412', '4406', '3044', '2498', '2468', '6706'],
    '衛星通訊':       ['3045', '2467', '6706', '3062'],

    # ── 工業與自動化 ──────────────────────────────────────────────────────────
    '機器人':         ['2308', '1590', '2395', '1525', '1522', '6415', '4958'],
    '工業 4.0':       ['2308', '2395', '6547', '3536', '3017', '1590', '1525'],
    '智慧城市':       ['3536', '3017', '6515', '3062', '5269', '6274', '3653'],

    # ── 航運與物流 ────────────────────────────────────────────────────────────
    '航運':           ['2603', '2615', '2609', '2610', '2634', '2608', '2618', '2617',
                       '2641', '5608'],

    # ── 金融 ──────────────────────────────────────────────────────────────────
    '金融':           ['2882', '2881', '2883', '2884', '2885', '2886', '2887', '2888',
                       '2891', '2892', '5880', '2890', '2880'],
    '證券期貨':       ['2883', '6008', '2885', '2890', '6005', '2887'],

    # ── 生技醫療 ──────────────────────────────────────────────────────────────
    '生技醫療':       ['4763', '4726', '1786', '4711', '4144', '6461', '4180', '6492',
                       '1789', '4743', '4147', '6547', '4154', '1729'],

    # ── 觀光休閒 ──────────────────────────────────────────────────────────────
    '運動休閒':       ['9914', '5904', '1795', '2707', '5706', '6290'],
    '觀光旅遊':       ['2701', '2712', '2704', '2706', '2710', '2717'],
}


# ─── 台股總覽：電子產業分組 ──────────────────────────────────────────────────
# TWSE 電子相關大類股名稱（用於 fallback 分組）
_ELEC_TWSE_INDUSTRIES: set[str] = {
    '半導體', '電腦週邊', '光電', '通信網路', '電零組', '電子通路', '資訊服務', '其它電子', '電子'
}

# 完整對應 wantgoo 電子產業 35 個子分類成分股（第一個命中的子類為主分組）
# 未列入的電子股會以 TWSE 大類別作為 fallback
_MKT_ELEC_GROUPS: dict[str, list[str]] = {
    # ── 半導體 ────────────────────────────────────────────────────────────────
    'IC設計服務':    ['4919', '3317', '2436', '6643', '2388', '3545', '6651', '4968',
                     '3006', '8227', '6679', '3228', '6693', '3034', '3150', '2458',
                     '3556', '6962', '7749', '3438', '6129', '3014', '2401', '2379',
                     '5272', '6103', '6563', '8102', '4961', '6756', '8040',
                     '2363', '3588', '6996', '6411', '6695', '6229', '6842', '5274',
                     '3122', '3288', '5302', '7712', '3227', '7770', '8081', '6202',
                     '6462', '5468', '3169', '5471', '6526', '3094', '5299', '6719',
                     '2454', '3035', '5236', '3527', '8024', '6531', '6799', '3141',
                     '8016', '6104', '3257', '6684', '6720', '4951', '8054', '3268',
                     '6732', '6138', '3530', '6568', '6233', '3661', '5351', '4923',
                     '6716', '3041', '7796', '3592', '6237', '8261', '6415', '6243',
                     '6423', '6435', '5269', '6921', '7707', '4952', '5487', '6485',
                     '3073', '2337', '2408', '6291', '6533', '3443', '2344', '4925',
                     '3529', '6708', '6494', '6927', '8299', '7872', '6907', '3259'],
    'IC生產製造':    ['6451', '6683', '4971', '3467', '3105', '6510', '2351', '3264',
                     '3260', '8086', '3265', '6187', '5347', '6271', '6548', '3374',
                     '5285', '3178', '5222', '2338', '3675', '4991', '6854', '3189',
                     '2434', '2441', '6257', '6532', '6208', '8091', '6525', '4967',
                     '5443', '3686', '8277', '8383', '2303', '6789', '3567', '3372',
                     '8162', '8271', '2330', '2369', '6147', '5344', '3711', '6488',
                     '7899', '2329', '6829', '6239', '7887', '2481', '6182', '7843',
                     '7730', '6573', '8088', '2302', '8131', '2451', '4749', '8028',
                     '3450', '7856', '2342', '7866', '3532', '7669', '2449', '5262',
                     '5483', '3135', '2340', '7880', '8150', '8110', '5425', '6786',
                     '7772', '4973', '5246', '3016', '3707', '7886', '6770', '6819',
                     '7815', '6830'],
    # ── PCB ───────────────────────────────────────────────────────────────────
    'PCB':          ['3715', '2313', '6213', '8291', '7419', '4958', '6153', '6407',
                     '2367', '6269', '6191', '8155', '2402', '6274', '6141', '8213',
                     '2355', '2316', '1815', '5340', '4927', '5469', '8039', '8074',
                     '5439', '3044', '6672', '6552', '2368', '6108', '3037', '6924',
                     '8438', '6835', '3645', '6210', '3321', '5464', '5475', '6194',
                     '4939', '5381', '5291', '3229', '6156', '5355', '3390', '2383',
                     '8046', '3354', '4989', '3276', '6920', '2429', '8358', '3631',
                     '3585'],
    # ── 電子零組件 ────────────────────────────────────────────────────────────
    '被動元件':      ['4760', '6155', '3090', '3026', '6127', '2472', '3117', '6224',
                     '6175', '6597', '6284', '2492', '6173', '2375', '8043', '6204',
                     '6449', '6432', '6862', '8042', '2428', '3236', '3624', '2478',
                     '2327', '5328', '6834', '8121', '3191', '3357', '5228', '6207'],
    '電子連接相關':  ['3550', '3710', '6715', '3023', '3689', '2440', '8103', '8147',
                     '6134', '3597', '5254', '6197', '6205', '3092', '3217', '2460',
                     '7861', '4943', '3511', '3646', '6220', '5488', '6185', '6913',
                     '3021', '6279', '6290', '6418', '6115', '2328', '3605', '6133',
                     '2392', '6158', '5457', '5271', '3432', '6126', '3533', '3526',
                     '3501', '3003', '3492', '3520', '3322', '2462', '3011', '6272',
                     '7893', '6833'],
    '機殼':          ['3013', '2474', '5426', '8210', '6235', '6117', '5465', '3325',
                     '3540', '2354', '9136'],
    '其他零組件':    ['2484', '8182', '8289', '1582', '2493', '7744', '6217', '6174',
                     '3338', '3653', '2476', '6831', '5223', '5243', '8071', '4912',
                     '3042', '6967', '2483', '3324', '3607', '2415', '6124', '3543',
                     '6266', '5460', '3548', '2467', '6805', '3679', '5356', '3484',
                     '4999', '1569', '6275', '6208', '2059', '3296', '3290', '4545',
                     '3388', '6591', '6156', '2421', '3294', '3376', '5309', '1336',
                     '3512', '3206', '4924', '7732', '7788', '4980', '3115', '6755',
                     '3310'],
    # ── 整機組裝 ──────────────────────────────────────────────────────────────
    '主機板':        ['2399', '7710', '2376', '2357', '3515', '6161', '2331', '2377',
                     '7711'],
    'PC/NB/平板':   ['3693', '3416', '2362', '2382', '2356', '4938', '2353', '3231',
                     '3706', '2324', '2352', '3005'],
    '組裝代工':      ['2312', '4938', '8183', '3231', '2317', '2352'],
    # ── 網通 ──────────────────────────────────────────────────────────────────
    '網通設備組件':  ['6820', '6818', '7717', '3221', '6821', '3062', '2314', '6588',
                     '6285', '6426', '2424', '3491', '2419', '4908', '3152', '3163',
                     '8045', '7812', '3694', '8011', '3138', '5388', '8059', '4905',
                     '3047', '7455', '3380', '2485', '6792', '4977', '3081', '6442',
                     '2332', '6980', '3596', '6674', '3209', '4906', '2444', '3466',
                     '3499', '3447', '3704', '6263', '2321', '3363', '4903', '6241',
                     '3419', '7805', '5353', '3684', '3306', '6216', '3473', '5348',
                     '8176', '8089', '2496', '3025', '3632', '2455', '3564', '3234',
                     '6470', '2345', '6142', '6546', '6530', '3672', '8034', '3558',
                     '6152', '3027', '6465', '6784', '4979', '8048'],
    '其他網通':      ['6190', '6486', '3669', '7689', '6245', '6218', '8097', '6136',
                     '6143', '6417', '4909', '6761', '6906', '4980', '6163'],
    # ── 電腦週邊 ──────────────────────────────────────────────────────────────
    '電腦週邊配件':  ['2305', '5289', '2402', '3693', '7455', '3071', '6150', '4915',
                     '2417', '3272', '3494', '3211', '2385', '2465', '7892', '2495',
                     '5474', '3287', '6805', '4987', '6228', '6161', '8410', '3483',
                     '5438', '3611', '6188', '5215', '2387', '5490', '2365', '2425',
                     '8163', '3323', '6121', '3349', '6230', '3625', '2380', '3017',
                     '3060', '3057'],
    '其他電腦週邊':  ['6277', '3701', '3213', '6128', '2405', '5258', '5450', '3046',
                     '1569', '9912', '2432', '3002', '3712'],
    # ── 手機相關 ──────────────────────────────────────────────────────────────
    '手機相關':      ['2498', '6283', '3311', '6170', '2357', '2439', '8101', '6109',
                     '8171', '3095'],
    # ── 光電/光學 ─────────────────────────────────────────────────────────────
    'LED':          ['2426', '3339', '6271', '6548', '3603', '2438', '6597', '2486',
                     '4972', '3346', '5450', '7753', '2393', '3031', '4956', '3066',
                     '2455', '5230', '8111', '6729', '6164', '3516', '5244', '6226',
                     '3714', '6168', '2340', '3437', '3591', '6559'],
    '光電設備':      ['6419', '6425', '5240', '3455', '3297', '3434', '4537', '8072',
                     '6234', '3535', '3413', '5484', '6664', '5443', '3356', '3583',
                     '3128', '6125', '3485', '2466', '3581', '5251', '3490', '6706',
                     '8064'],
    '光學元件或組裝':['6498', '4915', '3630', '3441', '3008', '4974', '8249', '6668',
                     '6517', '5230', '3362', '6789', '3504', '3019', '4976', '6560',
                     '6742', '6209', '3406', '6859', '6915', '3659', '6787', '7772',
                     '5248', '6858', '2256'],
    '其他光電':      ['3595', '3230', '3059', '3543', '3663', '6225', '5392', '3050',
                     '2374', '2491', '7402', '2349', '2323', '5267'],
    '面板業':        ['3678', '2489', '8240', '3615', '3545', '4960', '6698', '5245',
                     '6246', '8215', '4729', '3051', '8069', '8105', '3673', '3685',
                     '5432', '6176', '4933', '6278', '3168', '4942', '4935', '5371',
                     '3666', '7770', '3531', '6916', '7753', '6120', '6899', '5220',
                     '3633', '5234', '5315', '8016', '3622', '4995', '6577', '6222',
                     '3038', '8104', '8049', '6673', '3481', '6114', '3623', '5283',
                     '6167', '3149', '6116', '3523', '6434', '6456', '4749', '6775',
                     '2409', '6405', '6979', '3049'],
    # ── 綠能 ──────────────────────────────────────────────────────────────────
    '電池或電源':    ['7890', '6940', '1723', '6292', '8093', '1471', '4739', '8109',
                     '6276', '6203', '7742', '4721', '3609', '3308', '6259', '6509',
                     '3032', '3332', '7854', '2420', '4931', '2413', '6412', '2308',
                     '3078', '6781', '2431', '3207', '8038', '6282', '6558', '3015',
                     '5227', '2301', '7816', '2457', '8028', '6555', '3226', '3593',
                     '3058', '7758'],
    '太陽能':        ['4949', '6538', '6839', '1809', '1802', '4720', '1304', '3628',
                     '6806', '1711', '1611', '6987', '1343', '4707', '1717', '2434',
                     '6692', '3050', '5234', '3686', '6729', '6477', '1514', '4934',
                     '3691', '2481', '4582', '2406', '2349', '6873', '2342', '6443',
                     '3713', '7590', '5483', '6682', '6244', '3576'],
    # ── 特殊電子 ──────────────────────────────────────────────────────────────
    '工業電腦':      ['3577', '3416', '4916', '6166', '6416', '6441', '6928', '3088',
                     '6206', '8119', '6160', '8050', '7562', '6579', '3521', '2364',
                     '3652', '3022', '6536', '8114', '6922', '6825', '3594', '6570',
                     '3479', '6414', '2397', '8076', '8234', '6680', '2395', '6669',
                     '3005', '6599', '7875', '3097', '5386'],
    '安全監控':      ['6638', '5493', '5489', '2390', '6556'],
    '消費電子或電器':['3067', '6151', '3285', '2488', '3024', '3541', '5225', '2477',
                     '6275', '6201', '6743'],
    '其他電子或零件':['1785', '8021', '3305', '3465', '8431', '3665', '8047', '3628',
                     '3552', '3551', '7749', '6283', '3466', '4554', '3289', '2373',
                     '2433', '2461', '3219', '6863', '3617', '6409', '6192', '6512',
                     '8499', '3373', '3043', '7885', '2497', '6722', '6215', '6146',
                     '6642', '8201', '2482', '2423', '3587', '6840', '8085', '3518',
                     '6988', '8058', '3508', '3580', '7896'],
    # ── 服務/軟體 ─────────────────────────────────────────────────────────────
    '電信服務':      ['3045', '4904', '6561', '2412', '7901'],
    '軟體設計':      ['6865', '5203', '6593', '6738', '6428', '6910', '6882', '7724',
                     '6516', '8272', '7834', '8416', '7765', '6925', '5202', '6231',
                     '5211', '3555', '8298', '6565'],
    '遊戲軟體':      ['3629', '7584', '6428', '4994'],
    '系統整合':      ['6884', '3158', '6868', '6486', '8099', '5310', '6938', '7547',
                     '7785', '6874', '6690', '6148', '6112', '6140', '6816', '3664',
                     '5201', '2480', '6811', '5410', '7801', '6791', '7868', '6516',
                     '6590', '7767', '6123', '8416', '8284', '2471', '6214', '5403',
                     '2453', '6614', '3570', '3147', '3029', '2427', '6752', '4953',
                     '6751', '6221', '7819', '6697', '6689', '5210', '5212', '6240',
                     '2468', '7714', '6898'],
    '網站經營/電商': ['5321', '6870', '6473', '6984', '5278', '8472', '3130', '3687',
                     '5287', '6183', '7551', '6741', '6763', '5904', '8454', '8477',
                     '2949', '6878', '8044', '3085'],
    # ── 電子通路 ──────────────────────────────────────────────────────────────
    '電子設備買賣':  ['3131', '3093', '3680', '6261', '8072', '6223', '3413', '3055',
                     '3583', '3581'],
    '電子組件買賣':  ['8096', '8070', '3312', '3597', '3036', '3537', '8032', '3114',
                     '8455', '3528', '3048', '5452', '5434', '3232', '6189', '6113',
                     '6474', '3224', '2459', '6270', '8084', '3028', '3010', '5498',
                     '8068', '8112', '6227', '3033', '6265', '3360', '3444', '3702',
                     '3555', '6707'],
    '3C賣場':       ['8067', '3709', '6118', '6154', '2450', '2414', '2430', '6776',
                     '6281', '2347', '6908'],
    # ── 設備 ──────────────────────────────────────────────────────────────────
    '設備或廠務工程':['4949', '6640', '7703', '6727', '3563', '7810', '4744', '6895',
                     '6909', '6187', '6438', '3030', '6877', '6887', '4542', '7734',
                     '6613', '6735', '6812', '7728', '6196', '7751', '4537', '8092',
                     '6691', '7828', '7846', '6654', '5536', '6953', '6788', '5209',
                     '6664', '6667', '2359', '7813', '3055', '2464', '6750', '8383',
                     '3583', '6826', '7848', '2459', '3498', '7704', '7822', '1595',
                     '3485', '7631', '7853', '7556', '6903', '2360', '7769', '3018',
                     '1594', '3402', '7730', '6139', '6725', '2404', '6739', '6515',
                     '7825', '7795', '6823', '6658', '7880', '5297', '6937', '7870',
                     '6849', '5267', '7530', '6983', '7849'],
}


_CMONEY_CONCEPT_CACHE_PATH = os.path.join(BASE_DIR, '.cmoney_concept_cache.json')
_CMONEY_CONCEPT_CACHE: dict[str, list[str]] = {}

FEE_RATE = 0.001425
TAX_RATE = 0.003

# ─── 深色主題色盤 ─────────────────────────────────────────────────────────────
C_BG      = '#1e1e1e'
C_PANEL   = '#252526'
C_INPUT   = '#2d2d2d'
C_FG      = '#cccccc'
C_FG2     = '#888888'
C_ACCENT  = '#0078d4'
C_BORDER  = '#3e3e3e'
C_BUY_FG  = '#4ec94e'
C_SELL_FG = '#f07070'
C_BUY_BG  = '#1a2e1a'
C_SELL_BG = '#2e1a1a'

# ─── 中文字型 ────────────────────────────────────────────────────────────────
def _setup_font():
    candidates = ['Microsoft JhengHei', 'Microsoft YaHei', 'SimHei', 'Noto Sans CJK TC']
    available  = {f.name for f in fm.fontManager.ttflist}
    for c in candidates:
        if c in available:
            return c
    for c in candidates:
        for name in available:
            if c.lower().replace(' ', '') in name.lower().replace(' ', ''):
                return name
    return 'DejaVu Sans'

CHART_FONT  = _setup_font()
BODY_FONT   = ('Microsoft JhengHei', 10)
HDR_FONT    = ('Microsoft JhengHei', 13, 'bold')
INPUT_FONT  = ('Microsoft JhengHei', 12)        # Tab 1 輸入區較大字體

plt.rcParams['axes.unicode_minus'] = False
plt.rcParams['font.family']        = CHART_FONT

# ─── Excel 工具 ──────────────────────────────────────────────────────────────
def init_excel():
    if not os.path.exists(EXCEL_PATH):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(COLUMNS)
        col_widths = [12, 10, 12, 10, 6, 10, 10, 12, 12, 14]
        for i, cell in enumerate(ws[1], 1):
            cell.font      = Font(bold=True, color='FFFFFF')
            cell.fill      = PatternFill('solid', fgColor='1e5fa0')
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = col_widths[i - 1]
        wb.save(EXCEL_PATH)


def _migrate_add_category():
    """若舊版 Excel 缺少「分類」欄，自動插入"""
    if not os.path.exists(EXCEL_PATH):
        return
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    header = [cell.value for cell in ws[1]]
    if '分類' in header:
        return  # already up-to-date

    if '買賣' not in header:
        return

    insert_col = header.index('買賣') + 1   # 1-based column index
    max_col    = ws.max_column

    for r in range(1, ws.max_row + 1):
        for c in range(max_col, insert_col - 1, -1):
            ws.cell(row=r, column=c + 1).value = ws.cell(row=r, column=c).value
        ws.cell(row=r, column=insert_col).value = ('分類' if r == 1 else '未分類')

    wb.save(EXCEL_PATH)


def load_df() -> pd.DataFrame:
    init_excel()
    _migrate_add_category()
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME, dtype={'股票代號': str})
        if df.empty or '日期' not in df.columns:
            return pd.DataFrame(columns=COLUMNS)
        df['日期'] = pd.to_datetime(df['日期'])
        if '分類' not in df.columns:
            df.insert(df.columns.get_loc('買賣'), '分類', '未分類')
        return df.sort_values('日期').reset_index(drop=True)
    except Exception as e:
        print(f'[load_df] {e}')
        return pd.DataFrame(columns=COLUMNS)


def save_row(row: dict):
    init_excel()
    _migrate_add_category()
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    header = [cell.value for cell in ws[1]]
    values = []
    for col in header:
        v = row.get(col, '')
        if isinstance(v, datetime):
            v = v.date()
        values.append(v)
    ws.append(values)
    wb.save(EXCEL_PATH)


def rewrite_excel(df: pd.DataFrame):
    """以 DataFrame 完整重寫 Excel（用於刪除 / 修改後）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(COLUMNS)
    col_widths = [12, 10, 12, 10, 6, 10, 10, 12, 12, 14]
    for i, cell in enumerate(ws[1], 1):
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = PatternFill('solid', fgColor='1e5fa0')
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = col_widths[i - 1]
    for _, row in df.iterrows():
        values = []
        for col in COLUMNS:
            v = row.get(col, '')
            if hasattr(v, 'date') and callable(v.date):
                v = v.date()
            elif pd.isna(v):
                v = ''
            values.append(v)
        ws.append(values)
    wb.save(EXCEL_PATH)

# ─── 庫存計算 ────────────────────────────────────────────────────────────────
def _split_ratio(code: str, trade_date) -> float:
    """計算指定交易日的股票分割累積倍數（只計算生效日之前的分割）"""
    ratio = 1.0
    if hasattr(trade_date, 'date'):
        trade_date = trade_date.date()
    elif not isinstance(trade_date, _date):
        trade_date = pd.to_datetime(trade_date).date()
    today = _date.today()
    for split_date, r in STOCK_SPLITS.get(code, []):
        if split_date <= today and trade_date < split_date:
            ratio *= r
    return ratio


def calc_holdings(df: pd.DataFrame) -> dict:
    """Key = (股票代號, 分類)，同一股票不同分類分開計算"""
    holdings: dict = {}
    if df.empty:
        return holdings
    for _, r in df.iterrows():
        code = str(r['股票代號']).strip()
        cat  = str(r.get('分類', '未分類'))
        key  = (code, cat)
        qty, price, fee = float(r['數量(股)']), float(r['價格(元)']), float(r['手續費(元)'])
        # 若該交易發生在股票分割前，數量換算為分割後單位（成本不變，均價自動調整）
        eff_qty = qty * _split_ratio(code, r['日期'])
        if r['買賣'] == '買':
            if key not in holdings:
                holdings[key] = {'name': str(r['股票名稱']),
                                 'qty': 0.0, 'total_cost': 0.0, 'category': cat}
            holdings[key]['qty']        += eff_qty
            holdings[key]['total_cost'] += qty * price + fee   # 成本以原始金額計
        else:
            # 賣出：優先從同分類扣除；若找不到則退回同代碼任意一個持倉
            h = holdings.get(key)
            if h is None:
                for k, v in holdings.items():
                    if k[0] == code:
                        h = v
                        break
            if h and h['qty'] > 0:
                avg = h['total_cost'] / h['qty']
                h['total_cost'] -= avg * eff_qty
                h['qty']        -= eff_qty
                # 獲利賣出：剩餘庫存平均成本重置為獲利賣出價，讓未實現損益歸零
                if cat == '獲利賣出' and h['qty'] > 0.5:
                    h['total_cost'] = h['qty'] * price
    result = {}
    for key, h in holdings.items():
        if h['qty'] > 0.5:
            h['avg_cost'] = h['total_cost'] / h['qty']
            result[key] = h
    return result

# ─── 報價工具 ────────────────────────────────────────────────────────────────
def get_price(code: str):
    import logging as _log, io as _io, sys as _sys_gp
    code = str(code).strip()

    # 優先用 TWSE/TPEX 官方收盤價（rwd afterTrading，永遠回傳最新確認收盤日資料）
    try:
        pc = _ensure_twse_price_cache()
        if code in pc:
            return pc[code][0], code
    except Exception:
        pass

    # 備援：yfinance（美股、ETF、或 TWSE 快取未包含的股票）
    _yf_log = _log.getLogger('yfinance')
    _logged = False
    for s in ['.TW', '.TWO', '']:
        try:
            _prev_lvl = _yf_log.level
            _yf_log.setLevel(_log.CRITICAL)
            _old_stderr, _sys_gp.stderr = _sys_gp.stderr, _io.StringIO()
            try:
                hist = yf.Ticker(code + s).history(period='5d')
            finally:
                _sys_gp.stderr = _old_stderr
                _yf_log.setLevel(_prev_lvl)
            if not hist.empty:
                _raw = hist['Close'].dropna()
                if _raw.empty:
                    continue
                price = float(_raw.iloc[-1])
                import math as _math
                if _math.isnan(price) or _math.isinf(price) or price <= 0:
                    continue
                # 若 Yahoo Finance 尚未更新拆分後報價，手動修正
                today = _date.today()
                for split_date, ratio in STOCK_SPLITS.get(code, []):
                    if split_date <= today and price > ratio * 10:
                        price = round(price / ratio, 2)
                return price, code + s
        except Exception as _e:
            if not _logged:
                _net_log(f'get_price({code}{s}) FAIL: {type(_e).__name__}: {_e}')
                _logged = True
    return None, code


def get_price_on_date(code: str, date_str: str):
    code = str(code).strip()
    try:
        target = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return None, False, code
    end_str = (datetime.strptime(date_str, '%Y-%m-%d') + timedelta(days=7)).strftime('%Y-%m-%d')
    for s in ['.TW', '.TWO', '']:
        try:
            hist = yf.Ticker(code + s).history(
                start=date_str, end=end_str, auto_adjust=False)
            if hist.empty:
                continue
            close_col = 'Close' if 'Close' in hist.columns else hist.columns[3]
            return float(hist[close_col].iloc[0]), (hist.index[0].date() == target), code + s
        except Exception:
            pass
    return None, False, code


def get_history(code: str, start_date: str, auto_adjust: bool = False) -> 'pd.Series | None':
    """取得從 start_date 到今天的每日收盤價（pd.Series，index 為 DatetimeIndex）
    auto_adjust=True 時回傳還原後股價（含股票分割調整）。"""
    end_str = (datetime.today() + timedelta(days=1)).strftime('%Y-%m-%d')
    for s in ['.TW', '.TWO', '']:
        try:
            hist = yf.Ticker(code + s).history(
                start=start_date, end=end_str, auto_adjust=auto_adjust)
            if not hist.empty:
                close_col = 'Close' if 'Close' in hist.columns else hist.columns[3]
                return hist[close_col]
        except Exception:
            pass
    return None


def _has_cjk(s: str) -> bool:
    return any('\u4e00' <= c <= '\u9fff' for c in str(s))


# ── 證交所 / 櫃買中心名稱快取 ─────────────────────────────────────────────────
_TWSE_NAMES_CACHE: dict[str, str] = {}
_NAMES_CACHE_PATH = os.path.join(BASE_DIR, '.stock_names_cache.json')

# ── TWSE/TPEX 即時收盤價快取 {代號: (收盤價, 漲跌)} ─────────────────────────
_TWSE_PRICE_CACHE: dict[str, tuple[float, float]] = {}
_TWSE_PRICE_CACHE_TS: float = 0.0   # 上次成功抓取的 Unix timestamp
_TWSE_PRICE_CACHE_TTL = 15 * 60     # 15 分鐘後重新抓取

def _ensure_twse_price_cache() -> dict[str, tuple[float, float]]:
    """確保 TWSE/TPEX 收盤價快取是最新交易日的資料，否則重新抓取。"""
    global _TWSE_PRICE_CACHE, _TWSE_PRICE_CACHE_TS
    import time as _tm
    now = _tm.time()
    if _TWSE_PRICE_CACHE and (now - _TWSE_PRICE_CACHE_TS) < _TWSE_PRICE_CACHE_TTL:
        return _TWSE_PRICE_CACHE
    cache: dict[str, tuple[float, float]] = {}
    fetched_date = ''

    # TWSE 上市：優先用 rwd afterTrading（回傳最新確認收盤，不會返回盤中舊資料）
    # 欄位：[0]代號, [1]名稱, [2]成交量, [3]成交金額, [4]開盤, [5]最高, [6]最低, [7]收盤, [8]漲跌
    twse_ok = False
    try:
        data = _cffi_get_json(
            'https://www.twse.com.tw/rwd/zh/afterTrading/STOCK_DAY_ALL?response=json',
            timeout=15)
        if isinstance(data, dict) and data.get('stat') == 'OK':
            fetched_date = str(data.get('date', ''))
            for row in data.get('data', []):
                try:
                    code  = str(row[0]).strip()
                    price = float(str(row[7]).replace(',', ''))
                    chg   = float(str(row[8]).replace(',', '').replace('X', '0'))
                    if price > 0:
                        cache[code] = (price, chg)
                except (ValueError, TypeError, IndexError):
                    pass
            twse_ok = bool(cache)
    except Exception as _e:
        _net_log(f'_ensure_twse_price_cache rwd fail: {_e}')

    # 備援：openapi（有時資料較舊，但至少有資料）
    if not twse_ok:
        try:
            rows = _cffi_get_json(
                'https://openapi.twse.com.tw/v1/exchangeReport/STOCK_DAY_ALL', timeout=15)
            for row in rows:
                code = str(row.get('Code', '')).strip()
                try:
                    price = float(str(row.get('ClosingPrice', '')).replace(',', '') or '0')
                    chg   = float(str(row.get('Change', '0')).replace(',', '') or '0')
                    if price > 0:
                        cache[code] = (price, chg)
                except (ValueError, TypeError):
                    pass
        except Exception as _e:
            _net_log(f'_ensure_twse_price_cache openapi fail: {_e}')

    # TPEX 上櫃
    try:
        rows = _cffi_get_json(
            'https://www.tpex.org.tw/openapi/v1/tpex_mainboard_daily_close_quotes', timeout=15)
        for row in rows:
            code = str(row.get('SecuritiesCompanyCode', '')).strip()
            try:
                price = float(str(row.get('Close', '')).replace(',', '') or '0')
                chg_s = str(row.get('Change', '0')).replace(',', '').strip()
                chg   = float(chg_s) if chg_s else 0.0
                if price > 0 and code not in cache:
                    cache[code] = (price, chg)
            except (ValueError, TypeError):
                pass
    except Exception as _e:
        _net_log(f'_ensure_twse_price_cache TPEX fail: {_e}')

    if cache:
        _TWSE_PRICE_CACHE = cache
        _TWSE_PRICE_CACHE_TS = now
    return _TWSE_PRICE_CACHE

# ── 產業別快取 ────────────────────────────────────────────────────────────────
_TWSE_INDUSTRY_CACHE: dict[str, str] = {}
_INDUSTRY_CACHE_PATH = os.path.join(BASE_DIR, '.stock_industry_cache.json')

# TWSE 產業代碼 → 中文名稱（含 1 位與 2 位格式）
_TWSE_IND_CODE_MAP: dict[str, str] = {
    # 使用 wantgoo 網站相同的分類名稱
    '1': '水泥',      '01': '水泥',
    '2': '食品',      '02': '食品',
    '3': '塑膠',      '03': '塑膠',
    '4': '紡織',      '04': '紡織',
    '5': '電機',      '05': '電機',
    '6': '電器電纜',  '06': '電器電纜',
    '8': '玻璃',      '08': '玻璃',
    '9': '造紙',      '09': '造紙',
    '10': '鋼鐵',
    '11': '橡膠',
    '12': '汽車',
    '14': '營建',
    '15': '航運',
    '16': '觀光',
    '17': '金融',
    '18': '貿易百貨',
    '20': '其他',
    '21': '化學',
    '22': '生技醫療',
    '23': '油電燃氣',
    '24': '半導體',
    '25': '電腦週邊',
    '26': '光電',
    '27': '通信網路',
    '28': '電零組',
    '29': '電子通路',
    '30': '資訊服務',
    '31': '其它電子',
    '32': '綠能環保',
    '33': '數位雲端',
    '34': '運動休閒',
    '35': '居家生活',
    '36': '電子',
}


def _load_twse_stock_names() -> dict[str, str]:
    """從證交所取得 {代號: 中文名稱} 對照表。
    優先從磁碟快取讀取；成功抓取後寫回磁碟；確保名稱不重置為 {}。
    來源順序: STOCK_DAY_ALL → t51sb01 → 磁碟快取
    """
    global _TWSE_NAMES_CACHE
    if _TWSE_NAMES_CACHE:
        return _TWSE_NAMES_CACHE

    names: dict[str, str] = {}
    twse_hdr = {
        'If-Modified-Since': 'Mon, 26 Jul 1997 05:00:00 GMT',
        'Cache-Control': 'no-cache',
        'Pragma': 'no-cache',
    }

    # 方法1: STOCK_DAY_ALL（當日全部上市股票，含ETF）────────────────────────
    try:
        data = _cffi_get_json(
            'https://openapi.twse.com.tw/v1/exchangeReport/STOCK_DAY_ALL',
            headers=twse_hdr, timeout=15)
        for row in data:
            c = str(row.get('Code', '')).strip()
            n = str(row.get('Name', '')).strip()
            if c and n:
                names[c] = n
        if names:
            print(f'[TWSE names] STOCK_DAY_ALL {len(names)} 筆')
    except Exception as e:
        print(f'[TWSE names] STOCK_DAY_ALL 失敗: {type(e).__name__}')

    # 方法2: t51sb01 有價證券清單（備用）────────────────────────────────────
    if not names:
        try:
            data = _cffi_get_json(
                'https://openapi.twse.com.tw/v1/openData/t51sb01',
                headers=twse_hdr, timeout=15)
            for row in data:
                c = str(row.get('有價證券代號', '')).strip()
                n = str(row.get('有價證券名稱', '')).strip()
                if c and n:
                    names[c] = n
            if names:
                print(f'[TWSE names] t51sb01 {len(names)} 筆')
        except Exception as e:
            print(f'[TWSE names] t51sb01 失敗: {type(e).__name__}')

    # 方法2b: 補充 TPEx OTC 股票名稱（TWSE ISIN 查詢頁，Big5 HTML）──────────
    try:
        import re as _re
        _isin_url = 'https://isin.twse.com.tw/isin/C_public.jsp?strMode=4'
        if _CFFI_OK:
            _r = _cffi_req.get(_isin_url, timeout=15, verify=False,
                               headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/122.0.0.0'})
            _html = _r.content.decode('big5', errors='replace')
        else:
            import urllib.request, ssl as _ssl
            _ctx = _ssl.create_default_context()
            _ctx.check_hostname = False
            _ctx.verify_mode = _ssl.CERT_NONE
            with urllib.request.urlopen(_isin_url, context=_ctx, timeout=15) as _resp:
                _html = _resp.read().decode('big5', errors='replace')
        # 每筆格式: <td>6223　台灣聯合光纖</td>  (U+3000 全形空格分隔代號和名稱)
        _tpex_added = 0
        for _code, _name in _re.findall(r'<td>(\d{4,6})\u3000([^<\u3000\r\n]+)', _html):
            _code = _code.strip(); _name = _name.strip()
            if _code and _name and _code not in names:
                names[_code] = _name
                _tpex_added += 1
        if _tpex_added:
            print(f'[TPEX names] ISIN 補充 {_tpex_added} 筆 OTC 名稱')
    except Exception as _e:
        print(f'[TPEX names] ISIN 失敗: {type(_e).__name__}: {_e}')

    # 方法2c: TPEx opendata 上櫃每日收盤（含中文名稱，補 ISIN 遺漏的上櫃股）──
    try:
        _tpex_data = _cffi_get_json(
            'https://www.tpex.org.tw/openapi/v1/tpex_mainboard_daily_close_quotes',
            timeout=10)
        _tpex2_added = 0
        for _row in _tpex_data:
            _c = str(_row.get('SecuritiesCompanyCode', '')).strip()
            _n = str(_row.get('CompanyName', '')).strip()
            if _c and _n and _c not in names:
                names[_c] = _n
                _tpex2_added += 1
        if _tpex2_added:
            print(f'[TPEX names] opendata 補充 {_tpex2_added} 筆上櫃名稱')
    except Exception as _e:
        print(f'[TPEX names] opendata 失敗: {type(_e).__name__}')

    # 成功取得 → 寫磁碟快取
    if names:
        try:
            with open(_NAMES_CACHE_PATH, 'w', encoding='utf-8') as f:
                _json.dump(names, f, ensure_ascii=False)
        except Exception:
            pass
    else:
        # 方法3: 讀磁碟快取（上次成功的結果）───────────────────────────────
        try:
            with open(_NAMES_CACHE_PATH, encoding='utf-8') as f:
                names = _json.load(f)
            print(f'[TWSE names] 磁碟快取 {len(names)} 筆')
        except Exception:
            print('[TWSE names] 無法取得名稱資料（網路不通且無快取）')

    _TWSE_NAMES_CACHE = names
    return names


def _fetch_twse_industry_map() -> dict[str, str]:
    """回傳 {股票代號: 產業別} 對照表（TWSE 上市公司基本資料）。
    有磁碟快取，網路失敗時讀快取。
    """
    global _TWSE_INDUSTRY_CACHE
    if _TWSE_INDUSTRY_CACHE:
        return _TWSE_INDUSTRY_CACHE

    result: dict[str, str] = {}
    try:
        data = _cffi_get_json(
            'https://openapi.twse.com.tw/v1/openData/t187ap03_L',
            timeout=15)
        if data:
            first = data[0]
            # 動態找欄位名稱（防 API 改版）
            code_key = next((k for k in ['公司代號', '代號', 'Code'] if k in first), None)
            ind_key  = next((k for k in ['產業別', '行業別', 'Industry'] if k in first), None)
            if code_key and ind_key:
                for row in data:
                    c = str(row.get(code_key, '')).strip()
                    i = str(row.get(ind_key,  '')).strip()
                    if c and i:
                        # 若 API 回傳數字代碼，轉換成中文名稱
                        i = _TWSE_IND_CODE_MAP.get(i, i)
                        result[c] = i
                print(f'[TWSE industry] t187ap03_L {len(result)} 筆')
    except Exception as e:
        print(f'[TWSE industry] 抓取失敗: {type(e).__name__}')

    if result:
        _TWSE_INDUSTRY_CACHE = result
        try:
            with open(_INDUSTRY_CACHE_PATH, 'w', encoding='utf-8') as _f:
                _json.dump(result, _f, ensure_ascii=False)
        except Exception:
            pass
    else:
        try:
            with open(_INDUSTRY_CACHE_PATH, encoding='utf-8') as _f:
                cached = _json.load(_f)
            # 讀取快取後也補做代碼轉換（修正舊版快取）
            result = {c: _TWSE_IND_CODE_MAP.get(i, i) for c, i in cached.items()}
            _TWSE_INDUSTRY_CACHE = result
            print(f'[TWSE industry] 磁碟快取 {len(result)} 筆')
        except Exception:
            pass

    return result


_TPEX_INDUSTRY_CACHE: dict[str, str] = {}

def _fetch_tpex_industry_map() -> dict[str, str]:
    """回傳 {上櫃股票代號: 產業別} 對照表（TPEx 上櫃公司基本資料）。"""
    global _TPEX_INDUSTRY_CACHE
    if _TPEX_INDUSTRY_CACHE:
        return _TPEX_INDUSTRY_CACHE
    result: dict[str, str] = {}
    try:
        data = _cffi_get_json(
            'https://www.tpex.org.tw/openapi/v1/tpex_mainboard_listed_companies',
            timeout=12)
        for row in data:
            code = str(row.get('SecuritiesCompanyCode', '')).strip()
            ind  = str(row.get('IndustryGroup', row.get('IndustryGroupCode', ''))).strip()
            if code and ind:
                result[code] = ind
        if result:
            print(f'[TPEX industry] {len(result)} 筆')
    except Exception as e:
        print(f'[TPEX industry] 失敗: {type(e).__name__}')
    _TPEX_INDUSTRY_CACHE = result
    return result


def _fetch_cmoney_concept_map() -> dict[str, list[str]]:
    """CMoney 概念股 {概念名稱: [股票代號]}，7天磁碟快取，失敗時 fallback 內建資料。"""
    global _CMONEY_CONCEPT_CACHE
    if _CMONEY_CONCEPT_CACHE:
        return _CMONEY_CONCEPT_CACHE
    import time as _time, re as _re
    # 讀磁碟快取
    try:
        with open(_CMONEY_CONCEPT_CACHE_PATH, 'r', encoding='utf-8') as _f:
            _cached = _json.load(_f)
        if _cached.get('_ts', 0) > _time.time() - 7 * 86400:
            _CMONEY_CONCEPT_CACHE = _cached.get('data', {})
            print(f'[CMoney concept] 磁碟快取 {len(_CMONEY_CONCEPT_CACHE)} 筆')
            return _CMONEY_CONCEPT_CACHE
    except Exception:
        pass
    # 從 CMoney 抓取
    result: dict[str, list[str]] = {}
    try:
        html = _cffi_get_text('https://www.cmoney.tw/forum/concept', timeout=20)
        pairs = _re.findall(r'href="(/forum/category/C\d+)"[^>]*>\s*([^<]{2,30})\s*<', html)
        seen: dict[str, str] = {}
        for path, name in pairs:
            name = name.strip()
            if name and path not in seen:
                seen[path] = name
        for path, name in list(seen.items())[:60]:
            try:
                cat_html = _cffi_get_text(f'https://www.cmoney.tw{path}', timeout=12)
                codes = _re.findall(r'\b([1-9]\d{3})\b', cat_html)
                codes = list(dict.fromkeys(c for c in codes if 1000 <= int(c) <= 9999))[:40]
                if codes:
                    result[name] = codes
            except Exception:
                pass
        print(f'[CMoney concept] 抓取 {len(result)} 個概念股')
    except Exception as e:
        print(f'[CMoney concept] 失敗: {type(e).__name__}')
    if result:
        _CMONEY_CONCEPT_CACHE = result
        try:
            with open(_CMONEY_CONCEPT_CACHE_PATH, 'w', encoding='utf-8') as _f:
                _json.dump({'_ts': _time.time(), 'data': result}, _f, ensure_ascii=False)
        except Exception:
            pass
    else:
        _CMONEY_CONCEPT_CACHE = dict(_MKT_CONCEPT_GROUPS)
    return _CMONEY_CONCEPT_CACHE


# ── ETF 歷史 PCF 快取 ──────────────────────────────────────────────────────────
_ETF_PCF_HIST_CACHE: dict[str, dict] = {}   # code → {date_str: [holdings]}

# ── ETF 成分股歷史快照 ──────────────────────────────────────────────────────────
_ETF_HOLDINGS_HIST_CACHE: dict[str, dict] = {}   # code → {YYYYMMDD: [{code,weight}]}

# ── 主動型 ETF 清單 ────────────────────────────────────────────────────────────
ACTIVE_ETFS = [
    ('00981A', '統一增長'), ('00992A', '群益科技'), ('00990A', '元大AI'),
    ('00400A', '國泰動能'), ('00993A', '安聯台灣'),
]


def _squarify_layout(items, x0, y0, w, h):
    """Squarified treemap layout. items: [(value, tag)] sorted desc. Returns [(x1,y1,x2,y2,tag)]."""
    if not items or w <= 0 or h <= 0:
        return []
    total = sum(v for v, _ in items)
    if total == 0:
        return []
    area = w * h
    norm = [(v / total * area, t) for v, t in items]

    def _worst(vals, width):
        if not vals or width == 0:
            return float('inf')
        s = sum(vals)
        return max(width * width * max(vals) / (s * s),
                   s * s / (width * width * min(vals)))

    def _row_rects(rv, rt, x, y, dx, dy):
        s = sum(rv)
        out = []
        if dx >= dy:
            rw = s / dy if dy > 0 else 0
            cy = y
            for v, t in zip(rv, rt):
                rh = v / s * dy if s > 0 else 0
                out.append((x, cy, x + rw, cy + rh, t))
                cy += rh
            return out, x + rw, y, dx - rw, dy
        else:
            rh = s / dx if dx > 0 else 0
            cx = x
            for v, t in zip(rv, rt):
                rw = v / s * dx if s > 0 else 0
                out.append((cx, y, cx + rw, y + rh, t))
                cx += rw
            return out, x, y + rh, dx, dy - rh

    result = []

    def _rec(items, x, y, dx, dy):
        if not items or dx <= 0 or dy <= 0:
            return
        width = min(dx, dy)
        rv, rt = [], []
        for i, (v, t) in enumerate(items):
            cand = rv + [v]
            if not rv or _worst(cand, width) <= _worst(rv, width):
                rv.append(v)
                rt.append(t)
            else:
                rects, nx, ny, ndx, ndy = _row_rects(rv, rt, x, y, dx, dy)
                result.extend(rects)
                _rec(items[i:], nx, ny, ndx, ndy)
                return
        rects, _, _, _, _ = _row_rects(rv, rt, x, y, dx, dy)
        result.extend(rects)

    _rec(norm, x0, y0, w, h)
    return result


def _fetch_moneydj_etf_snapshot(code: str) -> list[dict]:
    """從 MoneyDJ Basic0007B 取得全部成分股，回傳 [{code, weight}, ...]。"""
    import re as _re
    _hdrs = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/122.0.0.0',
        'Referer': 'https://www.moneydj.com/',
    }
    _pat_full = _re.compile(
        r"etfid=(\d{4,6})\.TWO?[^'\"]*['\"][^>]*>[^<]*</a>"
        r"</td>\s*<td[^>]*>([\d.]+)</td>"
        r"\s*<td[^>]*>([\d,]+)</td>", _re.I)
    _pat_lite = _re.compile(
        r"etfid=(\d{4,6})\.TWO?[^'\"]*['\"][^>]*>[^<]*</a>"
        r"</td>\s*<td[^>]*>([\d.]+)</td>", _re.I)
    for _sfx in ['.TW', '.TWO']:
        _url = (f'https://www.moneydj.com/ETF/X/Basic/Basic0007B.xdjhtm'
                f'?etfid={code}{_sfx}')
        try:
            if _CFFI_OK:
                _r = _cffi_req.get(_url, headers=_hdrs, timeout=10, verify=False,
                                   impersonate='chrome110')
                _html = _r.content.decode('latin-1')
            else:
                import urllib.request as _ureq, ssl as _ssl
                _ctx = _ssl.create_default_context()
                _ctx.check_hostname = False
                _ctx.verify_mode = _ssl.CERT_NONE
                with _ureq.urlopen(_ureq.Request(_url, headers=_hdrs),
                                   context=_ctx, timeout=10) as _resp:
                    _html = _resp.read().decode('latin-1')
            _raw_full = _pat_full.findall(_html)
            if _raw_full:
                _hits = [(c, float(w), int(s.replace(',', '')))
                         for c, w, s in _raw_full if 0 < float(w) <= 100]
                if _hits:
                    return [{'code': c, 'weight': w, 'shares': s} for c, w, s in _hits]
            _hits = [(c, float(w)) for c, w in _pat_lite.findall(_html)
                     if 0 < float(w) <= 100]
            if _hits:
                return [{'code': c, 'weight': w} for c, w in _hits]
        except Exception:
            continue
    return []


def _load_etf_holdings_history(code: str) -> dict[str, list]:
    """讀取歷史快照，回傳 {YYYYMMDD: [{code, weight}, ...]}。"""
    global _ETF_HOLDINGS_HIST_CACHE
    if code in _ETF_HOLDINGS_HIST_CACHE:
        return _ETF_HOLDINGS_HIST_CACHE[code]
    _path = os.path.join(BASE_DIR, f'.etf_holdings_hist_{code}.json')
    if os.path.exists(_path):
        try:
            with open(_path, 'r', encoding='utf-8') as _f:
                _data = _json.load(_f)
            _ETF_HOLDINGS_HIST_CACHE[code] = _data
            return _data
        except Exception:
            pass
    _ETF_HOLDINGS_HIST_CACHE[code] = {}
    return {}


def _save_etf_holdings_snapshot(code: str, holdings: list[dict]) -> str:
    """儲存一份帶日期的快照到磁碟；回傳日期字串 YYYYMMDD。"""
    from datetime import date as _ddate
    _ds = _ddate.today().strftime('%Y%m%d')
    _hist = _load_etf_holdings_history(code)
    _hist[_ds] = holdings
    _ETF_HOLDINGS_HIST_CACHE[code] = _hist
    _path = os.path.join(BASE_DIR, f'.etf_holdings_hist_{code}.json')
    try:
        with open(_path, 'w', encoding='utf-8') as _f:
            _json.dump(_hist, _f, ensure_ascii=False, separators=(',', ':'))
    except Exception:
        pass
    return _ds


def _diff_etf_holdings(old: list[dict], new: list[dict]) -> dict:
    """比較兩份快照，回傳新進、退出、比例變化三類清單。"""
    _old = {h['code']: h for h in old}
    _new = {h['code']: h for h in new}
    def _w(h): return h.get('weight', 0.0)
    def _s(h): return h.get('shares', 0)
    _entered = sorted(
        [{'code': c, 'weight': _w(_new[c]), 'shares': _s(_new[c])}
         for c in set(_new) - set(_old)],
        key=lambda x: -x['weight'])
    _exited = sorted(
        [{'code': c, 'weight': _w(_old[c]), 'shares': _s(_old[c])}
         for c in set(_old) - set(_new)],
        key=lambda x: -x['weight'])
    _changed = []
    for c in set(_old) & set(_new):
        d  = round(_w(_new[c]) - _w(_old[c]), 2)
        ds = _s(_new[c]) - _s(_old[c])
        if abs(d) >= 0.01 or abs(ds) >= 1:
            _changed.append({'code': c, 'old': _w(_old[c]), 'new': _w(_new[c]),
                             'delta': d, 'delta_shares': ds})
    _changed.sort(key=lambda x: -(abs(x['delta_shares']) if abs(x['delta']) < 0.01 else abs(x['delta'])))
    return {'entered': _entered, 'exited': _exited, 'changed': _changed}


def _fetch_etf_pcf_history(code: str, months: int = 6,
                            progress_cb=None) -> dict[str, list[dict]]:
    """抓取 ETF 近 N 個月的每週 PCF（申購買回清單）歷史資料。

    以每週一個交易日（週五或往前推）為採樣點，約 26 筆。
    採樣點不在快取中才呼叫 TWSE API，之後永久快取（歷史日不變）。
    快取檔：<BASE_DIR>/.etf_pcf_history_<code>.json
    progress_cb(done, total) 於每次 API 呼叫後觸發。
    Returns {YYYYMMDD: [{'code','name','weight'}, ...]}
    """
    from datetime import datetime as _dt, timedelta as _td
    import time as _t

    cache_path = os.path.join(BASE_DIR, f'.etf_pcf_history_{code}.json')

    # 讀記憶體快取 → 磁碟快取
    existing: dict[str, list] = {}
    if code in _ETF_PCF_HIST_CACHE:
        existing = _ETF_PCF_HIST_CACHE[code]
    elif os.path.exists(cache_path):
        try:
            with open(cache_path, 'r', encoding='utf-8') as _f:
                existing = _json.load(_f).get('data', {})
        except Exception:
            existing = {}

    # ── 產生採樣日期 ──────────────────────────────────────────────────────
    # 策略：近 30 天每天採樣（確保 API 可查），更早的每週採樣（減少 API 呼叫）
    today = _dt.now().replace(hour=0, minute=0, second=0, microsecond=0)
    cutoff = today - _td(days=months * 31)
    today_str = today.strftime('%Y%m%d')
    recent_cutoff = today - _td(days=30)

    sample_dates: set[str] = set()
    cur = today - _td(days=1)
    while cur >= cutoff:
        if cur.weekday() < 5:   # 週一~週五
            ds = cur.strftime('%Y%m%d')
            if cur >= recent_cutoff:
                sample_dates.add(ds)       # 近 30 天：每天
            else:
                if cur.weekday() == 4:     # 更早：只保留週五
                    sample_dates.add(ds)
        cur -= _td(days=1)

    # 剔除快取中已有「非空」結果的日期；空結果不快取（下次重試）
    cached_ok = {d for d, v in existing.items() if v}
    missing = sorted(sample_dates - cached_ok)

    twse_hdr = {
        'Accept':           'application/json, text/javascript, */*; q=0.01',
        'Accept-Language':  'zh-TW,zh;q=0.9',
        'X-Requested-With': 'XMLHttpRequest',
        'Referer': f'https://www.twse.com.tw/zh/ETF/fund/{code}',
    }

    total = len(missing)
    done = 0

    def _parse_holdings(data: dict) -> list:
        fields  = data.get('fields', [])
        records = data.get('data',   [])
        out = []
        for rec in records:
            row    = dict(zip(fields, rec))
            s_code = str(next(
                (row[k] for k in row if any(t in k for t in ('代號', '代碼'))), ''
            )).strip()
            s_name = str(next(
                (row[k] for k in row if '名稱' in k), ''
            )).strip()
            w_raw  = next(
                (row[k] for k in row if any(t in k for t in ('比例', 'percent', 'Percent'))), '0'
            )
            try:
                weight = float(str(w_raw).replace('%', '').replace(',', ''))
            except (ValueError, TypeError):
                weight = 0.0
            if s_code:
                out.append({'code': s_code, 'name': s_name, 'weight': weight})
        return out

    for target_date in missing:
        # 對每個目標日期：往前找 5 天涵蓋假日，先試 DAILY 再試 MONTHLY
        _base = _dt.strptime(target_date, '%Y%m%d')
        fetched = False
        for q_type in ('DAILY', 'MONTHLY'):
            if fetched:
                break
            for _back in range(5):
                _try_date = (_base - _td(days=_back)).strftime('%Y%m%d')
                try:
                    data = _cffi_get_json(
                        'https://www.twse.com.tw/fund/etfFundHoldingData',
                        params={'fund': code, 'type': q_type, 'date': _try_date},
                        headers=twse_hdr, timeout=8)
                    if not isinstance(data, dict) or data.get('stat') != 'OK':
                        continue
                    holdings = _parse_holdings(data)
                    if holdings:
                        existing[target_date] = holdings
                        fetched = True
                        break
                except Exception:
                    pass
                _t.sleep(0.05)

        # 抓不到不快取（保留 None，下次重試）——不存 [] 避免永久遮蔽
        done += 1
        if progress_cb:
            progress_cb(done, total)
        _t.sleep(0.2)

    # 寫回磁碟快取（只寫有資料的日期）
    save_data = {d: v for d, v in existing.items() if v}
    _ETF_PCF_HIST_CACHE[code] = save_data
    if missing:
        try:
            with open(cache_path, 'w', encoding='utf-8') as _f:
                _json.dump({'data': save_data}, _f,
                           ensure_ascii=False, separators=(',', ':'))
        except Exception:
            pass

    return save_data


def _fetch_twse_etf_holdings(etf_code: str) -> tuple[list[dict], str]:
    """抓 ETF 完整成分股＋持股比例。
    Returns (holdings, debug_msg)
    方法1: 證交所 etfFundHoldingData（DAILY/MONTHLY）
    方法2: Yahoo Finance topHoldings（正確權重，最多10檔）
    """
    code  = str(etf_code).strip()
    debug = []

    # ── 方法 1: 證交所 etfFundHoldingData（DAILY/MONTHLY，近10天）──────────
    twse_hdr = {
        'Accept':           'application/json, text/javascript, */*; q=0.01',
        'Accept-Language':  'zh-TW,zh;q=0.9',
        'X-Requested-With': 'XMLHttpRequest',
        'Referer': f'https://www.twse.com.tw/zh/ETF/fund/{code}',
    }
    _twse_api_dead = False  # 偵測 API 已改版回 HTML，避免重複嘗試
    for q_type in ('DAILY', 'MONTHLY'):
        if _twse_api_dead:
            break
        for days_back in range(1, 15):  # 往前找14天，涵蓋連假期間
            date = (datetime.now() - timedelta(days=days_back)).strftime('%Y%m%d')
            try:
                data = _cffi_get_json(
                    'https://www.twse.com.tw/fund/etfFundHoldingData',
                    params={'fund': code, 'type': q_type, 'date': date},
                    headers=twse_hdr, timeout=5)
                if not isinstance(data, dict):
                    # API 回傳 HTML 或其他非 JSON：整個 TWSE 方法已失效，直接放棄
                    _twse_api_dead = True
                    debug.append('TWSE API 已改版（非JSON）')
                    break
                stat = data.get('stat', '?')
                if stat != 'OK':
                    continue
                fields  = data.get('fields', [])
                records = data.get('data',   [])
                if not records:
                    continue
                holdings = []
                for rec in records:
                    row    = dict(zip(fields, rec))
                    s_code = str(next(
                        (row[k] for k in row if any(t in k for t in ('代號', '代碼'))), ''
                    )).strip()
                    s_name = str(next(
                        (row[k] for k in row if '名稱' in k), ''
                    )).strip()
                    w_raw  = next(
                        (row[k] for k in row if any(t in k for t in ('比例', 'percent', 'Percent'))), '0'
                    )
                    try:
                        weight = float(str(w_raw).replace('%', '').replace(',', ''))
                    except (ValueError, TypeError):
                        weight = 0.0
                    if s_code:
                        holdings.append({'code': s_code, 'name': s_name,
                                         'weight': weight, 'sym': s_code + '.TW'})
                if holdings:
                    debug.append(f'TWSE OK: {len(holdings)} 檔')
                    return holdings, ' | '.join(debug)
            except _json.JSONDecodeError:
                # 第一次 JSONDecodeError 就代表 API 回傳 HTML，不必繼續嘗試
                _twse_api_dead = True
                debug.append('TWSE API 已改版（JSONDecodeError）')
                break
            except Exception as e:
                debug.append(f'TWSE({q_type},{date}): {type(e).__name__}')
                continue

    # ── 方法 2: MoneyDJ 完整成分股（優先，HTML 解析）────────────────────────
    import re as _re
    _mj_holdings = []
    try:
        _mj_hdrs = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/122.0.0.0',
            'Accept': 'text/html,application/xhtml+xml',
            'Accept-Language': 'zh-TW,zh;q=0.9',
            'Referer': 'https://www.moneydj.com/',
        }
        # Try .TW then .TWO
        for _sfx in ['.TW', '.TWO']:
            _mj_url = (f'https://www.moneydj.com/ETF/X/Basic/Basic0007B.xdjhtm'
                       f'?etfid={code}{_sfx}')
            try:
                if _CFFI_OK:
                    _r = _cffi_req.get(_mj_url, headers=_mj_hdrs, timeout=8, verify=False)
                    _html = _r.content.decode('latin-1')
                else:
                    import urllib.request as _ureq, ssl as _ssl
                    _ctx = _ssl.create_default_context()
                    _ctx.check_hostname = False; _ctx.verify_mode = _ssl.CERT_NONE
                    _req2 = _ureq.Request(_mj_url, headers=_mj_hdrs)
                    with _ureq.urlopen(_req2, context=_ctx, timeout=8) as _resp:
                        _html = _resp.read().decode('latin-1')

                _pat = _re.compile(
                    r"etfid=(\d{4,6})\.TWO?[^'\"]*['\"][^>]*>[^<]*</a>"
                    r"</td>\s*<td[^>]*>([\d.]+)</td>",
                    _re.I)
                _seen: set = set()
                _tmp = []
                for _m in _pat.finditer(_html):
                    _sc = _m.group(1).strip()
                    if _sc in _seen:
                        continue
                    try:
                        _w = float(_m.group(2))
                    except ValueError:
                        _w = 0.0
                    if _w <= 0 or _w > 100:
                        continue
                    _tmp.append({'code': _sc,
                                 'name': _TWSE_NAMES_CACHE.get(_sc, _sc),
                                 'weight': _w,
                                 'sym': _sc + '.TW'})
                    _seen.add(_sc)
                if len(_tmp) > len(_mj_holdings):
                    _mj_holdings = _tmp
                if _mj_holdings:
                    break   # .TW worked, no need to try .TWO
            except Exception:
                continue

        if _mj_holdings:
            debug.append(f'MoneyDJ: {len(_mj_holdings)} 檔')
        else:
            debug.append('MoneyDJ: 0 檔')
    except Exception as _e:
        debug.append(f'MoneyDJ err: {type(_e).__name__}')

    # ── 方法 3: Yahoo Finance topHoldings（備用，MoneyDJ 已有結果則跳過）──────
    if _mj_holdings:
        return _mj_holdings, ' | '.join(debug)

    _yf_holdings = []
    try:
        from yfinance.data import YfData as _YfData
        _yfdata = _YfData(session=None)
        for suffix in ['.TW', '.TWO']:
            try:
                r = _yfdata.cache_get(
                    url='https://query2.finance.yahoo.com/v10/finance/quoteSummary/' + code + suffix,
                    params={'modules': 'topHoldings'},
                    timeout=8)
                payload = _json.loads(r.content)
                result  = payload.get('quoteSummary', {}).get('result', [{}])
                if not result:
                    continue
                raw_h = result[0].get('topHoldings', {}).get('holdings', [])
                if not raw_h:
                    continue
                _tmp_yf = []
                for h in raw_h:
                    sym   = str(h.get('symbol', ''))
                    clean = sym.replace('.TWO', '').replace('.TW', '').strip()
                    name  = str(h.get('holdingName', clean))
                    pct   = h.get('holdingPercent', {})
                    w     = (pct.get('raw', 0) if isinstance(pct, dict) else float(pct or 0)) * 100
                    full  = sym if (sym.endswith('.TW') or sym.endswith('.TWO')) \
                            else clean + '.TW'
                    _tmp_yf.append({'code': clean, 'sym': full, 'name': name, 'weight': w})
                if len(_tmp_yf) > len(_yf_holdings):
                    _yf_holdings = _tmp_yf
                if _yf_holdings:
                    break
            except Exception:
                pass
        if _yf_holdings:
            debug.append(f'Yahoo topHoldings: {len(_yf_holdings)} 檔')
    except Exception as e:
        debug.append(f'Yahoo err: {type(e).__name__}')

    # 優先採用 MoneyDJ（筆數較多且完整），Yahoo 作為備用
    if _mj_holdings:
        return _mj_holdings, ' | '.join(debug)
    if _yf_holdings:
        return _yf_holdings, ' | '.join(debug)

    return [], ' | '.join(debug)


def fetch_etf_data(etf_code: str) -> tuple[str, list[dict], str]:
    """Fetch ETF name and all constituent stocks.
    Primary source: TWSE / FinMind. Fallback: yfinance funds_data.
    Returns (etf_display_name, components, debug_msg).
    components: list of {code, name, weight, price, change_pct}
    """
    code = str(etf_code).strip()

    # ── Step 1: TWSE 名稱表（含 ETF 自身名稱）───────────────────────────────
    name_table = _load_twse_stock_names()
    etf_name   = name_table.get(code, code)

    # ── Step 2: 成分股（TWSE 為主，Yahoo Finance topHoldings 為備）────────────
    raw_holdings, debug_msg = _fetch_twse_etf_holdings(code)
    print(f'[ETF DEBUG] {code}: {debug_msg}')

    if not raw_holdings:
        return etf_name, [], debug_msg

    # ── Step 3: 補齊中文名稱（優先用 TWSE 名稱表）───────────────────────────
    for h in raw_holdings:
        if not h.get('sym'):
            h['sym'] = h['code'] + '.TW'
        tw_name = name_table.get(h['code'])
        if tw_name:
            h['name'] = tw_name
        elif not _has_cjk(h.get('name', '')):
            h['name'] = h['code']

    # ── Step 4: 若持股比例全為 0，改用等權重（防 ZeroDivisionError）──────────
    total_w = sum(h['weight'] for h in raw_holdings)
    if total_w <= 0:
        eq = 100.0 / len(raw_holdings)
        for h in raw_holdings:
            h['weight'] = eq

    # ── Step 5: 批次抓即時股價＋漲跌幅 ────────────────────────────────────────
    all_syms:   list[str]        = [h['sym'] for h in raw_holdings]
    price_map:  dict[str, float] = {}
    change_map: dict[str, float] = {}

    # 優先用 TWSE STOCK_DAY_ALL（官方收盤價，含漲跌幅）
    try:
        twse_cache = _ensure_twse_price_cache()  # {代號: (收盤價, 漲跌額)}
        for h in raw_holdings:
            c = h['code']
            if c in twse_cache:
                close_p, chg_amt = twse_cache[c]
                price_map[h['sym']] = close_p
                if close_p and close_p - chg_amt > 0:
                    change_map[h['sym']] = chg_amt / (close_p - chg_amt) * 100
    except Exception:
        pass

    # 未取得資料的股票，用 Yahoo Finance 補充
    missing_syms = [h['sym'] for h in raw_holdings if h['sym'] not in price_map]
    if missing_syms:
        try:
            import concurrent.futures as _cff
            def _do_dl():
                import logging as _log, io, sys as _sys
                _yf_log = _log.getLogger('yfinance')
                _prev_lvl = _yf_log.level
                _yf_log.setLevel(_log.CRITICAL)
                _old_stderr, _sys.stderr = _sys.stderr, io.StringIO()
                try:
                    return yf.download(missing_syms, period='5d', auto_adjust=False,
                                       progress=False, threads=True, timeout=8)
                finally:
                    _sys.stderr = _old_stderr
                    _yf_log.setLevel(_prev_lvl)
            with _cff.ThreadPoolExecutor(max_workers=1) as _dl_ex:
                try:
                    hist = _dl_ex.submit(_do_dl).result(timeout=15)
                except Exception:
                    hist = pd.DataFrame()
            if not hist.empty:
                if isinstance(hist.columns, pd.MultiIndex):
                    close_df = hist['Close']
                elif 'Close' in hist.columns:
                    close_df = hist[['Close']].rename(columns={'Close': missing_syms[0]})
                else:
                    close_df = pd.DataFrame()
                for sym in missing_syms:
                    if sym in close_df.columns:
                        prices = close_df[sym].dropna()
                        if len(prices) >= 2:
                            price_map[sym]  = float(prices.iloc[-1])
                            change_map[sym] = ((float(prices.iloc[-1]) - float(prices.iloc[-2]))
                                               / float(prices.iloc[-2]) * 100)
                        elif len(prices) == 1:
                            price_map[sym]  = float(prices.iloc[0])
                            change_map[sym] = 0.0
        except Exception:
            pass

    # ── Step 6: 組合結果 ──────────────────────────────────────────────────────
    components = []
    for h in raw_holdings:
        components.append({
            'code':       h['code'],
            'name':       h['name'],
            'weight':     h['weight'],
            'price':      price_map.get(h['sym']),
            'change_pct': change_map.get(h['sym'], 0.0),
        })
    components.sort(key=lambda x: -x['weight'])
    return etf_name, components, debug_msg


def _fetch_etfortune_meta(code: str) -> dict:
    """Scrape TWSE ETFortune page for AUM, NAV, closing price, and discount/premium.
    Returns dict with keys: total_assets (元), nav, price, atmps (折溢價 %).
    """
    import re as _re
    meta: dict = {}
    try:
        url  = f'https://www.twse.com.tw/zh/ETFortune/etfInfo/{code}'
        html = _cffi_get_text(url)
        if not html or len(html) < 500:
            return meta

        # AUM: first <span>NUMBER</span>&nbsp;<b>億元</b> on page
        m = _re.search(r'<span>([\d,\.]+)</span>&nbsp;<b>億元</b>', html)
        if m:
            meta['total_assets'] = float(m.group(1).replace(',', '')) * 1e8  # 億元 → 元

        # chartData JS variable: netPrice=NAV, close1=market price, atmps=折溢價%
        idx = html.find('chartData = {')
        if idx >= 0:
            chart, _ = _json.JSONDecoder().raw_decode(html[idx + len('chartData = '):])
            net = chart.get('netPrice', [])
            if net:
                meta['nav'] = float(net[-1]['count'])
            close = chart.get('close1', [])
            if close:
                meta['price'] = float(close[-1]['count'])
            atmps = chart.get('atmps', [])
            if atmps:
                meta['atmps'] = float(atmps[-1]['count'])
    except Exception as _e:
        _net_log(f'_fetch_etfortune_meta({code}): {_e}')
    return meta


def fetch_etf_meta(etf_code: str) -> dict:
    """Fetch ETF AUM, yield, NAV, sector weights, asset allocation.
    Source 1: TWSE ETFortune page (authoritative for TW ETFs: AUM, NAV, price, 折溢價)
    Source 2: Yahoo Finance quoteSummary (yield_pct, sector_weights, asset_alloc)
    Returns dict with keys: total_assets, yield_pct, nav, price, atmps, sector_weights, asset_alloc.
    """
    code = str(etf_code).strip()
    meta: dict = {}

    # ── 方法1：TWSE ETFortune（AUM、NAV、市價、折溢價）────────────────────────
    meta.update(_fetch_etfortune_meta(code))

    # ── 方法2：Yahoo Finance quoteSummary（補充 yield、sector、asset allocation）
    try:
        from yfinance.data import YfData as _YfData
        _yfdata = _YfData(session=None)
        for suffix in ['.TW', '.TWO']:
            try:
                r = _yfdata.cache_get(
                    url=f'https://query2.finance.yahoo.com/v10/finance/quoteSummary/{code}{suffix}',
                    params={'modules': 'summaryDetail,etfProfile'},
                    timeout=8)
                if r is None:
                    continue
                payload = _json.loads(r.content)
                result  = payload.get('quoteSummary', {}).get('result', [{}])
                if not result:
                    continue
                res  = result[0]
                sd   = res.get('summaryDetail', {})
                for src_key, dst_key in [('totalAssets',  'total_assets'),
                                          ('yield',        'yield_pct'),
                                          ('navPrice',     'nav'),
                                          ('regularMarketPrice', 'price'),
                                          ('previousClose', 'prev_close')]:
                    if not meta.get(dst_key):
                        val = sd.get(src_key, {})
                        v   = val.get('raw') if isinstance(val, dict) else (val or None)
                        if v:
                            meta[dst_key] = v
                ep = res.get('etfProfile', {})
                if not meta.get('sector_weights'):
                    meta['sector_weights'] = ep.get('sectorWeightings', [])
                if not meta.get('asset_alloc'):
                    meta['asset_alloc']    = ep.get('assetAllocations', [])
                if any(v is not None for v in meta.values()):
                    break
            except Exception:
                pass
    except Exception:
        pass

    # ── 方法3：TWSE 收盤價快取（ETFortune 和 Yahoo 都無法提供市價時）──────────
    if not meta.get('price') and not meta.get('prev_close'):
        try:
            price_cache = _ensure_twse_price_cache()
            if code in price_cache:
                meta['price'] = price_cache[code][0]
        except Exception:
            pass

    return meta


def get_stock_name(code: str) -> str | None:
    code = str(code).strip()
    for s in ['.TW', '.TWO', '']:
        try:
            info = yf.Ticker(code + s).info
            name = info.get('longName') or info.get('shortName')
            if name:
                return name
        except Exception:
            pass
    return None

# ─── 樹狀圖顏色（finviz 風格）────────────────────────────────────────────────
def pnl_color(pct: float, max_abs: float = 10.0) -> str:
    """台灣慣例：正報酬紅色，負報酬綠色。
    顏色深淺依據 pct 佔 max_abs 的比例線性插值，而非固定門檻。"""
    try:
        pct = float(pct)
    except (TypeError, ValueError):
        pct = 0.0
    import math
    if math.isnan(pct) or math.isinf(pct):
        pct = 0.0
    if max_abs <= 0:
        return '#3a3a3a'
    t = min(abs(pct) / max_abs, 1.0)   # 0.0（接近0%）→ 1.0（極值）
    # 中性灰 → 深紅 / 深綠
    if pct >= 0:
        r = int(0x3a + t * (0xcc - 0x3a))
        g = int(0x3a + t * (0x1a - 0x3a))
        b = int(0x3a + t * (0x1a - 0x3a))
    else:
        r = int(0x3a + t * (0x0e - 0x3a))
        g = int(0x3a + t * (0x7a - 0x3a))
        b = int(0x3a + t * (0x1e - 0x3a))
    return f'#{r:02x}{g:02x}{b:02x}'

# ─── 美股顏色（US 慣例：漲綠跌紅，與台股相反）────────────────────────────────
def pnl_color_us(pct: float, max_abs: float = 10.0) -> str:
    """美國慣例：正報酬綠色，負報酬紅色。"""
    try:
        pct = float(pct)
    except (TypeError, ValueError):
        pct = 0.0
    import math
    t = min(abs(pct) / max(max_abs, 0.01), 1.0)
    t = math.sqrt(t)
    if pct >= 0:
        r = int(0x3a + t * (0x0e - 0x3a))
        g = int(0x3a + t * (0x7a - 0x3a))
        b = int(0x3a + t * (0x1e - 0x3a))
    else:
        r = int(0x3a + t * (0xc0 - 0x3a))
        g = int(0x3a + t * (0x20 - 0x3a))
        b = int(0x3a + t * (0x20 - 0x3a))
    return f'#{r:02x}{g:02x}{b:02x}'

# ─── 美股 S&P 500 類股代號對照表 ──────────────────────────────────────────────
_US_SECTOR_STOCKS: dict[str, list[str]] = {
    '科技': [
        'AAPL', 'MSFT', 'NVDA', 'AVGO', 'ORCL', 'AMD', 'QCOM', 'TXN', 'MU',
        'AMAT', 'LRCX', 'NOW', 'CRM', 'ADBE', 'INTU', 'PANW', 'NET', 'FTNT',
        'CRWD', 'ADI', 'KLAC', 'MCHP', 'ON', 'INTC', 'HPQ', 'IBM', 'ACN',
    ],
    '醫療保健': [
        'UNH', 'JNJ', 'LLY', 'ABBV', 'MRK', 'TMO', 'ABT', 'DHR', 'AMGN',
        'GILD', 'CI', 'ISRG', 'MDT', 'BSX', 'SYK', 'ELV', 'VRTX', 'REGN',
        'HUM', 'CVS', 'MCK', 'ZBH', 'BAX', 'BDX',
    ],
    '金融': [
        'BRK-B', 'JPM', 'BAC', 'WFC', 'GS', 'MS', 'BLK', 'SCHW', 'AXP',
        'USB', 'PNC', 'COF', 'TFC', 'MET', 'CME', 'ICE', 'SPGI', 'MCO',
        'PRU', 'AFL', 'ALL', 'CB', 'AIG', 'HIG',
    ],
    '非必需消費': [
        'AMZN', 'TSLA', 'HD', 'MCD', 'NKE', 'LOW', 'SBUX', 'BKNG', 'TJX',
        'ORLY', 'AZO', 'ROST', 'CMG', 'DG', 'DLTR', 'DHI', 'LEN', 'PHM',
        'F', 'GM', 'APTV', 'HMC',
    ],
    '通訊服務': [
        'META', 'GOOG', 'GOOGL', 'NFLX', 'DIS', 'CMCSA', 'T', 'VZ',
        'TMUS', 'CHTR', 'EA', 'TTWO', 'WBD', 'MTCH',
    ],
    '工業': [
        'GE', 'CAT', 'HON', 'UPS', 'RTX', 'BA', 'LMT', 'NOC', 'GD', 'MMM',
        'EMR', 'ITW', 'PH', 'ETN', 'DE', 'FDX', 'NSC', 'UNP', 'CSX', 'WM',
        'RSG', 'CTAS', 'FAST', 'GWW',
    ],
    '必需消費': [
        'PG', 'KO', 'PEP', 'COST', 'WMT', 'PM', 'MO', 'MDLZ', 'CL',
        'KMB', 'GIS', 'K', 'HSY', 'MKC', 'CAG', 'HRL', 'SJM',
    ],
    '能源': [
        'XOM', 'CVX', 'COP', 'EOG', 'SLB', 'MPC', 'PSX', 'VLO', 'OXY',
        'HAL', 'BKR', 'DVN', 'APA', 'HES', 'FANG',
    ],
    '原物料': [
        'LIN', 'SHW', 'APD', 'FCX', 'NEM', 'NUE', 'VMC', 'MLM', 'DOW',
        'DD', 'PPG', 'IFF', 'CF', 'MOS', 'ALB',
    ],
    '房地產': [
        'PLD', 'AMT', 'EQIX', 'CCI', 'PSA', 'EXR', 'WELL', 'DLR', 'O',
        'SPG', 'VTR', 'EQR', 'AVB', 'INVH', 'VICI',
    ],
    '公用事業': [
        'NEE', 'DUK', 'SO', 'D', 'AEP', 'XEL', 'SRE', 'ED', 'WEC', 'AWK',
        'ES', 'FE', 'ETR', 'AEE', 'PCG',
    ],
}

# ─── 自製日期選擇器（避免 DateEntry 導覽時自動關閉的 bug）────────────────────
class DatePickerEntry(ttk.Frame):
    """Entry + 📅 按鈕，點擊後開啟獨立 Calendar Toplevel，切換月份/年份不會關閉"""

    def __init__(self, parent, initial_date=None, on_date_selected=None, **_kw):
        super().__init__(parent, style='TFrame')
        self._callback = on_date_selected

        # 初始日期
        d = initial_date or datetime.today()
        self._date = d.date() if isinstance(d, datetime) else d

        self._var = tk.StringVar(value=self._date.strftime('%Y-%m-%d'))

        entry = ttk.Entry(self, textvariable=self._var, width=13)
        entry.pack(side='left', ipady=2)
        entry.bind('<FocusOut>', self._parse_entry)
        entry.bind('<Return>',   self._parse_entry)

        ttk.Button(self, text='📅', width=3,
                   command=self._open_popup).pack(side='left', padx=(3, 0))

    # ── 手動輸入解析 ──────────────────────────────────────────────────────────
    def _parse_entry(self, _=None):
        try:
            self._date = datetime.strptime(self._var.get().strip(), '%Y-%m-%d').date()
        except ValueError:
            self._var.set(self._date.strftime('%Y-%m-%d'))   # 格式錯誤還原

    # ── 開啟日曆彈窗 ──────────────────────────────────────────────────────────
    def _open_popup(self):
        parent_top = self.winfo_toplevel()   # 可能是主視窗或 EditDialog

        popup = tk.Toplevel()
        popup.title('選擇日期')
        popup.configure(bg=C_BG)
        popup.resizable(False, False)
        popup.transient(parent_top)

        # 定位在 Entry 下方
        self.update_idletasks()
        px = self.winfo_rootx()
        py = self.winfo_rooty() + self.winfo_height() + 2
        popup.geometry(f'+{px}+{py}')

        cal = Calendar(
            popup,
            selectmode='day',
            year=self._date.year, month=self._date.month, day=self._date.day,
            date_pattern='yyyy-mm-dd',
            background=C_ACCENT,       foreground='white',
            headersbackground='#1c2333', headersforeground='#8ab4d4',
            normalbackground=C_PANEL,  normalforeground=C_FG,
            weekendbackground='#2a2a2a', weekendforeground=C_FG2,
            othermonthforeground='#555', othermonthbackground=C_BG,
            othermonthweforeground='#444', othermonthwebackground=C_BG,
            selectbackground=C_ACCENT, selectforeground='white',
            bordercolor=C_BORDER,
            font=('Microsoft JhengHei', 10),
        )
        cal.pack(padx=8, pady=8)

        def _confirm(_=None):
            date_str = cal.get_date()
            try:
                self._date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
            self._var.set(self._date.strftime('%Y-%m-%d'))
            _close()
            if self._callback:
                self._callback()

        def _close():
            popup.grab_release()
            popup.destroy()
            # 若父視窗是 modal Toplevel，歸還 grab
            if isinstance(parent_top, tk.Toplevel):
                try:
                    parent_top.grab_set()
                except Exception:
                    pass

        popup.protocol('WM_DELETE_WINDOW', _close)

        btn_f = ttk.Frame(popup)
        btn_f.pack(pady=(0, 8))
        ttk.Button(btn_f, text='✅  確認', command=_confirm).pack(side='left', padx=5)
        ttk.Button(btn_f, text='✖  取消', command=_close).pack(side='left', padx=5)

        cal.bind('<Double-1>', _confirm)

        # 若父視窗有 grab，先暫時釋放再接管
        if isinstance(parent_top, tk.Toplevel):
            try:
                parent_top.grab_release()
            except Exception:
                pass
        popup.grab_set()
        popup.focus_set()

    # ── 公開介面（與 DateEntry 相容）─────────────────────────────────────────
    def get_date(self):
        """返回 datetime.date"""
        return self._date

    def set_date(self, d):
        if isinstance(d, datetime):
            d = d.date()
        elif isinstance(d, str):
            d = datetime.strptime(d, '%Y-%m-%d').date()
        self._date = d
        self._var.set(d.strftime('%Y-%m-%d'))


# ─── 主視窗 ──────────────────────────────────────────────────────────────────
class StockApp(tk.Tk):

    # sidebar nav items: (icon, label, page_index)
    _NAV_ITEMS = [
        ('🌳', '庫存樹狀圖'),
        ('📊', '庫存個股分析'),
        ('📈', 'ETF分析'),
        ('📝', '買賣紀錄'),
        ('🇹🇼', '台股總覽'),
        ('📉', '個股分析'),
        ('🇺🇸', '美股總覽'),
    ]
    _SIDEBAR_EXP = 200   # expanded width px
    _SIDEBAR_COL = 54    # collapsed width px
    _SB_BG       = '#16161e'
    _SB_SEL      = '#1f2d45'
    _SB_HOV      = '#1c2030'

    def __init__(self):
        super().__init__()
        self.title('個人股票管理工具')
        self.geometry('1200x780')
        self.configure(bg=C_BG)
        self._apply_dark_theme()

        # ── 跨執行緒 UI 回呼佇列（必須最先初始化，背景執行緒可能很早啟動）──────
        import queue as _q
        self._ui_queue: _q.SimpleQueue = _q.SimpleQueue()
        self._pump_job = self.after(50, self._pump_ui_queue)

        # ── 主框架：sidebar + content ─────────────────────────────────────────
        root_frame = tk.Frame(self, bg=C_BG)
        root_frame.pack(fill='both', expand=True)

        # Left sidebar
        self._sb_expanded = True
        self._sidebar = tk.Frame(root_frame, bg=self._SB_BG,
                                 width=self._SIDEBAR_EXP)
        self._sidebar.pack(side='left', fill='y')
        self._sidebar.pack_propagate(False)

        # Thin separator line
        tk.Frame(root_frame, bg='#2a2a3a', width=1).pack(side='left', fill='y')

        # Content area (stacked pages)
        self._content = tk.Frame(root_frame, bg=C_BG)
        self._content.pack(side='left', fill='both', expand=True)

        self.tab1 = tk.Frame(self._content, bg=C_BG)
        self.tab2 = tk.Frame(self._content, bg=C_BG)
        self.tab3 = tk.Frame(self._content, bg=C_BG)
        self.tab4 = tk.Frame(self._content, bg=C_BG)
        self.tab5 = tk.Frame(self._content, bg=C_BG)
        self.tab6 = tk.Frame(self._content, bg=C_BG)
        self.tab7 = tk.Frame(self._content, bg=C_BG)
        for f in (self.tab1, self.tab2, self.tab3, self.tab4, self.tab5, self.tab6, self.tab7):
            f.place(relx=0, rely=0, relwidth=1, relheight=1)

        self._build_sidebar()

        self._vars: dict[str, tk.StringVar] = {}
        self._build_tab1()
        self._build_tab2()
        self._build_tab3()
        self._build_tab4()
        self._build_tab5()
        self._build_tab6()
        self._build_tab7()

        self._current_page = -1
        self._show_page(0)

        self.protocol('WM_DELETE_WINDOW', self._on_close)

    def _pump_ui_queue(self):
        """主執行緒每 50 ms 執行背景執行緒排隊的 UI 回呼"""
        try:
            while True:
                fn = self._ui_queue.get_nowait()
                try:
                    fn()
                except Exception:
                    pass
        except Exception:
            pass
        self._pump_job = self.after(50, self._pump_ui_queue)

    def _ui_call(self, fn):
        """Python 3.13+ 相容的跨執行緒 UI 回呼。
        主執行緒：直接執行；背景執行緒：放入佇列由主執行緒輪詢處理。"""
        import threading
        if threading.current_thread() is threading.main_thread():
            fn()
        else:
            self._ui_queue.put(fn)

    def _on_close(self):
        """視窗關閉：清理 matplotlib 資源後結束程序"""
        try:
            self.after_cancel(self._pump_job)
        except Exception:
            pass
        try:
            plt.close('all')
        except Exception:
            pass
        self.destroy()
        import os
        os._exit(0)

    # ── Sidebar ───────────────────────────────────────────────────────────────
    def _build_sidebar(self):
        SB = self._SB_BG

        # ── 頂部：Logo + 收合按鈕 ──────────────────────────────────────────
        top = tk.Frame(self._sidebar, bg=SB)
        top.pack(fill='x', padx=6, pady=(12, 6))

        self._sb_title = tk.Label(top, text='  股票管理', bg=SB,
                                   fg='#aecde8',
                                   font=('Microsoft JhengHei', 12, 'bold'),
                                   anchor='w')
        self._sb_title.pack(side='left', fill='x', expand=True)

        self._sb_toggle = tk.Button(top, text='◀', bg=SB, fg='#5a7fa8',
                                     font=('Microsoft JhengHei', 11),
                                     relief='flat', bd=0, cursor='hand2',
                                     activebackground=SB, activeforeground='#aecde8',
                                     command=self._toggle_sidebar)
        self._sb_toggle.pack(side='right', padx=4)

        # Divider
        tk.Frame(self._sidebar, bg='#252535', height=1).pack(fill='x',
                                                               padx=8, pady=4)

        # ── 導覽項目 ────────────────────────────────────────────────────────
        self._nav_rows   = []   # outer frame per item
        self._nav_labels = []   # text Label per item (hidden when collapsed)

        for i, (icon, label) in enumerate(self._NAV_ITEMS):
            row = tk.Frame(self._sidebar, bg=SB, cursor='hand2')
            row.pack(fill='x', padx=6, pady=2)

            icon_lbl = tk.Label(row, text=icon, bg=SB, fg=C_FG,
                                 font=('Segoe UI Emoji', 15),
                                 cursor='hand2', width=2)
            icon_lbl.pack(side='left', padx=(8, 4), pady=8)

            txt_lbl = tk.Label(row, text=label, bg=SB, fg='#9090a0',
                                font=('Microsoft JhengHei', 11),
                                anchor='w', cursor='hand2')
            txt_lbl.pack(side='left', fill='x', expand=True, pady=8)

            for w in (row, icon_lbl, txt_lbl):
                w.bind('<Button-1>', lambda e, n=i: self._show_page(n))
                w.bind('<Enter>',    lambda e, r=row, n=i: self._sb_hover(r, n, True))
                w.bind('<Leave>',    lambda e, r=row, n=i: self._sb_hover(r, n, False))

            self._nav_rows.append((row, icon_lbl, txt_lbl))
            self._nav_labels.append(txt_lbl)

    def _toggle_sidebar(self):
        self._sb_expanded = not self._sb_expanded
        if self._sb_expanded:
            self._sidebar.config(width=self._SIDEBAR_EXP)
            self._sb_title.pack(side='left', fill='x', expand=True)
            for _, _, txt in self._nav_rows:
                txt.pack(side='left', fill='x', expand=True, pady=8)
            self._sb_toggle.config(text='◀')
        else:
            self._sidebar.config(width=self._SIDEBAR_COL)
            self._sb_title.pack_forget()
            for _, _, txt in self._nav_rows:
                txt.pack_forget()
            self._sb_toggle.config(text='▶')

    def _sb_hover(self, row, idx, entering):
        if idx == self._current_page:
            return
        bg = self._SB_HOV if entering else self._SB_BG
        for w in [row] + list(row.winfo_children()):
            w.config(bg=bg)

    def _show_page(self, idx):
        pages = [self.tab2, self.tab3, self.tab4, self.tab1, self.tab5, self.tab6, self.tab7]
        pages[idx].tkraise()

        # Update nav highlight
        for i, (row, icon_lbl, txt_lbl) in enumerate(self._nav_rows):
            if i == idx:
                bg = self._SB_SEL
                fg_txt = C_FG
            else:
                bg = self._SB_BG
                fg_txt = '#9090a0'
            row.config(bg=bg)
            icon_lbl.config(bg=bg)
            txt_lbl.config(bg=bg, fg=fg_txt)

        prev = self._current_page
        self._current_page = idx

        # 離開 Tab 5 時關閉懸浮提示
        if prev == 4 and idx != 4:
            self._mkt_tooltip.withdraw()
        # 離開 Tab 7（美股）時關閉懸浮提示
        if prev == 6 and idx != 6:
            if hasattr(self, '_us_mkt_tooltip'):
                self._us_mkt_tooltip.withdraw()

        # Side-effects
        if prev == 0 and idx != 0:
            self._cancel_tm_refresh()
        if idx == 0:
            if prev != 0:
                self._draw_treemap()
            self._schedule_tm_refresh()
        elif idx == 1 and prev != 1:
            self._update_stock_list()
        elif idx == 4 and prev != 4:
            self._draw_market_map()
        elif idx == 6 and prev != 6:
            self._draw_us_market_map()
        elif idx == 5 and prev != 5:
            code = self._stk_code.get().strip()
            if code and self._stk_current_code != code:
                self._load_stk_chart()
            elif self._stk_current_code is None:
                # 第一次顯示：畫提示文字
                fig = self._stk_fig
                fig.clear()
                fig.patch.set_facecolor('#111111')
                ax = fig.add_axes([0, 0, 1, 1])
                ax.set_facecolor('#111111')
                ax.axis('off')
                fp = fm.FontProperties(family=CHART_FONT)
                ax.text(0.5, 0.5, '請輸入股票代號後點擊查詢',
                        ha='center', va='center', color='#888',
                        fontsize=13, fontproperties=fp, transform=ax.transAxes)
                self._stk_canvas.draw_idle()

    # ── 深色主題 ──────────────────────────────────────────────────────────────
    def _apply_dark_theme(self):
        style = ttk.Style()
        style.theme_use('clam')

        style.configure('.',
            background=C_BG, foreground=C_FG,
            fieldbackground=C_INPUT, troughcolor=C_PANEL,
            bordercolor=C_BORDER, darkcolor=C_PANEL, lightcolor=C_PANEL,
            font=BODY_FONT)

        style.configure('TFrame',     background=C_BG)
        style.configure('TLabel',     background=C_BG, foreground=C_FG)
        style.configure('Hdr.TLabel', background=C_BG, foreground=C_FG, font=HDR_FONT)

        style.configure('TLabelframe',
            background=C_BG, bordercolor=C_BORDER, relief='solid')
        style.configure('TLabelframe.Label',
            background=C_BG, foreground=C_ACCENT, font=BODY_FONT)

        style.configure('TNotebook',
            background=C_BG, bordercolor=C_BORDER, tabmargins=[2, 5, 0, 0])
        style.configure('TNotebook.Tab',
            background=C_PANEL, foreground=C_FG2,
            padding=[14, 6], font=('Microsoft JhengHei', 11, 'bold'))
        style.map('TNotebook.Tab',
            background=[('selected', C_BG),    ('active', '#303030')],
            foreground=[('selected', C_FG),     ('active', C_FG)])

        style.configure('TEntry',
            fieldbackground=C_INPUT, foreground=C_FG,
            insertcolor=C_FG, bordercolor=C_BORDER, relief='flat', padding=4)
        style.map('TEntry',
            fieldbackground=[('focus', '#3a3a3a')],
            bordercolor=[('focus', C_ACCENT)])

        style.configure('TCombobox',
            fieldbackground=C_INPUT, foreground=C_FG,
            background=C_PANEL, arrowcolor=C_FG,
            bordercolor=C_BORDER, relief='flat', padding=4)
        style.map('TCombobox',
            fieldbackground=[('readonly', C_INPUT)],
            foreground=[('readonly', C_FG)],
            selectbackground=[('readonly', C_ACCENT)],
            selectforeground=[('readonly', 'white')])

        style.configure('TButton',
            background=C_ACCENT, foreground='white',
            bordercolor=C_ACCENT, relief='flat', padding=[10, 5])
        style.map('TButton',
            background=[('active', '#1a90e0'), ('pressed', '#005ba1')])

        style.configure('Treeview',
            background=C_PANEL, foreground=C_FG,
            fieldbackground=C_PANEL, rowheight=26,
            font=('Microsoft JhengHei UI', 10),
            bordercolor=C_BORDER, relief='flat')
        style.configure('Treeview.Heading',
            background='#333333', foreground=C_FG,
            relief='flat', bordercolor=C_BORDER)
        style.map('Treeview',
            background=[('selected', C_ACCENT)],
            foreground=[('selected', 'white')])

        style.configure('TScrollbar',
            background=C_PANEL, troughcolor=C_BG,
            arrowcolor=C_FG2, bordercolor=C_BORDER)

        # 買賣 Combobox 顏色（全域小字）
        style.configure('Buy.TCombobox',  foreground=C_BUY_FG,  font=BODY_FONT)
        style.configure('Sell.TCombobox', foreground=C_SELL_FG, font=BODY_FONT)
        style.map('Buy.TCombobox',  fieldbackground=[('readonly', C_BUY_BG)])
        style.map('Sell.TCombobox', fieldbackground=[('readonly', C_SELL_BG)])

        # ── 輸入區專用樣式（INPUT_FONT 12pt）──────────────────────────────────
        style.configure('In.TLabelframe',
            background=C_BG, bordercolor=C_BORDER, relief='solid')
        style.configure('In.TLabelframe.Label',
            background=C_BG, foreground=C_ACCENT, font=INPUT_FONT)

        style.configure('In.TLabel',
            background=C_BG, foreground=C_FG, font=INPUT_FONT)

        style.configure('In.TEntry',
            fieldbackground=C_INPUT, foreground=C_FG,
            insertcolor=C_FG, bordercolor=C_BORDER,
            relief='flat', padding=5, font=INPUT_FONT)
        style.map('In.TEntry',
            fieldbackground=[('focus', '#3a3a3a')],
            bordercolor=[('focus', C_ACCENT)])

        style.configure('In.TCombobox',
            fieldbackground=C_INPUT, foreground=C_FG,
            background=C_PANEL, arrowcolor=C_FG,
            bordercolor=C_BORDER, relief='flat', padding=5, font=INPUT_FONT)
        style.map('In.TCombobox',
            fieldbackground=[('readonly', C_INPUT)],
            foreground=[('readonly', C_FG)],
            selectbackground=[('readonly', C_ACCENT)],
            selectforeground=[('readonly', 'white')])

        style.configure('In.TButton',
            background=C_ACCENT, foreground='white',
            bordercolor=C_ACCENT, relief='flat',
            padding=[12, 6], font=INPUT_FONT)
        style.map('In.TButton',
            background=[('active', '#1a90e0'), ('pressed', '#005ba1')])

        style.configure('Nav.TButton',
            background='#1e1e2e', foreground=C_FG,
            bordercolor='#2a2a3a', relief='flat',
            padding=[12, 7], font=('Microsoft JhengHei', 10))
        style.map('Nav.TButton',
            background=[('active', '#252538'), ('pressed', '#1a253a')],
            foreground=[('active', '#ffffff'), ('pressed', '#ffffff')])

        style.configure('In.TCheckbutton',
            background=C_INPUT, foreground=C_FG,
            font=INPUT_FONT, focuscolor=C_INPUT)
        style.map('In.TCheckbutton',
            background=[('active', C_INPUT)],
            foreground=[('active', C_FG)])

        # 買賣 Combobox（輸入區大字）
        style.configure('InBuy.TCombobox',  foreground=C_BUY_FG,  font=INPUT_FONT)
        style.configure('InSell.TCombobox', foreground=C_SELL_FG, font=INPUT_FONT)
        style.map('InBuy.TCombobox',  fieldbackground=[('readonly', C_BUY_BG)])
        style.map('InSell.TCombobox', fieldbackground=[('readonly', C_SELL_BG)])

    # ── Tab 切換 ──────────────────────────────────────────────────────────────
    def _on_tab(self, event):
        pass  # replaced by _show_page

    # ═══════════════════════════════════════════════════════════════════════════
    # Tab 1：買賣紀錄
    # ═══════════════════════════════════════════════════════════════════════════
    def _build_tab1(self):
        f = self.tab1

        box = ttk.LabelFrame(f, text='  新增交易  ', padding=14, style='In.TLabelframe')
        box.pack(fill='x', padx=14, pady=(12, 6))

        lkw = dict(sticky='e', padx=(12, 4), pady=8)
        ekw = dict(sticky='w', padx=(0, 18), pady=8)

        today = datetime.today()

        # Row 0: 日期 | 股票代號
        ttk.Label(box, text='日期', style='In.TLabel').grid(row=0, column=0, **lkw)
        self._date_entry = DatePickerEntry(
            box, on_date_selected=self._on_date_change)
        self._date_entry.grid(row=0, column=1, **ekw)

        ttk.Label(box, text='股票代號', style='In.TLabel').grid(row=0, column=2, **lkw)
        self._vars['code'] = tk.StringVar()
        self._code_entry = ttk.Combobox(box, textvariable=self._vars['code'],
                                         width=20, style='In.TCombobox')
        self._code_entry.grid(row=0, column=3, **ekw)
        self._code_entry.bind('<FocusOut>',           self._on_code_focusout)
        self._code_entry.bind('<Return>',
            lambda e: (self._on_code_focusout(e), self._name_entry.focus()))
        self._code_entry.bind('<<ComboboxSelected>>', self._on_code_selected)

        # Row 1: 股票名稱 | 分類
        ttk.Label(box, text='股票名稱', style='In.TLabel').grid(row=1, column=0, **lkw)
        self._vars['name'] = tk.StringVar()
        self._name_entry = ttk.Combobox(box, textvariable=self._vars['name'],
                                         width=20, style='In.TCombobox')
        self._name_entry.grid(row=1, column=1, **ekw)
        self._name_entry.bind('<FocusOut>',           self._on_name_focusout)
        self._name_entry.bind('<<ComboboxSelected>>', self._on_name_selected)

        ttk.Label(box, text='投資分類', style='In.TLabel').grid(row=1, column=2, **lkw)
        self._vars['category'] = tk.StringVar(value=CATEGORIES[0])
        ttk.Combobox(box, textvariable=self._vars['category'],
                     values=CATEGORIES, state='readonly', width=18,
                     style='In.TCombobox').grid(row=1, column=3, **ekw)

        # Row 2: 買/賣 | 數量
        ttk.Label(box, text='買 / 賣', style='In.TLabel').grid(row=2, column=0, **lkw)
        self._vars['side'] = tk.StringVar(value='買')
        self._side_combo = ttk.Combobox(
            box, textvariable=self._vars['side'],
            values=['買', '賣'], state='readonly', width=18,
            style='InBuy.TCombobox')
        self._side_combo.grid(row=2, column=1, **ekw)
        self._vars['side'].trace_add('write', self._on_side_change)

        self._qty_label_var = tk.StringVar(value='數量 (股)')
        ttk.Label(box, textvariable=self._qty_label_var,
                  style='In.TLabel').grid(row=2, column=2, **lkw)
        self._vars['qty'] = tk.StringVar()
        _qty_frame = ttk.Frame(box)
        _qty_frame.grid(row=2, column=3, **ekw)
        ttk.Entry(_qty_frame, textvariable=self._vars['qty'],
                  width=14, style='In.TEntry').pack(side='left')
        self._lot_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(_qty_frame, text='整張', variable=self._lot_var,
                        style='In.TCheckbutton',
                        command=self._on_lot_toggle).pack(side='left', padx=(8, 0))

        # Row 3: 成交價格 | 手續費
        ttk.Label(box, text='成交價格 (元)', style='In.TLabel').grid(row=3, column=0, **lkw)
        self._vars['price'] = tk.StringVar()
        _price_frame = ttk.Frame(box)
        _price_frame.grid(row=3, column=1, **ekw)
        ttk.Entry(_price_frame, textvariable=self._vars['price'],
                  width=14, style='In.TEntry').pack(side='left')
        _tick_btn_frame = tk.Frame(_price_frame, bg=C_INPUT)
        _tick_btn_frame.pack(side='left', padx=(4, 0))
        tk.Button(_tick_btn_frame, text='▲', font=('Microsoft JhengHei', 8, 'bold'),
                  bg=C_INPUT, fg=C_FG, activebackground=C_BORDER, relief='flat',
                  bd=0, width=2, cursor='hand2',
                  command=lambda: self._adjust_price(+1)).pack(side='top')
        tk.Button(_tick_btn_frame, text='▼', font=('Microsoft JhengHei', 8, 'bold'),
                  bg=C_INPUT, fg=C_FG, activebackground=C_BORDER, relief='flat',
                  bd=0, width=2, cursor='hand2',
                  command=lambda: self._adjust_price(-1)).pack(side='top')

        ttk.Label(box, text='手續費 (留空自動)', style='In.TLabel').grid(row=3, column=2, **lkw)
        self._vars['fee'] = tk.StringVar()
        ttk.Entry(box, textvariable=self._vars['fee'],
                  width=20, style='In.TEntry').grid(row=3, column=3, **ekw)

        # Row 4: 交易稅（賣出時顯示）
        self._tax_label = ttk.Label(box, text='交易稅 (留空自動)', style='In.TLabel')
        self._tax_label.grid(row=4, column=2, **lkw)
        self._vars['tax'] = tk.StringVar()
        self._tax_entry = ttk.Entry(box, textvariable=self._vars['tax'],
                                    width=20, style='In.TEntry')
        self._tax_entry.grid(row=4, column=3, **ekw)
        # 買入時隱藏交易稅欄位
        if self._vars['side'].get() == '買':
            self._tax_label.grid_remove()
            self._tax_entry.grid_remove()

        # Row 5: 狀態列 + 按鈕
        self._status_var = tk.StringVar(value='')
        ttk.Label(box, textvariable=self._status_var,
                  foreground=C_FG2, font=('Microsoft JhengHei', 10)
                  ).grid(row=5, column=0, columnspan=3, sticky='w', padx=14)

        btn_f = ttk.Frame(box)
        btn_f.grid(row=5, column=3, sticky='e', padx=(0, 4), pady=(6, 0))
        ttk.Button(btn_f, text='✅  新增交易', style='Nav.TButton',
                   command=self._add_tx).pack(side='left', padx=4)
        ttk.Button(btn_f, text='🔄  重新整理', style='Nav.TButton',
                   command=self._refresh_table).pack(side='left', padx=4)

        # ── 歷史紀錄 ──────────────────────────────────────────────────────────
        list_box = ttk.LabelFrame(f, text='  歷史交易紀錄  ', padding=8)
        list_box.pack(fill='both', expand=True, padx=14, pady=(0, 12))

        # 主表格：日期 ~ 淨金額，套用買/賣配色
        self._sort_col = '日期'
        self._sort_asc = False
        self._pnl_gen  = 0
        # 總計 ≤ 755px，確保淨金額不需橫向捲動 (P&L 178 + vsb 17 = 195, 可用 ~955-195=760)
        _main_widths = [84, 66, 78, 70, 44, 64, 72, 70, 68, 86]
        self._tree = ttk.Treeview(list_box, columns=COLUMNS, show='headings', height=11)
        for col, w in zip(COLUMNS, _main_widths):
            self._tree.heading(col, text=col,
                               command=lambda c=col: self._sort_by_col(c))
            self._tree.column(col, width=w, anchor='center', stretch=False)
        self._tree.tag_configure('buy',  foreground=C_BUY_FG)
        self._tree.tag_configure('sell', foreground=C_SELL_FG)
        self._tree.tag_configure('row_even', background='#252526')
        self._tree.tag_configure('row_odd',  background='#202030')

        # 損益子表格：獨立配色，紅底=獲利，綠底=虧損
        _PNL_COLS = ['損益(元)', '損益率(%)']
        _pnl_widths = [92, 86]
        self._tree_pnl = ttk.Treeview(list_box, columns=_PNL_COLS, show='headings',
                                       height=11, selectmode='none')
        for col, w in zip(_PNL_COLS, _pnl_widths):
            self._tree_pnl.heading(col, text=col)
            self._tree_pnl.column(col, width=w, anchor='center', stretch=False,
                                   minwidth=w)
        self._tree_pnl.tag_configure('pnl_gain', background='#6b0000', foreground='#ffffff')
        self._tree_pnl.tag_configure('pnl_loss', background='#0a3d0a', foreground='#ffffff')
        self._tree_pnl.tag_configure('row_even', background='#252526')
        self._tree_pnl.tag_configure('row_odd',  background='#202030')

        def _ysync(*args):
            self._tree.yview(*args)
            self._tree_pnl.yview(*args)
        vsb = ttk.Scrollbar(list_box, orient='vertical', command=_ysync)
        hsb = ttk.Scrollbar(list_box, orient='horizontal', command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._tree_pnl.configure(yscrollcommand=vsb.set)

        def _on_wheel(event):
            delta = -1 * (event.delta // 120)
            self._tree.yview_scroll(delta, 'units')
            self._tree_pnl.yview_scroll(delta, 'units')
            return 'break'
        self._tree.bind('<MouseWheel>', _on_wheel)
        self._tree_pnl.bind('<MouseWheel>', _on_wheel)

        self._tree.grid(row=0, column=0, sticky='nsew')
        self._tree_pnl.grid(row=0, column=1, sticky='ns')
        vsb.grid(row=0, column=2, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        list_box.rowconfigure(0, weight=1)
        list_box.columnconfigure(0, weight=1)

        # ── 操作按鈕列 ────────────────────────────────────────────────────────
        btn_row = ttk.Frame(list_box)
        btn_row.grid(row=2, column=0, columnspan=3, sticky='w', pady=(6, 2))
        ttk.Button(btn_row, text='✏️  修改選取', style='Nav.TButton',
                   command=self._edit_selected).pack(side='left', padx=(0, 6))
        ttk.Button(btn_row, text='🗑️  刪除選取', style='Nav.TButton',
                   command=self._delete_selected).pack(side='left')

        # 右鍵選單
        self._ctx_menu = tk.Menu(self, tearoff=0, bg=C_PANEL, fg=C_FG,
                                  activebackground=C_ACCENT, activeforeground='white',
                                  font=BODY_FONT, bd=0)
        self._ctx_menu.add_command(label='✏️  修改', command=self._edit_selected)
        self._ctx_menu.add_command(label='🗑️  刪除', command=self._delete_selected)
        self._tree.bind('<Button-3>', self._show_ctx_menu)
        self._tree.bind('<Double-1>', lambda e: self._edit_selected())

        self._refresh_table()
        self._refresh_stock_options()

    # ── 股票代號 / 名稱快選選項 ────────────────────────────────────────────────
    def _refresh_stock_options(self):
        """從歷史紀錄更新 Tab 1 代號 / 名稱的快選清單"""
        df = load_df()
        if df.empty:
            return
        # 依最近使用排序（DataFrame 已按日期排序，取 unique 保留最後出現順序）
        codes = list(dict.fromkeys(df['股票代號'].astype(str).tolist()))[::-1]
        names = list(dict.fromkeys(df['股票名稱'].astype(str).tolist()))[::-1]
        self._code_entry['values'] = codes
        self._name_entry['values'] = names

    # ── 代號 / 名稱下拉選取後的處理 ───────────────────────────────────────────
    def _on_code_selected(self, _=None):
        """從下拉選單選取代號：清空名稱後觸發自動帶入"""
        self._vars['name'].set('')
        self._on_code_focusout()

    def _on_name_selected(self, _=None):
        """從下拉選單選取名稱：清空代號後觸發反向查詢"""
        self._vars['code'].set('')
        self._on_name_focusout()

    # ── 事件 ──────────────────────────────────────────────────────────────────
    def _on_side_change(self, *_):
        s = self._vars['side'].get()
        self._side_combo.configure(
            style='InBuy.TCombobox' if s == '買' else 'InSell.TCombobox')
        if s == '賣':
            self._tax_label.grid()
            self._tax_entry.grid()
        else:
            self._vars['tax'].set('')
            self._tax_label.grid_remove()
            self._tax_entry.grid_remove()

    def _on_lot_toggle(self):
        if self._lot_var.get():
            self._qty_label_var.set('數量 (張)')
        else:
            self._qty_label_var.set('數量 (股)')

    def _on_date_change(self, _=None):
        code = self._vars['code'].get().strip().upper()
        if code:
            self._trigger_price_lookup(code)

    def _on_code_focusout(self, _=None):
        code = self._vars['code'].get().strip().upper()
        if not code:
            return
        self._vars['code'].set(code)
        name = self._hist_name(code)
        if name:
            if not self._vars['name'].get():
                self._vars['name'].set(name)
            self._trigger_price_lookup(code)
        else:
            self._set_status('🔍 查詢股票資訊中…')
            threading.Thread(target=self._bg_name, args=(code,), daemon=True).start()

    def _on_name_focusout(self, _=None):
        name = self._vars['name'].get().strip()
        if not name or self._vars['code'].get().strip():
            return
        code = self._hist_code(name)
        if code:
            self._vars['code'].set(code)
            self._set_status(f'✅ 從歷史找到代號：{code}')
            self._trigger_price_lookup(code)

    # ── 台股跳檔輔助 ─────────────────────────────────────────────────────────
    @staticmethod
    def _tick_size(price: float) -> float:
        return 0.01 if price < 10 else 0.05

    def _adjust_price(self, direction: int):
        """direction: +1 上調一檔, -1 下調一檔
        先對齊到最近的 tick 邊界，再往指定方向移動一檔。
        例：100.02 → 上 → 100.05；100.05 → 上 → 100.10
        """
        try:
            p = float(self._vars['price'].get() or 0)
        except ValueError:
            return
        tick = self._tick_size(p)
        eps = 1e-9
        if direction == 1:
            # 先floor到當前tick，再加一格
            p = math.floor(p / tick + eps) * tick + tick
        else:
            # 先ceil到當前tick，再減一格
            p = math.ceil(p / tick - eps) * tick - tick
        p = round(max(p, tick), 2)
        self._vars['price'].set(f'{p:.2f}')

    def _trigger_price_lookup(self, code: str):
        date_str = self._date_entry.get_date().strftime('%Y-%m-%d')
        self._set_status('🔍 查詢歷史收盤價…')
        threading.Thread(target=self._bg_price,
                         args=(code, date_str), daemon=True).start()

    def _bg_name(self, code: str):
        name = get_stock_name(code)
        def _ui():
            if name:
                if not self._vars['name'].get():
                    self._vars['name'].set(name)
                self._set_status(f'✅ 股票名稱：{name}')
            else:
                self._set_status('⚠ 查無名稱，請手動填入')
            self._trigger_price_lookup(code)
        self._ui_call(_ui)

    def _bg_price(self, code: str, date_str: str):
        price, exact, ticker = get_price_on_date(code, date_str)
        def _ui():
            if price is not None:
                self._vars['price'].set(f'{price:.2f}')
                if exact:
                    self._set_status(f'✅ {date_str} 收盤價：{price:.2f} 元（{ticker}）')
                else:
                    self._set_status(
                        f'⚠ {date_str} 非交易日，填入最近收盤價 {price:.2f} 元（{ticker}）')
            else:
                self._set_status('⚠ 查無歷史價格，請手動填入')
        self._ui_call(_ui)

    def _hist_name(self, code: str) -> str | None:
        df = load_df()
        rows = df[df['股票代號'].astype(str) == code]
        return str(rows['股票名稱'].iloc[0]) if not rows.empty else None

    def _hist_code(self, name: str) -> str | None:
        df = load_df()
        rows = df[df['股票名稱'].astype(str).str.contains(name, na=False)]
        return str(rows['股票代號'].iloc[0]) if not rows.empty else None

    def _set_status(self, msg: str):
        self._ui_call(lambda: self._status_var.set(msg))

    def _add_tx(self):
        try:
            date_str = self._date_entry.get_date().strftime('%Y-%m-%d')
            code     = self._vars['code'].get().strip().upper()
            name     = self._vars['name'].get().strip()
            cat      = self._vars['category'].get()
            side     = self._vars['side'].get()
            qty_s    = self._vars['qty'].get().strip()
            price_s  = self._vars['price'].get().strip()
            fee_s    = self._vars['fee'].get().strip()
            tax_s    = self._vars['tax'].get().strip()

            if not code or not name:
                raise ValueError('請填寫股票代號和名稱')
            if not qty_s or not price_s:
                raise ValueError('請填寫數量和價格')

            tx_date = datetime.strptime(date_str, '%Y-%m-%d')
            qty, price = float(qty_s), float(price_s)
            if self._lot_var.get():
                qty *= 1000          # 整張：1 張 = 1000 股
            if qty <= 0 or price <= 0:
                raise ValueError('數量和價格必須大於 0')

            gross = qty * price
            fee   = float(fee_s) if fee_s else max(20, round(gross * FEE_RATE))
            tax   = float(tax_s) if tax_s else (round(gross * TAX_RATE) if side == '賣' else 0)
            net   = round(gross - fee - tax) if side == '賣' else -round(gross + fee)

            save_row({
                '日期': tx_date, '股票代號': code, '股票名稱': name,
                '分類': cat, '買賣': side,
                '數量(股)': int(qty), '價格(元)': price,
                '手續費(元)': int(fee), '交易稅(元)': tax, '淨金額(元)': net,
            })
            self._refresh_table()
            self._refresh_stock_options()
            messagebox.showinfo('新增成功',
                f'[{cat}]  {side} {code} {name}\n'
                f'{int(qty):,} 股 @ {price:.2f} 元\n'
                f'手續費 {fee:,.0f}  稅 {tax:,}  淨額 {net:,}')
            for k in ('qty', 'price', 'fee', 'tax'):
                self._vars[k].set('')
            self._status_var.set('')

        except ValueError as e:
            messagebox.showerror('輸入錯誤', str(e))
        except Exception as e:
            messagebox.showerror('錯誤', str(e))

    def _show_ctx_menu(self, event):
        item = self._tree.identify_row(event.y)
        if item:
            self._tree.selection_set(item)
            self._ctx_menu.post(event.x_root, event.y_root)

    def _sort_by_col(self, col: str):
        if self._sort_col == col:
            self._sort_asc = not self._sort_asc
        else:
            self._sort_col = col
            self._sort_asc = True
        self._refresh_table()

    def _refresh_table(self):
        for item in self._tree.get_children():
            self._tree.delete(item)
        df = load_df()

        # 同步清空損益子表格
        for item in self._tree_pnl.get_children():
            self._tree_pnl.delete(item)

        # 更新欄位標題排序指示符
        for c in COLUMNS:
            label = c + ('  ▲' if self._sort_asc else '  ▼') if c == self._sort_col else c
            self._tree.heading(c, text=label)

        # 依選取欄位排序（保留 iid = 原始行索引，供 edit/delete 使用）
        if self._sort_col and self._sort_col in df.columns:
            df = df.sort_values(self._sort_col, ascending=self._sort_asc,
                                kind='stable')

        # 股票名稱欄自動寬度
        if not df.empty:
            max_chars = df['股票名稱'].astype(str).str.len().max()
            name_w = max(78, min(160, int(max_chars * 10) + 10))
            self._tree.column('股票名稱', width=name_w)

        for pos, (orig_idx, r) in enumerate(df.iterrows()):
            row_tag = 'row_even' if pos % 2 == 0 else 'row_odd'
            vals = [
                r['日期'].strftime('%Y-%m-%d'),
                r['股票代號'], r['股票名稱'],
                r.get('分類', ''),
                r['買賣'],
                f"{int(r['數量(股)']):,}",
                f"{r['價格(元)']:.2f}",
                f"{int(r['手續費(元)']):,}",
                f"{int(r['交易稅(元)']):,}",
                f"{int(r['淨金額(元)']):,}",
            ]
            side_tag = 'buy' if r['買賣'] == '買' else 'sell'
            self._tree.insert('', 'end', iid=str(orig_idx), values=vals,
                              tags=(side_tag, row_tag))
            self._tree_pnl.insert('', 'end', iid=str(orig_idx), values=('—', '—'),
                                  tags=(row_tag,))

        # 背景執行緒非同步抓即時報價並填入損益
        self._pnl_gen += 1
        gen = self._pnl_gen
        meta = df[['股票代號', '買賣', '價格(元)', '數量(股)']].copy()
        threading.Thread(target=self._bg_fill_pnl,
                         args=(meta, gen), daemon=True).start()

    def _bg_fill_pnl(self, meta_df, gen):
        prices = {}
        for code in meta_df['股票代號'].unique():
            p, _ = get_price(str(code))
            prices[str(code)] = p
        self._ui_call(lambda: self._apply_pnl_to_table(meta_df, prices, gen))

    def _apply_pnl_to_table(self, meta_df, prices, gen):
        if gen != self._pnl_gen:
            return  # stale update, discard
        for orig_idx, r in meta_df.iterrows():
            iid = str(orig_idx)
            if not self._tree_pnl.exists(iid):
                continue
            if r['買賣'] != '買':
                continue
            cur = prices.get(str(r['股票代號']))
            if not cur:
                continue
            trade_p = float(r['價格(元)'])
            qty = float(r['數量(股)'])
            pnl_amt = (cur - trade_p) * qty
            pnl_pct = (cur / trade_p - 1) * 100 if trade_p else 0
            sign = '+' if pnl_amt >= 0 else ''
            pnl_tag = 'pnl_gain' if pnl_amt >= 0 else 'pnl_loss'
            self._tree_pnl.item(iid,
                                values=(f'{sign}{pnl_amt:,.0f}', f'{sign}{pnl_pct:.2f}%'),
                                tags=(pnl_tag,))

    def _delete_selected(self):
        sel = self._tree.selection()
        if not sel:
            messagebox.showinfo('提示', '請先選取一筆紀錄')
            return
        idx  = int(sel[0])
        vals = self._tree.item(sel[0], 'values')
        date_str, code, name = vals[0], vals[1], vals[2]
        if not messagebox.askyesno('確認刪除',
                f'確定要刪除以下紀錄？\n\n'
                f'{date_str}  {code} {name}  {vals[4]}  {vals[5]}股 @ {vals[6]}元',
                icon='warning'):
            return
        df = load_df()
        df = df.drop(index=idx).reset_index(drop=True)
        rewrite_excel(df)
        self._refresh_table()

    def _edit_selected(self):
        sel = self._tree.selection()
        if not sel:
            messagebox.showinfo('提示', '請先選取一筆紀錄')
            return
        idx = int(sel[0])
        df  = load_df()
        row = df.iloc[idx]
        EditDialog(self, idx, row, on_save=self._on_edit_save)

    def _on_edit_save(self, idx: int, new_row: dict):
        df = load_df()
        for col, val in new_row.items():
            df.at[idx, col] = val
        rewrite_excel(df)
        self._refresh_table()
        self._refresh_stock_options()

    # ═══════════════════════════════════════════════════════════════════════════
    # Tab 2：庫存樹狀圖（finviz 風格）
    # ═══════════════════════════════════════════════════════════════════════════
    def _build_tab2(self):
        f = self.tab2

        ctrl = ttk.Frame(f)
        ctrl.pack(fill='x', padx=14, pady=8)
        ttk.Label(ctrl, text='庫存現值分佈', style='Hdr.TLabel').pack(side='left')
        self._tm_status = tk.StringVar(value='切換至此頁面自動更新 | 也可手動點擊更新')
        ttk.Label(ctrl, textvariable=self._tm_status,
                  foreground=C_FG2, font=('Microsoft JhengHei', 9)).pack(side='left', padx=14)
        ttk.Button(ctrl, text='💾  另存圖檔', style='Nav.TButton', command=self._save_treemap).pack(side='right', padx=(4, 0))
        ttk.Button(ctrl, text='🔄  更新報價', style='Nav.TButton', command=self._draw_treemap).pack(side='right')

        # ── 整體庫存摘要列 ────────────────────────────────────────────────────
        summary_bar = tk.Frame(f, bg=C_PANEL)
        summary_bar.pack(fill='x', padx=8, pady=(0, 4))

        card = tk.Frame(summary_bar, bg='#1a1a2e', padx=16, pady=10)
        card.pack(fill='x')
        self._summary_card = card

        # 第一行：標題 + 持股檔數
        _top = tk.Frame(card, bg='#1a1a2e')
        _top.pack(fill='x')
        tk.Label(_top, text='庫存總市值', bg='#1a1a2e',
                 fg='#6a8faf', font=('Microsoft JhengHei', 13)).pack(side='left')
        self._sv_count = tk.StringVar(value='—')
        tk.Label(_top, textvariable=self._sv_count, bg='#1a1a2e',
                 fg='#6a8faf', font=('Microsoft JhengHei', 11)).pack(side='right')

        # 市值 + 右側三項並排
        _mid = tk.Frame(card, bg='#1a1a2e')
        _mid.pack(fill='x', pady=(4, 0))

        # 左：大字總市值
        self._sv_mktval = tk.StringVar(value='—')
        tk.Label(_mid, textvariable=self._sv_mktval, bg='#1a1a2e',
                 fg=C_FG, font=('Microsoft JhengHei', 30, 'bold')).pack(side='left', padx=(0, 24))

        # 右：損益試算 / 報酬率 / 總付出成本
        def _sub(parent, title):
            fr = tk.Frame(parent, bg='#1a1a2e', padx=16)
            fr.pack(side='left', anchor='center')
            tk.Label(fr, text=title, bg='#1a1a2e',
                     fg='#6a8faf', font=('Microsoft JhengHei', 9)).pack(anchor='w')
            var = tk.StringVar(value='—')
            lbl = tk.Label(fr, textvariable=var, bg='#1a1a2e',
                           fg=C_FG, font=('Microsoft JhengHei', 16, 'bold'))
            lbl.pack(anchor='w')
            return var, lbl

        self._sv_pnl,    self._sl_pnl    = _sub(_mid, '損益試算')
        self._sv_pnlpct, self._sl_pnlpct = _sub(_mid, '報酬率')
        self._sv_cost,   self._sl_cost   = _sub(_mid, '總付出成本')

        self._tm_fig = plt.Figure(figsize=(9.5, 5.5), dpi=100, facecolor='#111111')
        self._tm_canvas = FigureCanvasTkAgg(self._tm_fig, master=f)
        w = self._tm_canvas.get_tk_widget()
        w.configure(bg='#111111')
        w.pack(fill='both', expand=True, padx=8, pady=(0, 8))

        self._tm_rects      = []      # stock block rects for hover detection
        self._tm_cat_rects  = []      # category rects for hover detection
        self._tm_tooltip    = None    # tooltip Toplevel
        self._tm_ax         = None    # saved axes reference for coord transform
        self._tm_drill_cat  = None    # None=全覽, str=鑽入的分類名稱
        self._tm_fetching   = False   # background fetch guard
        self._tm_refresh_job = None   # after() job id for auto-refresh
        self._tm_canvas.mpl_connect('motion_notify_event', self._on_tm_motion)
        self._tm_canvas.mpl_connect('button_press_event',  self._on_tm_click)
        self._tm_canvas.mpl_connect('figure_leave_event',  lambda e: self._hide_tooltip())

    # ── treemap click: 鑽入分類 / 返回全覽 / 右鍵開個股分析 ──────────────────────
    def _on_tm_click(self, event):
        if event.inaxes is None or self._tm_ax is None:
            return
        xd, yd = event.xdata, event.ydata

        # 右鍵：在個股方塊上點擊 → 開啟個股分析
        if event.button == 3:
            for r in self._tm_rects:
                if (r['rx'] <= xd <= r['rx'] + r['rw'] and
                        r['ry'] <= yd <= r['ry'] + r['rh']):
                    self._open_stk_analysis(r['code'], r['name'])
                    return
            return

        # 鑽入模式：左鍵點任意個股區塊返回全覽
        if self._tm_drill_cat is not None:
            self._tm_drill_cat = None
            self._draw_treemap()
            return

        # 全覽模式：點任意個股方塊或分類區域 → 進入該分類鑽入
        # 先嘗試個股方塊（找出所屬分類）
        for r in self._tm_rects:
            if (r['rx'] <= xd <= r['rx'] + r['rw'] and
                    r['ry'] <= yd <= r['ry'] + r['rh']):
                self._tm_drill_cat = r.get('category')
                self._draw_treemap()
                return

        # 再嘗試分類標題列
        for c in self._tm_cat_rects:
            hdr_y = c['cy'] + c['ch'] - c['hdr_h']
            if (c['cx'] <= xd <= c['cx'] + c['cw'] and
                    hdr_y <= yd <= c['cy'] + c['ch']):
                self._tm_drill_cat = c['cat_name']
                self._draw_treemap()
                return

    # ── treemap hover helpers ──────────────────────────────────────────────────
    def _on_tm_motion(self, event):
        if event.inaxes is None or self._tm_ax is None:
            self._hide_tooltip()
            return
        xd, yd = event.xdata, event.ydata
        widget = self._tm_canvas.get_tk_widget()
        cx = widget.winfo_rootx() + int(event.x)
        cy = widget.winfo_rooty() + int(widget.winfo_height() - event.y)

        # 先嘗試股票方塊
        for r in self._tm_rects:
            if r['rx'] <= xd <= r['rx'] + r['rw'] and r['ry'] <= yd <= r['ry'] + r['rh']:
                self._show_tooltip(cx, cy, r)
                return

        # 再嘗試分類區域
        for c in self._tm_cat_rects:
            if c['cx'] <= xd <= c['cx'] + c['cw'] and c['cy'] <= yd <= c['cy'] + c['ch']:
                self._show_cat_tooltip(cx, cy, c)
                return

        self._hide_tooltip()

    def _show_tooltip(self, sx, sy, data):
        expected_keys = {'title', 'price', 'qty', 'mktval', 'ratio', 'avgcost', 'pnl'}
        needs_rebuild = (self._tm_tooltip is None or not self._tm_tooltip.winfo_exists()
                         or not expected_keys.issubset(self._tm_tip_widgets))
        if needs_rebuild:
            if self._tm_tooltip and self._tm_tooltip.winfo_exists():
                self._tm_tooltip.destroy()
            tip = tk.Toplevel(self)
            tip.overrideredirect(True)
            tip.attributes('-topmost', True)
            tip.configure(bg='#1e1e1e')
            # thin border frame
            border = tk.Frame(tip, bg='#3e3e3e', padx=1, pady=1)
            border.pack(fill='both', expand=True)
            inner = tk.Frame(border, bg='#252526', padx=10, pady=8)
            inner.pack(fill='both', expand=True)
            self._tm_tip_widgets = {}
            for key in ('title', 'price', 'qty', 'mktval', 'ratio', 'avgcost', 'pnl'):
                lbl = tk.Label(inner, bg='#252526',
                               font=('Microsoft JhengHei', 10), anchor='w')
                lbl.pack(fill='x')
                self._tm_tip_widgets[key] = lbl
            self._tm_tooltip = tip

        w = self._tm_tip_widgets
        code  = data['code']
        name  = data['name']
        cat   = data['category']
        price = data['price']
        qty   = data['qty']
        avg   = data['avg_cost']
        val   = data['value']
        pct   = data['pnl_pct']
        pnl_a = (price - avg) * qty if price else None

        w['title'].config(text=f'{code}  {name}  ({cat})',
                          fg='#8ab4d4', font=('Microsoft JhengHei', 11, 'bold'))
        w['price'].config(text=f'現價：{"—" if not price else f"{price:,.2f} 元"}',
                          fg='#cccccc')
        w['qty'].config(text=f'持股：{qty:,.0f} 股', fg='#cccccc')
        w['mktval'].config(text=f'市值：{val:,.0f} 元', fg='#cccccc')
        portfolio_pct = data.get('portfolio_pct', 0.0)
        w['ratio'].config(text=f'占比：{portfolio_pct:.2f}%', fg='#cccccc')
        w['avgcost'].config(text=f'均價：{avg:,.2f} 元', fg='#cccccc')
        pnl_color_tip = '#f07070' if pct >= 0 else '#4ec94e'
        sign = '+' if pct >= 0 else ''
        pnl_txt = (f'損益：{sign}{pnl_a:,.0f} 元  ({sign}{pct:.2f}%)'
                   if pnl_a is not None else '損益：—')
        w['pnl'].config(text=pnl_txt, fg=pnl_color_tip)

        # keep tooltip on screen
        self._place_tooltip(self._tm_tooltip, sx, sy)
        self._tm_tooltip.deiconify()

    def _save_treemap(self):
        from tkinter import filedialog
        import io
        from PIL import Image

        path = filedialog.asksaveasfilename(
            title='儲存庫存圖',
            defaultextension='.png',
            filetypes=[('PNG 圖片', '*.png'), ('JPEG 圖片', '*.jpg'), ('所有檔案', '*.*')],
            initialfile='treemap.png')
        if not path:
            return

        # 1. 取得樹狀圖（matplotlib figure → PIL），決定輸出寬度
        buf = io.BytesIO()
        self._tm_fig.savefig(buf, format='png', dpi=150, bbox_inches='tight',
                             facecolor=self._tm_fig.get_facecolor())
        buf.seek(0)
        img_chart = Image.open(buf).copy()
        W = img_chart.width

        # 2. 用 PIL 繪製 summary card
        from PIL import ImageDraw, ImageFont
        PAD, CARD_H = 28, 110
        BG      = (26, 26, 46)    # #1a1a2e
        FG      = (220, 220, 220)
        SUB_FG  = (106, 143, 175) # #6a8faf

        pnl_val = self._sv_pnl.get()
        pnl_fg  = (240, 112, 112) if self._sl_pnl.cget('fg') == '#f07070' else (78, 201, 78)

        img_card = Image.new('RGB', (W, CARD_H), BG)
        draw = ImageDraw.Draw(img_card)

        # 嘗試載入系統字型，fallback 用預設
        def _font(size, bold=False):
            for name in (['msjhbd.ttc'] if bold else ['msjh.ttc', 'msjhbd.ttc']):
                for folder in [r'C:\Windows\Fonts']:
                    fp = os.path.join(folder, name)
                    if os.path.exists(fp):
                        try: return ImageFont.truetype(fp, size)
                        except: pass
            return ImageFont.load_default()

        f_title  = _font(15)
        f_count  = _font(13)
        f_val    = _font(36, bold=True)
        f_sublbl = _font(11)
        f_subval = _font(20, bold=True)

        # 標題行
        draw.text((PAD, 10),      '庫存總市值',            font=f_title,  fill=SUB_FG)
        draw.text((W - PAD, 10),  self._sv_count.get(),   font=f_count,  fill=SUB_FG, anchor='ra')

        # 大字市值
        draw.text((PAD, 30), self._sv_mktval.get(), font=f_val, fill=FG)

        # 三項：損益試算 / 報酬率 / 總付出成本
        items = [
            ('損益試算',  pnl_val,                pnl_fg),
            ('報酬率',    self._sv_pnlpct.get(),  pnl_fg),
            ('總付出成本', self._sv_cost.get(),    FG),
        ]
        col_w = (W - PAD * 2) // 3
        for i, (lbl, val, fg) in enumerate(items):
            x = PAD + i * col_w
            draw.text((x, 72), lbl, font=f_sublbl, fill=SUB_FG)
            draw.text((x, 86), val, font=f_subval, fill=fg)

        # 3. 垂直合併並儲存
        combined = Image.new('RGB', (W, CARD_H + img_chart.height))
        combined.paste(img_card,  (0, 0))
        combined.paste(img_chart, (0, CARD_H))
        combined.save(path)
        self._tm_status.set(f'已儲存：{path}')

    def _save_analysis(self):
        from tkinter import filedialog
        sel = self._stock_var.get().strip()
        default_name = sel.split()[0] if sel and not sel.startswith('──') else 'analysis'
        path = filedialog.asksaveasfilename(
            title='儲存個股分析圖',
            defaultextension='.png',
            filetypes=[('PNG 圖片', '*.png'), ('JPEG 圖片', '*.jpg'), ('所有檔案', '*.*')],
            initialfile=f'{default_name}.png')
        if path:
            self._an_fig.savefig(path, dpi=150, bbox_inches='tight',
                                 facecolor=self._an_fig.get_facecolor())

    def _show_cat_tooltip(self, sx, sy, cat):
        expected_keys = {'title', 'price', 'qty', 'mktval', 'ratio', 'avgcost', 'pnl'}
        needs_rebuild = (self._tm_tooltip is None or not self._tm_tooltip.winfo_exists()
                         or not expected_keys.issubset(self._tm_tip_widgets))
        if needs_rebuild:
            if self._tm_tooltip and self._tm_tooltip.winfo_exists():
                self._tm_tooltip.destroy()
            tip = tk.Toplevel(self)
            tip.overrideredirect(True)
            tip.attributes('-topmost', True)
            tip.configure(bg='#1e1e1e')
            border = tk.Frame(tip, bg='#3e3e3e', padx=1, pady=1)
            border.pack(fill='both', expand=True)
            inner = tk.Frame(border, bg='#252526', padx=10, pady=8)
            inner.pack(fill='both', expand=True)
            self._tm_tip_widgets = {}
            for key in ('title', 'price', 'qty', 'mktval', 'ratio', 'avgcost', 'pnl'):
                lbl = tk.Label(inner, bg='#252526',
                               font=('Microsoft JhengHei', 10), anchor='w')
                lbl.pack(fill='x')
                self._tm_tip_widgets[key] = lbl
            self._tm_tooltip = tip

        w    = self._tm_tip_widgets
        sign = '+' if cat['pnl_amt'] >= 0 else ''
        pnl_fg = '#f07070' if cat['pnl_amt'] >= 0 else '#4ec94e'
        w['title'].config(text=cat['cat_name'],
                          fg='#aecde8', font=('Microsoft JhengHei', 11, 'bold'))
        w['price'].config(text=f"持股標的：{cat['count']} 檔", fg='#cccccc')
        w['qty'].config(  text=f"市值：{cat['value']:,.0f} 元",  fg='#cccccc')
        w['mktval'].config(text=f"成本：{cat['cost']:,.0f} 元",  fg='#cccccc')
        w['ratio'].config(text='', fg='#cccccc')
        w['avgcost'].config(text='', fg='#cccccc')
        w['pnl'].config(
            text=f"損益：{sign}{cat['pnl_amt']:,.0f} 元  ({sign}{cat['pnl_pct']:.2f}%)",
            fg=pnl_fg)
        self._place_tooltip(self._tm_tooltip, sx, sy)
        self._tm_tooltip.deiconify()

    def _place_tooltip(self, tip: tk.Toplevel, cursor_x: int, cursor_y: int,
                       offset: int = 16) -> None:
        """將 tooltip 定位在游標右下方；若超出螢幕右/下邊界則改到左/上方。
        多螢幕延伸時，以游標所在螢幕的邊界為準。"""
        tip.update_idletasks()
        tw = tip.winfo_width()
        th = tip.winfo_height()

        # 預設：主螢幕邊界
        mon_left, mon_top = 0, 0
        mon_right  = self.winfo_screenwidth()
        mon_bottom = self.winfo_screenheight()

        # 用 ctypes 定義正確的 MONITORINFO 結構，取得游標所在螢幕工作區
        try:
            import ctypes, ctypes.wintypes

            class _RECT(ctypes.Structure):
                _fields_ = [('left',   ctypes.c_long), ('top',    ctypes.c_long),
                            ('right',  ctypes.c_long), ('bottom', ctypes.c_long)]

            class _MONITORINFO(ctypes.Structure):
                _fields_ = [('cbSize',    ctypes.c_ulong),
                            ('rcMonitor', _RECT),
                            ('rcWork',    _RECT),
                            ('dwFlags',   ctypes.c_ulong)]

            pt    = ctypes.wintypes.POINT(cursor_x, cursor_y)
            hmon  = ctypes.windll.user32.MonitorFromPoint(pt, 2)   # DEFAULTTONEAREST
            minfo = _MONITORINFO()
            minfo.cbSize = ctypes.sizeof(_MONITORINFO)
            ctypes.windll.user32.GetMonitorInfoW(hmon, ctypes.byref(minfo))
            mon_left   = minfo.rcWork.left
            mon_top    = minfo.rcWork.top
            mon_right  = minfo.rcWork.right
            mon_bottom = minfo.rcWork.bottom
        except Exception:
            pass

        x = cursor_x + offset
        y = cursor_y + offset
        if x + tw > mon_right:
            x = cursor_x - tw - offset
        if y + th > mon_bottom:
            y = cursor_y - th - offset
        x = max(mon_left, x)
        y = max(mon_top,  y)
        tip.geometry(f'+{x}+{y}')

    def _hide_tooltip(self):
        if self._tm_tooltip and self._tm_tooltip.winfo_exists():
            self._tm_tooltip.withdraw()

    _TM_REFRESH_MS = 30_000  # 30 秒自動更新間隔

    def _draw_treemap(self):
        if self._tm_fetching:
            return
        self._tm_fetching = True
        self._tm_status.set('更新中…')
        import threading as _th
        _th.Thread(target=self._draw_treemap_bg, daemon=True).start()

    def _draw_treemap_bg(self):
        try:
            self._draw_treemap_impl()
        except Exception as e:
            import traceback; traceback.print_exc()
            self._ui_call(lambda: self._tm_status.set(f'繪製錯誤：{e}'))
        finally:
            self._tm_fetching = False

    def _schedule_tm_refresh(self):
        self._cancel_tm_refresh()
        self._tm_refresh_job = self.after(self._TM_REFRESH_MS, self._on_tm_auto_refresh)

    def _cancel_tm_refresh(self):
        if self._tm_refresh_job:
            self.after_cancel(self._tm_refresh_job)
            self._tm_refresh_job = None

    def _on_tm_auto_refresh(self):
        self._tm_refresh_job = None
        if self._current_page == 0:
            self._draw_treemap()
            self._schedule_tm_refresh()

    def _draw_treemap_impl(self):
        TREE_BG = '#111111'
        SEP_COL = '#111111'

        df       = load_df()
        holdings = calc_holdings(df)

        self._tm_rects     = []
        self._tm_cat_rects = []
        self._hide_tooltip()
        self._tm_fig.clear()
        self._tm_fig.patch.set_facecolor(TREE_BG)

        # 主畫布（保留底部 5% 給狀態列）
        ax = self._tm_fig.add_axes([0, 0.04, 1, 0.96])
        ax.set_facecolor(TREE_BG)
        ax.set_xlim(0, 100)
        ax.set_ylim(0, 100)
        ax.axis('off')
        self._tm_ax = ax

        # 底部狀態軸
        sax = self._tm_fig.add_axes([0, 0, 1, 0.04])
        sax.set_facecolor('#0d0d0d')
        sax.axis('off')

        if not holdings:
            for sv in (self._sv_count, self._sv_mktval, self._sv_cost,
                       self._sv_pnl, self._sv_pnlpct):
                sv.set('—')
            ax.text(50, 50, '目前沒有庫存', ha='center', va='center',
                    color='#888', fontsize=14, fontfamily=CHART_FONT)
            self._tm_canvas.draw()
            return

        # 批次抓取 MIS 即時報價（tse/otc 各試一次），找不到的 fallback 到 yfinance
        _uniq_codes = list({code for (code, _) in holdings.keys()})
        _mis_pairs  = [(c, ex) for c in _uniq_codes for ex in ('tse', 'otc')]
        try:
            _mis_prices = self._fetch_mis_prices(_mis_pairs)
        except Exception:
            _mis_prices = {}

        # 依分類分組
        groups: dict[str, list] = {}
        no_price_list: list[str] = []
        total_val = 0.0

        for (code, _cat), h in holdings.items():
            if code in _mis_prices:
                p = _mis_prices[code][0]
            else:
                p, _ = get_price(code)
            cat  = h.get('category', '未分類')
            avg  = h['avg_cost']
            if p:
                val = h['qty'] * p
                pnl = (p - avg) / avg * 100 if avg else 0.0
            else:
                val = h['total_cost']
                pnl = 0.0
                no_price_list.append(code)
            total_val += val
            groups.setdefault(cat, []).append({
                'code': code, 'name': h['name'],
                'value': val, 'pnl_pct': pnl, 'price': p,
                'avg_cost': avg, 'qty': h['qty'],
                'total_cost': h['total_cost'],
            })

        # ── 更新摘要列 ────────────────────────────────────────────────────────
        total_cost    = sum(h['total_cost'] for h in holdings.values())
        pnl_amt       = total_val - total_cost
        pnl_pct_total = pnl_amt / total_cost * 100 if total_cost else 0
        pnl_fg        = '#f07070' if pnl_amt >= 0 else '#4ec94e'
        sign          = '+' if pnl_amt >= 0 else ''
        self._sv_count .set(f'持股標的  {len(holdings)} 檔')
        self._sv_mktval.set(f'{total_val:,.0f} 元')
        self._sv_cost  .set(f'{total_cost:,.0f} 元')
        self._sv_pnl   .set(f'{sign}{pnl_amt:,.0f} 元')
        self._sv_pnlpct.set(f'{sign}{pnl_pct_total:.2f}%')
        self._sl_pnl   .config(fg=pnl_fg)
        self._sl_pnlpct.config(fg=pnl_fg)
        self._sl_cost  .config(fg=C_FG)

        # 動態計算報酬率範圍，作為顏色插值基準
        import math as _math
        all_pcts = [s['pnl_pct'] for stocks in groups.values() for s in stocks
                    if not _math.isnan(s['pnl_pct'])]
        _max_abs = max((abs(p) for p in all_pcts), default=10.0) or 10.0

        # 各分類依市值降冪排列
        for cat in groups:
            groups[cat].sort(key=lambda x: -x['value'])

        # Pixel conversion factors for font sizing (needed by both drill-down and full view)
        _dpi       = self._tm_fig.dpi
        _fig_w_px  = self._tm_fig.get_size_inches()[0] * _dpi
        _fig_h_px  = self._tm_fig.get_size_inches()[1] * _dpi
        _px_per_ux = _fig_w_px / 100
        _px_per_uy = _fig_h_px * 0.96 / 100
        _pt_per_px = 72 / _dpi
        GAP = 0.35

        # ── 鑽入模式：只顯示單一分類 ──────────────────────────────────────────
        if self._tm_drill_cat and self._tm_drill_cat in groups:
            _drill_name  = self._tm_drill_cat
            _drill_stocks = groups[_drill_name]

            # 返回提示列（頂部 5%）
            BACK_H = 5.5
            back_y = 100 - BACK_H
            ax.add_patch(plt.Rectangle((0, back_y), 100, BACK_H,
                facecolor='#1a253a', edgecolor='none', zorder=10))
            ax.text(2, back_y + BACK_H / 2,
                    '← 點擊任意處返回全覽',
                    ha='left', va='center', color='#6a9fcf',
                    fontsize=9, fontfamily=CHART_FONT, zorder=11)
            _drill_cat_val = sum(s['value'] for s in _drill_stocks)
            _drill_pct = _drill_cat_val / total_val * 100 if total_val else 0
            ax.text(50, back_y + BACK_H / 2,
                    f'{_drill_name}  {_drill_pct:.1f}%',
                    ha='center', va='center', color='#aecde8',
                    fontsize=11, fontweight='bold', fontfamily=CHART_FONT, zorder=11)

            # 股票佔滿剩餘區域
            _di_y = 0
            _di_h = back_y
            _di_x = 0
            _di_w = 100
            _stock_vals = [s['value'] for s in _drill_stocks]
            if _stock_vals and _di_h > 0:
                _norm = squarify.normalize_sizes(_stock_vals, _di_w, _di_h)
                _raw  = squarify.squarify(_norm, _di_x, _di_y, _di_w, _di_h)
                _srects = [{'x': r['x'],
                             'y': _di_y + _di_h - (r['y'] - _di_y) - r['dy'],
                             'dx': r['dx'], 'dy': r['dy']} for r in _raw]
                for _stk, _sr in zip(_drill_stocks, _srects):
                    _sx, _sy, _sw, _sh = _sr['x'], _sr['y'], _sr['dx'], _sr['dy']
                    _rx, _ry = _sx + GAP / 2, _sy + GAP / 2
                    _rw, _rh = _sw - GAP, _sh - GAP
                    if _rw <= 0 or _rh <= 0:
                        continue
                    ax.add_patch(plt.Rectangle(
                        (_rx, _ry), _rw, _rh,
                        facecolor=pnl_color(_stk['pnl_pct'], _max_abs),
                        edgecolor=SEP_COL, linewidth=0.4, zorder=1))
                    self._tm_rects.append({
                        'rx': _rx, 'ry': _ry, 'rw': _rw, 'rh': _rh,
                        'code':     _stk['code'],  'name':     _stk['name'],
                        'value':    _stk['value'], 'price':    _stk['price'],
                        'pnl_pct':  _stk['pnl_pct'], 'avg_cost': _stk['avg_cost'],
                        'qty':      _stk['qty'],   'category': _drill_name,
                        'portfolio_pct': _stk['value'] / _drill_cat_val * 100 if _drill_cat_val else 0,
                    })
                    _min_dim = min(_rw, _rh)
                    if _min_dim < 2:
                        continue
                    _rw_px     = _rw * _px_per_ux
                    _rh_px     = _rh * _px_per_uy
                    _chars     = max(len(_stk['code']), 4)
                    _name_chars = max(len(_stk['name']), 2)
                    _fs_by_w  = 0.58 * _rw_px * _pt_per_px / (_chars * 0.60)
                    _fs_by_h  = 0.65 * _rh_px * _pt_per_px / 3.6
                    _fs_code  = max(min(_fs_by_w, _fs_by_h, 44), 6)
                    _fs_name  = max(min(0.80 * _rw_px * _pt_per_px / (_name_chars * 0.95),
                                        _fs_code * 0.55, 22), 5)
                    _fs_sub   = max(_fs_code * 0.55, 5)
                    _pnl_str  = f"{'+' if _stk['pnl_pct'] >= 0 else ''}{_stk['pnl_pct']:.2f}%"
                    _cx_t     = _rx + _rw / 2
                    _mid_y    = _ry + _rh * 0.50
                    _fs_code_uy = _fs_code / _pt_per_px / _px_per_uy
                    _fs_name_uy = _fs_name / _pt_per_px / _px_per_uy
                    _fs_sub_uy  = _fs_sub  / _pt_per_px / _px_per_uy
                    if _min_dim >= 8:
                        _gap_cn = (_fs_code_uy + _fs_name_uy) * 0.55
                        _gap_np = (_fs_name_uy + _fs_sub_uy)  * 0.55
                        _half   = (_gap_cn + _gap_np) / 2
                        ax.text(_cx_t, _mid_y + _half, _stk['code'],
                                ha='center', va='center', color='white',
                                fontsize=_fs_code, fontweight='bold',
                                fontfamily=CHART_FONT, clip_on=True, zorder=4)
                        ax.text(_cx_t, _mid_y + _half - _gap_cn, _stk['name'],
                                ha='center', va='center', color='#dddddd',
                                fontsize=_fs_name, fontfamily=CHART_FONT,
                                clip_on=True, zorder=4)
                        ax.text(_cx_t, _mid_y - _half, _pnl_str,
                                ha='center', va='center', color='white',
                                fontsize=_fs_sub, fontfamily=CHART_FONT,
                                clip_on=True, zorder=4)
                    elif _min_dim >= 5:
                        _gap = (_fs_code_uy + _fs_sub_uy) * 0.55
                        ax.text(_cx_t, _mid_y + _gap * 0.5, _stk['code'],
                                ha='center', va='center', color='white',
                                fontsize=_fs_code, fontweight='bold',
                                fontfamily=CHART_FONT, clip_on=True, zorder=4)
                        ax.text(_cx_t, _mid_y - _gap * 0.5, _pnl_str,
                                ha='center', va='center', color='white',
                                fontsize=_fs_sub, fontfamily=CHART_FONT,
                                clip_on=True, zorder=4)
                    else:
                        ax.text(_cx_t, _mid_y, _stk['code'],
                                ha='center', va='center', color='white',
                                fontsize=_fs_code, fontweight='bold',
                                fontfamily=CHART_FONT, clip_on=True, zorder=4)

            # 同步摘要列為此分類數據
            _d_val  = sum(s['value']      for s in _drill_stocks)
            _d_cost = sum(s['total_cost'] for s in _drill_stocks)
            _d_pnl  = _d_val - _d_cost
            _d_pct  = _d_pnl / _d_cost * 100 if _d_cost else 0
            _d_sign = '+' if _d_pnl >= 0 else ''
            _d_fg   = '#f07070' if _d_pnl >= 0 else '#4ec94e'
            self._sv_count .set(f'持股標的  {len(_drill_stocks)} 檔')
            self._sv_mktval.set(f'{_d_val:,.0f} 元')
            self._sv_cost  .set(f'{_d_cost:,.0f} 元')
            self._sv_pnl   .set(f'{_d_sign}{_d_pnl:,.0f} 元')
            self._sv_pnlpct.set(f'{_d_sign}{_d_pct:.2f}%')
            self._sl_pnl   .config(fg=_d_fg)
            self._sl_pnlpct.config(fg=_d_fg)
            self._sl_cost  .config(fg=C_FG)

            warn = f'  ⚠ {", ".join(no_price_list)} 無即時報價' if no_price_list else ''
            sax.text(0.5, 0.5,
                     f'{_drill_name}：{len(_drill_stocks)} 檔  ·  '
                     f'市值 {_d_val:,.0f} 元{warn}',
                     ha='center', va='center', color='#888',
                     fontsize=8, fontfamily=CHART_FONT, transform=sax.transAxes)
            self._tm_canvas.draw()
            return

        cat_names  = list(groups.keys())
        cat_values = [sum(s['value'] for s in groups[c]) for c in cat_names]

        norm_cats  = squarify.normalize_sizes(cat_values, 100, 100)
        _cat_raw   = squarify.squarify(norm_cats, 0, 0, 100, 100)
        # 翻轉 y：squarify 從 y=0（畫面底部）填起，翻轉後最大方塊在左上
        cat_rects  = [{'x': r['x'], 'y': 100 - r['y'] - r['dy'],
                        'dx': r['dx'], 'dy': r['dy']} for r in _cat_raw]

        for cat_name, crect in zip(cat_names, cat_rects):
            cx, cy, cw, ch = crect['x'], crect['y'], crect['dx'], crect['dy']
            stocks = groups[cat_name]

            # 分類標題列（實色橫條，Finviz 風格）
            # 以像素為單位決定高度，再換算回 data-unit，字體隨之縮放
            _target_hdr_px = 20                          # 希望的標題高度 px
            _max_hdr_frac  = 0.40                        # 最多佔分類高度的 40%
            hdr_h_px = min(_target_hdr_px, ch * _px_per_uy * _max_hdr_frac)
            hdr_h    = hdr_h_px / _px_per_uy
            hdr_y    = cy + ch - hdr_h
            fs_hdr   = min(10, max(6, hdr_h_px * _pt_per_px * 0.62))
            ax.add_patch(plt.Rectangle(
                (cx, hdr_y), cw, hdr_h,
                facecolor='#1a1a2e', edgecolor='none', linewidth=0, zorder=6, clip_on=True))
            cat_pct = sum(s['value'] for s in stocks) / total_val * 100 if total_val else 0
            ax.text(cx + 0.8, hdr_y + hdr_h / 2,
                    f'{cat_name}  {cat_pct:.1f}%',
                    ha='left', va='center', color='#aecde8',
                    fontsize=fs_hdr, fontweight='bold', fontfamily=CHART_FONT,
                    clip_on=True, zorder=7)

            # 分類外框
            ax.add_patch(plt.Rectangle(
                (cx, cy), cw, ch,
                facecolor='none', edgecolor='#2a2a2a', linewidth=2.0, zorder=8))

            # 股票填滿標題列以下的區域
            inner_y = cy
            inner_h = ch - hdr_h
            if inner_h < 1 or not stocks:
                continue

            # 儲存分類 rect 供 hover 偵測
            cat_cost = sum(s['total_cost'] for s in stocks)
            cat_val  = sum(s['value']      for s in stocks)
            self._tm_cat_rects.append({
                'cx': cx, 'cy': cy, 'cw': cw, 'ch': ch,
                'cat_name': cat_name,
                'count':    len(stocks),
                'value':    cat_val,
                'cost':     cat_cost,
                'pnl_amt':  cat_val - cat_cost,
                'pnl_pct':  (cat_val - cat_cost) / cat_cost * 100 if cat_cost else 0,
                'hdr_h':    hdr_h,
            })

            stock_vals = [s['value'] for s in stocks]
            if cw >= inner_h:
                # 橫式容器：squarify 後翻轉 y，最大方塊在左上
                norm_stocks = squarify.normalize_sizes(stock_vals, cw, inner_h)
                _sr_raw = squarify.squarify(norm_stocks, cx, inner_y, cw, inner_h)
                srects = [{'x': r['x'],
                           'y': inner_y + inner_h - (r['y'] - inner_y) - r['dy'],
                           'dx': r['dx'], 'dy': r['dy']} for r in _sr_raw]
            else:
                # 直式容器：轉 90° 後翻轉 y，方塊更接近正方形且左上最大
                norm_stocks = squarify.normalize_sizes(stock_vals, inner_h, cw)
                srects_t = squarify.squarify(norm_stocks, 0, 0, inner_h, cw)
                srects = [{'x': cx + r['y'],
                           'y': inner_y + inner_h - r['x'] - r['dx'],
                           'dx': r['dy'], 'dy': r['dx']} for r in srects_t]

            for stock, sr in zip(stocks, srects):
                sx, sy, sw, sh = sr['x'], sr['y'], sr['dx'], sr['dy']
                rx, ry = sx + GAP / 2, sy + GAP / 2
                rw, rh = sw - GAP, sh - GAP
                if rw <= 0 or rh <= 0:
                    continue

                ax.add_patch(plt.Rectangle(
                    (rx, ry), rw, rh,
                    facecolor=pnl_color(stock['pnl_pct'], _max_abs),
                    edgecolor=SEP_COL, linewidth=0.4, zorder=1))

                # 儲存 rect 供 hover 偵測
                self._tm_rects.append({
                    'rx': rx, 'ry': ry, 'rw': rw, 'rh': rh,
                    'code':     stock['code'],
                    'name':     stock['name'],
                    'value':    stock['value'],
                    'price':    stock['price'],
                    'pnl_pct':  stock['pnl_pct'],
                    'avg_cost': stock['avg_cost'],
                    'qty':      stock['qty'],
                    'category': cat_name,
                    'portfolio_pct': stock['value'] / total_val * 100 if total_val else 0,
                })

                # 字體依方塊像素大小計算（方塊越大=占比越高=字越大）
                min_dim = min(rw, rh)
                if min_dim < 2:
                    continue

                rw_px      = rw * _px_per_ux
                rh_px      = rh * _px_per_uy
                chars      = max(len(stock['code']), 4)
                name_chars = max(len(stock['name']), 2)
                fs_by_w  = 0.58 * rw_px * _pt_per_px / (chars * 0.60)
                fs_by_h  = 0.65 * rh_px * _pt_per_px / 3.6
                fs_code  = max(min(fs_by_w, fs_by_h, 44), 6)
                fs_name  = max(min(0.80 * rw_px * _pt_per_px / (name_chars * 0.95),
                                   fs_code * 0.55, 22), 5)
                fs_sub   = max(fs_code * 0.55, 5)

                pnl_str = f"{'+' if stock['pnl_pct'] >= 0 else ''}{stock['pnl_pct']:.2f}%"
                cx_t    = rx + rw / 2
                mid_y   = ry + rh * 0.50

                fs_code_uy = fs_code / _pt_per_px / _px_per_uy
                fs_name_uy = fs_name / _pt_per_px / _px_per_uy
                fs_sub_uy  = fs_sub  / _pt_per_px / _px_per_uy

                if min_dim >= 8:   # 三行：代號 / 名稱 / 漲跌幅
                    gap_cn = (fs_code_uy + fs_name_uy) * 0.55
                    gap_np = (fs_name_uy + fs_sub_uy)  * 0.55
                    half   = (gap_cn + gap_np) / 2
                    ax.text(cx_t, mid_y + half, stock['code'],
                            ha='center', va='center', color='white',
                            fontsize=fs_code, fontweight='bold',
                            fontfamily=CHART_FONT, clip_on=True, zorder=4)
                    ax.text(cx_t, mid_y + half - gap_cn, stock['name'],
                            ha='center', va='center', color='#dddddd',
                            fontsize=fs_name, fontfamily=CHART_FONT,
                            clip_on=True, zorder=4)
                    ax.text(cx_t, mid_y - half, pnl_str,
                            ha='center', va='center', color='white',
                            fontsize=fs_sub, fontfamily=CHART_FONT,
                            clip_on=True, zorder=4)
                elif min_dim >= 5:   # 兩行：代號 / 漲跌幅
                    gap = (fs_code_uy + fs_sub_uy) * 0.55
                    ax.text(cx_t, mid_y + gap * 0.5, stock['code'],
                            ha='center', va='center', color='white',
                            fontsize=fs_code, fontweight='bold',
                            fontfamily=CHART_FONT, clip_on=True, zorder=4)
                    ax.text(cx_t, mid_y - gap * 0.5, pnl_str,
                            ha='center', va='center', color='white',
                            fontsize=fs_sub, fontfamily=CHART_FONT,
                            clip_on=True, zorder=4)
                else:               # 一行：代號
                    ax.text(cx_t, mid_y, stock['code'],
                            ha='center', va='center', color='white',
                            fontsize=fs_code, fontweight='bold',
                            fontfamily=CHART_FONT, clip_on=True, zorder=4)


        # 狀態文字
        warn = f'  ⚠ {", ".join(no_price_list)} 無即時報價' if no_price_list else ''
        sax.text(0.5, 0.5,
                 f'估值合計 {total_val:,.0f} 元  ·  '
                 f'{datetime.now().strftime("%Y-%m-%d %H:%M")} 更新{warn}',
                 ha='center', va='center', color='#666',
                 fontsize=8, fontfamily=CHART_FONT,
                 transform=sax.transAxes)

        self._ui_call(lambda: self._tm_status.set(
            f'更新時間：{datetime.now().strftime("%H:%M:%S")}  ·  每 30 秒自動更新'))
        self._ui_call(self._tm_canvas.draw)

    # ═══════════════════════════════════════════════════════════════════════════
    # Tab 3：個股買賣分析
    # ═══════════════════════════════════════════════════════════════════════════
    def _build_tab3(self):
        f = self.tab3

        ctrl = ttk.Frame(f)
        ctrl.pack(fill='x', padx=14, pady=8)
        ttk.Label(ctrl, text='個股買賣分析', style='Hdr.TLabel').pack(side='left')
        ttk.Label(ctrl, text='庫存快選 / 輸入代號：').pack(side='left', padx=(20, 4))
        self._stock_var   = tk.StringVar()
        self._stock_combo = ttk.Combobox(ctrl, textvariable=self._stock_var,
                                          state='normal', width=26)
        self._stock_combo.pack(side='left', padx=4)
        self._stock_combo.bind('<Return>', lambda _: self._draw_analysis())
        ttk.Label(ctrl, text='（↵ 或點按鈕）',
                  foreground=C_FG2, font=('Microsoft JhengHei', 8)).pack(side='left', padx=(0, 6))
        ttk.Button(ctrl, text='📊  繪製圖表', style='Nav.TButton', command=self._draw_analysis).pack(side='left')
        ttk.Button(ctrl, text='💾  另存圖檔', style='Nav.TButton', command=self._save_analysis).pack(side='left', padx=(6, 0))
        self._adj_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(ctrl, text='還原股價（含股票分割）',
                        variable=self._adj_var,
                        command=self._draw_analysis).pack(side='left', padx=(14, 0))

        self._an_fig    = plt.Figure(figsize=(9.5, 5.5), dpi=100, facecolor=C_BG)
        self._an_canvas = FigureCanvasTkAgg(self._an_fig, master=f)
        w = self._an_canvas.get_tk_widget()
        w.configure(bg=C_BG)
        w.pack(fill='both', expand=True, padx=8, pady=(0, 8))

        self._an_bar_data = []      # list of bar dicts for hover detection
        self._an_ax1      = None    # saved axes ref
        self._an_tooltip  = None    # tooltip Toplevel
        self._an_canvas.mpl_connect('motion_notify_event', self._on_an_motion)
        self._an_canvas.mpl_connect('figure_leave_event',  lambda e: self._hide_an_tooltip())

        self._update_stock_list()

    def _update_stock_list(self):
        df = load_df()
        if df.empty:
            return
        holdings = calc_holdings(df)
        holding_keys = set(holdings.keys())   # set of (code, cat)

        # 取得每個 (代號, 分類) 組合的股票名稱
        combos = (df.groupby(['股票代號', '分類'])['股票名稱']
                    .last().reset_index())

        def _opt(row):
            code = str(row['股票代號'])
            cat  = str(row['分類'])
            name = str(row['股票名稱'])
            return f"{code}  {name}  [{cat}]"

        hold_opts, sold_opts = [], []
        for _, row in combos.iterrows():
            key = (str(row['股票代號']), str(row['分類']))
            (hold_opts if key in holding_keys else sold_opts).append(_opt(row))

        options = hold_opts + (['── 已賣出 ──'] if sold_opts else []) + sold_opts
        self._stock_combo['values'] = options
        if not self._stock_var.get() and hold_opts:
            self._stock_var.set(hold_opts[0])

    # ── 個股分析長條 hover helpers ─────────────────────────────────────────────
    def _on_an_motion(self, event):
        if self._an_ax1 is None or not self._an_bar_data or event.x is None:
            self._hide_an_tooltip()
            return
        # 不依賴 event.inaxes（twinx 時可能回傳 ax2），直接轉換顯示座標到 ax1
        try:
            xd, yd = self._an_ax1.transData.inverted().transform((event.x, event.y))
        except Exception:
            self._hide_an_tooltip()
            return
        found = None
        for b in self._an_bar_data:
            if abs(xd - b['x']) <= b['w'] / 2 and 0 <= yd <= b['price']:
                found = b
                break
        if found is None:
            self._hide_an_tooltip()
            return
        widget = self._an_canvas.get_tk_widget()
        sx = widget.winfo_rootx() + int(event.x)
        sy = widget.winfo_rooty() + int(widget.winfo_height() - event.y)
        self._show_an_tooltip(sx, sy, found)

    def _show_an_tooltip(self, sx, sy, data):
        if self._an_tooltip is None or not self._an_tooltip.winfo_exists():
            tip = tk.Toplevel(self)
            tip.overrideredirect(True)
            tip.attributes('-topmost', True)
            tip.configure(bg='#1e1e1e')
            border = tk.Frame(tip, bg='#3e3e3e', padx=1, pady=1)
            border.pack(fill='both', expand=True)
            inner = tk.Frame(border, bg='#252526', padx=10, pady=8)
            inner.pack(fill='both', expand=True)
            self._an_tip_widgets = {}
            for key in ('title', 'price', 'qty', 'fee', 'tax', 'net', 'avg'):
                lbl = tk.Label(inner, bg='#252526',
                               font=('Microsoft JhengHei', 10), anchor='w')
                lbl.pack(fill='x')
                self._an_tip_widgets[key] = lbl
            self._an_tooltip = tip

        w   = self._an_tip_widgets
        if data['side'] == '買':
            side_color = C_BUY_FG
        elif data['category'] == '獲利賣出':
            side_color = '#ffd700'
        else:
            side_color = C_SELL_FG
        w['title'].config(
            text=f"{data['date']}　{data['side']}　({data['category']})",
            fg=side_color, font=('Microsoft JhengHei', 11, 'bold'))
        w['price'].config(text=f"成交價：{data['price']:,.2f} 元", fg='#cccccc')
        w['qty'].config(  text=f"數　量：{data['qty']:,.0f} 股",   fg='#cccccc')
        w['fee'].config(  text=f"手續費：{data['fee']:,.0f} 元",   fg='#cccccc')
        tax_txt = f"交易稅：{data['tax']:,.0f} 元" if data['tax'] else ''
        w['tax'].config(  text=tax_txt, fg='#cccccc')
        net_sign = '+' if data['net'] >= 0 else ''
        w['net'].config(  text=f"淨金額：{net_sign}{data['net']:,.0f} 元",
                          fg='#4ec94e' if data['net'] >= 0 else '#f07070')
        if not np.isnan(data['avg_cost']):
            w['avg'].config(text=f"當時均價：{data['avg_cost']:,.2f} 元", fg='#8ab4d4')
        else:
            w['avg'].config(text='', fg='#cccccc')

        self._place_tooltip(self._an_tooltip, sx, sy)
        self._an_tooltip.deiconify()

    def _hide_an_tooltip(self):
        if self._an_tooltip and self._an_tooltip.winfo_exists():
            self._an_tooltip.withdraw()

    def _draw_analysis(self):
        sel = self._stock_var.get().strip()
        if not sel or sel.startswith('──'):
            messagebox.showinfo('提示', '請輸入或選擇股票代號')
            return

        # 格式："CODE  NAME  [分類]" 或 "CODE  NAME" 或 "CODE"
        cat_filter = None
        if sel.endswith(']') and '[' in sel:
            cat_filter = sel[sel.rfind('[') + 1:-1]
            sel = sel[:sel.rfind('[')].strip()
        code = sel.split()[0].upper()
        df   = load_df()
        sdf  = df[df['股票代號'].astype(str) == code].copy()
        if cat_filter:
            # 獲利賣出賣出不屬於任何買入分類，但需要一併顯示以反映成本重置
            sdf = sdf[(sdf['分類'].astype(str) == cat_filter) |
                      ((sdf['分類'].astype(str) == '獲利賣出') & (sdf['買賣'].astype(str) == '賣'))]
        sdf = sdf.sort_values('日期').reset_index(drop=True)

        if sdf.empty:
            messagebox.showinfo('提示', f'找不到 {code} 的交易紀錄\n請確認代號，或先在「買賣紀錄」新增交易。')
            return

        # ── 計算持股均價（追蹤均價：獲利賣出後重置；真實均價：實際投入成本）─────────
        has_takep = any(str(c) == '獲利賣出' for c in sdf['分類'].values)
        qty_cum,      cost_cum      = 0.0, 0.0
        real_qty_cum, real_cost_cum = 0.0, 0.0
        avg_costs, real_avg_costs   = [], []
        for _, r in sdf.iterrows():
            q, p, fee = float(r['數量(股)']), float(r['價格(元)']), float(r['手續費(元)'])
            cat = str(r.get('分類', ''))
            if r['買賣'] == '買':
                cost_cum      += q * p + fee
                qty_cum       += q
                real_cost_cum += q * p + fee
                real_qty_cum  += q
            else:
                if qty_cum > 0:
                    cost_cum -= (cost_cum / qty_cum) * q
                    qty_cum  -= q
                    if cat == '獲利賣出' and qty_cum > 0.5:
                        cost_cum = qty_cum * p
                if real_qty_cum > 0:
                    real_cost_cum -= (real_cost_cum / real_qty_cum) * q
                    real_qty_cum  -= q
            avg_costs.append(cost_cum / qty_cum if qty_cum > 0.5 else np.nan)
            real_avg_costs.append(real_cost_cum / real_qty_cum if real_qty_cum > 0.5 else np.nan)

        sdf['avg_cost']      = avg_costs
        sdf['real_avg_cost'] = real_avg_costs

        dates  = sdf['日期'].values
        prices = sdf['價格(元)'].values.astype(float)
        sides  = sdf['買賣'].values
        qtys   = sdf['數量(股)'].values.astype(float)

        # ── 抓歷史收盤價 ──────────────────────────────────────────────────────
        adj_mode  = self._adj_var.get()
        start_str = pd.to_datetime(dates[0]).strftime('%Y-%m-%d')
        hist_close = get_history(code, start_str, auto_adjust=adj_mode)

        # ── 還原模式：用 adj/raw 收盤比值推算分割比例，同步調整價格與股數 ──────
        if adj_mode:
            _hi_raw = get_history(code, start_str, auto_adjust=False)
            _hi_adj = hist_close  # already auto_adjust=True

            sdf = sdf.copy()
            adj_prices, adj_qtys = [], []

            if (_hi_raw is not None and not _hi_raw.empty and
                    _hi_adj is not None and not _hi_adj.empty):
                # 去除 timezone
                _raw_idx = _hi_raw.copy()
                _adj_idx = _hi_adj.copy()
                if getattr(_raw_idx.index, 'tz', None):
                    _raw_idx.index = _raw_idx.index.tz_convert(None)
                if getattr(_adj_idx.index, 'tz', None):
                    _adj_idx.index = _adj_idx.index.tz_convert(None)

                for orig_p, orig_q, d in zip(prices, qtys, dates):
                    ts = pd.Timestamp(d).normalize()
                    try:
                        # 在「交易日當天或之後最近交易日」取得 raw 與 adj 收盤
                        pos = min(_raw_idx.index.searchsorted(ts), len(_raw_idx) - 1)
                        raw_v = float(_raw_idx.iloc[pos])
                        adj_v = float(_adj_idx.iloc[pos])
                        # ratio < 1 表示有拆股（adj 比 raw 小）
                        ratio = adj_v / raw_v if raw_v > 0 else 1.0
                    except Exception:
                        ratio = 1.0
                    # 還原後：成交價 × ratio（縮小）、股數 ÷ ratio（放大）
                    adj_prices.append(orig_p * ratio)
                    adj_qtys.append(orig_q / ratio)
            else:
                adj_prices = list(prices)
                adj_qtys   = list(qtys)

            prices = np.array(adj_prices)
            qtys   = np.array(adj_qtys)
            sdf['價格(元)']  = prices
            sdf['數量(股)'] = qtys

            # 重算均價（用調整後的價格與股數）
            qty_cum2,      cost_cum2      = 0.0, 0.0
            real_qty_cum2, real_cost_cum2 = 0.0, 0.0
            new_avg, new_real_avg         = [], []
            for _, r in sdf.iterrows():
                q   = float(r['數量(股)'])
                p   = float(r['價格(元)'])
                fee = float(r['手續費(元)'])
                cat = str(r.get('分類', ''))
                if r['買賣'] == '買':
                    cost_cum2      += q * p + fee
                    qty_cum2       += q
                    real_cost_cum2 += q * p + fee
                    real_qty_cum2  += q
                else:
                    if qty_cum2 > 0:
                        cost_cum2 -= (cost_cum2 / qty_cum2) * q
                        qty_cum2  -= q
                        if cat == '獲利賣出' and qty_cum2 > 0.5:
                            cost_cum2 = qty_cum2 * p
                    if real_qty_cum2 > 0:
                        real_cost_cum2 -= (real_cost_cum2 / real_qty_cum2) * q
                        real_qty_cum2  -= q
                new_avg.append(cost_cum2 / qty_cum2 if qty_cum2 > 0.5 else np.nan)
                new_real_avg.append(real_cost_cum2 / real_qty_cum2 if real_qty_cum2 > 0.5 else np.nan)
            sdf['avg_cost']      = new_avg
            sdf['real_avg_cost'] = new_real_avg
            avg_costs      = new_avg
            real_avg_costs = new_real_avg
            cost_cum       = cost_cum2
            qty_cum        = qty_cum2
            real_cost_cum  = real_cost_cum2
            real_qty_cum   = real_qty_cum2

        # ── 繪圖 ──────────────────────────────────────────────────────────────
        self._an_fig.clear()
        ax1 = self._an_fig.add_subplot(111, facecolor=C_PANEL)
        ax2 = ax1.twinx()

        for spine in ax1.spines.values():
            spine.set_edgecolor(C_BORDER)
        ax1.tick_params(colors=C_FG2)
        ax1.xaxis.label.set_color(C_FG)
        ax1.yaxis.label.set_color(C_FG)

        x_num = mdates.date2num(pd.to_datetime(dates))
        bar_w = max(0.6, (x_num[-1] - x_num[0]) / max(len(x_num) * 2, 10)) if len(x_num) > 1 else 1.5

        # 同日多筆：計算各 bar 的並排 x 位置與寬度
        from collections import defaultdict
        date_groups: dict = defaultdict(list)
        for i, xn in enumerate(x_num):
            date_groups[round(xn, 4)].append(i)

        per_x = np.empty(len(x_num))
        per_w = np.empty(len(x_num))
        for indices in date_groups.values():
            n     = len(indices)
            sub_w = bar_w / n * 0.92          # 留 8% 間隙
            for rank, idx in enumerate(indices):
                per_x[idx] = x_num[idx] - bar_w / 2 + (rank + 0.5) * (bar_w / n)
                per_w[idx] = sub_w

        # 歷史收盤價折線（最底層）
        hist_plotted = False
        if hist_close is not None and not hist_close.empty:
            try:
                hidx = hist_close.index
                if getattr(hidx, 'tz', None) is not None:
                    hidx = hidx.tz_localize(None)
                hx = mdates.date2num(pd.to_datetime(hidx))
                hy = hist_close.values.astype(float)
                ax1.plot(hx, hy, color='#b0bec5', linewidth=1.4,
                         alpha=0.75, zorder=2, label='收盤價走勢')
                hist_plotted = True
            except Exception:
                pass

        # 買賣成交價長條（並排）
        cats = sdf['分類'].values
        bar_colors = [
            C_BUY_FG if s == '買' else ('#ffd700' if str(c) == '獲利賣出' else C_SELL_FG)
            for s, c in zip(sides, cats)
        ]
        ax1.bar(per_x, prices, width=per_w, color=bar_colors, alpha=0.70, zorder=3)

        # 追蹤均價折線（獲利賣出後重置，用原始 x_num 保持連續）
        valid = ~np.isnan(sdf['avg_cost'].values)
        if valid.any():
            ax1.plot(x_num[valid], sdf['avg_cost'].values[valid],
                     color='#64b5f6', linewidth=2.2, marker='o', markersize=5, zorder=5)

        # 真實均價折線（灰色虛線，僅在有獲利賣出紀錄時顯示）
        if has_takep:
            real_valid = ~np.isnan(sdf['real_avg_cost'].values)
            if real_valid.any():
                ax1.plot(x_num[real_valid], sdf['real_avg_cost'].values[real_valid],
                         color='#90a4ae', linewidth=1.6, linestyle='--',
                         marker='o', markersize=3, zorder=5)

        # 即時現價橫線
        cur_price, _ = get_price(code)
        if cur_price:
            ax1.axhline(cur_price, color='#ffa726', linestyle='--', linewidth=1.8, zorder=4)

        # 次軸：交易股數（並排）
        vol_colors = ['#4caf50' if s == '買' else '#f44336' for s in sides]
        ax2.bar(per_x, qtys, width=per_w, color=vol_colors, alpha=0.18, zorder=1)
        ax2.set_ylabel('交易股數', color=C_FG2, fontsize=9)
        ax2.tick_params(axis='y', colors=C_FG2)
        ax2.yaxis.set_major_formatter(
            matplotlib.ticker.FuncFormatter(lambda v, _: f'{int(v):,}'))
        for sp in ax2.spines.values():
            sp.set_edgecolor(C_BORDER)

        ax1.xaxis_date()
        ax1.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
        self._an_fig.autofmt_xdate(rotation=28, ha='right')
        ax1.yaxis.set_major_formatter(
            matplotlib.ticker.FuncFormatter(lambda v, _: f'{v:,.1f}'))
        ax1.grid(axis='y', alpha=0.15, linestyle='--', color=C_FG2)

        # 圖例
        legend_handles = [
            Patch(color=C_BUY_FG,  alpha=0.8, label='買入'),
            Patch(color=C_SELL_FG, alpha=0.8, label='賣出'),
        ]
        if has_takep:
            legend_handles.append(Patch(color='#ffd700', alpha=0.8, label='獲利賣出'))
        if hist_plotted:
            legend_handles.append(
                Line2D([0], [0], color='#b0bec5', linewidth=1.4, label='收盤價走勢'))
        legend_handles.append(
            Line2D([0], [0], color='#64b5f6', linewidth=2.2,
                   marker='o', markersize=5, label='追蹤均價' if has_takep else '持股均價'))
        if has_takep:
            legend_handles.append(
                Line2D([0], [0], color='#90a4ae', linewidth=1.6, linestyle='--',
                       marker='o', markersize=3, label='真實均價'))
        if cur_price:
            legend_handles.append(
                Line2D([0], [0], color='#ffa726', linestyle='--',
                       linewidth=1.8, label=f'現價 {cur_price:.1f}'))
        legend = ax1.legend(handles=legend_handles, loc='best',
                            fontsize=9, framealpha=0.6)
        legend.get_frame().set_facecolor(C_PANEL)
        for text in legend.get_texts():
            text.set_color(C_FG)

        ax1.set_ylabel('價格 (元)', fontsize=10, color=C_FG)

        name      = sdf['股票名稱'].iloc[0]
        last_avg  = cost_cum / qty_cum if qty_cum > 0.5 else None
        real_last = real_cost_cum / real_qty_cum if real_qty_cum > 0.5 else None
        subtitle  = ''
        if last_avg:
            pnl_str   = (f'{cur_price - last_avg:+.2f} ({(cur_price/last_avg - 1)*100:+.1f}%)'
                         if cur_price else '')
            avg_label = '追蹤均價' if has_takep else '均價'
            subtitle  = (f'庫存 {qty_cum:,.0f} 股  |  {avg_label} {last_avg:.2f} 元'
                         + (f'  |  損益 {pnl_str}' if pnl_str else ''))
            if has_takep and real_last and abs(real_last - last_avg) > 0.01:
                real_pnl = (f'{cur_price - real_last:+.2f} ({(cur_price/real_last - 1)*100:+.1f}%)'
                            if cur_price else '')
                subtitle += (f'\n真實均價 {real_last:.2f} 元'
                             + (f'  |  實際損益 {real_pnl}' if real_pnl else ''))

        cat_label = f'  [{cat_filter}]' if cat_filter else ''
        adj_label = '  ［還原股價］' if adj_mode else ''
        ax1.set_title(f'{code}  {name}{cat_label}{adj_label}\n{subtitle}',
                      fontsize=12, fontweight='bold', pad=8, color=C_FG)

        # 儲存長條資料供 hover 偵測
        self._an_ax1     = ax1
        self._an_bar_data = []
        self._hide_an_tooltip()
        for i in range(len(sdf)):
            row = sdf.iloc[i]
            self._an_bar_data.append({
                'x':        per_x[i],
                'w':        per_w[i],
                'price':    float(row['價格(元)']),
                'side':     str(row['買賣']),
                'date':     str(row['日期']),
                'qty':      float(row['數量(股)']),
                'fee':      float(row['手續費(元)']),
                'tax':      float(row['交易稅(元)']),
                'net':      float(row['淨金額(元)']),
                'avg_cost': sdf['avg_cost'].iloc[i],
                'category': str(row['分類']),
            })

        self._an_fig.tight_layout()
        self._an_canvas.draw()


    # ═══════════════════════════════════════════════════════════════════════════
    # Tab 4：ETF 分析（成分股分析 / ETF 比較）
    # ═══════════════════════════════════════════════════════════════════════════
    def _build_tab4(self):
        f = self.tab4

        # ── 子頁籤切換列 ──────────────────────────────────────────────────────
        sub_bar = tk.Frame(f, bg='#0d0d1a')
        sub_bar.pack(fill='x', padx=0, pady=0)

        self._etf_sub_btns: dict[str, tk.Label] = {}
        for sid, slbl in [('analysis', 'ETF 成分股分析'),
                           ('compare',  'ETF 比較'),
                           ('change',   '主動型ETF分析')]:
            btn = tk.Label(sub_bar, text=slbl,
                           bg='#1a1a2e', fg='#9090a0',
                           font=('Microsoft JhengHei', 10, 'bold'),
                           padx=18, pady=7, cursor='hand2')
            btn.pack(side='left')
            btn.bind('<Button-1>', lambda e, s=sid: self._show_etf_sub(s))
            self._etf_sub_btns[sid] = btn

        # ── 子頁面容器 ────────────────────────────────────────────────────────
        self._etf_sub_analysis = tk.Frame(f, bg=C_BG)
        self._etf_sub_analysis.place(relx=0, rely=0.045, relwidth=1, relheight=0.955)
        self._etf_sub_compare  = tk.Frame(f, bg=C_BG)
        self._etf_sub_compare.place(relx=0, rely=0.045, relwidth=1, relheight=0.955)
        self._etf_sub_change   = tk.Frame(f, bg=C_BG)
        self._etf_sub_change.place(relx=0, rely=0.045, relwidth=1, relheight=0.955)

        self._build_etf_analysis_sub(self._etf_sub_analysis)
        self._build_etf_compare_sub(self._etf_sub_compare)
        self._build_etf_change_sub(self._etf_sub_change)
        self._show_etf_sub('analysis')

    def _show_etf_sub(self, sid: str):
        _frames = {'analysis': self._etf_sub_analysis,
                   'compare':  self._etf_sub_compare,
                   'change':   self._etf_sub_change}
        _frames.get(sid, self._etf_sub_analysis).tkraise()
        for k, btn in self._etf_sub_btns.items():
            btn.config(bg='#2a3f6f' if k == sid else '#1a1a2e',
                       fg='white'   if k == sid else '#9090a0')
        if sid == 'change':
            self.after(200, self._chg_fetch_all)

    def _build_etf_analysis_sub(self, f):
        # ── 控制列 ────────────────────────────────────────────────────────────
        ctrl = ttk.Frame(f)
        ctrl.pack(fill='x', padx=14, pady=8)
        ttk.Label(ctrl, text='ETF 成分股分析', style='Hdr.TLabel').pack(side='left')

        ttk.Label(ctrl, text='選擇 ETF：').pack(side='left', padx=(20, 4))
        self._etf_var = tk.StringVar()
        # 持有中的 ETF 排在前面，加 ★ 標記
        try:
            _held = set(load_df()['股票代號'].dropna().astype(str).unique())
        except Exception:
            _held = set()
        _held_opts  = [f'{c}  ★{n}' for c, n in ETF_LIST if c in _held]
        _other_opts = [f'{c}  {n}'  for c, n in ETF_LIST if c not in _held]
        etf_options = _held_opts + _other_opts
        self._etf_combo = ttk.Combobox(ctrl, textvariable=self._etf_var,
                                        values=etf_options, state='normal', width=26)
        self._etf_combo.pack(side='left', padx=4)
        def _on_combo_select(_=None):
            sel = self._etf_var.get().strip()
            if sel:
                self._etf_code_var.set(sel.split()[0])
            self._draw_etf_map()

        self._etf_combo.bind('<<ComboboxSelected>>', _on_combo_select)
        self._etf_combo.bind('<Return>', _on_combo_select)

        ttk.Label(ctrl, text='或輸入代號：',
                  foreground=C_FG2, font=('Microsoft JhengHei', 9)).pack(side='left', padx=(10, 4))
        self._etf_code_var = tk.StringVar()
        etf_entry = ttk.Entry(ctrl, textvariable=self._etf_code_var, width=10)
        etf_entry.pack(side='left', padx=4)

        def _on_entry_change(*_):
            typed = self._etf_code_var.get().strip()
            # 找下拉清單中有無對應項目並同步
            for opt in self._etf_combo['values']:
                if opt.split()[0] == typed:
                    self._etf_var.set(opt)
                    return
            self._etf_var.set('')

        self._etf_code_var.trace_add('write', _on_entry_change)
        etf_entry.bind('<Return>', lambda _: self._draw_etf_map())

        ttk.Button(ctrl, text='📈  繪製', style='Nav.TButton',
                   command=self._draw_etf_map).pack(side='left', padx=(6, 0))
        ttk.Button(ctrl, text='💾  另存圖檔', style='Nav.TButton',
                   command=self._save_etf_map).pack(side='right', padx=(4, 0))

        # 狀態列（獨立一行，避免被截斷）
        self._etf_status = tk.StringVar(value='請選擇 ETF 或輸入代號後點擊繪製')
        status_bar = tk.Frame(f, bg=C_BG)
        status_bar.pack(fill='x', padx=14, pady=(0, 2))
        ttk.Label(status_bar, textvariable=self._etf_status,
                  foreground=C_FG2, font=('Microsoft JhengHei', 9),
                  wraplength=1100, justify='left').pack(anchor='w')

        # ── 摘要列（兩行卡片）────────────────────────────────────────────────
        summary_bar = tk.Frame(f, bg=C_PANEL)
        summary_bar.pack(fill='x', padx=8, pady=(0, 4))

        row1 = tk.Frame(summary_bar, bg=C_PANEL)
        row1.pack(fill='x')
        row2 = tk.Frame(summary_bar, bg=C_PANEL)
        row2.pack(fill='x', pady=(2, 0))

        def _card(parent, title):
            fr = tk.Frame(parent, bg='#1a1a2e', padx=14, pady=5)
            fr.pack(side='left', fill='x', expand=True, padx=3)
            tk.Label(fr, text=title, bg='#1a1a2e',
                     fg='#6a8faf', font=('Microsoft JhengHei', 8)).pack(anchor='w')
            var = tk.StringVar(value='—')
            lbl = tk.Label(fr, textvariable=var, bg='#1a1a2e',
                           fg=C_FG, font=('Microsoft JhengHei', 11, 'bold'))
            lbl.pack(anchor='w')
            return var, lbl

        self._etf_sv_name,   self._etf_sl_name   = _card(row1, 'ETF 名稱')
        self._etf_sv_count,  self._etf_sl_count  = _card(row1, '成分股 / 前10大占比')
        self._etf_sv_top,    self._etf_sl_top    = _card(row1, '最大持股')
        self._etf_sv_change, self._etf_sl_change = _card(row1, '加權漲跌幅')
        self._etf_sv_aum,    self._etf_sl_aum    = _card(row2, '基金規模')
        self._etf_sv_yield,  self._etf_sl_yield  = _card(row2, '年化殖利率')
        self._etf_sv_nav,    self._etf_sl_nav    = _card(row2, 'ETF 淨值(NAV)')
        self._etf_sv_alloc,  self._etf_sl_alloc  = _card(row2, '折溢價')

        # ── 可捲動主區域（樹狀圖 + 分析圖）──────────────────────────────────
        sc_outer = tk.Frame(f, bg=C_BG)
        sc_outer.pack(fill='both', expand=True, padx=0, pady=(0, 0))

        vsb = ttk.Scrollbar(sc_outer, orient='vertical')
        vsb.pack(side='right', fill='y')

        self._etf_sc = tk.Canvas(sc_outer, bg='#111111',
                                  yscrollcommand=vsb.set, highlightthickness=0)
        self._etf_sc.pack(side='left', fill='both', expand=True)
        vsb.config(command=self._etf_sc.yview)

        self._etf_inner = tk.Frame(self._etf_sc, bg='#111111')
        _sc_win = self._etf_sc.create_window(0, 0, anchor='nw', window=self._etf_inner)

        def _on_inner_cfg(e):
            self._etf_sc.configure(scrollregion=self._etf_sc.bbox('all'))
        self._etf_inner.bind('<Configure>', _on_inner_cfg)

        def _on_sc_cfg(e):
            self._etf_sc.itemconfig(_sc_win, width=e.width)
        self._etf_sc.bind('<Configure>', _on_sc_cfg)

        def _mw(e):
            self._etf_sc.yview_scroll(-1 if e.delta > 0 else 1, 'units')
        for _w in (self._etf_sc, self._etf_inner):
            _w.bind('<MouseWheel>', _mw)

        # ── K 線圖（上）：水平控制列 + 畫布 ─────────────────────────────
        kline_outer = tk.Frame(self._etf_inner, bg='#111111')
        kline_outer.pack(fill='x', padx=8, pady=(4, 2))
        kline_outer.bind('<MouseWheel>', _mw)

        # ── 水平控制列 ────────────────────────────────────────────────────
        kctrl = tk.Frame(kline_outer, bg='#1c1c28', pady=4)
        kctrl.pack(fill='x', pady=(0, 2))
        kctrl.bind('<MouseWheel>', _mw)

        _CTRL_BG = '#1c1c28'
        _BTN_OFF = dict(bg='#e8e8ee', fg='#222233',
                        font=('Microsoft JhengHei', 8, 'bold'),
                        relief='flat', bd=0, padx=10, pady=3,
                        cursor='hand2',
                        highlightbackground='#aaaacc', highlightthickness=1,
                        activebackground='#d0d0e0', activeforeground='#111122')
        _BTN_ON  = dict(bg='#3a5fcd', fg='#ffffff',
                        font=('Microsoft JhengHei', 8, 'bold'),
                        relief='flat', bd=0, padx=10, pady=3,
                        cursor='hand2',
                        highlightbackground='#2a4fbd', highlightthickness=1,
                        activebackground='#4a70e0', activeforeground='#ffffff')

        def _kbtn_style(btn: tk.Button, on: bool) -> None:
            btn.configure(**(_BTN_ON if on else _BTN_OFF))

        # ── 區間 ─────────────────────────────────────────────────────────
        self._kline_period = '3M'
        self._kline_period_btns: dict[str, tk.Button] = {}

        tk.Label(kctrl, text='區間:', bg=_CTRL_BG, fg='#8899cc',
                 font=('Microsoft JhengHei', 8, 'bold')).pack(side='left', padx=(8, 4))

        def _set_period(lbl: str) -> None:
            self._kline_period = lbl
            for _l, _b in self._kline_period_btns.items():
                _kbtn_style(_b, _l == lbl)
            self._redraw_etf_kline()

        for _pl in ['1M', '3M', '6M', '1Y', '全部']:
            _b = tk.Button(kctrl, text=_pl, **_BTN_OFF,
                           command=lambda l=_pl: _set_period(l))
            _b.pack(side='left', padx=2)
            self._kline_period_btns[_pl] = _b
        _kbtn_style(self._kline_period_btns['3M'], True)

        # 分隔
        tk.Label(kctrl, text='│', bg=_CTRL_BG, fg='#444466').pack(side='left', padx=8)

        # ── 指標 ─────────────────────────────────────────────────────────
        self._kline_ind_state: dict[str, bool] = {
            'MA': True, 'BB': False, 'VOL': False,
            'MACD': True, 'RSI': False, 'KD': False}
        self._kline_ind_btns: dict[str, tk.Button] = {}

        tk.Label(kctrl, text='指標:', bg=_CTRL_BG, fg='#8899cc',
                 font=('Microsoft JhengHei', 8, 'bold')).pack(side='left', padx=(0, 4))

        def _toggle_ind(lbl: str) -> None:
            self._kline_ind_state[lbl] = not self._kline_ind_state[lbl]
            _kbtn_style(self._kline_ind_btns[lbl], self._kline_ind_state[lbl])
            self._redraw_etf_kline()

        for _il in ['MA', 'BB', 'VOL', 'MACD', 'RSI', 'KD']:
            _on = self._kline_ind_state[_il]
            _b = tk.Button(kctrl, text=_il,
                           **(_BTN_ON if _on else _BTN_OFF),
                           command=lambda l=_il: _toggle_ind(l))
            _b.pack(side='left', padx=2)
            self._kline_ind_btns[_il] = _b

        # K 線畫布
        self._etf_kline_fig = plt.Figure(figsize=(9.5, 5.3), dpi=100, facecolor='#111111')
        self._etf_kline_canvas = FigureCanvasTkAgg(self._etf_kline_fig, master=kline_outer)
        wk = self._etf_kline_canvas.get_tk_widget()
        wk.configure(bg='#111111')
        wk.pack(fill='both', expand=True)
        wk.bind('<MouseWheel>', _mw)

        # ── 樹狀圖畫布（K 線之後）────────────────────────────────────────
        self._etf_fig    = plt.Figure(figsize=(9.5, 5.3), dpi=100, facecolor='#111111')
        self._etf_canvas = FigureCanvasTkAgg(self._etf_fig, master=self._etf_inner)
        w = self._etf_canvas.get_tk_widget()
        w.configure(bg='#111111')
        w.pack(fill='x', padx=8, pady=(2, 2))
        w.bind('<MouseWheel>', _mw)

        # ── 分析圖畫布（環形圖）────────────────────────────────────────────
        self._etf_info_fig = plt.Figure(figsize=(9.5, 4.2), dpi=100, facecolor='#1a1a2e')
        self._etf_info_canvas = FigureCanvasTkAgg(self._etf_info_fig, master=self._etf_inner)
        w2 = self._etf_info_canvas.get_tk_widget()
        w2.configure(bg='#1a1a2e')
        w2.pack(fill='x', padx=8, pady=(2, 8))
        w2.bind('<MouseWheel>', _mw)

        # ── 歷史熱力圖畫布 ─────────────────────────────────────────────────
        self._etf_heatmap_fig = plt.Figure(figsize=(9.5, 5.5), dpi=100,
                                            facecolor='#111111')
        self._etf_heatmap_canvas = FigureCanvasTkAgg(
            self._etf_heatmap_fig, master=self._etf_inner)
        w3 = self._etf_heatmap_canvas.get_tk_widget()
        w3.configure(bg='#111111')
        w3.pack(fill='x', padx=8, pady=(2, 2))
        w3.bind('<MouseWheel>', _mw)

        # ── 成分股變化列表 ──────────────────────────────────────────────
        self._etf_change_outer = tk.Frame(self._etf_inner, bg='#111111')
        self._etf_change_outer.pack(fill='x', padx=8, pady=(2, 8))
        self._etf_change_outer.bind('<MouseWheel>', _mw)

        self._etf_rects         = []
        self._etf_tooltip       = None
        self._etf_ax            = None
        self._etf_info_cid      = None
        self._etf_info_charts   = []
        self._kline_ax          = None
        self._kline_ohlcv       = None
        self._kline_n           = 0
        self._kline_header      = None
        self._kline_vline       = None
        self._kline_hline       = None
        self._kline_sub_vlines  = []
        self._kline_price_ann   = None
        self._kline_all_axes    = []
        self._current_kline_code = None
        self._current_kline_name = ''
        self._etf_heatmap_expanded = False   # "其他" 列是否展開
        self._etf_treemap_expanded = False   # 樹狀圖是否顯示全部
        self._etf_canvas.mpl_connect('motion_notify_event', self._on_etf_motion)
        self._etf_canvas.mpl_connect('figure_leave_event',  lambda e: self._hide_etf_tooltip())
        self._etf_kline_canvas.mpl_connect('motion_notify_event', self._on_kline_hover)

    # ── ETF 比較子頁面 ────────────────────────────────────────────────────────
    def _build_etf_compare_sub(self, f):
        C_PNL = '#1a1a2e'
        # ── 頂部控制列 ────────────────────────────────────────────────────────
        ctrl = ttk.Frame(f)
        ctrl.pack(fill='x', padx=14, pady=8)
        ttk.Label(ctrl, text='ETF 比較', style='Hdr.TLabel').pack(side='left')

        ttk.Label(ctrl, text='搜尋 ETF：').pack(side='left', padx=(20, 4))
        self._cmp_search_var = tk.StringVar()
        cmp_entry = ttk.Entry(ctrl, textvariable=self._cmp_search_var, width=16)
        cmp_entry.pack(side='left', padx=4)
        cmp_entry.bind('<Return>', lambda _: self._cmp_add_etf())

        # 搜尋下拉
        etf_opts = [f'{c}  {n}' for c, n in ETF_LIST]
        self._cmp_combo = ttk.Combobox(ctrl, values=etf_opts, width=24, state='normal')
        self._cmp_combo.pack(side='left', padx=4)
        self._cmp_combo.bind('<<ComboboxSelected>>', lambda _: self._cmp_select_combo())

        ttk.Button(ctrl, text='＋ 加入比較', style='Nav.TButton',
                   command=self._cmp_add_etf).pack(side='left', padx=(4, 0))
        ttk.Button(ctrl, text='🗑 清除全部', style='Nav.TButton',
                   command=self._cmp_clear_all).pack(side='left', padx=(6, 0))
        ttk.Button(ctrl, text='📊 更新比較', style='Nav.TButton',
                   command=self._cmp_run).pack(side='left', padx=(6, 0))

        # ── 預設分類快選列 ────────────────────────────────────────────────────
        # 每個分類都自動包含 0050 作為大盤基準
        _BENCHMARK = [('0050', '元大台灣50')]
        _CMP_PRESETS = [
            ('指數型',   [('006208','富邦台灣50'), ('0051','元大中型100'),
                          ('00733','富邦臺灣中小'), ('009816','街口臺灣高成長'),
                          ('009804','統一台灣動能')]),
            ('高股息型', [('0056','元大高股息'), ('00878','國泰永續高股息'),
                          ('00919','群益台灣精選高息'), ('00713','元大台灣高息低波'),
                          ('00929','復華台灣科技優息'), ('00940','元大台灣價值高息'),
                          ('00939','統一台灣高息動能')]),
            ('科技/半導體',[('0052','富邦科技'), ('00891','中信關鍵半導體'),
                           ('00892','富邦台灣半導體'), ('00881','國泰台灣5G+')]),
            ('ESG/永續', [('00692','富邦公司治理'), ('00757','統一MSCI台灣ESG'),
                          ('00850','元大MSCI台灣'), ('00896','中信綠能及電動車')]),
            ('主動型',   [('00981A','統一台股增長'), ('00992A','中信優先順位'),
                          ('00994A','台新臺灣動能'), ('00987A','野村優息'),
                          ('00985A','第一金優選'), ('00991A','群益多因子'),
                          ('00995A','統一台灣動力')]),
            ('正2/反1',  [('00631L','元大台灣50正2'), ('00632R','元大台灣50反1')]),
            ('海外',     [('00646','元大S&P500'), ('00864B','中信美國公債20年'),
                          ('00679B','元大美債20年')]),
        ]

        preset_bar = tk.Frame(f, bg='#0d0d1a')
        preset_bar.pack(fill='x', padx=8, pady=(0, 2))
        tk.Label(preset_bar, text='快選分類：', bg='#0d0d1a', fg='#6a8faf',
                 font=('Microsoft JhengHei', 8)).pack(side='left', padx=(4, 4), pady=4)
        for p_name, p_etfs in _CMP_PRESETS:
            btn = tk.Label(preset_bar, text=p_name,
                           bg='#1e2d4e', fg='#aecde8',
                           font=('Microsoft JhengHei', 8),
                           padx=10, pady=3, cursor='hand2', relief='flat')
            btn.pack(side='left', padx=2, pady=3)
            # capture by value
            def _preset_click(e, etfs=p_etfs):
                self._cmp_clear_all()
                for code, name in _BENCHMARK + etfs:
                    if not any(c == code for c, _ in self._cmp_etf_list):
                        self._cmp_etf_list.append((code, name))
                self._cmp_rebuild_tags()
                self._cmp_run()
            btn.bind('<Button-1>', _preset_click)
            btn.bind('<Enter>', lambda e, b=btn: b.config(bg='#2a4a7f'))
            btn.bind('<Leave>', lambda e, b=btn: b.config(bg='#1e2d4e'))

        # ── 已選 ETF 標籤列 ───────────────────────────────────────────────────
        tag_frame_outer = tk.Frame(f, bg=C_PNL)
        tag_frame_outer.pack(fill='x', padx=8, pady=(0, 4))
        tk.Label(tag_frame_outer, text='比較清單：', bg=C_PNL, fg='#6a8faf',
                 font=('Microsoft JhengHei', 8)).pack(side='left', padx=(8, 4), pady=4)
        self._cmp_tag_frame = tk.Frame(tag_frame_outer, bg=C_PNL)
        self._cmp_tag_frame.pack(side='left', fill='x', expand=True, pady=4)
        self._cmp_status_var = tk.StringVar(value='請加入 ETF 後點擊「更新比較」')
        tk.Label(tag_frame_outer, textvariable=self._cmp_status_var,
                 bg=C_PNL, fg='#6a8faf', font=('Microsoft JhengHei', 8)).pack(
            side='right', padx=8)

        # ── 比較表格（可捲動）────────────────────────────────────────────────
        sc_outer = tk.Frame(f, bg=C_BG)
        sc_outer.pack(fill='both', expand=True, padx=8, pady=(0, 8))

        vsb = ttk.Scrollbar(sc_outer, orient='vertical')
        vsb.pack(side='right', fill='y')
        hsb = ttk.Scrollbar(sc_outer, orient='horizontal')
        hsb.pack(side='bottom', fill='x')

        self._cmp_canvas = tk.Canvas(sc_outer, bg=C_BG,
                                     yscrollcommand=vsb.set,
                                     xscrollcommand=hsb.set,
                                     highlightthickness=0)
        self._cmp_canvas.pack(side='left', fill='both', expand=True)
        vsb.config(command=self._cmp_canvas.yview)
        hsb.config(command=self._cmp_canvas.xview)

        self._cmp_inner = tk.Frame(self._cmp_canvas, bg=C_BG)
        _win = self._cmp_canvas.create_window(0, 0, anchor='nw', window=self._cmp_inner)

        def _cfg_inner(e):
            self._cmp_canvas.configure(scrollregion=self._cmp_canvas.bbox('all'))
        self._cmp_inner.bind('<Configure>', _cfg_inner)
        def _cfg_canvas(e):
            self._cmp_canvas.itemconfig(_win, width=max(e.width, self._cmp_inner.winfo_reqwidth()))
        self._cmp_canvas.bind('<Configure>', _cfg_canvas)
        def _mw_cmp(e):
            self._cmp_canvas.yview_scroll(-1 if e.delta > 0 else 1, 'units')
        self._cmp_canvas.bind('<MouseWheel>', _mw_cmp)
        self._cmp_inner.bind('<MouseWheel>', _mw_cmp)

        def _bind_cmp_scroll(widget):
            """遞迴將 MouseWheel 綁定到 _cmp_inner 下所有子元件。"""
            widget.bind('<MouseWheel>', _mw_cmp)
            for child in widget.winfo_children():
                _bind_cmp_scroll(child)
        self._bind_cmp_scroll = _bind_cmp_scroll

        self._cmp_etf_list: list[tuple[str, str]] = []  # [(code, name), ...]
        self._cmp_tags:     list[tk.Frame] = []

        # 預先顯示欄位說明
        self._cmp_show_placeholder()

    def _cmp_show_placeholder(self):
        for w in self._cmp_inner.winfo_children():
            w.destroy()
        tk.Label(self._cmp_inner, text='請加入至少 2 個 ETF 後點擊「更新比較」',
                 bg=C_BG, fg='#555577', font=('Microsoft JhengHei', 11)).pack(
            pady=60, padx=40)

    # ── ETF 成分股變化子頁面 ──────────────────────────────────────────────────
    def _build_etf_change_sub(self, f):
        C_PNL   = '#1a1a2e'
        self._chg_etf_selected = {code for code, _ in ACTIVE_ETFS}  # 預設全選
        self._chg_period       = tk.StringVar(value='1d')
        self._chg_status       = tk.StringVar(value='')
        self._chg_rects_buy    = []
        self._chg_rects_sell   = []
        self._chg_fetching     = False
        self._chg_agg_buy      = {}
        self._chg_agg_sell     = {}
        self._chg_resize_after = None
        self._chg_metric       = tk.StringVar(value='weight')

        # ── ETF 選擇列（2 列各 5 顆切換按鈕）────────────────────────────────
        etf_bar = tk.Frame(f, bg=C_BG)
        etf_bar.pack(fill='x', padx=10, pady=(8, 2))
        tk.Label(etf_bar, text='主動型 ETF：', bg=C_BG, fg=C_FG,
                 font=('Microsoft JhengHei', 10)
                 ).grid(row=0, column=0, rowspan=2, sticky='ns', padx=(0, 8))
        self._chg_etf_btns = {}
        for i, (code, _lbl) in enumerate(ACTIVE_ETFS):
            r, c = divmod(i, 5)
            btn = tk.Label(etf_bar, text=code, bg='#2a5c8f', fg='white',  # 預設選取色
                           font=('Consolas', 9, 'bold'),
                           padx=8, pady=4, cursor='hand2', bd=1, relief='solid')
            btn.grid(row=r, column=c + 1, padx=3, pady=2)
            self._chg_etf_btns[code] = btn
            btn.bind('<Button-1>', lambda e, cd=code: self._chg_toggle_etf(cd))

        # ── 操作列 ────────────────────────────────────────────────────────────
        ctrl = tk.Frame(f, bg=C_BG)
        ctrl.pack(fill='x', padx=10, pady=(4, 6))
        tk.Label(ctrl, text='期間：', bg=C_BG, fg=C_FG,
                 font=('Microsoft JhengHei', 10)).pack(side='left')
        for pval, plbl in [('1d', '1日'), ('1w', '1週'), ('1m', '1月')]:
            tk.Radiobutton(ctrl, text=plbl, variable=self._chg_period, value=pval,
                           bg=C_BG, fg=C_FG, selectcolor='#2a3f6f',
                           activebackground=C_BG,
                           font=('Microsoft JhengHei', 10)).pack(side='left', padx=4)
        tk.Label(ctrl, text='  指標：', bg=C_BG, fg=C_FG,
                 font=('Microsoft JhengHei', 10)).pack(side='left')
        for mval, mlbl in [('weight', '比重(%)'), ('shares', '張數')]:
            tk.Radiobutton(ctrl, text=mlbl, variable=self._chg_metric, value=mval,
                           bg=C_BG, fg=C_FG, selectcolor='#2a3f6f',
                           activebackground=C_BG,
                           font=('Microsoft JhengHei', 10),
                           command=self._chg_refresh).pack(side='left', padx=4)
        ttk.Button(ctrl, text='擷取所有快照', style='Nav.TButton',
                   command=self._chg_fetch_all).pack(side='left', padx=(16, 4))
        ttk.Button(ctrl, text='顯示變化', style='Nav.TButton',
                   command=self._chg_refresh).pack(side='left', padx=4)
        ttk.Label(ctrl, textvariable=self._chg_status,
                  foreground='#7090c0').pack(side='left', padx=(10, 0))

        # ── 主要內容區（上下分割）────────────────────────────────────────────
        content = tk.Frame(f, bg=C_BG)
        content.pack(fill='both', expand=True, padx=8, pady=(0, 6))
        content.rowconfigure(0, weight=6)
        content.rowconfigure(1, weight=5)
        content.columnconfigure(0, weight=1)

        # 上半：買入 / 賣出 Treemap Canvas
        top_pane = tk.Frame(content, bg=C_BG)
        top_pane.grid(row=0, column=0, sticky='nsew', pady=(0, 4))
        top_pane.columnconfigure(0, weight=1)
        top_pane.columnconfigure(1, weight=1)
        top_pane.rowconfigure(1, weight=1)

        tk.Label(top_pane, text='▲  買入 / 增加比例', bg='#0e3018', fg='#80ff80',
                 font=('Microsoft JhengHei', 10, 'bold'), pady=4
                 ).grid(row=0, column=0, sticky='ew', padx=(0, 2))
        tk.Label(top_pane, text='▼  賣出 / 減少比例', bg='#300e0e', fg='#ff8080',
                 font=('Microsoft JhengHei', 10, 'bold'), pady=4
                 ).grid(row=0, column=1, sticky='ew', padx=(2, 0))

        self._chg_cv_buy  = tk.Canvas(top_pane, bg='#0a1a0e', highlightthickness=0)
        self._chg_cv_buy.grid(row=1, column=0, sticky='nsew', padx=(0, 2))
        self._chg_cv_sell = tk.Canvas(top_pane, bg='#1a0a0a', highlightthickness=0)
        self._chg_cv_sell.grid(row=1, column=1, sticky='nsew', padx=(2, 0))

        self._chg_cv_buy.bind('<Motion>',    lambda e: self._chg_on_motion(e, 'buy'))
        self._chg_cv_buy.bind('<Leave>',     lambda e: self._chg_hide_tip())
        self._chg_cv_sell.bind('<Motion>',   lambda e: self._chg_on_motion(e, 'sell'))
        self._chg_cv_sell.bind('<Leave>',    lambda e: self._chg_hide_tip())
        self._chg_cv_buy.bind('<Configure>',  lambda e: self._chg_on_resize())
        self._chg_cv_sell.bind('<Configure>', lambda e: self._chg_on_resize())

        # 下半：異動統計 Treeview
        bot_pane = tk.Frame(content, bg=C_BG)
        bot_pane.grid(row=1, column=0, sticky='nsew')
        tk.Label(bot_pane, text='ETF 成分股異動統計（依活躍程度排序）',
                 bg='#2a3f6f', fg='white',
                 font=('Microsoft JhengHei', 10, 'bold'), pady=4
                 ).pack(fill='x')
        tv_fr = tk.Frame(bot_pane, bg=C_PNL)
        tv_fr.pack(fill='both', expand=True)
        _cols = [('code', '代號', 58, 'center'),
                 ('name', '名稱', 95, 'w'),
                 ('buy_etfs',  '買入 ETF（比例/張數）', 280, 'w'),
                 ('sell_etfs', '賣出 ETF（比例/張數）', 280, 'w'),
                 ('net', '淨變化', 120, 'center')]
        ttk.Style().configure('Chg.Treeview', rowheight=20)
        self._chg_sum_tv = ttk.Treeview(tv_fr,
                                         columns=[c for c, _n, _w, _a in _cols],
                                         show='headings', height=8,
                                         style='Chg.Treeview')
        for cid, cname, cw, anchor in _cols:
            self._chg_sum_tv.heading(cid, text=cname)
            self._chg_sum_tv.column(cid, width=cw, anchor=anchor)
        vsb = ttk.Scrollbar(tv_fr, orient='vertical',
                             command=self._chg_sum_tv.yview)
        self._chg_sum_tv.configure(yscrollcommand=vsb.set)
        vsb.pack(side='right', fill='y')
        self._chg_sum_tv.pack(fill='both', expand=True)
        self._chg_sum_tv.tag_configure('buy',  background='#0a1e0e', foreground='#80ff80')
        self._chg_sum_tv.tag_configure('sell', background='#1e0a0a', foreground='#ff8080')
        self._chg_sum_tv.tag_configure('both', background='#1a1a2e', foreground='#c0c0c0')

    # ── ETF 成分股變化 helpers ────────────────────────────────────────────────

    def _chg_toggle_etf(self, code: str):
        btn = self._chg_etf_btns[code]
        if code in self._chg_etf_selected:
            self._chg_etf_selected.discard(code)
            btn.config(bg='#2a3040', fg='#8090a0')
        else:
            self._chg_etf_selected.add(code)
            btn.config(bg='#2a5c8f', fg='white')

    def _chg_fetch_all(self):
        if self._chg_fetching:
            return
        sel = list(self._chg_etf_selected)
        if not sel:
            self._chg_status.set('請先選擇 ETF')
            return
        self._chg_fetching = True
        self._chg_status.set(f'擷取中 0/{len(sel)}…')
        import threading
        done = [0]
        lock = threading.Lock()

        def _fetch_one(code):
            holdings = _fetch_moneydj_etf_snapshot(code)
            if holdings:
                _save_etf_holdings_snapshot(code, holdings)
            with lock:
                done[0] += 1
                n = done[0]
            if n < len(sel):
                self.after(0, lambda n=n: self._chg_status.set(f'擷取中 {n}/{len(sel)}…'))
            else:
                self._chg_fetching = False
                self.after(0, lambda: self._chg_status.set(f'擷取完成，分析中…'))
                self.after(100, self._chg_refresh)

        for code in sel:
            threading.Thread(target=_fetch_one, args=(code,), daemon=True).start()

    def _chg_refresh(self):
        from datetime import date as _ddate, timedelta as _td
        _load_twse_stock_names()
        sel = list(self._chg_etf_selected)
        if not sel:
            self._chg_status.set('請先選擇 ETF')
            return
        days   = {'1d': 1, '1w': 7, '1m': 30}[self._chg_period.get()]
        cutoff = (_ddate.today() - _td(days=days)).strftime('%Y%m%d')

        agg_buy  = {}   # stock_code → [(etf, weight, shares)]
        agg_sell = {}
        missing  = []

        for etf in sel:
            hist  = _load_etf_holdings_history(etf)
            if not hist:
                missing.append(etf)
                continue
            dates = sorted(hist.keys(), reverse=True)
            new_d = dates[0]
            old_d = next((d for d in dates[1:] if d <= cutoff), None)
            if old_d is None:
                old_d = dates[-1] if len(dates) >= 2 else None
            if not old_d or old_d == new_d:
                missing.append(etf)
                continue
            diff = _diff_etf_holdings(hist[old_d], hist[new_d])
            for item in diff['entered']:
                w = item['weight']
                s = item.get('shares', 0)
                if w > 0 or s > 0:
                    agg_buy.setdefault(item['code'], []).append((etf, w, s))
            for item in diff['exited']:
                w = item['weight']
                s = item.get('shares', 0)
                if w > 0 or s > 0:
                    agg_sell.setdefault(item['code'], []).append((etf, w, s))
            for item in diff['changed']:
                dw = item['delta']
                ds = item.get('delta_shares', 0)
                if dw > 0 or ds > 0:
                    agg_buy.setdefault(item['code'], []).append(
                        (etf, max(dw, 0.0), max(ds, 0)))
                if dw < 0 or ds < 0:
                    agg_sell.setdefault(item['code'], []).append(
                        (etf, abs(min(dw, 0.0)), abs(min(ds, 0))))

        ok = len(sel) - len(missing)
        status = f'已分析 {ok}/{len(sel)} 支'
        if missing:
            status += f'（{", ".join(missing)} 快照不足）'
        self._chg_status.set(status)

        self._chg_agg_buy  = agg_buy
        self._chg_agg_sell = agg_sell
        self._chg_draw_treemap(self._chg_cv_buy,  agg_buy,  'buy')
        self._chg_draw_treemap(self._chg_cv_sell, agg_sell, 'sell')
        self._chg_update_summary(agg_buy, agg_sell)

    def _chg_draw_treemap(self, canvas, agg, side):
        canvas.delete('all')
        if not agg:
            return
        canvas.update_idletasks()
        cw, ch = canvas.winfo_width(), canvas.winfo_height()
        if cw < 10 or ch < 10:
            return
        is_shares = self._chg_metric.get() == 'shares'
        def _sz(w, s): return s / 1000 if is_shares else w
        items = sorted([(sum(_sz(w, s) for _, w, s in lst), code)
                        for code, lst in agg.items()
                        if sum(_sz(w, s) for _, w, s in lst) > 0], reverse=True)
        rects_raw = _squarify_layout(items, 0, 0, cw, ch)
        names = _TWSE_NAMES_CACHE
        rects_data = []
        for x1, y1, x2, y2, code in rects_raw:
            etf_list  = agg[code]
            total_w   = sum(_sz(w, s) for _, w, s in etf_list)
            intensity = min(1.0, total_w / (2000.0 if is_shares else 8.0))
            if side == 'buy':
                fill = '#{:02x}{:02x}{:02x}'.format(
                    int(14 + intensity * 30),
                    int(80 + intensity * 150),
                    int(20 + intensity * 30))
            else:
                fill = '#{:02x}{:02x}{:02x}'.format(
                    int(80 + intensity * 150),
                    int(14 + intensity * 30),
                    int(20 + intensity * 30))
            canvas.create_rectangle(x1 + 1, y1 + 1, x2 - 1, y2 - 1,
                                     fill=fill, outline='#0a0a14', width=1)
            bw, bh = x2 - x1, y2 - y1
            cy = (y1 + y2) / 2
            if bw > 38 and bh > 18:
                canvas.create_text((x1 + x2) / 2, cy - (8 if bh > 34 else 0),
                                    text=code, fill='white',
                                    font=('Consolas', 8, 'bold'), anchor='center')
            if bw > 38 and bh > 34:
                canvas.create_text((x1 + x2) / 2, cy + 9,
                                    text=names.get(code, '')[:5],
                                    fill='#b0f0b0' if side == 'buy' else '#f0b0b0',
                                    font=('Microsoft JhengHei', 7), anchor='center')
            rects_data.append((x1, y1, x2, y2, code, etf_list))
        if side == 'buy':
            self._chg_rects_buy  = rects_data
        else:
            self._chg_rects_sell = rects_data

    def _chg_on_motion(self, event, side):
        rects = self._chg_rects_buy if side == 'buy' else self._chg_rects_sell
        mx, my = event.x, event.y
        is_shares = self._chg_metric.get() == 'shares'
        def _sz(w, s): return s / 1000 if is_shares else w
        def _fmt_ws(w, s):
            parts = []
            if w > 0.005: parts.append(f'{w:.2f}%')
            if s > 0:     parts.append(f'{s // 1000:.0f}張')
            return ' / '.join(parts) or '-'
        for x1, y1, x2, y2, code, etf_list in rects:
            if x1 <= mx <= x2 and y1 <= my <= y2:
                names = _TWSE_NAMES_CACHE
                lbl   = '買入' if side == 'buy' else '賣出'
                lines = [f'{code}  {names.get(code, "")}', '─' * 24]
                for etf, w, s in sorted(etf_list, key=lambda x: -_sz(x[1], x[2])):
                    lines.append(f'  {etf}   {lbl} {_fmt_ws(w, s)}')
                total_w = sum(w for _, w, s in etf_list)
                total_s = sum(s for _, w, s in etf_list)
                lines += ['─' * 24,
                          f'合計 {lbl}：{_fmt_ws(total_w, total_s)}']
                rx = event.widget.winfo_rootx() + mx
                ry = event.widget.winfo_rooty() + my
                self._chg_show_tip(rx, ry, '\n'.join(lines))
                return
        self._chg_hide_tip()

    def _chg_show_tip(self, x, y, text):
        if not hasattr(self, '_chg_tip_win') or self._chg_tip_win is None:
            self._chg_tip_win = tk.Toplevel(self)
            self._chg_tip_win.wm_overrideredirect(True)
            self._chg_tip_win.attributes('-topmost', True)
            self._chg_tip_lbl = tk.Label(
                self._chg_tip_win, bg='#1c2a3a', fg='#c0d8f0',
                font=('Microsoft JhengHei', 9), justify='left',
                padx=10, pady=8, relief='solid', bd=1)
            self._chg_tip_lbl.pack()
        self._chg_tip_lbl.config(text=text)
        self._chg_tip_win.wm_geometry(f'+{x + 16}+{y + 16}')
        self._chg_tip_win.deiconify()

    def _chg_hide_tip(self):
        if hasattr(self, '_chg_tip_win') and self._chg_tip_win:
            self._chg_tip_win.withdraw()

    def _chg_on_resize(self):
        if self._chg_resize_after:
            self.after_cancel(self._chg_resize_after)
        self._chg_resize_after = self.after(150, self._chg_redraw_on_resize)

    def _chg_redraw_on_resize(self):
        self._chg_resize_after = None
        if self._chg_agg_buy or self._chg_agg_sell:
            self._chg_draw_treemap(self._chg_cv_buy,  self._chg_agg_buy,  'buy')
            self._chg_draw_treemap(self._chg_cv_sell, self._chg_agg_sell, 'sell')

    def _chg_update_summary(self, agg_buy, agg_sell):
        tv = self._chg_sum_tv
        for row in tv.get_children():
            tv.delete(row)
        all_codes = set(agg_buy) | set(agg_sell)
        names = _TWSE_NAMES_CACHE

        def _score(c):
            bl = agg_buy.get(c, [])
            sl = agg_sell.get(c, [])
            return (len(bl) + len(sl),
                    sum(w for _, w, s in bl) + sum(w for _, w, s in sl))

        def _wrap(items, per_line=3):
            lines = ['  '.join(items[i:i+per_line])
                     for i in range(0, len(items), per_line)]
            return '\n'.join(lines)

        def _fmt_ws(w, s):
            parts = []
            if w > 0.005: parts.append(f'{w:.1f}%')
            if s > 0:     parts.append(f'{s // 1000:.0f}張')
            return '/'.join(parts) or '-'

        is_shares = self._chg_metric.get() == 'shares'
        def _sort_key(t): return t[2] / 1000 if is_shares else t[1]

        max_lines = 1
        rows = []
        for code in sorted(all_codes, key=_score, reverse=True):
            bl = agg_buy.get(code, [])
            sl = agg_sell.get(code, [])
            buy_items  = [f'{e}(+{_fmt_ws(w, s)})' for e, w, s in sorted(bl, key=_sort_key, reverse=True)]
            sell_items = [f'{e}(-{_fmt_ws(w, s)})' for e, w, s in sorted(sl, key=_sort_key, reverse=True)]
            buy_str  = _wrap(buy_items)
            sell_str = _wrap(sell_items)
            row_lines = max(buy_str.count('\n') + 1 if buy_str else 1,
                            sell_str.count('\n') + 1 if sell_str else 1)
            max_lines = max(max_lines, row_lines)
            net_w = sum(w for _, w, s in bl) - sum(w for _, w, s in sl)
            net_s = (sum(s for _, w, s in bl) - sum(s for _, w, s in sl)) // 1000
            nparts = []
            if abs(net_w) > 0.005: nparts.append(f'{net_w:+.1f}%')
            if net_s != 0:         nparts.append(f'{net_s:+.0f}張')
            nstr = ' / '.join(nparts) or '0'
            tag  = ('buy'  if bl and not sl else
                    'sell' if sl and not bl else 'both')
            rows.append((code, names.get(code, code), buy_str, sell_str, nstr, tag))

        ttk.Style().configure('Chg.Treeview', rowheight=max_lines * 20)
        for code, name, buy_str, sell_str, nstr, tag in rows:
            tv.insert('', 'end',
                      values=(code, name, buy_str, sell_str, nstr),
                      tags=(tag,))

    def _cmp_select_combo(self):
        sel = self._cmp_combo.get().strip()
        if sel:
            self._cmp_search_var.set(sel.split()[0])

    def _cmp_add_etf(self):
        raw  = self._cmp_search_var.get().strip() or self._cmp_combo.get().strip()
        code = raw.split()[0] if raw else ''
        if not code:
            return
        # 找名稱
        name = next((n for c, n in ETF_LIST if c == code), code)
        # 已在清單中則跳過
        if any(c == code for c, _ in self._cmp_etf_list):
            return
        self._cmp_etf_list.append((code, name))
        self._cmp_rebuild_tags()
        self._cmp_search_var.set('')
        self._cmp_combo.set('')

    def _cmp_remove_etf(self, code: str):
        self._cmp_etf_list = [(c, n) for c, n in self._cmp_etf_list if c != code]
        self._cmp_rebuild_tags()
        if not self._cmp_etf_list:
            self._cmp_show_placeholder()

    def _cmp_clear_all(self):
        self._cmp_etf_list.clear()
        self._cmp_rebuild_tags()
        self._cmp_show_placeholder()

    def _cmp_rebuild_tags(self):
        for w in self._cmp_tag_frame.winfo_children():
            w.destroy()
        self._cmp_tags.clear()
        for code, name in self._cmp_etf_list:
            tag = tk.Frame(self._cmp_tag_frame, bg='#2a3f6f',
                           padx=6, pady=2, relief='flat')
            tag.pack(side='left', padx=3, pady=2)
            tk.Label(tag, text=f'{code} {name}', bg='#2a3f6f', fg='white',
                     font=('Microsoft JhengHei', 8)).pack(side='left')
            _x = tk.Label(tag, text=' ×', bg='#2a3f6f', fg='#f07070',
                          font=('Microsoft JhengHei', 9, 'bold'), cursor='hand2')
            _x.pack(side='left', padx=(2, 0))
            _x.bind('<Button-1>', lambda e, c=code: self._cmp_remove_etf(c))
            self._cmp_tags.append(tag)

    def _cmp_run(self):
        if len(self._cmp_etf_list) < 1:
            self._cmp_status_var.set('請先加入 ETF')
            return
        self._cmp_status_var.set('載入中…')
        codes = list(self._cmp_etf_list)
        threading.Thread(target=self._cmp_fetch_and_render,
                         args=(codes,), daemon=True).start()

    def _cmp_fetch_and_render(self, etf_pairs: list):
        import concurrent.futures as _cf
        results = {}

        def _yf_history_safe(code: str, period: str = '3y'):
            """yfinance 1.2.0 有 TypeError bug，逐個 suffix 嘗試並捕捉所有例外。
            auto_adjust=True 確保遇到股票分割時使用還原股價（adjusted close）。
            每個 suffix 最多等 12 秒，避免 yfinance 掛住整個 fetch。"""
            def _fix_unadjusted_splits(s):
                """偵測單日跌幅 >50% 的異常（yfinance 未補齊的拆分），補正拆分前歷史價格。"""
                if s is None or len(s) < 2:
                    return s
                import numpy as _np2
                pct = s.pct_change()
                split_locs = [i for i, v in enumerate(pct) if i > 0 and v < -0.5]
                for i in split_locs:
                    pre = float(s.iloc[i - 1])
                    post = float(s.iloc[i])
                    if post <= 0:
                        continue
                    ratio = pre / post
                    s = s.copy()
                    s.iloc[:i] = s.iloc[:i] / ratio
                return s

            def _fetch_sfx(sfx):
                h = yf.Ticker(code + sfx).history(period=period, auto_adjust=True)
                if h is not None and not h.empty:
                    return _fix_unadjusted_splits(h['Close'].dropna())
                return None

            for sfx in ['.TW', '.TWO']:
                _ex = _cf.ThreadPoolExecutor(max_workers=1)
                try:
                    fut = _ex.submit(_fetch_sfx, sfx)
                    result = fut.result(timeout=12)
                    _ex.shutdown(wait=False)
                    if result is not None:
                        return result
                except Exception:
                    _ex.shutdown(wait=False)
            return None

        def _fetch_one(code, name):
            etf_name, components, _ = fetch_etf_data(code)
            meta = fetch_etf_meta(code)
            perf = {}
            hist_3y = _yf_history_safe(code, '3y')
            if hist_3y is not None and len(hist_3y) >= 2:
                for label, trade_days in [('1M', 21), ('3M', 63), ('6M', 126),
                                          ('1Y', 252), ('3Y', 756)]:
                    # 資料不足該期間就跳過，不用第一筆湊數
                    if len(hist_3y) < trade_days:
                        continue
                    idx = -(trade_days + 1)
                    try:
                        p0 = float(hist_3y.iloc[idx])
                        p1 = float(hist_3y.iloc[-1])
                        if p0 > 0:
                            r = (p1 - p0) / p0 * 100
                            if trade_days >= 252:
                                r = ((1 + r / 100) ** (252 / abs(idx)) - 1) * 100
                            perf[label] = r
                    except Exception:
                        pass
            results[code] = {
                'code': code, 'name': etf_name or name,
                'components': components,
                'meta': meta, 'perf': perf,
                'hist': hist_3y,
            }

        def _safe_fetch(pair):
            code, name = pair
            try:
                _fetch_one(code, name)
            except Exception as _e:
                _net_log(f'ETF compare fetch {code}: {_e}')
                results[code] = {'code': code, 'name': name, 'components': [],
                                 'meta': {}, 'perf': {}, 'hist': None}

        pool = _cf.ThreadPoolExecutor(max_workers=8)
        futs = [pool.submit(_safe_fetch, p) for p in etf_pairs]
        _cf.wait(futs, timeout=45)          # 最多等 45 秒
        pool.shutdown(wait=False)           # 不等卡住的 future，直接繼續

        self._ui_call(lambda: self._cmp_render(etf_pairs, results))

    def _cmp_render(self, etf_pairs: list, results: dict):
        import matplotlib.pyplot as _plt
        import matplotlib.ticker as _mticker
        import numpy as _np

        for w in self._cmp_inner.winfo_children():
            w.destroy()

        C_HDR  = '#1a1a2e'
        C_ROW1 = '#111122'
        C_ROW2 = '#0e0e1a'
        C_SEP  = '#2a2a3e'
        FNT    = ('Microsoft JhengHei', 9)
        FNT_B  = ('Microsoft JhengHei', 9, 'bold')
        FNT_H  = ('Microsoft JhengHei', 8)

        codes   = [c for c, _ in etf_pairs]
        ordered = [results[c] for c in codes if c in results]

        # 補齊中文名稱：若 fetch 時 TWSE 快取尚未就緒，名稱可能等於代號，在此補查
        for _d in ordered:
            for _h in _d.get('components', []):
                if not _has_cjk(_h.get('name', '')):
                    _cached = _TWSE_NAMES_CACHE.get(_h.get('code', ''))
                    if _cached:
                        _h['name'] = _cached
        if not ordered:
            self._cmp_show_placeholder()
            self._cmp_status_var.set('無法取得資料')
            return

        # ── 圖形化比較區 ──────────────────────────────────────────────────────
        COLORS = ['#5b9cf6', '#f07070', '#4ec94e', '#f0c060',
                  '#c07af0', '#60d0e0', '#f0a060', '#80e0a0']

        chart_frame = tk.Frame(self._cmp_inner, bg='#0d0d1a')
        chart_frame.pack(fill='x', padx=0, pady=(0, 4))

        import matplotlib.dates as _mdates
        FS_TITLE = 11   # chart title font size
        FS_TICK  = 9    # tick label font size
        FS_LEG   = 9    # legend font size

        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg as _FCA

        # ── 上方圖：淨值走勢（獨立 figure，工具列接在下方）─────────────────
        fig_line = plt.Figure(figsize=(12, 4.5), dpi=100, facecolor='#0d0d1a')
        fig_line.subplots_adjust(left=0.07, right=0.97, top=0.90, bottom=0.12)
        ax_line = fig_line.add_subplot(1, 1, 1)
        ax_line.set_facecolor('#111122')
        ax_line.tick_params(colors='#888899', labelsize=FS_TICK)
        ax_line.yaxis.set_major_formatter(_mticker.FormatStrFormatter('%.0f'))
        for spine in ax_line.spines.values():
            spine.set_edgecolor('#333355')
        ax_line.xaxis.set_major_formatter(_mdates.DateFormatter('%m/%y'))

        # 儲存各 ETF 完整原始價格，供期間切換時重算基準
        _hist_data = []
        for i, d in enumerate(ordered):
            hist = d.get('hist')
            if hist is None or len(hist) < 5:
                continue
            if hasattr(hist.index, 'tz') and hist.index.tz is not None:
                hist = hist.copy()
                hist.index = hist.index.tz_localize(None)
            _hist_data.append({
                'code': d['code'], 'name': d['name'],
                'hist': hist, 'color': COLORS[i % len(COLORS)],
            })

        line_canvas = _FCA(fig_line, master=chart_frame)
        line_cw = line_canvas.get_tk_widget()
        line_cw.configure(bg='#0d0d1a')
        line_cw.pack(fill='x', padx=0, pady=0)

        # ── 期間選擇按鈕 ──────────────────────────────────────────────────────
        import datetime as _datetime
        import pandas as _pd_line
        _period_frame = tk.Frame(chart_frame, bg='#1a1a2e')
        _period_frame.pack(fill='x', pady=(2, 4))
        _cur_pid = ['1Y']   # mutable container so closure can read current period

        # 查價線容器：用 list 讓 _draw_period 在 ax_line.cla() 後可更新引用
        _vline = [ax_line.axvline(x=ax_line.get_xlim()[0],
                                  color='#ffffff', linewidth=0.8,
                                  linestyle='--', alpha=0.5, visible=False)]
        _hline = [ax_line.axhline(y=ax_line.get_ylim()[0],
                                  color='#ffffff', linewidth=0.8,
                                  linestyle='--', alpha=0.5, visible=False)]

        def _draw_period(pid, btn_ref):
            _cur_pid[0] = pid
            for _b in _period_btns:
                _b.config(bg='#2a2a4e', fg='#8899cc')
            btn_ref.config(bg='#4a6fa5', fg='white')

            now = _datetime.datetime.now()
            days_map = {'1M': 30, '3M': 90, '6M': 180, '1Y': 365}
            if pid == 'ALL':
                x_start = None
            else:
                x_start = _pd_line.Timestamp(now - _datetime.timedelta(days=days_map[pid]))

            ax_line.cla()
            ax_line.set_facecolor('#111122')
            ax_line.tick_params(colors='#888899', labelsize=FS_TICK)
            ax_line.yaxis.set_major_formatter(_mticker.FormatStrFormatter('%.0f'))
            for spine in ax_line.spines.values():
                spine.set_edgecolor('#333355')
            ax_line.xaxis.set_major_formatter(_mdates.DateFormatter('%m/%y'))
            lbl_suffix = '' if pid == 'ALL' else {'1M':'近1個月','3M':'近3個月',
                          '6M':'近6個月','1Y':'近1年'}[pid]
            ax_line.set_title(f'淨值走勢比較（基準＝100，{lbl_suffix}）' if lbl_suffix
                              else '淨值走勢比較（基準＝100，全部）',
                              color='#aecde8', fontsize=FS_TITLE,
                              fontfamily=CHART_FONT, pad=8)

            any_line = False
            for hd in _hist_data:
                h = hd['hist']
                # 截取期間內資料
                if x_start is not None:
                    h_slice = h[h.index >= x_start]
                    if len(h_slice) < 2:
                        h_slice = h.iloc[-2:]
                else:
                    h_slice = h
                # 以期間起點為基準100
                normed = h_slice / float(h_slice.iloc[0]) * 100
                ax_line.plot(normed.index, normed.values,
                             color=hd['color'], linewidth=2.0,
                             label=f'{hd["code"]} {hd["name"][:8]}')
                any_line = True

            if any_line:
                ax_line.axhline(100, color='#444466', linewidth=0.8, linestyle='--')
                ax_line.legend(loc='upper left', fontsize=FS_LEG,
                               facecolor='#1a1a2e', edgecolor='#333355',
                               labelcolor='white')
                interval = {'1M': 1, '3M': 1, '6M': 2, '1Y': 2, 'ALL': 6}.get(pid, 2)
                ax_line.xaxis.set_major_locator(_mdates.MonthLocator(interval=interval))
                fig_line.autofmt_xdate(rotation=30, ha='right')
            else:
                ax_line.text(0.5, 0.5, '無歷史資料', ha='center', va='center',
                             color='#555577', fontsize=12, transform=ax_line.transAxes)
            # 重建查價線（ax_line.cla() 清掉了舊的）
            _vline[0] = ax_line.axvline(x=ax_line.get_xlim()[0],
                                        color='#ffffff', linewidth=0.8,
                                        linestyle='--', alpha=0.5, visible=False)
            _hline[0] = ax_line.axhline(y=ax_line.get_ylim()[0],
                                        color='#ffffff', linewidth=0.8,
                                        linestyle='--', alpha=0.5, visible=False)
            line_canvas.draw_idle()

        _period_btns = []
        for _pid, _plbl in [('1M','1個月'), ('3M','3個月'), ('6M','6個月'),
                              ('1Y','1年'), ('ALL','全部')]:
            _b = tk.Button(_period_frame, text=_plbl,
                           bg='#4a6fa5' if _pid == '1Y' else '#2a2a4e',
                           fg='white' if _pid == '1Y' else '#8899cc',
                           font=('Microsoft JhengHei', 9),
                           relief='flat', padx=10, pady=3, cursor='hand2')
            _b.pack(side='left', padx=3)
            _period_btns.append(_b)
            _b.config(command=lambda p=_pid, b=_b: _draw_period(p, b))

        # 預設顯示近1年
        _draw_period('1Y', _period_btns[3])
        line_canvas.draw()

        # ── 下方圖：績效長條 + 熱力圖（獨立 figure）─────────────────────────
        fig = plt.Figure(figsize=(12, 5.0), dpi=100, facecolor='#0d0d1a')
        fig.subplots_adjust(left=0.07, right=0.97, top=0.90, bottom=0.12,
                            wspace=0.32)

        # ── 左：績效長條圖 ───────────────────────────────────────────────────
        ax_bar = fig.add_subplot(1, 2, 1)
        ax_bar.set_facecolor('#111122')
        ax_bar.set_title('各期間報酬率 (%)',
                          color='#aecde8', fontsize=FS_TITLE,
                          fontfamily=CHART_FONT, pad=8)
        perf_keys = ['1M', '3M', '6M', '1Y', '3Y']
        perf_lbls = ['1月', '3月', '6月', '1年\n(年化)', '3年\n(年化)']
        x         = _np.arange(len(perf_keys))
        n         = len(ordered)
        bar_w     = 0.7 / max(n, 1)

        _bar_rects: list[tuple] = []
        for i, d in enumerate(ordered):
            xpos = x + (i - n / 2 + 0.5) * bar_w
            for xi, pk in zip(xpos, perf_keys):
                v = d['perf'].get(pk)
                if v is None:
                    continue
                bars = ax_bar.bar(xi, v, width=bar_w * 0.85,
                                  color=COLORS[i % len(COLORS)], alpha=0.85)
                _bar_rects.append((bars[0], i, pk))

        ax_bar.axhline(0, color='#555577', linewidth=0.8)
        ax_bar.set_xticks(x)
        ax_bar.set_xticklabels(perf_lbls, fontsize=FS_TICK, color='#888899')
        ax_bar.tick_params(colors='#888899', labelsize=FS_TICK)
        ax_bar.yaxis.set_major_formatter(_mticker.FormatStrFormatter('%.0f%%'))
        for spine in ax_bar.spines.values():
            spine.set_edgecolor('#333355')

        # ── 右：成分股相似度熱力圖 ──────────────────────────────────────────
        ax_heat = fig.add_subplot(1, 2, 2)
        ax_heat.set_facecolor('#111122')
        ax_heat.set_title('成分股相似度 % (加權Jaccard)',
                           color='#aecde8', fontsize=FS_TITLE,
                           fontfamily=CHART_FONT, pad=8)
        n_etf = len(ordered)
        matrix = _np.zeros((n_etf, n_etf))

        # 加權 Jaccard：Σ min(w_a, w_b) / Σ max(w_a, w_b)
        # 比單純集合 Jaccard 更能反映持倉比例的實質相似程度
        def _wjac(comps_a, comps_b):
            wa = {h['code']: h.get('weight', 0) for h in comps_a}
            wb = {h['code']: h.get('weight', 0) for h in comps_b}
            all_c = set(wa) | set(wb)
            if not all_c:
                return 0.0
            inter = sum(min(wa.get(c, 0), wb.get(c, 0)) for c in all_c)
            union_w = sum(max(wa.get(c, 0), wb.get(c, 0)) for c in all_c)
            return inter / union_w * 100 if union_w > 0 else 0.0

        for i, da in enumerate(ordered):
            for j, db in enumerate(ordered):
                if i == j:
                    matrix[i][j] = 100.0
                else:
                    matrix[i][j] = _wjac(da['components'], db['components'])
        short_names = [d['code'] for d in ordered]
        im = ax_heat.imshow(matrix, cmap='Blues', vmin=0, vmax=100, aspect='auto')
        ax_heat.set_xticks(range(n_etf))
        ax_heat.set_xticklabels(short_names, fontsize=FS_TICK, color='#888899', rotation=30)
        ax_heat.set_yticks(range(n_etf))
        ax_heat.set_yticklabels(short_names, fontsize=FS_TICK, color='#888899')
        _heat_fs = max(7, 10 - n_etf)
        for i in range(n_etf):
            for j in range(n_etf):
                ax_heat.text(j, i, f'{matrix[i][j]:.0f}%',
                             ha='center', va='center', fontsize=_heat_fs,
                             color='white' if matrix[i][j] > 50 else '#333344')
        fig.colorbar(im, ax=ax_heat, fraction=0.046, pad=0.04).ax.tick_params(
            labelsize=8, colors='#888899')

        cmp_fig_canvas = _FCA(fig, master=chart_frame)
        cw = cmp_fig_canvas.get_tk_widget()
        cw.configure(bg='#0d0d1a')
        cw.pack(fill='x', padx=0, pady=0)
        cmp_fig_canvas.draw()

        # store refs
        self._cmp_fig        = fig
        self._cmp_fig_line   = fig_line
        self._cmp_fig_canvas = cmp_fig_canvas
        self._cmp_line_canvas = line_canvas

        # ── Hover tooltip（每次 render 重建，確保欄數正確）─────────────────────
        if hasattr(self, '_cmp_bar_tip') and self._cmp_bar_tip and \
                self._cmp_bar_tip.winfo_exists():
            self._cmp_bar_tip.destroy()
        self._cmp_bar_tip = tk.Toplevel(self)
        self._cmp_bar_tip.overrideredirect(True)
        self._cmp_bar_tip.attributes('-topmost', True)
        self._cmp_bar_tip.configure(bg='#1e1e2e')
        _tip_inner = tk.Frame(self._cmp_bar_tip, bg='#252538', padx=8, pady=6)
        _tip_inner.pack(fill='both', expand=True, padx=1, pady=1)
        # 用 grid 支援雙欄：每欄放一個 Label
        _TIP_COLS = 2 if len(ordered) > 4 else 1
        self._cmp_tip_labels = []
        for _ci in range(_TIP_COLS):
            _lbl = tk.Label(_tip_inner, text='', bg='#252538', fg=C_FG,
                            font=('Microsoft JhengHei', 9),
                            justify='left', anchor='nw', padx=6)
            _lbl.grid(row=0, column=_ci, sticky='nw')
            if _ci < _TIP_COLS - 1:
                tk.Frame(_tip_inner, bg='#3a3a5e', width=1).grid(
                    row=0, column=_ci * 2 + 1, sticky='ns', padx=2)
            self._cmp_tip_labels.append(_lbl)
        self._cmp_bar_tip_lbl = self._cmp_tip_labels[0]   # backward compat
        self._cmp_tip_ncols   = _TIP_COLS
        self._cmp_bar_tip.withdraw()

        # ── Build line data lookup for hover ─────────────────────────────────
        # _line_data 保存完整歷史；hover 時顯示的漲幅與當前選取期間一致
        _line_data: list[dict] = []
        for hd in _hist_data:
            hist_full = hd['hist']
            if hasattr(hist_full.index, 'tz') and hist_full.index.tz is not None:
                hist_full = hist_full.copy()
                hist_full.index = hist_full.index.tz_localize(None)
            _line_data.append({
                'code': hd['code'], 'name': hd['name'],
                'hist_full': hist_full,
                'perf': next((d['perf'] for d in ordered if d['code'] == hd['code']), {}),
            })

        # ── 熱圖 hover 用的持股資料（etf_code → {stock_code: (name, weight)}）─
        _heat_holdings = {
            d['code']: {h['code']: (h.get('name', h['code']), h.get('weight', 0))
                        for h in d['components']}
            for d in ordered
        }

        def _on_cmp_motion(event):
            import matplotlib.dates as _md
            import pandas as _pd

            # ── Line chart hover ─────────────────────────────────────────────
            if event.inaxes is ax_line and event.xdata is not None and _line_data:
                try:
                    cursor_dt = _md.num2date(event.xdata).replace(tzinfo=None)
                except Exception:
                    self._cmp_bar_tip.withdraw()
                    return
                # 計算當前期間的起點，用來重算基準100
                import datetime as _dtm
                _pid_now = _cur_pid[0]
                _days_map = {'1M': 30, '3M': 90, '6M': 180, '1Y': 365}
                if _pid_now == 'ALL':
                    _period_start = None
                else:
                    _period_start = _pd.Timestamp(
                        _dtm.datetime.now() - _dtm.timedelta(days=_days_map[_pid_now]))

                lines_txt = []
                for ld in _line_data:
                    try:
                        h = ld['hist_full']
                        ts = _pd.Timestamp(cursor_dt)
                        # 截取期間資料並重算基準
                        if _period_start is not None:
                            h_slice = h[h.index >= _period_start]
                            if len(h_slice) < 2:
                                h_slice = h.iloc[-2:]
                        else:
                            h_slice = h
                        base   = float(h_slice.iloc[0])
                        pos    = min(h_slice.index.searchsorted(ts), len(h_slice) - 1)
                        price  = float(h_slice.iloc[pos])
                        chg    = (price - base) / base * 100
                        parts = [f'{ld["code"]} {ld["name"][:10]}',
                                 f'  現價 {price:.2f}  ({chg:+.1f}%)']
                        v3m = ld['perf'].get('3M')
                        if v3m is not None:
                            parts.append(f'  近3月 {v3m:+.2f}%')
                        lines_txt.append('\n'.join(parts))
                    except Exception:
                        pass
                if lines_txt:
                    # 分欄：超過4支時左右各半
                    _nc = getattr(self, '_cmp_tip_ncols', 1)
                    _lbls = getattr(self, '_cmp_tip_labels', [self._cmp_bar_tip_lbl])
                    _half = (len(lines_txt) + _nc - 1) // _nc
                    for _ci, _lbl in enumerate(_lbls):
                        _chunk = lines_txt[_ci * _half: (_ci + 1) * _half]
                        _lbl.config(text='\n─────\n'.join(_chunk))
                    wx = line_cw.winfo_rootx() + int(event.x) + 16
                    wy = line_cw.winfo_rooty() + int(line_cw.winfo_height() - event.y) + 16
                    self._cmp_bar_tip.deiconify()
                    self._place_tooltip(self._cmp_bar_tip, wx, wy, offset=0)
                _vline[0].set_xdata([event.xdata, event.xdata])
                _vline[0].set_visible(True)
                _hline[0].set_ydata([event.ydata, event.ydata])
                _hline[0].set_visible(True)
                line_canvas.draw_idle()
                return

            # ── Bar chart hover ──────────────────────────────────────────────
            if event.inaxes is ax_bar and event.xdata is not None:
                for rect, etf_idx, pk in _bar_rects:
                    if rect.get_x() <= event.xdata <= rect.get_x() + rect.get_width():
                        d = ordered[etf_idx]
                        hist = d.get('hist')
                        price_str = f'{float(hist.iloc[-1]):.2f}' if hist is not None and len(hist) > 0 else '—'
                        lines = [f'{d["code"]} {d["name"]}', f'現價：{price_str}']
                        for lbl, key in [('近1月','1M'),('近3月','3M'),('近6月','6M'),('近1年','1Y'),('近3年','3Y')]:
                            v = d['perf'].get(key)
                            if v is not None:
                                lines.append(f'{lbl}：{v:+.2f}%')
                        self._cmp_bar_tip_lbl.config(text='\n'.join(lines))
                        wx = cw.winfo_rootx() + int(event.x) + 12
                        wy = cw.winfo_rooty() + int(cw.winfo_height() - event.y) + 12
                        self._cmp_bar_tip.deiconify()
                        self._place_tooltip(self._cmp_bar_tip, wx, wy, offset=0)
                        _vline[0].set_visible(False)
                        cmp_fig_canvas.draw_idle()
                        return

            # ── Heatmap hover ────────────────────────────────────────────────
            if event.inaxes is ax_heat and event.xdata is not None and event.ydata is not None:
                col = int(round(event.xdata))
                row = int(round(event.ydata))
                if 0 <= row < n_etf and 0 <= col < n_etf:
                    ca = ordered[row]['code']
                    cb = ordered[col]['code']
                    ha = _heat_holdings.get(ca, {})
                    hb = _heat_holdings.get(cb, {})
                    common = sorted(ha.keys() & hb.keys(),
                                    key=lambda c: -(ha[c][1] + hb[c][1]))
                    if ca == cb:
                        tip_txt = f'{ca}（自身）\n持股：{len(ha)} 檔'
                    else:
                        j_val = matrix[row][col]
                        # 計算加權重疊比例
                        w_inter = sum(min(ha.get(c,(None,0))[1], hb.get(c,(None,0))[1])
                                      for c in common)
                        tip_txt = (f'{ca} vs {cb}\n'
                                   f'加權相似度：{j_val:.0f}%\n'
                                   f'共同持股 {len(common)} 檔，加權重疊 {w_inter:.1f}%\n'
                                   f'（共同持股，依比例排序）：')
                        if common:
                            lines = []
                            for c in common[:12]:
                                wa_v = ha.get(c, (c, 0))[1]
                                wb_v = hb.get(c, (c, 0))[1]
                                nm   = ha.get(c, hb.get(c, (c, 0)))[0][:5]
                                lines.append(f'{c} {nm}  {ca}:{wa_v:.1f}% / {cb}:{wb_v:.1f}%')
                            tip_txt += '\n' + '\n'.join(lines)
                            if len(common) > 12:
                                tip_txt += f'\n  …共 {len(common)} 檔'
                        else:
                            tip_txt += '\n  無重疊'
                    self._cmp_bar_tip_lbl.config(text=tip_txt)
                    wx = cw.winfo_rootx() + int(event.x) + 14
                    wy = cw.winfo_rooty() + int(cw.winfo_height() - event.y) + 14
                    self._cmp_bar_tip.deiconify()
                    self._place_tooltip(self._cmp_bar_tip, wx, wy, offset=0)
                    _vline[0].set_visible(False)
                    _hline[0].set_visible(False)
                    return

            # ── Neither axis ─────────────────────────────────────────────────
            _vline[0].set_visible(False)
            _hline[0].set_visible(False)
            self._cmp_bar_tip.withdraw()
            line_canvas.draw_idle()
            cmp_fig_canvas.draw_idle()

        def _on_line_leave(_event):
            _vline[0].set_visible(False)
            _hline[0].set_visible(False)
            self._cmp_bar_tip.withdraw()
            line_canvas.draw_idle()

        def _on_cmp_leave(_event):
            self._cmp_bar_tip.withdraw()
            cmp_fig_canvas.draw_idle()

        def _on_heat_click(event):
            if event.inaxes is not ax_heat or event.xdata is None or event.ydata is None:
                return
            col = int(round(event.xdata))
            row = int(round(event.ydata))
            if not (0 <= row < n_etf and 0 <= col < n_etf) or row == col:
                return
            ca = ordered[row]['code']
            cb = ordered[col]['code']
            ha = _heat_holdings.get(ca, {})
            hb = _heat_holdings.get(cb, {})
            all_c = set(ha) | set(hb)
            inter    = sum(min(ha.get(c,(None,0))[1], hb.get(c,(None,0))[1])     for c in all_c)
            a_excess = sum(ha.get(c,(None,0))[1] - min(ha.get(c,(None,0))[1], hb.get(c,(None,0))[1]) for c in all_c)
            b_excess = sum(hb.get(c,(None,0))[1] - min(ha.get(c,(None,0))[1], hb.get(c,(None,0))[1]) for c in all_c)
            j_val = matrix[row][col]
            common = sorted(ha.keys() & hb.keys(), key=lambda c: -(ha[c][1] + hb[c][1]))

            popup = tk.Toplevel(self)
            popup.title(f'相似度解析：{ca} vs {cb}')
            popup.configure(bg='#0d0d1a')
            popup.geometry('860x440')
            popup.attributes('-topmost', True)

            fig2, (ax_pie, ax_bar) = _plt.subplots(1, 2, figsize=(11, 5.2))
            fig2.patch.set_facecolor('#111122')

            # ── 圓餅圖：加權 Jaccard 拆解 ────────────────────────────────────
            pie_vals   = [inter, a_excess, b_excess]
            pie_labels = [f'共同重疊\n({inter:.1f}%)',
                          f'{ca} 獨有\n({a_excess:.1f}%)',
                          f'{cb} 獨有\n({b_excess:.1f}%)']
            pie_colors = ['#4a90d9', '#5b9cf6', '#f07070']
            non_zero = [(v, l, c) for v, l, c in zip(pie_vals, pie_labels, pie_colors) if v > 0]
            if non_zero:
                pv, pl, pc = zip(*non_zero)
                wedges, texts, autotexts = ax_pie.pie(
                    pv, labels=pl, colors=pc, autopct='%1.1f%%', startangle=90,
                    textprops={'color': 'white', 'fontsize': 8,
                               'fontfamily': CHART_FONT})
                for at in autotexts:
                    at.set_color('white'); at.set_fontsize(8)
            ax_pie.set_facecolor('#111122')
            ax_pie.set_title(f'加權 Jaccard＝{j_val:.0f}%\n共同重疊 ÷ 全聯集',
                             color='#aecde8', fontsize=10, fontfamily=CHART_FONT, pad=10)

            # ── 水平長條圖：前 12 共同持股 ────────────────────────────────────
            ax_bar.set_facecolor('#111122')
            if common:
                top_c  = common[:12]
                labels = [f'{c} {ha.get(c, hb.get(c,(c,0)))[0][:5]}' for c in top_c]
                wa_v   = [ha.get(c,(None,0))[1] for c in top_c]
                wb_v   = [hb.get(c,(None,0))[1] for c in top_c]
                bh, ys = 0.35, _np.arange(len(top_c))
                ax_bar.barh(ys + bh/2, wa_v, bh, color='#5b9cf6', label=ca)
                ax_bar.barh(ys - bh/2, wb_v, bh, color='#f07070', label=cb)
                ax_bar.set_yticks(ys)
                ax_bar.set_yticklabels(labels, fontsize=7.5, color='#aaa',
                                       fontfamily=CHART_FONT)
                ax_bar.invert_yaxis()
                ax_bar.set_xlabel('持股比重 %', color='#888899', fontsize=8)
                ax_bar.set_title('共同持股比重比較', color='#aecde8', fontsize=10,
                                 fontfamily=CHART_FONT, pad=10)
                ax_bar.tick_params(colors='#888899', labelsize=7)
                ax_bar.legend(fontsize=8, facecolor='#1a1a2e',
                              edgecolor='#333355', labelcolor='white')
                for spine in ax_bar.spines.values():
                    spine.set_edgecolor('#333355')
            else:
                ax_bar.text(0.5, 0.5, '無共同持股', ha='center', va='center',
                            color='#555577', fontsize=12, transform=ax_bar.transAxes)
                for spine in ax_bar.spines.values():
                    spine.set_edgecolor('#333355')

            fig2.tight_layout(pad=2.5)
            popup_canvas = _FCA(fig2, master=popup)
            popup_canvas.get_tk_widget().pack(fill='both', expand=True, padx=8, pady=8)
            popup_canvas.draw()

            def _close_popup():
                _plt.close(fig2)
                popup.destroy()
            popup.protocol('WM_DELETE_WINDOW', _close_popup)

        # 折線圖 canvas 負責 line hover + leave
        line_canvas.mpl_connect('motion_notify_event', _on_cmp_motion)
        line_canvas.mpl_connect('figure_leave_event',  _on_line_leave)
        # 下方 canvas 負責 bar + heatmap hover 及熱圖點擊
        cmp_fig_canvas.mpl_connect('motion_notify_event', _on_cmp_motion)
        cmp_fig_canvas.mpl_connect('figure_leave_event',  _on_cmp_leave)
        cmp_fig_canvas.mpl_connect('button_press_event',  _on_heat_click)

        tk.Frame(self._cmp_inner, bg='#2a2a3e', height=2).pack(fill='x', pady=(4, 0))

        # ── Header row ──────────────────────────────────────────────────────
        COL_W   = 120   # px per ETF column
        LABEL_W = 160   # px for row label

        # 每個 ETF 欄位的所有 cell widgets，用於切換 highlight
        _col_cells: list[list[tk.Label]] = [[] for _ in ordered]
        _col_base_bgs: list[list[str]]   = [[] for _ in ordered]  # 各 cell 原始底色
        _selected_col = [-1]             # 目前被選取的欄位 index，-1 = 無
        C_HL = '#2a3a5e'                 # highlight 底色

        def _toggle_col(col_idx):
            if _selected_col[0] == col_idx:
                # 取消選取
                for lbl, orig_bg in zip(_col_cells[col_idx], _col_base_bgs[col_idx]):
                    lbl.configure(bg=orig_bg)
                _selected_col[0] = -1
            else:
                # 取消上一個選取
                if _selected_col[0] >= 0:
                    prev = _selected_col[0]
                    for lbl, orig_bg in zip(_col_cells[prev], _col_base_bgs[prev]):
                        lbl.configure(bg=orig_bg)
                # 套用新選取
                for lbl in _col_cells[col_idx]:
                    lbl.configure(bg=C_HL)
                _selected_col[0] = col_idx

        def _cell(parent, text, bg, fg, font, width, anchor='center', bold=False):
            lbl = tk.Label(parent, text=text, bg=bg, fg=fg, font=font,
                           width=width // 8, anchor=anchor,
                           relief='flat', pady=4, padx=4)
            lbl.pack(side='left', fill='y')
            tk.Frame(parent, bg=C_SEP, width=1).pack(side='left', fill='y')
            return lbl

        perf_labels = [('1M', '近1月'), ('3M', '近3月'), ('6M', '近6月'),
                       ('1Y', '近1年(年化)'), ('3Y', '近3年(年化)')]

        def _row(label, values_fn, row_idx, is_pct=False, compare=True):
            bg = C_ROW1 if row_idx % 2 == 0 else C_ROW2
            row = tk.Frame(self._cmp_inner, bg=bg)
            row.pack(fill='x')
            _cell(row, label, bg, '#8899cc', FNT_H, LABEL_W, 'w')
            vals = [values_fn(d) for d in ordered]
            # 找最佳值（數字比較）
            nums = []
            for v in vals:
                try:    nums.append(float(str(v).replace('%','').replace('億','').replace('—','')))
                except: nums.append(None)
            best = None
            if compare and any(n is not None for n in nums):
                valid = [n for n in nums if n is not None]
                best = max(valid) if valid else None

            for ci, (d, val, num) in enumerate(zip(ordered, vals, nums)):
                if is_pct and num is not None:
                    fg = '#f07070' if num >= 0 else '#4ec94e'
                    txt = f'{num:+.2f}%'
                elif val == '—' or val is None:
                    fg = '#555577'; txt = '—'
                else:
                    fg = C_FG; txt = str(val)
                if compare and num is not None and num == best:
                    fg = '#ffd700'  # 最佳值金色
                cell_bg = C_HL if _selected_col[0] == ci else bg
                lbl = _cell(row, txt, cell_bg, fg, FNT, COL_W)
                _col_cells[ci].append(lbl)
                _col_base_bgs[ci].append(bg)
                def _on_cell_click(e, i=ci, code=d['code']):
                    if e.state & 0x4:  # Ctrl 鍵
                        self._etf_code_var.set(code)
                        self._show_etf_sub('analysis')
                        self._draw_etf_map()
                    else:
                        _toggle_col(i)
                lbl.bind('<Button-1>', _on_cell_click)
                lbl.configure(cursor='hand2')
            tk.Frame(self._cmp_inner, bg=C_SEP, height=1).pack(fill='x')

        # ── 分組標題 ────────────────────────────────────────────────────────
        def _section(title):
            s = tk.Frame(self._cmp_inner, bg='#1a1a2e')
            s.pack(fill='x')
            tk.Label(s, text=f'  {title}', bg='#1a1a2e', fg='#6a8faf',
                     font=('Microsoft JhengHei', 8, 'bold'),
                     anchor='w', pady=3).pack(fill='x')
            tk.Frame(self._cmp_inner, bg='#2a2a4e', height=1).pack(fill='x')

        row_idx = [0]
        def _r(label, fn, is_pct=False, compare=True):
            _row(label, fn, row_idx[0], is_pct, compare)
            row_idx[0] += 1

        # ── 基本資訊 ────────────────────────────────────────────────────────
        _section('基本資訊')
        _r('ETF 代號',   lambda d: d['code'], compare=False)
        _r('ETF 名稱',   lambda d: d['name'], compare=False)
        _r('成分股數量', lambda d: f"{len(d['components'])} 檔" if d['components'] else '—')
        def _fmt_aum(d):
            v = d['meta'].get('total_assets')
            if v is None: return '—'
            return f'{v/1e8:.1f} 億' if v >= 1e6 else str(v)
        _r('基金規模',   _fmt_aum, compare=False)
        def _fmt_yield(d):
            v = d['meta'].get('yield_pct')
            return f'{v*100:.2f}%' if v is not None else '—'
        _r('年化殖利率', _fmt_yield, compare=False)
        def _fmt_nav(d):
            v = d['meta'].get('nav') or d['meta'].get('price')
            return f'{v:.2f}' if v is not None else '—'
        _r('ETF 淨值',   _fmt_nav, compare=False)

        # ── 績效表現 ────────────────────────────────────────────────────────
        _section('績效表現（績效最佳標示金色）')
        for key, label in perf_labels:
            _r(label,
               lambda d, k=key: d['perf'].get(k),
               is_pct=True)

        # ── 前10大持股 ──────────────────────────────────────────────────────
        _section('前10大持股')
        max_h = max((len(d['components']) for d in ordered), default=0)
        for i in range(min(10, max_h)):
            def _top_holding(d, idx=i):
                if idx < len(d['components']):
                    h = d['components'][idx]
                    w = h.get('weight', 0)
                    return f"{h.get('code','')} {h.get('name','')[:6]} {w:.1f}%"
                return '—'
            _r(f'#{i+1}', _top_holding, compare=False)

        self._cmp_status_var.set(f'已比較 {len(ordered)} 支 ETF')
        self._cmp_canvas.update_idletasks()
        self._cmp_canvas.configure(scrollregion=self._cmp_canvas.bbox('all'))
        self._bind_cmp_scroll(self._cmp_inner)

    def _get_etf_code(self) -> str:
        code = self._etf_code_var.get().strip()
        if not code:
            sel = self._etf_var.get().strip()
            if sel:
                code = sel.split()[0]
        return code

    def _draw_etf_map(self):
        code = self._get_etf_code()
        if not code:
            self._etf_status.set('請選擇 ETF 或輸入代號')
            return

        self._etf_status.set(f'載入 {code} 中…')
        self._etf_rects = []
        self._hide_etf_tooltip()
        self._etf_fig.clear()
        self._etf_fig.patch.set_facecolor('#111111')
        ax = self._etf_fig.add_axes([0, 0, 1, 1])
        ax.set_facecolor('#111111')
        ax.axis('off')
        ax.text(0.5, 0.5, f'載入 {code} 成分股中，請稍候…',
                ha='center', va='center', color='#888',
                fontsize=13, fontfamily=CHART_FONT, transform=ax.transAxes)
        self._etf_canvas.draw()

        def _worker():
            try:
                import concurrent.futures
                with concurrent.futures.ThreadPoolExecutor(max_workers=3) as _ex:
                    _f1 = _ex.submit(fetch_etf_data, code)
                    _f2 = _ex.submit(fetch_etf_meta, code)
                    _f3 = _ex.submit(_fetch_twse_industry_map)
                    etf_name, components, dbg = _f1.result()
                    meta    = _f2.result()
                    ind_map = _f3.result()
                self._ui_call(lambda: self._draw_etf_map_impl(
                    code, etf_name, components, dbg, meta, ind_map))
            except Exception as e:
                _e = e
                self._ui_call(lambda: self._etf_status.set(f'錯誤：{_e}'))

        threading.Thread(target=_worker, daemon=True).start()

    def _draw_etf_map_impl(self, code: str, etf_name: str, components: list,
                           debug_msg: str = '', meta: dict = None,
                           ind_map: dict = None):
        if meta is None:
            meta = {}
        TREE_BG = '#111111'
        SEP_COL = '#111111'
        is_partial = 'Yahoo topHoldings' in debug_msg  # only top-10

        self._etf_rects = []
        self._hide_etf_tooltip()
        self._etf_fig.clear()
        self._etf_fig.patch.set_facecolor(TREE_BG)

        ax = self._etf_fig.add_axes([0, 0.04, 1, 0.96])
        ax.set_facecolor(TREE_BG)
        ax.set_xlim(0, 100)
        ax.set_ylim(0, 100)
        ax.axis('off')
        self._etf_ax = ax

        sax = self._etf_fig.add_axes([0, 0, 1, 0.04])
        sax.set_facecolor('#0d0d0d')
        sax.axis('off')

        if not components:
            ax.text(50, 50,
                    f'{code} 無法取得成分股資料\n（請查看下方狀態列的診斷訊息）',
                    ha='center', va='center', color='#888',
                    fontsize=12, fontfamily=CHART_FONT)
            self._etf_canvas.draw()
            self._etf_status.set(f'{code} 無法取得成分股資料（網路連線問題，請稍後再試）')
            return

        # ── 更新摘要列 ────────────────────────────────────────────────────────
        total_w = sum(c['weight'] for c in components) or 1.0
        weighted_chg = sum(c['change_pct'] * c['weight'] for c in components) / total_w
        top_stk = components[0]
        chg_fg   = '#f07070' if weighted_chg >= 0 else '#4ec94e'
        chg_sign = '+' if weighted_chg >= 0 else ''
        short_name = etf_name[:18] if len(etf_name) > 18 else etf_name
        self._etf_sv_name  .set(short_name)
        # 請求3: 若資料不完整, 改顯示前N大占比而非成分股數
        if is_partial:
            top_cov = sum(c['weight'] for c in components)
            self._etf_sv_count.set(f'前{len(components)}大  {top_cov:.1f}%')
        else:
            self._etf_sv_count.set(f'{len(components)} 檔')
        self._etf_sv_top   .set(f'{top_stk["code"]}  {top_stk["weight"]:.2f}%')
        self._etf_sv_change.set(f'{chg_sign}{weighted_chg:.2f}%')
        self._etf_sl_change.config(fg=chg_fg)
        # 規模 / 殖利率 / NAV / 配置
        aum = meta.get('total_assets')
        self._etf_sv_aum.set(
            f'NT$ {aum/1e8:.0f} 億' if aum and aum > 1e8
            else (f'US$ {aum/1e9:.2f} B' if aum else '—'))
        yld = meta.get('yield_pct')
        self._etf_sv_yield.set(f'{yld*100:.2f}%' if yld else '—')
        nav = meta.get('nav')
        self._etf_sv_nav.set(f'{nav:.2f}' if nav else '—')
        # 折溢價：優先用 TWSE ETFortune 直接計算的 atmps，否則自行計算
        atmps = meta.get('atmps')
        if atmps is not None:
            prem = atmps
        else:
            nav   = meta.get('nav')
            price = meta.get('price') or meta.get('prev_close')
            prem  = (price - nav) / nav * 100 if nav and price and nav > 0 else None
        if prem is not None:
            sign  = '+' if prem >= 0 else ''
            label = '溢價' if prem >= 0 else '折價'
            fg    = '#f07070' if prem >= 0 else '#4ec94e'
            self._etf_sv_alloc.set(f'{sign}{prem:.3f}%（{label}）')
            self._etf_sl_alloc.config(fg=fg)
        else:
            self._etf_sv_alloc.set('—')

        # ── 顏色基準 ──────────────────────────────────────────────────────────
        all_pcts = [c['change_pct'] for c in components]
        _max_abs = max((abs(p) for p in all_pcts), default=10.0) or 10.0

        # ── 計算繪圖轉換係數 ──────────────────────────────────────────────────
        _dpi      = self._etf_fig.dpi
        _fig_w_px = self._etf_fig.get_size_inches()[0] * _dpi
        _fig_h_px = self._etf_fig.get_size_inches()[1] * _dpi
        _px_per_ux = _fig_w_px / 100
        _px_per_uy = _fig_h_px * 0.96 / 100
        _pt_per_px = 72 / _dpi
        GAP = 0.3

        # ── 樹狀圖（依 ETF 權重，無分類）────────────────────────────────────
        weights   = [max(c['weight'], 0.001) for c in components]
        norm_w    = squarify.normalize_sizes(weights, 100, 100)
        raw_rects = squarify.squarify(norm_w, 0, 0, 100, 100)
        rects     = [{'x': r['x'], 'y': 100 - r['y'] - r['dy'],
                      'dx': r['dx'], 'dy': r['dy']} for r in raw_rects]

        for comp, sr in zip(components, rects):
            rx = sr['x'] + GAP / 2
            ry = sr['y'] + GAP / 2
            rw = sr['dx'] - GAP
            rh = sr['dy'] - GAP
            if rw <= 0 or rh <= 0:
                continue

            ax.add_patch(plt.Rectangle(
                (rx, ry), rw, rh,
                facecolor=pnl_color(comp['change_pct'], _max_abs),
                edgecolor=SEP_COL, linewidth=0.4, zorder=1))

            self._etf_rects.append({
                'rx': rx, 'ry': ry, 'rw': rw, 'rh': rh,
                'code':       comp['code'],
                'name':       comp['name'],
                'weight':     comp['weight'],
                'price':      comp['price'],
                'change_pct': comp['change_pct'],
            })

            min_dim = min(rw, rh)
            if min_dim < 2:
                continue

            rw_px = rw * _px_per_ux
            rh_px = rh * _px_per_uy
            name_disp = comp['name'][:8] if len(comp['name']) > 8 else comp['name']
            code_disp = comp['code']
            pnl_str   = f"{'+' if comp['change_pct'] >= 0 else ''}{comp['change_pct']:.2f}%"

            chars    = max(len(name_disp), len(code_disp), 4)
            fs_by_w  = 0.55 * rw_px * _pt_per_px / (chars * 0.60)
            fs_by_h  = 0.60 * rh_px * _pt_per_px / 3.8
            fs_name  = max(min(fs_by_w, fs_by_h, 36), 5.5)
            fs_code  = max(fs_name * 0.80, 5)
            fs_pct   = max(fs_name * 0.72, 5)
            cx_t     = rx + rw / 2
            mid_y    = ry + rh * 0.50

            if min_dim >= 6:
                line_gap = fs_name / _pt_per_px / _px_per_uy * 1.15
                ax.text(cx_t, mid_y + line_gap, name_disp,
                        ha='center', va='center', color='white',
                        fontsize=fs_name, fontweight='bold',
                        fontfamily=CHART_FONT, clip_on=True, zorder=4)
                ax.text(cx_t, mid_y, code_disp,
                        ha='center', va='center', color='#cccccc',
                        fontsize=fs_code, fontfamily=CHART_FONT,
                        clip_on=True, zorder=4)
                ax.text(cx_t, mid_y - line_gap, pnl_str,
                        ha='center', va='center', color='white',
                        fontsize=fs_pct, fontfamily=CHART_FONT,
                        clip_on=True, zorder=4)
            elif min_dim >= 4:
                gap = (fs_name / _pt_per_px / _px_per_uy +
                       fs_pct  / _pt_per_px / _px_per_uy) * 0.55
                ax.text(cx_t, mid_y + gap * 0.5, name_disp,
                        ha='center', va='center', color='white',
                        fontsize=fs_name, fontweight='bold',
                        fontfamily=CHART_FONT, clip_on=True, zorder=4)
                ax.text(cx_t, mid_y - gap * 0.5, pnl_str,
                        ha='center', va='center', color='white',
                        fontsize=fs_pct, fontfamily=CHART_FONT,
                        clip_on=True, zorder=4)
            else:
                ax.text(cx_t, mid_y, name_disp,
                        ha='center', va='center', color='white',
                        fontsize=fs_name, fontweight='bold',
                        fontfamily=CHART_FONT, clip_on=True, zorder=4)

        count_txt = (f'前{len(components)}大持股' if is_partial
                     else f'{len(components)} 檔成分股')
        sax.text(0.5, 0.5,
                 f'{etf_name}  ·  {count_txt}  ·  '
                 f'{datetime.now().strftime("%Y-%m-%d %H:%M")} 更新',
                 ha='center', va='center', color='#666',
                 fontsize=8, fontfamily=CHART_FONT, transform=sax.transAxes)

        if 'TWSE OK' in debug_msg:
            src_note = '  ✓ 資料來源：證交所'
        elif 'MoneyDJ' in debug_msg:
            src_note = '  ✓ 資料來源：MoneyDJ（完整成分股）'
        elif is_partial:
            top_cov = sum(c['weight'] for c in components)
            src_note = f'  ⚠ Yahoo Finance（前{len(components)}大，合計 {top_cov:.1f}%）'
        else:
            src_note = ''
        count_label = (f'前{len(components)}大' if is_partial else f'{len(components)} 檔')
        self._etf_status.set(f'{code}  {etf_name}  ·  {count_label}  ·  '
                              f'更新：{datetime.now().strftime("%H:%M:%S")}{src_note}')
        self._etf_canvas.draw()
        self._draw_etf_kline(code, etf_name)
        self._draw_etf_info(components, meta, debug_msg, ind_map or {})
        self._draw_etf_history(code, etf_name)

    # ── ETF 分析圖（互動式環形圖 × 2）────────────────────────────────────────
    def _draw_etf_info(self, components: list, meta: dict,
                       debug_msg: str = '', ind_map: dict = None):
        INFO_BG = '#1a1a2e'
        PANEL_BG = '#22223a'

        import colorsys as _cs

        def _grad(n, hue=0.60):
            """n 種同色系漸層色（深→淺）。"""
            if n <= 0:
                return []
            out = []
            for i in range(n):
                t = i / max(n - 1, 1)
                r, g, b = _cs.hsv_to_rgb(hue, 0.85 - 0.30 * t, 0.48 + 0.42 * t)
                out.append(f'#{int(r*255):02x}{int(g*255):02x}{int(b*255):02x}')
            return out

        fig = self._etf_info_fig
        fig.clear()
        fig.patch.set_facecolor(INFO_BG)

        # Disconnect previous hover callback
        if self._etf_info_cid is not None:
            try:
                self._etf_info_canvas.mpl_disconnect(self._etf_info_cid)
            except Exception:
                pass
            self._etf_info_cid = None
        self._etf_info_charts = []

        # ── Helper: draw one donut + right-side legend ─────────────────────
        def _donut(pie_ax, leg_ax, vals, lbls, title, hue=0.60):
            """Draw donut in pie_ax, legend in leg_ax. Returns (wedges, center_text)."""
            colors = _grad(len(vals), hue)
            # Pie axes
            pie_ax.set_facecolor(PANEL_BG)
            pie_ax.set_title(title, color='#9ab8d8', fontsize=10,
                             fontfamily=CHART_FONT, pad=5)
            wedges, _ = pie_ax.pie(
                vals, colors=colors,
                wedgeprops={'width': 0.50, 'edgecolor': INFO_BG, 'linewidth': 1.5},
                startangle=90)
            pie_ax.set_aspect('equal')   # 確保是正圓
            # Center text (updated on hover)
            ct = pie_ax.text(0, 0, '', ha='center', va='center',
                             color='white', fontsize=9, fontfamily=CHART_FONT,
                             fontweight='bold', multialignment='center', zorder=10)
            # Legend axes（超過 6 項自動兩欄）
            leg_ax.set_facecolor(PANEL_BG)
            leg_ax.axis('off')
            leg_ax.set_xlim(0, 1)
            n = len(lbls)
            leg_ax.set_ylim(0, 1)
            n_cols   = 2 if n > 6 else 1
            per_col  = (n + n_cols - 1) // n_cols      # ceil(n / n_cols)
            col_w    = 1.0 / n_cols
            row_h    = min(0.13, 0.92 / max(per_col, 1))
            y0       = 0.5 + (per_col - 1) * row_h / 2  # 垂直居中
            for i, (lbl, val) in enumerate(zip(lbls, vals)):
                col_idx = i // per_col
                row_idx = i % per_col
                y       = y0 - row_idx * row_h
                xo      = col_idx * col_w               # 欄偏移
                c       = colors[i % len(colors)]
                # scatter 用螢幕像素單位，不受 axes 長寬比影響，永遠是正圓
                dot_sz  = 55 if n_cols == 2 else 70
                leg_ax.scatter([xo + 0.04], [y], s=dot_sz, c=[c],
                               zorder=5, linewidths=0, clip_on=False)
                fs_name = 8.5 if n_cols == 2 else 9.5
                fs_pct  = 7.5 if n_cols == 2 else 8.5
                leg_ax.text(xo + 0.11, y + 0.018, lbl[:9],
                            ha='left', va='center', color='#e0e0e0',
                            fontsize=fs_name, fontfamily=CHART_FONT,
                            transform=leg_ax.transData)
                leg_ax.text(xo + 0.11, y - 0.018, f'{val:.2f}%',
                            ha='left', va='center', color='#aaaaaa',
                            fontsize=fs_pct, fontfamily=CHART_FONT,
                            transform=leg_ax.transData)
            return wedges, ct

        # ── Chart 1: Top holdings ──────────────────────────────────────────
        top_n  = min(10, len(components))
        top    = components[:top_n]
        others = sum(c['weight'] for c in components[top_n:])
        vals1  = [c['weight'] for c in top]
        lbls1  = [c['name'][:10] for c in top]
        if others > 0.01:
            vals1.append(others)
            lbls1.append('其他')

        ax1  = fig.add_axes([0.01, 0.06, 0.23, 0.88])
        axl1 = fig.add_axes([0.25, 0.06, 0.21, 0.88])
        w1, ct1 = _donut(ax1, axl1, vals1, lbls1, f'前 {top_n} 大持股', hue=0.60)
        self._etf_info_charts.append((ax1, w1, lbls1, vals1, ct1))

        # ── Chart 2: Sector distribution ──────────────────────────────────
        # 方法 1: Yahoo Finance sector_weights（英文 key → 中文）
        sw = meta.get('sector_weights', []) if meta else []
        sector_data: dict = {}
        for item in sw:
            if isinstance(item, dict):
                for k, v in item.items():
                    if v and v > 0.001:
                        sector_data[SECTOR_ZH.get(k, k)] = v * 100

        # 方法 2: 由成分股代號 × 產業別對照表計算（fallback）
        if not sector_data and ind_map and components:
            for comp in components:
                ind = ind_map.get(comp['code'], '其他')
                if not ind:
                    ind = '其他'
                sector_data[ind] = sector_data.get(ind, 0.0) + comp['weight']

        ax2  = fig.add_axes([0.51, 0.06, 0.23, 0.88])
        axl2 = fig.add_axes([0.75, 0.06, 0.23, 0.88])

        if sector_data:
            sorted_s = sorted(sector_data.items(), key=lambda x: -x[1])
            top_s    = sorted_s[:5]
            other_s  = sum(v for _, v in sorted_s[5:])
            lbls2    = [k for k, _ in top_s]
            vals2    = [v for _, v in top_s]
            if other_s > 0.01:
                lbls2.append('其他')
                vals2.append(other_s)
            w2, ct2 = _donut(ax2, axl2, vals2, lbls2, '產業分布', hue=0.45)
            self._etf_info_charts.append((ax2, w2, lbls2, vals2, ct2))
        else:
            ax2.set_facecolor(PANEL_BG)
            ax2.axis('off')
            ax2.set_title('產業分布', color='#9ab8d8', fontsize=9,
                          fontfamily=CHART_FONT, pad=5)
            ax2.text(0.5, 0.5, '產業分布資料\n暫不可用', ha='center', va='center',
                     color='#555', fontsize=9, fontfamily=CHART_FONT,
                     transform=ax2.transAxes)
            axl2.axis('off')
            axl2.set_facecolor(PANEL_BG)

        # ── Hover interaction ──────────────────────────────────────────────
        def _on_info_hover(event):
            if event.inaxes is None:
                return
            changed = False
            for pie_ax, wedges, lbls, vals, ct in self._etf_info_charts:
                if event.inaxes is not pie_ax:
                    continue
                found = False
                for wedge, lbl, val in zip(wedges, lbls, vals):
                    try:
                        hit, _ = wedge.contains(event)
                    except Exception:
                        hit = False
                    if hit:
                        new_txt = f'{lbl}\n{val:.2f}%'
                        if ct.get_text() != new_txt:
                            ct.set_text(new_txt)
                            changed = True
                        found = True
                        break
                if not found and ct.get_text():
                    ct.set_text('')
                    changed = True
            if changed:
                self._etf_info_canvas.draw_idle()

        self._etf_info_cid = self._etf_info_canvas.mpl_connect(
            'motion_notify_event', _on_info_hover)
        self._etf_info_canvas.draw()

    # ── ETF K 線圖 ────────────────────────────────────────────────────────────
    def _redraw_etf_kline(self):
        """指標 toggle 後重新繪製 K 線圖。"""
        if self._current_kline_code:
            self._draw_etf_kline(self._current_kline_code, self._current_kline_name)

    def _draw_etf_kline(self, code: str, etf_name: str):
        """Draw K-line (candlestick) + selectable indicators for ETF.
        Fetches OHLCV in a background thread to avoid blocking the UI."""
        CHART_BG = '#111111'
        PANEL_BG = '#1a1a2e'

        # 儲存供指標 toggle redraw 使用
        self._current_kline_code = code
        self._current_kline_name = etf_name

        fig = self._etf_kline_fig
        fig.clear()
        fig.patch.set_facecolor(CHART_BG)
        self._kline_ax = None
        self._kline_ohlcv = None

        # 顯示「載入中」佔位，立即 render
        _ax_tmp = fig.add_subplot(111, facecolor=PANEL_BG)
        _ax_tmp.text(0.5, 0.5, f'{code} K 線載入中…',
                     ha='center', va='center', color='#888',
                     fontsize=11, fontfamily=CHART_FONT, transform=_ax_tmp.transAxes)
        _ax_tmp.axis('off')
        self._etf_kline_canvas.draw()

        _period_map = {'1M': '1mo', '3M': '3mo', '6M': '6mo', '1Y': '1y', '全部': '5y'}
        _yf_period = _period_map.get(getattr(self, '_kline_period', '3M'), '3mo')

        def _fetch():
            import concurrent.futures as _cf_k
            ohlcv = None
            for s in ['.TW', '.TWO', '']:
                def _do(sfx=s):
                    h = yf.Ticker(code + sfx).history(period=_yf_period, auto_adjust=False)
                    return h if (h is not None and not h.empty) else None
                _ex = _cf_k.ThreadPoolExecutor(max_workers=1)
                try:
                    result = _ex.submit(_do).result(timeout=15)
                    _ex.shutdown(wait=False)
                    if result is not None:
                        ohlcv = result
                        break
                except Exception:
                    _ex.shutdown(wait=False)
            self._ui_call(lambda: self._render_etf_kline(code, ohlcv))

        threading.Thread(target=_fetch, daemon=True).start()

    def _render_etf_kline(self, code: str, ohlcv):
        """Render K-line chart on the UI thread after data is fetched."""
        import numpy as np

        CHART_BG = '#131722'
        PANEL_BG = '#1e2133'
        UP_CLR, DN_CLR = '#ef5350', '#26a69a'

        fig = self._etf_kline_fig
        fig.clear()
        fig.patch.set_facecolor(CHART_BG)
        self._kline_ax         = None
        self._kline_ohlcv      = None
        self._kline_sub_vlines = []
        self._kline_price_ann  = None
        self._kline_all_axes   = []

        if ohlcv is None or ohlcv.empty:
            ax = fig.add_subplot(111, facecolor=PANEL_BG)
            ax.text(0.5, 0.5, f'{code} 無法取得 K 線資料',
                    ha='center', va='center', color='#888',
                    fontsize=11, fontfamily=CHART_FONT, transform=ax.transAxes)
            ax.axis('off')
            self._etf_kline_canvas.draw()
            return

        # ── 讀取指標開關 ──────────────────────────────────────────────────
        ind = dict(getattr(self, '_kline_ind_state', {
            'MA': True, 'BB': False, 'VOL': False,
            'MACD': True, 'RSI': False, 'KD': False}))
        close = ohlcv['Close']
        ohlcv = ohlcv.copy()

        if ind.get('MA', True):
            ohlcv['MA5']   = close.rolling(5,   min_periods=1).mean()
            ohlcv['MA10']  = close.rolling(10,  min_periods=1).mean()
            ohlcv['MA20']  = close.rolling(20,  min_periods=1).mean()
            ohlcv['MA60']  = close.rolling(60,  min_periods=1).mean()
            ohlcv['MA120'] = close.rolling(120, min_periods=1).mean()
            ohlcv['MA240'] = close.rolling(240, min_periods=1).mean()
        if ind.get('BB', False):
            _bb_m = close.rolling(20, min_periods=5).mean()
            _bb_s = close.rolling(20, min_periods=5).std(ddof=0)
            ohlcv['BB_U'] = _bb_m + 2 * _bb_s
            ohlcv['BB_M'] = _bb_m
            ohlcv['BB_L'] = _bb_m - 2 * _bb_s
        if ind.get('MACD', True):
            _e12 = close.ewm(span=12, adjust=False).mean()
            _e26 = close.ewm(span=26, adjust=False).mean()
            ohlcv['MACD_L'] = _e12 - _e26
            ohlcv['MACD_S'] = ohlcv['MACD_L'].ewm(span=9, adjust=False).mean()
            ohlcv['MACD_H'] = ohlcv['MACD_L'] - ohlcv['MACD_S']
        if ind.get('RSI', False):
            _d  = close.diff()
            _g  = _d.clip(lower=0)
            _l  = (-_d).clip(lower=0)
            _rs = _g.ewm(com=13, adjust=False).mean() / \
                  _l.ewm(com=13, adjust=False).mean().replace(0, float('nan'))
            ohlcv['RSI'] = 100 - 100 / (1 + _rs)
        if ind.get('KD', False):
            _lo = ohlcv.get('Low',  ohlcv['Close']).rolling(9, min_periods=1).min()
            _hi = ohlcv.get('High', ohlcv['Close']).rolling(9, min_periods=1).max()
            _rk = 100 * (close - _lo) / (_hi - _lo + 1e-10)
            ohlcv['KD_K'] = _rk.ewm(com=2, adjust=False).mean()
            ohlcv['KD_D'] = ohlcv['KD_K'].ewm(com=2, adjust=False).mean()

        n  = len(ohlcv)
        xs = np.arange(n)

        # ── 動態佈局 ─────────────────────────────────────────────────────
        sub_list = []
        if ind.get('VOL',  False): sub_list.append('VOL')
        if ind.get('KD',   False): sub_list.append('KD')
        if ind.get('RSI',  False): sub_list.append('RSI')
        if ind.get('MACD', True):  sub_list.append('MACD')

        L, W  = 0.07, 0.88
        TOP   = 0.94
        BOT   = 0.07
        GAP   = 0.012
        n_sub = len(sub_list)
        avail = TOP - BOT

        if n_sub == 0:
            sub_h = 0.0
            p_bot = BOT
            p_h   = avail
        else:
            sub_h = min(0.19, (avail * 0.44) / n_sub)
            p_bot = BOT + n_sub * (sub_h + GAP)
            p_h   = TOP - p_bot

        ax_price = fig.add_axes([L, p_bot, W, max(p_h, 0.25)])
        sub_axes: dict = {}
        _y = BOT
        for _name in reversed(sub_list):
            sub_axes[_name] = fig.add_axes([L, _y, W, sub_h - GAP * 0.5])
            _y += sub_h + GAP

        def _style(ax, xticks=False):
            ax.set_facecolor(PANEL_BG)
            ax.tick_params(colors='#888', labelsize=7, length=2)
            ax.yaxis.tick_right()
            ax.set_xlim(-0.5, n - 0.5)
            for sp in ax.spines.values():
                sp.set_color('#333')
            ax.grid(axis='y', color='#2a2a3a', linewidth=0.4, linestyle='-')
            ax.yaxis.set_major_locator(
                matplotlib.ticker.MaxNLocator(nbins=4, prune='both', integer=False))
            ax.yaxis.set_major_formatter(
                matplotlib.ticker.FuncFormatter(
                    lambda v, _: (f'{v/1e6:.1f}M' if abs(v) >= 1e6
                                  else f'{v/1e3:.0f}K' if abs(v) >= 1e3
                                  else f'{v:.1f}')))
            if not xticks:
                ax.set_xticks([])

        # ── 蠟燭圖（動態寬度，WantGoo 配色）────────────────────────────
        cw = max(0.3, min(0.75, 0.72 - n * 0.001))
        for i, (_, row) in enumerate(ohlcv.iterrows()):
            o_  = float(row.get('Open',  row['Close']))
            h_  = float(row.get('High',  row['Close']))
            l_  = float(row.get('Low',   row['Close']))
            c_  = float(row['Close'])
            clr = UP_CLR if c_ >= o_ else DN_CLR
            ax_price.plot([i, i], [l_, h_], color=clr, linewidth=0.8, zorder=2)
            ax_price.add_patch(plt.Rectangle(
                (i - cw/2, min(o_, c_)), cw, max(abs(c_ - o_), 0.005 * c_),
                facecolor=clr, edgecolor='none', linewidth=0, zorder=3))

        # ── MA 線（MA5/10/20/60/120/240）────────────────────────────────
        _leg_handles = []
        if ind.get('MA', True):
            for col, clr, lbl, lw in [
                    ('MA5',   '#f0e44a', 'MA5',   1.2),
                    ('MA10',  '#4a9cf0', 'MA10',  1.1),
                    ('MA20',  '#f0a04a', 'MA20',  1.1),
                    ('MA60',  '#e060e0', 'MA60',  1.0),
                    ('MA120', '#60c060', 'MA120', 1.0),
                    ('MA240', '#f08040', 'MA240', 1.0)]:
                if col in ohlcv:
                    ln, = ax_price.plot(xs, ohlcv[col].values, color=clr,
                                        linewidth=lw, label=lbl, zorder=4)
                    _leg_handles.append(ln)

        # ── Bollinger Bands ──────────────────────────────────────────────
        if ind.get('BB', False) and 'BB_U' in ohlcv:
            ax_price.plot(xs, ohlcv['BB_U'].values, color='#88aaff',
                          linewidth=0.8, linestyle='--', label='BB上', zorder=4)
            ax_price.plot(xs, ohlcv['BB_M'].values, color='#ffffff',
                          linewidth=0.6, linestyle='--', label='BB中', zorder=4)
            ax_price.plot(xs, ohlcv['BB_L'].values, color='#88aaff',
                          linewidth=0.8, linestyle='--', label='BB下', zorder=4)
            ax_price.fill_between(xs, ohlcv['BB_U'].values, ohlcv['BB_L'].values,
                                  color='#4a9cf0', alpha=0.06, zorder=1)

        # ── Y 軸縮放 ─────────────────────────────────────────────────────
        _price_vals = [ohlcv['High'].dropna().values, ohlcv['Low'].dropna().values]
        if 'BB_U' in ohlcv: _price_vals.append(ohlcv['BB_U'].dropna().values)
        if 'BB_L' in ohlcv: _price_vals.append(ohlcv['BB_L'].dropna().values)
        _all_p = np.concatenate(_price_vals)
        _pmin, _pmax = float(np.nanmin(_all_p)), float(np.nanmax(_all_p))
        _mg = (_pmax - _pmin) * 0.05 or 0.5
        ax_price.set_ylim(_pmin - _mg, _pmax + _mg * 2.0)

        _style(ax_price)
        ax_price.tick_params(labelsize=7.5)
        ax_price.yaxis.set_major_locator(
            matplotlib.ticker.MaxNLocator(nbins=6, prune='upper', integer=False))

        if _leg_handles:
            ax_price.legend(handles=_leg_handles, loc='upper left',
                            fontsize=7.5, facecolor='#1e2133',
                            edgecolor='#2a2a4a', labelcolor='white',
                            handlelength=1.2, framealpha=0.80, borderpad=0.4,
                            ncol=3)

        # ── 副圖 ──────────────────────────────────────────────────────────
        if 'VOL' in sub_axes:
            ax_v = sub_axes['VOL']
            vol = ohlcv.get('Volume', None)
            if vol is not None:
                v_clrs = [UP_CLR if ohlcv['Close'].iloc[i] >= ohlcv['Open'].iloc[i]
                          else DN_CLR for i in range(n)]
                ax_v.bar(xs, vol.values, color=v_clrs, width=0.7, zorder=2)
            ax_v.set_ylabel('VOL', color='#888', fontsize=7, labelpad=2)
            ax_v.yaxis.set_label_position('left')
            _style(ax_v)

        if 'MACD' in sub_axes and 'MACD_L' in ohlcv:
            ax_m = sub_axes['MACD']
            h_clrs = ['#f07070' if v >= 0 else '#4ec94e' for v in ohlcv['MACD_H'].values]
            ax_m.bar(xs, ohlcv['MACD_H'].values, color=h_clrs, width=0.6, zorder=2)
            ax_m.plot(xs, ohlcv['MACD_L'].values, color='#4a9cf0', linewidth=0.9, zorder=3)
            ax_m.plot(xs, ohlcv['MACD_S'].values, color='#f0a04a', linewidth=0.9, zorder=3)
            ax_m.axhline(0, color='#444', linewidth=0.5, linestyle='--')
            ax_m.set_ylabel('MACD', color='#888', fontsize=7, labelpad=2)
            ax_m.yaxis.set_label_position('left')
            _style(ax_m)

        if 'RSI' in sub_axes and 'RSI' in ohlcv:
            ax_r = sub_axes['RSI']
            ax_r.plot(xs, ohlcv['RSI'].values, color='#c46af0', linewidth=0.9)
            ax_r.axhline(70, color='#f07070', linewidth=0.5, linestyle='--')
            ax_r.axhline(30, color='#4ec94e', linewidth=0.5, linestyle='--')
            ax_r.set_ylim(0, 100)
            ax_r.set_ylabel('RSI', color='#888', fontsize=7, labelpad=2)
            ax_r.yaxis.set_label_position('left')
            _style(ax_r)

        if 'KD' in sub_axes and 'KD_K' in ohlcv:
            ax_k = sub_axes['KD']
            ax_k.plot(xs, ohlcv['KD_K'].values, color='#f0e44a', linewidth=0.9, label='K')
            ax_k.plot(xs, ohlcv['KD_D'].values, color='#4a9cf0', linewidth=0.9, label='D')
            ax_k.axhline(80, color='#f07070', linewidth=0.5, linestyle='--')
            ax_k.axhline(20, color='#4ec94e', linewidth=0.5, linestyle='--')
            ax_k.set_ylim(0, 100)
            ax_k.set_ylabel('KD', color='#888', fontsize=7, labelpad=2)
            ax_k.yaxis.set_label_position('left')
            _style(ax_k)

        # ── X 軸日期 ─────────────────────────────────────────────────────
        _bottom_ax = sub_axes[sub_list[-1]] if sub_list else ax_price
        step  = max(1, n // 8)
        xtks  = list(range(0, n, step))
        xlbls = [ohlcv.index[i].strftime('%Y-%m-%d') for i in xtks]
        _bottom_ax.set_xticks(xtks)
        _bottom_ax.set_xticklabels(xlbls, rotation=25, ha='right', fontsize=7, color='#888')

        # ── 頂部資訊文字 ─────────────────────────────────────────────────
        hdr = fig.text(
            0.07, 0.997,
            '日期：—    開：—    高：—    低：—    收：—    量：—',
            va='top', ha='left', color='#cccccc', fontsize=10,
            fontfamily=CHART_FONT,
            bbox=dict(facecolor=CHART_BG, alpha=0.0, edgecolor='none', pad=1))
        self._kline_header = hdr

        # ── 游標線（主圖 + 所有副圖共用同一 X）─────────────────────────
        self._kline_vline = ax_price.axvline(
            -999, color='#888', linewidth=0.8, linestyle='--', zorder=5)
        self._kline_hline = ax_price.axhline(
            -999, color='#aaa', linewidth=0.7, linestyle=':', zorder=5, alpha=0.8)
        self._kline_sub_vlines = [
            ax.axvline(-999, color='#666', linewidth=0.8, linestyle='--', zorder=5)
            for ax in sub_axes.values()]

        # ── 右側 Y 軸股價標籤 ────────────────────────────────────────────
        self._kline_price_ann = ax_price.annotate(
            '', xy=(1.0, 0), xycoords=('axes fraction', 'data'),
            xytext=(4, 0), textcoords='offset points',
            ha='left', va='center', fontsize=8, color='white',
            zorder=10, clip_on=False,
            bbox=dict(boxstyle='square,pad=0.25', facecolor=UP_CLR,
                      edgecolor='none', alpha=0.92))

        # 儲存供 hover 使用
        self._kline_all_axes = [ax_price] + list(sub_axes.values())
        self._kline_ax    = ax_price
        self._kline_ohlcv = ohlcv
        self._kline_n     = n

        self._etf_kline_canvas.draw()

    def _on_kline_hover(self, event):
        """K 線圖 hover：更新頂部資訊文字、十字游標（延伸副圖）、右側價格標籤。"""
        if self._kline_ax is None or self._kline_header is None:
            return
        all_ax = getattr(self, '_kline_all_axes', [self._kline_ax])
        if event.inaxes not in all_ax:
            return
        if event.xdata is None:
            return
        xi = int(round(event.xdata))
        if not (0 <= xi < self._kline_n):
            return

        row      = self._kline_ohlcv.iloc[xi]
        date_str = self._kline_ohlcv.index[xi].strftime('%Y-%m-%d')
        o  = float(row.get('Open',   row['Close']))
        h_ = float(row.get('High',   row['Close']))
        l_ = float(row.get('Low',    row['Close']))
        c  = float(row['Close'])
        vol = int(row.get('Volume', 0)) // 1000

        txt = (f'日期：{date_str}    開：{o:.2f}    高：{h_:.2f}'
               f'    低：{l_:.2f}    收：{c:.2f}    量：{vol}K')
        self._kline_header.set_text(txt)
        if self._kline_vline is not None:
            self._kline_vline.set_xdata([xi, xi])
        if self._kline_hline is not None:
            self._kline_hline.set_ydata([c, c])
        for vl in getattr(self, '_kline_sub_vlines', []):
            vl.set_xdata([xi, xi])
        # 右側價格標籤
        ann = getattr(self, '_kline_price_ann', None)
        if ann is not None:
            prev_c = float(self._kline_ohlcv['Close'].iloc[xi - 1]) if xi > 0 else c
            is_up  = c >= prev_c
            clr    = '#ef5350' if is_up else '#26a69a'
            ann.set_text(f'{c:.2f}')
            ann.xy = (1.0, c)
            ann.get_bbox_patch().set_facecolor(clr)
        self._etf_kline_canvas.draw_idle()

    # ── ETF 歷史成分股熱力圖 ──────────────────────────────────────────────────
    def _draw_etf_history(self, code: str, etf_name: str):
        """顯示佔位後在背景 thread 載入 PCF 歷史，完成後渲染熱力圖與變化列表。"""
        fig = self._etf_heatmap_fig
        fig.clear()
        fig.patch.set_facecolor('#111111')
        _ax = fig.add_subplot(111, facecolor='#1a1a2e')
        _ax.text(0.5, 0.5, f'{code} 歷史成分股資料載入中…',
                 ha='center', va='center', color='#888',
                 fontsize=11, fontfamily=CHART_FONT, transform=_ax.transAxes)
        _ax.axis('off')
        self._etf_heatmap_canvas.draw()

        # 清空舊的變化列表
        for w in self._etf_change_outer.winfo_children():
            w.destroy()

        def _worker():
            try:
                def _prog(done, total):
                    self._ui_call(lambda: self._etf_status.set(
                        f'載入 {code} 歷史資料 {done}/{total}…'))
                history = _fetch_etf_pcf_history(code, months=6, progress_cb=_prog)
                self._ui_call(lambda: self._draw_etf_heatmap_impl(
                    code, etf_name, history))
            except Exception as _e:
                self._ui_call(lambda: (
                    self._etf_heatmap_fig.clear(),
                    self._etf_heatmap_canvas.draw()))

        threading.Thread(target=_worker, daemon=True).start()

    def _draw_etf_heatmap_impl(self, code: str, etf_name: str,
                                history: dict):
        """在主執行緒繪製成分股歷史熱力圖（imshow）。

        history: {YYYYMMDD: [{'code','name','weight'}, ...]}
        X 軸 = 週次（最多 26 週）；Y 軸 = 前 20 大成分股 + 其他列。
        """
        import numpy as np
        import matplotlib.colors as _mcolors

        HM_BG   = '#111111'
        PANEL   = '#1a1a2e'

        fig = self._etf_heatmap_fig
        fig.clear()
        fig.patch.set_facecolor(HM_BG)

        # 過濾掉空日期
        dated = {d: v for d, v in history.items() if v}
        if not dated:
            ax = fig.add_subplot(111, facecolor=PANEL)
            ax.text(0.5, 0.55,
                    f'{code} 暫無歷史成分股資料',
                    ha='center', va='center', color='#aaa',
                    fontsize=12, fontfamily=CHART_FONT, transform=ax.transAxes)
            ax.text(0.5, 0.40,
                    '可能原因：ETF 上市未滿 1 個月 / 非主動型 ETF / TWSE API 限制\n'
                    '歷史資料將從今日起逐步累積，請明日再試。',
                    ha='center', va='center', color='#666',
                    fontsize=9, fontfamily=CHART_FONT, transform=ax.transAxes)
            ax.axis('off')
            self._etf_heatmap_canvas.draw()
            return

        sorted_dates = sorted(dated.keys())

        # ── 決定前 TOP_N 大成分股（以各期平均權重排序）─────────────────────
        TOP_N = 20
        weight_sum: dict[str, float] = {}
        name_map:   dict[str, str]   = {}
        for d, holdings in dated.items():
            for h in holdings:
                c = h['code']
                weight_sum[c] = weight_sum.get(c, 0.0) + h['weight']
                if c not in name_map:
                    name_map[c] = h['name']

        sorted_by_avg = sorted(weight_sum, key=weight_sum.get, reverse=True)
        top_codes = sorted_by_avg[:TOP_N]
        top_names = [f"{name_map.get(c, c)}\n{c}" for c in top_codes]

        # ── 建立熱力圖矩陣 ────────────────────────────────────────────────
        # rows = top stocks (index 0..TOP_N-1) + "其他" (index TOP_N)
        # cols = dates
        n_rows = TOP_N + 1
        n_cols = len(sorted_dates)
        matrix    = np.zeros((n_rows, n_cols))
        heatmap_expanded = getattr(self, '_etf_heatmap_expanded', False)

        for ci, d in enumerate(sorted_dates):
            holdings = dated.get(d, [])
            wmap = {h['code']: h['weight'] for h in holdings}
            others_w = 0.0
            covered  = set()
            for ri, c in enumerate(top_codes):
                matrix[ri, ci] = wmap.get(c, 0.0)
                if c in wmap:
                    covered.add(c)
            for c, w in wmap.items():
                if c not in covered:
                    others_w += w
            matrix[TOP_N, ci] = others_w

        # ── 繪製 ─────────────────────────────────────────────────────────
        ax = fig.add_axes([0.16, 0.06, 0.80, 0.86])
        ax.set_facecolor(PANEL)

        cmap = _mcolors.LinearSegmentedColormap.from_list(
            'etf_hm', ['#1a1a2e', '#1a4a8a', '#2a7ad0', '#58c0f0', '#ffffff'])

        im = ax.imshow(matrix, aspect='auto', cmap=cmap,
                       vmin=0, vmax=max(matrix.max(), 1),
                       interpolation='nearest', origin='upper')

        # Y 軸標籤（股票名稱 + 代號）
        all_row_labels = top_names + ['其他']
        ax.set_yticks(range(n_rows))
        ax.set_yticklabels(all_row_labels, fontsize=7, fontfamily=CHART_FONT,
                           color='#cccccc', va='center')
        ax.tick_params(axis='y', length=0, pad=4)

        # X 軸標籤（每隔幾週顯示日期）
        step = max(1, n_cols // 10)
        xtick_pos = list(range(0, n_cols, step))
        xtick_lbl = [sorted_dates[i][4:6] + '/' + sorted_dates[i][6:] for i in xtick_pos]
        ax.set_xticks(xtick_pos)
        ax.set_xticklabels(xtick_lbl, fontsize=8, fontfamily=CHART_FONT,
                            color='#aaaaaa', rotation=0)
        ax.tick_params(axis='x', length=0, pad=3)

        # 分隔線：在「其他」列上方畫虛線
        ax.axhline(TOP_N - 0.5, color='#555577', linewidth=0.8, linestyle='--')

        # 在每個格子內顯示數值（僅對大格子）
        if n_cols <= 30:
            for ri in range(n_rows):
                for ci in range(n_cols):
                    v = matrix[ri, ci]
                    if v > 0.5:
                        ax.text(ci, ri, f'{v:.1f}', ha='center', va='center',
                                fontsize=6.5, fontfamily=CHART_FONT,
                                color='white' if v < matrix.max() * 0.6 else '#333')

        # 色條
        cbar_ax = fig.add_axes([0.97, 0.06, 0.015, 0.86])
        cb = fig.colorbar(im, cax=cbar_ax)
        cb.ax.tick_params(labelsize=7, colors='#aaaaaa', length=0)
        cb.outline.set_edgecolor('#333355')
        cb.set_label('%', color='#aaaaaa', fontsize=8)

        # 標題
        fig.text(0.50, 0.99,
                 f'{etf_name} ({code})  成分股歷史權重（近 6 個月，週採樣）',
                 ha='center', va='top', color='#9ab8d8',
                 fontsize=10, fontfamily=CHART_FONT)

        # 「其他」列點擊提示
        expand_hint = '▴ 點「其他」收合' if heatmap_expanded else '▾ 點「其他」展開'
        fig.text(0.99, 0.02, expand_hint, ha='right', va='bottom',
                 color='#666688', fontsize=8, fontfamily=CHART_FONT)

        self._etf_heatmap_canvas.draw()

        # 畫完熱力圖後，繪製變化列表
        self._draw_etf_change_table_impl(code, dated, sorted_dates, top_codes, name_map)

    def _draw_etf_change_table_impl(self, code: str, dated: dict,
                                     sorted_dates: list, top_codes: list,
                                     name_map: dict):
        """比較最近兩個採樣日，列出成分股異動（新增 / 移除 / 權重變化 ≥ 1%）。"""
        # 清空
        for w in self._etf_change_outer.winfo_children():
            w.destroy()

        if len(sorted_dates) < 2:
            return

        d_new = sorted_dates[-1]
        d_old = sorted_dates[-2]
        h_new = {h['code']: h for h in dated.get(d_new, [])}
        h_old = {h['code']: h for h in dated.get(d_old, [])}

        added   = []
        removed = []
        changed = []

        all_codes = set(h_new) | set(h_old)
        for c in all_codes:
            n = h_new.get(c)
            o = h_old.get(c)
            nm = (n or o or {}).get('name', name_map.get(c, c))
            if n and not o:
                added.append((c, nm, n['weight']))
            elif o and not n:
                removed.append((c, nm, o['weight']))
            elif n and o:
                delta = n['weight'] - o['weight']
                if abs(delta) >= 1.0:
                    changed.append((c, nm, o['weight'], n['weight'], delta))

        # 如果全部為空則不顯示
        if not added and not removed and not changed:
            return

        d_new_fmt = f"{d_new[:4]}/{d_new[4:6]}/{d_new[6:]}"
        d_old_fmt = f"{d_old[:4]}/{d_old[4:6]}/{d_old[6:]}"

        C_TBL = '#1a1a2e'
        C_HDR = '#2a3f6f'

        outer = self._etf_change_outer

        # 標題列
        hdr = tk.Frame(outer, bg=C_HDR, pady=4)
        hdr.pack(fill='x', pady=(4, 0))
        tk.Label(hdr, text=f'成分股異動  {d_old_fmt} → {d_new_fmt}',
                 bg=C_HDR, fg='#9ab8d8',
                 font=('Microsoft JhengHei', 10, 'bold')).pack(side='left', padx=10)

        def _section(title, rows, title_color):
            if not rows:
                return
            sf = tk.Frame(outer, bg=C_TBL, pady=2)
            sf.pack(fill='x', pady=(2, 0))
            tk.Label(sf, text=title, bg=C_TBL, fg=title_color,
                     font=('Microsoft JhengHei', 9, 'bold')).grid(
                row=0, column=0, columnspan=4, sticky='w', padx=8, pady=(4, 2))
            for ri, row in enumerate(rows, start=1):
                for ci, (txt, anchor, width) in enumerate(row):
                    tk.Label(sf, text=txt, bg=C_TBL, fg='#cccccc',
                             font=('Microsoft JhengHei', 9),
                             anchor=anchor, width=width).grid(
                        row=ri, column=ci, sticky='w', padx=(8 if ci == 0 else 4), pady=1)

        _section('▲ 新加入', [
            [(c, 'w', 8), (nm[:12], 'w', 14), (f'{w:.2f}%', 'e', 8), ('', 'w', 1)]
            for c, nm, w in sorted(added, key=lambda x: -x[2])
        ], '#4ec94e')

        _section('▼ 移除', [
            [(c, 'w', 8), (nm[:12], 'w', 14), (f'{w:.2f}%', 'e', 8), ('', 'w', 1)]
            for c, nm, w in sorted(removed, key=lambda x: -x[2])
        ], '#f07070')

        _section('⇅ 權重變化 ≥ 1%', [
            [(c, 'w', 8), (nm[:12], 'w', 14),
             (f'{ow:.2f}% → {nw:.2f}%', 'e', 16),
             (f"{'+' if d >= 0 else ''}{d:.2f}%", 'e', 10)]
            for c, nm, ow, nw, d in sorted(changed, key=lambda x: -abs(x[4]))
        ], '#8ab4d4')

    # ── ETF 樹狀圖 hover ──────────────────────────────────────────────────────
    def _on_etf_motion(self, event):
        if event.inaxes is None or self._etf_ax is None:
            self._hide_etf_tooltip()
            return
        xd, yd = event.xdata, event.ydata
        widget = self._etf_canvas.get_tk_widget()
        sx = widget.winfo_rootx() + int(event.x)
        sy = widget.winfo_rooty() + int(widget.winfo_height() - event.y)
        for r in self._etf_rects:
            if (r['rx'] <= xd <= r['rx'] + r['rw'] and
                    r['ry'] <= yd <= r['ry'] + r['rh']):
                self._show_etf_tooltip(sx, sy, r)
                return
        self._hide_etf_tooltip()

    def _show_etf_tooltip(self, sx, sy, data):
        if self._etf_tooltip is None or not self._etf_tooltip.winfo_exists():
            tip = tk.Toplevel(self)
            tip.overrideredirect(True)
            tip.attributes('-topmost', True)
            tip.configure(bg='#1e1e1e')
            border = tk.Frame(tip, bg='#3e3e3e', padx=1, pady=1)
            border.pack(fill='both', expand=True)
            inner = tk.Frame(border, bg='#252526', padx=10, pady=8)
            inner.pack(fill='both', expand=True)
            self._etf_tip_widgets = {}
            for key in ('title', 'weight', 'price', 'change'):
                lbl = tk.Label(inner, bg='#252526',
                               font=('Microsoft JhengHei', 10), anchor='w')
                lbl.pack(fill='x')
                self._etf_tip_widgets[key] = lbl
            self._etf_tooltip = tip

        w    = self._etf_tip_widgets
        pct  = data['change_pct']
        fg   = '#f07070' if pct >= 0 else '#4ec94e'
        sign = '+' if pct >= 0 else ''
        w['title'].config(text=f'{data["name"]}  {data["code"]}',
                          fg='#8ab4d4', font=('Microsoft JhengHei', 11, 'bold'))
        price_txt = '—' if data['price'] is None else f'{data["price"]:,.2f} 元'
        w['weight'].config(text=f'ETF 占比：{data["weight"]:.4f}%', fg='#cccccc')
        w['price'].config(text=f'現價：{price_txt}', fg='#cccccc')
        w['change'].config(text=f'今日漲跌：{sign}{pct:.2f}%', fg=fg)
        self._place_tooltip(self._etf_tooltip, sx, sy)
        self._etf_tooltip.deiconify()

    def _hide_etf_tooltip(self):
        if self._etf_tooltip and self._etf_tooltip.winfo_exists():
            self._etf_tooltip.withdraw()

    def _save_etf_map(self):
        from tkinter import filedialog
        code = self._get_etf_code() or 'etf'
        path = filedialog.asksaveasfilename(
            title='儲存 ETF 樹狀圖',
            defaultextension='.png',
            filetypes=[('PNG 圖片', '*.png'), ('JPEG 圖片', '*.jpg'), ('所有檔案', '*.*')],
            initialfile=f'{code}_treemap.png')
        if path:
            self._etf_fig.savefig(path, dpi=150, bbox_inches='tight',
                                   facecolor=self._etf_fig.get_facecolor())
            self._etf_status.set(f'已儲存：{path}')


    # ═══════════════════════════════════════════════════════════════════════════
    # Tab 5：台股市場總覽
    # ═══════════════════════════════════════════════════════════════════════════

    def _build_tab5(self):
        f = self.tab5

        # ── 控制列 ────────────────────────────────────────────────────────────
        ctrl = ttk.Frame(f)
        ctrl.pack(fill='x', padx=14, pady=8)
        ttk.Label(ctrl, text='台股市場總覽', style='Hdr.TLabel').pack(side='left')
        ttk.Button(ctrl, text='💾  另存圖檔', style='Nav.TButton',
                   command=self._save_market_map).pack(side='right', padx=(4, 0))
        ttk.Button(ctrl, text='🔄  更新', style='Nav.TButton',
                   command=self._draw_market_map).pack(side='right')

        # ── 分類模式下拉選單 ───────────────────────────────────────────────────
        ttk.Label(ctrl, text='分類方式：',
                  font=('Microsoft JhengHei', 9), foreground=C_FG2).pack(side='left', padx=(18, 2))
        self._mkt_group_mode = tk.StringVar(value='上市類股')
        _mode_cb = ttk.Combobox(
            ctrl, textvariable=self._mkt_group_mode,
            values=['上市類股', '上市+上櫃類股', '細分類', '概念股'],
            state='readonly', width=12,
            font=('Microsoft JhengHei', 9))
        _mode_cb.pack(side='left', padx=(0, 6))
        _mode_cb.bind('<<ComboboxSelected>>', lambda e: self._on_mkt_mode_change())

        self._mkt_status = tk.StringVar(value='切換至此頁面自動更新')
        ttk.Label(ctrl, textvariable=self._mkt_status,
                  foreground=C_FG2, font=('Microsoft JhengHei', 9)).pack(side='left', padx=14)

        # ── 期間選擇 ──────────────────────────────────────────────────────────
        period_bar = tk.Frame(f, bg=C_BG)
        period_bar.pack(fill='x', padx=14, pady=(0, 4))
        tk.Label(period_bar, text='期間：', bg=C_BG, fg=C_FG2,
                 font=('Microsoft JhengHei', 9)).pack(side='left')
        self._mkt_period     = '1D'
        self._mkt_period_btns: dict[str, tk.Label] = {}
        for pid, plbl in [('RT', '即時'), ('1D', '1日'), ('5D', '5日'), ('10D', '10日'), ('1M', '1月')]:
            btn = tk.Label(period_bar, text=plbl, bg='#1a1a2e', fg=C_FG2,
                           font=('Microsoft JhengHei', 9), padx=10, pady=3,
                           relief='flat', cursor='hand2')
            btn.pack(side='left', padx=2)
            btn.bind('<Button-1>', lambda e, p=pid: self._set_mkt_period(p))
            self._mkt_period_btns[pid] = btn
        self._update_period_btn_style()

        # ── 摘要列 ────────────────────────────────────────────────────────────
        stats_bar = tk.Frame(f, bg=C_PANEL)
        stats_bar.pack(fill='x', padx=8, pady=(0, 4))

        def _sc(parent, title):
            fr = tk.Frame(parent, bg='#1a1a2e', padx=14, pady=6)
            fr.pack(side='left', padx=3)
            tk.Label(fr, text=title, bg='#1a1a2e', fg='#6a8faf',
                     font=('Microsoft JhengHei', 8)).pack(anchor='w')
            var = tk.StringVar(value='—')
            lbl = tk.Label(fr, textvariable=var, bg='#1a1a2e',
                           fg=C_FG, font=('Microsoft JhengHei', 11, 'bold'))
            lbl.pack(anchor='w')
            return var, lbl

        self._mkt_sv_taiex, self._mkt_sl_taiex = _sc(stats_bar, '加權指數')
        self._mkt_sv_up,    self._mkt_sl_up    = _sc(stats_bar, '上漲')
        self._mkt_sv_down,  self._mkt_sl_down  = _sc(stats_bar, '下跌')
        self._mkt_sv_flat,  self._mkt_sl_flat  = _sc(stats_bar, '平盤')
        self._mkt_sv_total, self._mkt_sl_total = _sc(stats_bar, '總計')

        # ── 強勢類股排行（右側）──────────────────────────────────────────────
        self._mkt_hot_frame = tk.Frame(stats_bar, bg='#1a1a2e')
        self._mkt_hot_frame.pack(side='left', fill='both', expand=True, padx=(8, 4), pady=3)
        tk.Label(self._mkt_hot_frame, text='強勢類股', bg='#1a1a2e', fg='#6a8faf',
                 font=('Microsoft JhengHei', 8)).pack(anchor='nw', padx=8, pady=(4, 0))
        self._mkt_hot_lbl = tk.Label(
            self._mkt_hot_frame, text='', bg='#1a1a2e', fg='#f07070',
            font=('Microsoft JhengHei', 11, 'bold'),
            anchor='w', justify='left', padx=8, pady=0)
        self._mkt_hot_lbl.pack(fill='x', expand=True)
        self._mkt_hot_chip_labels: list[tk.Label] = []  # 保留，避免舊參照錯誤

        # ── Treemap canvas ────────────────────────────────────────────────────
        self._mkt_fig = plt.Figure(figsize=(9.5, 5.5), dpi=100, facecolor='#111111')
        self._mkt_canvas = FigureCanvasTkAgg(self._mkt_fig, master=f)
        w = self._mkt_canvas.get_tk_widget()
        w.configure(bg='#111111')
        w.pack(fill='both', expand=True, padx=8, pady=(0, 8))
        w.bind('<Leave>', lambda e: self._mkt_tooltip.withdraw()
               if hasattr(self, '_mkt_tooltip') else None)

        # ── Tooltip ───────────────────────────────────────────────────────────
        self._mkt_tooltip = tk.Toplevel(self)
        self._mkt_tooltip.withdraw()
        self._mkt_tooltip.overrideredirect(True)
        self._mkt_tooltip.configure(bg='#1a1a2e')
        self._mkt_tt_lbl = tk.Label(
            self._mkt_tooltip, bg='#1a1a2e', fg=C_FG,
            font=('Microsoft JhengHei', 9), justify='left', padx=8, pady=6)
        self._mkt_tt_lbl.pack()

        # ── 內部狀態 ──────────────────────────────────────────────────────────
        self._mkt_rects:     list = []
        self._mkt_cat_rects: list = []
        self._mkt_groups:    dict = {}      # 類股 → 股票清單（帶 chg_pct）
        self._mkt_drill:     str | None = None   # None = 總覽；str = 類股名
        self._mkt_sub_drill: str | None = None   # 其他業展開後的子類股名
        self._mkt_hist_cache: dict = {}     # (period, date_str) → {code: chg_pct}
        self._mkt_streaks:    dict = {}     # ind_name → (direction, count)  1=漲 -1=跌 0=平
        self._mkt_inst_today: dict = {}     # code → {foreign, trust, dealer} 三大法人今日（張）
        self._mkt_up = self._mkt_down = self._mkt_flat = 0
        self._mkt_taiex_price = self._mkt_taiex_chg = None

        self._mkt_canvas.mpl_connect('motion_notify_event', self._on_mkt_motion)
        self._mkt_canvas.mpl_connect('button_press_event', self._on_mkt_click)

    # ── 期間切換 ──────────────────────────────────────────────────────────────

    def _on_mkt_mode_change(self):
        """分類模式切換：重置 drill 狀態並重繪"""
        self._mkt_drill     = None
        self._mkt_sub_drill = None
        self._mkt_hist_cache.clear()   # 清快取，確保重新抓取
        self._draw_market_map()

    def _set_mkt_period(self, period: str):
        if period == self._mkt_period:
            return
        self._mkt_period = period
        self._update_period_btn_style()
        self._draw_market_map()

    def _update_period_btn_style(self):
        for pid, btn in self._mkt_period_btns.items():
            if pid == self._mkt_period:
                btn.config(bg='#2a4a6e', fg='white')
            else:
                btn.config(bg='#1a1a2e', fg=C_FG2)

    # ── 資料抓取 + 繪製（背景執行緒）─────────────────────────────────────────

    # ── 即時報價輔助 ──────────────────────────────────────────────────────────

    @staticmethod
    def _is_tw_market_open() -> bool:
        """判斷台股是否正在交易中（週一～五 09:00–13:30，台北時間）。"""
        from datetime import datetime as _dt, timezone, timedelta, time as _time
        tw = timezone(timedelta(hours=8))
        now = _dt.now(tw)
        if now.weekday() >= 5:
            return False
        t = now.time()
        return _time(9, 0) <= t <= _time(13, 30)

    @staticmethod
    def _fetch_mis_prices(code_ex_pairs: list) -> dict:
        """
        從 TWSE MIS 即時 API 抓取報價。
        code_ex_pairs: [(code, 'tse'|'otc'), ...]
        回傳: {code: (現價, 昨收)}
        """
        import concurrent.futures as _cf
        results: dict = {}
        lock = threading.Lock()
        BATCH = 80

        def _fetch_one(pairs):
            ex_ch = '|'.join(f'{ex}_{c}.tw' for c, ex in pairs)
            url   = f'https://mis.twse.com.tw/stock/api/getStockInfo.jsp?ex_ch={ex_ch}'
            hdrs  = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/122.0.0.0',
                'Referer':    'https://mis.twse.com.tw/stock/fibest.jsp',
                'Accept':     'application/json',
            }
            try:
                data = _cffi_get_json(url, headers=hdrs, timeout=10)
                batch: dict = {}
                for s in data.get('msgArray', []):
                    c = str(s.get('c', '')).strip()
                    z = str(s.get('z', '-')).strip()   # 成交價（'-'=尚未成交）
                    o = str(s.get('o', '-')).strip()   # 開盤價（'-'=尚未開盤）
                    y = str(s.get('y', '')).strip()    # 昨收
                    try:
                        prev = float(y)
                        if z and z != '-':
                            curr = float(z)            # 優先用成交價
                        elif o and o != '-':
                            curr = float(o)            # 其次用開盤價
                        else:
                            curr = prev                # 尚未開盤，視為平盤
                        if prev > 0:
                            batch[c] = (curr, prev)
                    except (ValueError, TypeError):
                        pass
                with lock:
                    results.update(batch)
            except Exception as _e:
                _net_log(f'MIS batch fail: {_e}')

        batches = [code_ex_pairs[i:i+BATCH]
                   for i in range(0, len(code_ex_pairs), BATCH)]
        with _cf.ThreadPoolExecutor(max_workers=6) as _pool:
            list(_pool.map(_fetch_one, batches))
        return results

    def _draw_market_map(self):
        self._mkt_status.set('🔄 載入資料中…')
        self._mkt_rects     = []
        self._mkt_cat_rects = []
        self._mkt_drill     = None
        self._mkt_sub_drill = None
        self._mkt_hist_cache.clear()   # 強制重新抓取所有快取資料
        threading.Thread(target=self._draw_market_map_impl, daemon=True).start()

    def _draw_market_map_impl(self):
        import time as _time
        try:
            group_mode = self._mkt_group_mode.get()   # '上市類股' / '上市+上櫃類股' / '細分類' / '概念股'
            _ts = int(_time.time())   # 時間戳記，用於破除 CDN 快取
            hdr = {'If-Modified-Since': 'Mon, 26 Jul 1997 05:00:00 GMT',
                   'Cache-Control': 'no-cache, no-store, must-revalidate',
                   'Pragma': 'no-cache'}
            from datetime import date as _today_date
            _today_str = _today_date.today().strftime('%Y%m%d')

            # ── TWSE 上市日行情 ──────────────────────────────────────────────────
            # 優先使用 RWD 版本（收盤後更新較快），其次 OpenAPI（盤中亦可用）
            day_data = []
            _data_is_today = False
            try:
                _rwd = _cffi_get_json(
                    'https://www.twse.com.tw/rwd/zh/afterTrading/STOCK_DAY_ALL?response=json',
                    headers=hdr, timeout=15)
                if isinstance(_rwd, dict) and _rwd.get('stat') == 'OK' and _rwd.get('data'):
                    _fields = _rwd.get('fields', [])
                    _rwd_rows = []
                    for _row in _rwd['data']:
                        if len(_row) >= 9:
                            _rwd_rows.append({
                                'Code':         _row[0],
                                'Name':         _row[1],
                                'TradeValue':   _row[3].replace(',', ''),
                                'ClosingPrice': _row[7].replace(',', ''),
                                'Change':       _row[8].replace(',', '').replace('+', ''),
                            })
                    if _rwd_rows:
                        day_data = _rwd_rows
                        _data_is_today = (_rwd.get('date', '') == _today_str)
            except Exception:
                pass
            if not day_data:
                try:
                    day_data = _cffi_get_json(
                        f'https://openapi.twse.com.tw/v1/exchangeReport/STOCK_DAY_ALL?_={_ts}',
                        headers=hdr, timeout=15)
                    if day_data:
                        _data_is_today = str(day_data[0].get('Date', '')).endswith(
                            _today_date.today().strftime('%m%d'))
                except Exception:
                    pass

            ind_map    = _fetch_twse_industry_map()
            name_table = _load_twse_stock_names()

            # 上市+上櫃 模式：額外抓 TPEX 資料
            tpex_day_data = []
            tpex_ind_map: dict[str, str] = {}
            if group_mode == '上市+上櫃類股':
                try:
                    tpex_day_data = _cffi_get_json(
                        f'https://www.tpex.org.tw/openapi/v1/tpex_mainboard_daily_close_quotes?_={_ts}',
                        headers=hdr, timeout=12)
                except Exception:
                    pass
                try:
                    tpex_ind_map = _fetch_tpex_industry_map()
                except Exception:
                    pass

            # 加權指數（1日漲跌）
            taiex_price = taiex_chg = None
            try:
                hist = yf.Ticker('^TWII').history(period='5d')
                if len(hist) >= 2:
                    taiex_price = float(hist['Close'].iloc[-1])
                    taiex_chg   = (hist['Close'].iloc[-1] - hist['Close'].iloc[-2]) \
                                  / hist['Close'].iloc[-2] * 100
            except Exception:
                pass

            # ── 整理 1D 基礎資料（個股今日漲跌，成交金額）─────────────────────
            base_stocks: dict[str, dict] = {}
            up = down = flat = 0

            def _add_stock(code, name, close, change, trade_val, industry, exchange='tse'):
                nonlocal up, down, flat
                if close <= 0 or trade_val <= 0:
                    return
                prev   = close - change
                chg_1d = change / prev * 100 if prev > 0 else 0.0
                if change > 0:   up   += 1
                elif change < 0: down += 1
                else:            flat += 1
                base_stocks[code] = {
                    'code': code, 'name': name, 'industry': industry,
                    'close': close, 'change': change,
                    'chg_1d': chg_1d, 'trade_val': trade_val,
                    'exchange': exchange,
                }

            # 上市 TWSE
            for row in day_data:
                code = str(row.get('Code', '')).strip()
                if not (code.isdigit() and len(code) == 4):
                    continue
                try:
                    close     = float(row.get('ClosingPrice', '0') or '0')
                    change    = float(row.get('Change',       '0') or '0')
                    trade_val = float(row.get('TradeValue',   '0') or '0')
                except (ValueError, TypeError):
                    continue
                industry = ind_map.get(code, '其他')
                if not industry or industry.isdigit():
                    industry = '其他'
                name = name_table.get(code, row.get('Name', code))
                _add_stock(code, name, close, change, trade_val, industry, 'tse')

            # 上櫃 TPEX（僅在「上市+上櫃類股」模式下加入）
            if group_mode == '上市+上櫃類股' and tpex_day_data:
                for row in tpex_day_data:
                    code = str(row.get('SecuritiesCompanyCode', '')).strip()
                    if not (code.isdigit() and len(code) == 4):
                        continue
                    if code in base_stocks:   # 已有上市資料則跳過
                        continue
                    try:
                        close     = float(str(row.get('Close', '0') or '0').replace(',', ''))
                        chg_str   = str(row.get('Change', '0') or '0').replace(',', '').replace('+', '').strip()
                        change    = float(chg_str) if chg_str else 0.0
                        trade_val = float(str(row.get('TransactionAmount', '0') or '0').replace(',', ''))
                    except (ValueError, TypeError):
                        continue
                    industry = (tpex_ind_map.get(code)
                                or ind_map.get(code)
                                or '其他')
                    if not industry or industry.isdigit():
                        industry = '其他'
                    name = name_table.get(code, str(row.get('CompanyName', code)))
                    _add_stock(code, name, close, change, trade_val, industry, 'otc')

            # ── 即時模式：用 MIS 即時報價覆蓋今日漲跌（1日固定用 STOCK_DAY_ALL）──
            _use_rt = (self._mkt_period == 'RT')
            _price_label = '即時' if _use_rt else ('今收' if _data_is_today else '昨收')
            if _use_rt and base_stocks:
                try:
                    code_ex = [(c, s['exchange']) for c, s in base_stocks.items()]
                    mis = self._fetch_mis_prices(code_ex)
                    _mis_ok = 0
                    for code, (curr, prev) in mis.items():
                        if code in base_stocks and prev > 0 and curr > 0:
                            change = curr - prev
                            base_stocks[code]['close']  = curr
                            base_stocks[code]['change'] = change
                            base_stocks[code]['chg_1d'] = change / prev * 100
                            _mis_ok += 1
                    if _mis_ok > 0:
                        up = down = flat = 0
                        for s in base_stocks.values():
                            if s['change'] > 0:   up   += 1
                            elif s['change'] < 0: down += 1
                            else:                 flat += 1
                except Exception as _mis_e:
                    _net_log(f'MIS realtime update fail: {_mis_e}')

            # ── 歷史漲跌（非 1D / 非即時 期間）─────────────────────────────────
            period = self._mkt_period
            hist_chg: dict[str, float] = {}   # code → N日漲跌%
            if period not in ('1D', 'RT'):
                today_str = datetime.now().strftime('%Y%m%d')
                cache_key = (period, today_str)
                if cache_key in self._mkt_hist_cache:
                    hist_chg = self._mkt_hist_cache[cache_key]
                else:
                    n_map = {'5D': 5, '10D': 10, '1M': 21}
                    n = n_map.get(period, 5)
                    codes = list(base_stocks.keys())
                    tickers = [c + '.TW' for c in codes]
                    try:
                        import logging as _log, io, sys as _sys
                        _yf_log = _log.getLogger('yfinance')
                        _prev_lvl = _yf_log.level
                        _yf_log.setLevel(_log.CRITICAL)
                        _old_stderr, _sys.stderr = _sys.stderr, io.StringIO()
                        try:
                            raw = yf.download(
                                tickers, period=f'{n + 10}d',
                                progress=False, auto_adjust=True,
                                group_by='ticker', threads=True)
                        finally:
                            _sys.stderr = _old_stderr
                            _yf_log.setLevel(_prev_lvl)
                        for code, tkr in zip(codes, tickers):
                            try:
                                if len(tickers) == 1:
                                    closes = raw['Close']
                                else:
                                    closes = raw[tkr]['Close']
                                closes = closes.dropna()
                                if len(closes) >= 2:
                                    start_price = float(closes.iloc[max(-n-1, -len(closes))])
                                    end_price   = float(closes.iloc[-1])
                                    if start_price > 0:
                                        hist_chg[code] = (end_price - start_price) / start_price * 100
                            except Exception:
                                pass
                    except Exception:
                        pass
                    self._mkt_hist_cache[cache_key] = hist_chg

            # ── 組裝 groups dict ───────────────────────────────────────────────
            MERGE_THRESHOLD = 0.5   # 占比低於此值（%）的類股合併到「其他業」
            raw_groups: dict[str, list] = {}
            total_tv = sum(s['trade_val'] for s in base_stocks.values())

            if group_mode == '概念股':
                # CMoney 概念股：股票代號 → 第一個命中的概念組
                concept_assign: dict[str, str] = {}
                for cgrp, codes in _fetch_cmoney_concept_map().items():
                    for c in codes:
                        if c not in concept_assign:
                            concept_assign[c] = cgrp
                for stk in base_stocks.values():
                    chg_pct = hist_chg.get(stk['code'], stk['chg_1d']) if period not in ('1D', 'RT') else stk['chg_1d']
                    grp = concept_assign.get(stk['code'], '其他概念')
                    industry = stk['industry'] if grp == '其他概念' else grp
                    raw_groups.setdefault(grp, []).append(
                        {**stk, 'chg_pct': chg_pct, 'industry': industry})

            elif group_mode == '細分類':
                # CMoney 細分類：電子股用 _MKT_ELEC_GROUPS 細分，非電子股用 TWSE 大類
                elec_assign: dict[str, str] = {}
                for subcat, codes in _MKT_ELEC_GROUPS.items():
                    for c in codes:
                        if c not in elec_assign:
                            elec_assign[c] = subcat
                for stk in base_stocks.values():
                    chg_pct = hist_chg.get(stk['code'], stk['chg_1d']) if period not in ('1D', 'RT') else stk['chg_1d']
                    grp = elec_assign.get(stk['code'], stk['industry'])
                    raw_groups.setdefault(grp, []).append(
                        {**stk, 'chg_pct': chg_pct, 'industry': grp})

            else:
                for stk in base_stocks.values():
                    chg_pct = hist_chg.get(stk['code'], stk['chg_1d']) if period not in ('1D', 'RT') else stk['chg_1d']
                    raw_groups.setdefault(stk['industry'], []).append({**stk, 'chg_pct': chg_pct})

            groups: dict[str, list] = {}
            other_stocks: list = []
            for ind, stks in raw_groups.items():
                ind_tv = sum(s['trade_val'] for s in stks)
                pct = ind_tv / total_tv * 100 if total_tv else 0
                if pct < MERGE_THRESHOLD:
                    other_stocks.extend(stks)
                else:
                    groups[ind] = stks
            if other_stocks:
                groups.setdefault('其他業', []).extend(other_stocks)

            # ── 計算各類股連漲/跌天數 ────────────────────────────────────────
            today_str2 = datetime.now().strftime('%Y%m%d')
            streak_key = ('streak', today_str2)
            if streak_key in self._mkt_hist_cache:
                streaks = self._mkt_hist_cache[streak_key]
            else:
                streaks = {}
                # 用 raw_groups（合併前的原始分類）計算，才能涵蓋其他業的子類股
                streak_codes_map: dict[str, list] = {}
                for ind_name, stks in raw_groups.items():
                    top = sorted(stks, key=lambda s: -s['trade_val'])[:15]
                    streak_codes_map[ind_name] = top
                all_streak_codes = list({s['code'] for tops in streak_codes_map.values()
                                         for s in tops})
                streak_tickers = [c + '.TW' for c in all_streak_codes]
                try:
                    import logging as _log, io, sys as _sys
                    _yf_log = _log.getLogger('yfinance')
                    _prev_lvl = _yf_log.level
                    _yf_log.setLevel(_log.CRITICAL)
                    _old_stderr, _sys.stderr = _sys.stderr, io.StringIO()
                    try:
                        raw_s = yf.download(
                            streak_tickers, period='25d',
                            progress=False, auto_adjust=True,
                            group_by='ticker', threads=True)
                    finally:
                        _sys.stderr = _old_stderr
                        _yf_log.setLevel(_prev_lvl)
                    # 個股每日漲跌幅
                    daily_ret: dict[str, list] = {}
                    for code, tkr in zip(all_streak_codes, streak_tickers):
                        try:
                            closes = (raw_s['Close'] if len(streak_tickers) == 1
                                      else raw_s[tkr]['Close'])
                            closes = closes.dropna()
                            if len(closes) >= 2:
                                rets = closes.pct_change().dropna().tolist()
                                daily_ret[code] = rets
                        except Exception:
                            pass
                    # 類股加權日報酬 → 連漲/跌天數
                    for ind_name, top_stks in streak_codes_map.items():
                        max_days = max(
                            (len(daily_ret.get(s['code'], [])) for s in top_stks),
                            default=0)
                        if max_days == 0:
                            streaks[ind_name] = (0, 0)
                            continue
                        sector_daily = []
                        for day_i in range(max_days):
                            w_ret = w_tv = 0.0
                            for s in top_stks:
                                rets = daily_ret.get(s['code'], [])
                                if day_i < len(rets):
                                    w_ret += rets[day_i] * s['trade_val']
                                    w_tv  += s['trade_val']
                            sector_daily.append(w_ret / w_tv if w_tv else 0.0)
                        last_dir = (1 if sector_daily[-1] > 0
                                    else -1 if sector_daily[-1] < 0 else 0)
                        cnt = 0
                        for ret in reversed(sector_daily):
                            cur = 1 if ret > 0 else (-1 if ret < 0 else 0)
                            if cur == last_dir and cur != 0:
                                cnt += 1
                            else:
                                break
                        streaks[ind_name] = (last_dir, cnt)
                except Exception:
                    pass
                self._mkt_hist_cache[streak_key] = streaks
            self._mkt_streaks = streaks

            # ── 三大法人今日資料（for tooltip）─────────────────────────────
            inst_today: dict[str, dict] = {}
            try:
                inst_raw = _cffi_get_json(
                    'https://openapi.twse.com.tw/v1/fund/TWT38U',
                    headers=hdr, timeout=15)
                for _ir in inst_raw:
                    _c = str(_ir.get('Code', '')).strip()
                    if not _c:
                        continue
                    def _parse_lots(row, *keys):
                        for k in keys:
                            v = row.get(k)
                            if v is not None:
                                try:
                                    return int(str(v).replace(',', '') or '0') // 1000
                                except Exception:
                                    pass
                        return 0
                    inst_today[_c] = {
                        'foreign': _parse_lots(_ir,
                            'ForeignInvestmentNetBuyShares',
                            'foreignNetBuyShares'),
                        'trust': _parse_lots(_ir,
                            'InvestmentTrustNetBuyShares',
                            'trustNetBuyShares'),
                        'dealer': _parse_lots(_ir,
                            'DealerProprietaryNetBuyShares',
                            'dealerNetBuyShares'),
                        'total': _parse_lots(_ir,
                            'TotalInstNetBuyShares',
                            'totalNetBuyShares'),
                    }
            except Exception:
                pass
            self._mkt_inst_today = inst_today

            _pl = _price_label
            self._ui_call(lambda: self._render_market_map(
                groups, raw_groups, up, down, flat, taiex_price, taiex_chg, _pl))
        except Exception as e:
            _e = e
            self._ui_call(lambda: self._mkt_status.set(f'⚠ 載入失敗：{_e}'))

    def _render_market_map(self, groups, raw_groups, up, down, flat,
                           taiex_price, taiex_chg, price_label='今收'):
        # 儲存狀態供 drill-down 使用
        self._mkt_groups      = groups
        self._mkt_up          = up
        self._mkt_down        = down
        self._mkt_flat        = flat
        self._mkt_taiex_price = taiex_price
        self._mkt_taiex_chg   = taiex_chg

        # ── 更新摘要列 ────────────────────────────────────────────────────────
        if taiex_price:
            sign = '+' if (taiex_chg or 0) >= 0 else ''
            chg_str = f'  {sign}{taiex_chg:.2f}%' if taiex_chg is not None else ''
            self._mkt_sv_taiex.set(f'{taiex_price:,.2f}{chg_str}')
            self._mkt_sl_taiex.config(
                fg='#f07070' if (taiex_chg or 0) >= 0 else '#4ec94e')
        self._mkt_sv_up   .set(f'{up} 檔');  self._mkt_sl_up  .config(fg='#f07070')
        self._mkt_sv_down .set(f'{down} 檔'); self._mkt_sl_down.config(fg='#4ec94e')
        self._mkt_sv_flat .set(f'{flat} 檔'); self._mkt_sl_flat.config(fg=C_FG)
        self._mkt_sv_total.set(f'{up+down+flat} 檔')

        self._update_mkt_hot_sectors(raw_groups)

        self._mkt_price_label = price_label   # 儲存供 back-navigation 使用
        if self._mkt_drill:
            self._render_mkt_drill(self._mkt_drill)
        else:
            self._render_mkt_overview()

    # ── 強勢類股排行 ──────────────────────────────────────────────────────────

    def _update_mkt_hot_sectors(self, raw_groups: dict):
        """計算漲幅最大且連漲的類股，顯示在摘要列右側。"""
        # 計算各類股加權平均漲跌幅（以成交額為權重）
        sector_chg: list[tuple[float, int, str]] = []  # (chg_pct, streak_days, name)
        for ind_name, stks in raw_groups.items():
            if not stks:
                continue
            total_tv = sum(s['trade_val'] for s in stks)
            if total_tv <= 0:
                continue
            w_chg = sum(s['chg_pct'] * s['trade_val'] for s in stks) / total_tv
            direction, cnt = self._mkt_streaks.get(ind_name, (0, 0))
            streak_days = cnt if direction == 1 else 0
            sector_chg.append((w_chg, streak_days, ind_name))

        # 只取上漲的，先依連漲天數排序、再依漲幅排序，取前 3 名
        rising = [(chg, days, name) for chg, days, name in sector_chg if chg > 0]
        rising.sort(key=lambda x: (-x[1], -x[0]))
        top3 = rising[:3]

        parts = []
        for chg, days, name in top3:
            sign = '+' if chg >= 0 else ''
            streak_txt = f' ▲{days}日' if days >= 1 else ''
            parts.append(f'{name}  {sign}{chg:.1f}%{streak_txt}')

        self._mkt_hot_lbl.config(text='    '.join(parts) if parts else '—')

    # ── 總覽繪製 ──────────────────────────────────────────────────────────────

    def _render_mkt_overview(self):
        groups   = self._mkt_groups
        up       = self._mkt_up
        down     = self._mkt_down
        flat     = self._mkt_flat

        fig = self._mkt_fig
        fig.clf()
        ax  = fig.add_axes([0, 0.03, 1, 0.97])
        ax.set_xlim(0, 100); ax.set_ylim(0, 100); ax.axis('off')
        sax = fig.add_axes([0, 0, 1, 0.03]); sax.axis('off')

        dpi        = fig.dpi
        _px_per_ux = fig.get_size_inches()[0] * dpi / 100
        _px_per_uy = fig.get_size_inches()[1] * dpi * 0.97 / 100
        _pt_per_px = 72 / dpi
        SEP_COL    = '#111111'
        GAP        = 0.3
        HDR_H      = 3.2

        ind_vals    = {ind: sum(s['trade_val'] for s in stks)
                       for ind, stks in groups.items()}
        sorted_inds = sorted(groups, key=lambda x: -ind_vals[x])
        all_pcts    = [s['chg_pct'] for stks in groups.values() for s in stks]
        _max_abs    = max((abs(p) for p in all_pcts), default=10.0) or 10.0
        total_val   = sum(ind_vals.values())

        if not total_val:
            ax.text(50, 50, '無資料', ha='center', va='center',
                    color='#888', fontsize=14, fontfamily=CHART_FONT)
            self._mkt_canvas.draw()
            self._mkt_status.set('無法取得資料')
            return

        # 類股方塊佈局
        cat_sizes = [ind_vals[i] for i in sorted_inds]
        norm_cats = squarify.normalize_sizes(cat_sizes, 100, 96)
        cats_raw  = squarify.squarify(norm_cats, 0, 4, 100, 96)
        cat_rects = [{'x': r['x'],
                      'y': 4 + 96 - (r['y'] - 4) - r['dy'],
                      'dx': r['dx'], 'dy': r['dy']} for r in cats_raw]

        self._mkt_rects     = []
        self._mkt_cat_rects = []

        for ind_name, cr in zip(sorted_inds, cat_rects):
            stocks = groups[ind_name]
            cx, cy, cw, ch = cr['x'], cr['y'], cr['dx'], cr['dy']

            # 類股標題列（可點擊）
            hdr_y    = cy + ch - HDR_H
            hdr_h_px = HDR_H * _px_per_uy
            fs_hdr   = min(10, max(6, hdr_h_px * _pt_per_px * 0.62))
            ax.add_patch(plt.Rectangle(
                (cx, hdr_y), cw, HDR_H,
                facecolor='#1a1a2e', edgecolor='none', zorder=6, clip_on=True))
            ax.text(cx + 0.6, hdr_y + HDR_H / 2,
                    f'{ind_name}  {ind_vals[ind_name]/total_val*100:.1f}%',
                    ha='left', va='center', color='#aecde8',
                    fontsize=fs_hdr, fontweight='bold',
                    fontfamily=CHART_FONT, clip_on=True, zorder=7)

            # 連漲/跌天數指標（右側）
            s_dir, s_cnt = self._mkt_streaks.get(ind_name, (0, 0))
            if s_cnt >= 2 and cw > 2:
                sym       = '▲' if s_dir > 0 else '▼'
                sym_color = '#ff8080' if s_dir > 0 else '#80e080'
                ax.text(cx + cw - 0.5, hdr_y + HDR_H / 2,
                        f'{sym}{s_cnt}',
                        ha='right', va='center', color=sym_color,
                        fontsize=fs_hdr, fontweight='bold',
                        fontfamily=CHART_FONT, clip_on=True, zorder=7)

            ax.add_patch(plt.Rectangle(
                (cx, cy), cw, ch,
                facecolor='none', edgecolor='#2a2a2a', linewidth=1.5, zorder=8))

            ind_tv  = ind_vals[ind_name]
            ind_avg = (sum(s['chg_pct'] * s['trade_val'] for s in stocks) / ind_tv
                       if ind_tv else 0)
            self._mkt_cat_rects.append({
                'cx': cx, 'cy': cy, 'cw': cw, 'ch': ch,
                'name': ind_name, 'count': len(stocks),
                'trade_val': ind_tv, 'avg_chg': ind_avg, 'hdr_h': HDR_H,
                'streak': (s_dir, s_cnt),
            })

            # 個股方塊
            inner_h = ch - HDR_H
            if inner_h < 1:
                continue

            # 類股內部合併：占類股 < STK_MERGE_THR% 的小股合為「其他 N檔」
            STK_MERGE_THR = 2.0
            stks_sorted = sorted(stocks, key=lambda s: -s['trade_val'])
            sec_tv = sum(s['trade_val'] for s in stks_sorted) or 1
            main_stks, tiny_stks = [], []
            for s in stks_sorted:
                if s['trade_val'] / sec_tv * 100 >= STK_MERGE_THR:
                    main_stks.append(s)
                else:
                    tiny_stks.append(s)
            if tiny_stks:
                tiny_tv  = sum(s['trade_val'] for s in tiny_stks)
                tiny_avg = (sum(s['chg_pct'] * s['trade_val'] for s in tiny_stks) / tiny_tv
                            if tiny_tv else 0)
                main_stks.append({
                    'code': '', 'name': f'其他 {len(tiny_stks)}檔',
                    'close': 0, 'change': 0,
                    'chg_1d': tiny_avg, 'chg_pct': tiny_avg,
                    'trade_val': tiny_tv, 'industry': ind_name,
                    '_is_other': True, '_other_list': tiny_stks,
                })
            stks_sorted = main_stks

            svals = [s['trade_val'] for s in stks_sorted]
            if cw >= inner_h:
                norm_s = squarify.normalize_sizes(svals, cw, inner_h)
                sr_raw = squarify.squarify(norm_s, cx, cy, cw, inner_h)
                srects = [{'x': r['x'],
                           'y': cy + inner_h - (r['y'] - cy) - r['dy'],
                           'dx': r['dx'], 'dy': r['dy']} for r in sr_raw]
            else:
                norm_s = squarify.normalize_sizes(svals, inner_h, cw)
                sr_t   = squarify.squarify(norm_s, 0, 0, inner_h, cw)
                srects = [{'x': cx + r['y'],
                           'y': cy + inner_h - r['x'] - r['dx'],
                           'dx': r['dy'], 'dy': r['dx']} for r in sr_t]

            for stk, sr in zip(stks_sorted, srects):
                rx = sr['x'] + GAP / 2;  ry = sr['y'] + GAP / 2
                rw = sr['dx'] - GAP;     rh = sr['dy'] - GAP
                if rw <= 0 or rh <= 0:
                    continue
                ax.add_patch(plt.Rectangle(
                    (rx, ry), rw, rh,
                    facecolor=pnl_color(stk['chg_pct'], _max_abs),
                    edgecolor=SEP_COL, linewidth=0.3, zorder=1))
                self._mkt_rects.append({
                    'rx': rx, 'ry': ry, 'rw': rw, 'rh': rh,
                    **stk, 'industry': ind_name,
                })
                self._draw_stock_label(ax, stk, rx, ry, rw, rh,
                                       _px_per_ux, _px_per_uy, _pt_per_px)

        period_lbl = {'RT': '即時', '1D': '今日', '5D': '5日', '10D': '10日', '1M': '1月'}.get(
            self._mkt_period, self._mkt_period)
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        _pl = getattr(self, '_mkt_price_label', '今收')
        sax.text(0.5, 0.5,
                 f'台股上市  ·  {period_lbl}漲跌（{_pl}）  ·  上漲 {up}  下跌 {down}  平盤 {flat}'
                 f'  ·  點擊類股進入 · {now} 更新',
                 ha='center', va='center', color='#888', fontsize=8,
                 fontfamily=CHART_FONT, transform=sax.transAxes)
        self._mkt_canvas.draw()
        self._mkt_status.set(f'更新時間：{now}  【{_pl}】  （點擊類股區域可放大檢視）')

    # ── 類股 drill-down 繪製 ──────────────────────────────────────────────────

    def _render_mkt_drill(self, ind_name: str,
                          stocks: list | None = None,
                          back_label: str = '返回總覽'):
        if stocks is None:
            stocks = self._mkt_groups.get(ind_name, [])
        if not stocks:
            self._mkt_drill    = None
            self._mkt_sub_drill = None
            self._render_mkt_overview()
            return

        # 「其他業」/「其他概念」且非子類股展開：按原始 industry 分組，顯示各小類股方塊
        if ind_name in ('其他業', '其他概念') and back_label == '返回總覽':
            self._render_mkt_drill_other(stocks)
            return

        fig = self._mkt_fig
        fig.clf()
        # 頂部返回列
        bax = fig.add_axes([0, 0.93, 1, 0.07]); bax.axis('off')
        bax.set_facecolor('#1a1a2e')
        bax.text(0.012, 0.5, f'◀ {back_label}', va='center', color='#aecde8',
                 fontsize=10, fontweight='bold', fontfamily=CHART_FONT,
                 transform=bax.transAxes)
        ind_vals_all = {ind: sum(s['trade_val'] for s in stks)
                        for ind, stks in self._mkt_groups.items()}
        total_val = sum(ind_vals_all.values())
        ind_tv    = sum(s['trade_val'] for s in stocks)
        all_pcts  = [s['chg_pct'] for s in stocks]
        ind_avg   = (sum(s['chg_pct'] * s['trade_val'] for s in stocks) / ind_tv
                     if ind_tv else 0)
        sign_avg  = '+' if ind_avg >= 0 else ''
        bax.text(0.5, 0.5,
                 f'{ind_name}  ·  {len(stocks)} 檔  ·  '
                 f'占比 {ind_tv/total_val*100:.1f}%  ·  '
                 f'加權均漲跌 {sign_avg}{ind_avg:.2f}%',
                 va='center', ha='center', color='#dddddd',
                 fontsize=9, fontfamily=CHART_FONT, transform=bax.transAxes)

        ax  = fig.add_axes([0, 0.03, 1, 0.90])
        ax.set_xlim(0, 100); ax.set_ylim(0, 100); ax.axis('off')
        sax = fig.add_axes([0, 0, 1, 0.03]); sax.axis('off')

        dpi        = fig.dpi
        _px_per_ux = fig.get_size_inches()[0] * dpi / 100
        _px_per_uy = fig.get_size_inches()[1] * dpi * 0.90 / 100
        _pt_per_px = 72 / dpi
        SEP_COL    = '#111111'
        GAP        = 0.3
        _max_abs   = max((abs(p) for p in all_pcts), default=10.0) or 10.0

        stks_sorted = sorted(stocks, key=lambda s: -s['trade_val'])
        svals       = [s['trade_val'] for s in stks_sorted]
        norm_s      = squarify.normalize_sizes(svals, 100, 100)
        sr_raw      = squarify.squarify(norm_s, 0, 0, 100, 100)
        srects      = [{'x': r['x'],
                        'y': 100 - r['y'] - r['dy'],
                        'dx': r['dx'], 'dy': r['dy']} for r in sr_raw]

        self._mkt_rects     = []
        self._mkt_cat_rects = []

        for stk, sr in zip(stks_sorted, srects):
            rx = sr['x'] + GAP / 2;  ry = sr['y'] + GAP / 2
            rw = sr['dx'] - GAP;     rh = sr['dy'] - GAP
            if rw <= 0 or rh <= 0:
                continue
            ax.add_patch(plt.Rectangle(
                (rx, ry), rw, rh,
                facecolor=pnl_color(stk['chg_pct'], _max_abs),
                edgecolor=SEP_COL, linewidth=0.4, zorder=1))
            self._mkt_rects.append({
                'rx': rx, 'ry': ry, 'rw': rw, 'rh': rh,
                **stk, 'industry': ind_name,
            })
            self._draw_stock_label(ax, stk, rx, ry, rw, rh,
                                   _px_per_ux, _px_per_uy, _pt_per_px)

        period_lbl = {'RT': '即時', '1D': '今日', '5D': '5日', '10D': '10日', '1M': '1月'}.get(
            self._mkt_period, self._mkt_period)
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        sax.text(0.5, 0.5,
                 f'{ind_name}  ·  {period_lbl}漲跌  ·  點擊返回  ·  {now}',
                 ha='center', va='center', color='#888', fontsize=8,
                 fontfamily=CHART_FONT, transform=sax.transAxes)
        self._mkt_canvas.draw()
        self._mkt_status.set(f'{ind_name}  ·  {len(stocks)} 檔  ·  點擊任意處{back_label}')

    def _render_mkt_drill_other(self, stocks: list):
        """其他業/其他概念 drill-down：按原始 industry 分組，顯示各小類股方塊"""
        parent_name = self._mkt_drill or '其他業'

        # 按 industry 分組
        sub: dict[str, list] = {}
        for s in stocks:
            sub.setdefault(s['industry'], []).append(s)

        ind_vals_all = {ind: sum(s['trade_val'] for s in stks)
                        for ind, stks in self._mkt_groups.items()}
        total_val    = sum(ind_vals_all.values())
        other_tv     = sum(s['trade_val'] for s in stocks)

        # 每個小類股彙總資料
        sub_summary = []
        for sname, stks in sub.items():
            tv  = sum(s['trade_val'] for s in stks)
            avg = sum(s['chg_pct'] * s['trade_val'] for s in stks) / tv if tv else 0
            sub_summary.append({
                'name': sname, 'count': len(stks),
                'trade_val': tv, 'chg_pct': avg,
                'stocks': stks,
            })
        sub_summary.sort(key=lambda x: -x['trade_val'])

        # 更新強勢類股：改用子分類資料
        sub_raw_groups = {g['name']: g['stocks'] for g in sub_summary}
        self._update_mkt_hot_sectors(sub_raw_groups)

        fig = self._mkt_fig
        fig.clf()
        # 頂部返回列
        bax = fig.add_axes([0, 0.93, 1, 0.07]); bax.axis('off')
        bax.set_facecolor('#1a1a2e')
        bax.text(0.012, 0.5, '◀ 返回總覽', va='center', color='#aecde8',
                 fontsize=10, fontweight='bold', fontfamily=CHART_FONT,
                 transform=bax.transAxes)
        other_avg  = sum(s['chg_pct'] * s['trade_val'] for s in stocks) / other_tv if other_tv else 0
        sign_avg   = '+' if other_avg >= 0 else ''
        bax.text(0.5, 0.5,
                 f'{parent_name}  ·  {len(sub_summary)} 類  {len(stocks)} 檔  ·  '
                 f'占比 {other_tv/total_val*100:.1f}%  ·  '
                 f'加權均漲跌 {sign_avg}{other_avg:.2f}%',
                 va='center', ha='center', color='#dddddd',
                 fontsize=9, fontfamily=CHART_FONT, transform=bax.transAxes)

        ax  = fig.add_axes([0, 0.03, 1, 0.90])
        ax.set_xlim(0, 100); ax.set_ylim(0, 100); ax.axis('off')
        sax = fig.add_axes([0, 0, 1, 0.03]); sax.axis('off')

        dpi        = fig.dpi
        _px_per_ux = fig.get_size_inches()[0] * dpi / 100
        _px_per_uy = fig.get_size_inches()[1] * dpi * 0.90 / 100
        _pt_per_px = 72 / dpi
        SEP_COL    = '#111111'
        GAP        = 0.5
        HDR_H      = 4.0
        all_pcts   = [x['chg_pct'] for x in sub_summary]
        _max_abs   = max((abs(p) for p in all_pcts), default=10.0) or 10.0

        svals    = [x['trade_val'] for x in sub_summary]
        norm_s   = squarify.normalize_sizes(svals, 100, 100)
        sr_raw   = squarify.squarify(norm_s, 0, 0, 100, 100)
        srects   = [{'x': r['x'],
                     'y': 100 - r['y'] - r['dy'],
                     'dx': r['dx'], 'dy': r['dy']} for r in sr_raw]

        self._mkt_rects     = []
        self._mkt_cat_rects = []

        for grp, sr in zip(sub_summary, srects):
            cx = sr['x'] + GAP / 2;  cy = sr['y'] + GAP / 2
            cw = sr['dx'] - GAP;     ch = sr['dy'] - GAP
            if cw <= 0 or ch <= 0:
                continue

            # 類股背景（以加權平均漲跌上色）
            ax.add_patch(plt.Rectangle(
                (cx, cy), cw, ch,
                facecolor=pnl_color(grp['chg_pct'], _max_abs),
                edgecolor=SEP_COL, linewidth=0.8, zorder=1))

            # 類股標題列
            hdr_h    = min(HDR_H, ch * 0.35)
            hdr_y    = cy + ch - hdr_h
            hdr_h_px = hdr_h * _px_per_uy
            fs_hdr   = min(10, max(6, hdr_h_px * _pt_per_px * 0.55))
            ax.add_patch(plt.Rectangle(
                (cx, hdr_y), cw, hdr_h,
                facecolor='#1a1a2e', edgecolor='none', zorder=3))
            ax.text(cx + 0.4, hdr_y + hdr_h / 2,
                    f"{grp['name']}  {grp['trade_val']/other_tv*100:.1f}%",
                    ha='left', va='center', color='#aecde8',
                    fontsize=fs_hdr, fontweight='bold',
                    fontfamily=CHART_FONT, clip_on=True, zorder=4)

            # 連漲/跌天數指標（右側）
            gs_dir, gs_cnt = self._mkt_streaks.get(grp['name'], (0, 0))
            if gs_cnt >= 2 and cw > 2:
                g_sym   = '▲' if gs_dir > 0 else '▼'
                g_color = '#ff8080' if gs_dir > 0 else '#80e080'
                ax.text(cx + cw - 0.4, hdr_y + hdr_h / 2,
                        f'{g_sym}{gs_cnt}',
                        ha='right', va='center', color=g_color,
                        fontsize=fs_hdr, fontweight='bold',
                        fontfamily=CHART_FONT, clip_on=True, zorder=4)

            ax.add_patch(plt.Rectangle(
                (cx, cy), cw, ch,
                facecolor='none', edgecolor='#333333', linewidth=1.0, zorder=5))

            # 儲存類股 rect 供 hover
            self._mkt_cat_rects.append({
                'cx': cx, 'cy': cy, 'cw': cw, 'ch': ch,
                'name': grp['name'], 'count': grp['count'],
                'trade_val': grp['trade_val'],
                'avg_chg': grp['chg_pct'], 'hdr_h': hdr_h,
                'streak': (gs_dir, gs_cnt),
                '_is_other_sub': True,
            })

            # 內部個股小方塊
            inner_h = ch - hdr_h
            if inner_h < 1:
                continue
            stks_s  = sorted(grp['stocks'], key=lambda s: -s['trade_val'])
            svals_i = [s['trade_val'] for s in stks_s]
            if cw >= inner_h:
                norm_i = squarify.normalize_sizes(svals_i, cw, inner_h)
                ri_raw = squarify.squarify(norm_i, cx, cy, cw, inner_h)
                irect  = [{'x': r['x'],
                            'y': cy + inner_h - (r['y'] - cy) - r['dy'],
                            'dx': r['dx'], 'dy': r['dy']} for r in ri_raw]
            else:
                norm_i = squarify.normalize_sizes(svals_i, inner_h, cw)
                ri_t   = squarify.squarify(norm_i, 0, 0, inner_h, cw)
                irect  = [{'x': cx + r['y'],
                            'y': cy + inner_h - r['x'] - r['dx'],
                            'dx': r['dy'], 'dy': r['dx']} for r in ri_t]

            stk_max = max((s['trade_val'] for s in stks_s), default=1) or 1
            for stk, ir in zip(stks_s, irect):
                rx = ir['x'] + 0.15;  ry = ir['y'] + 0.15
                rw = ir['dx'] - 0.3;  rh = ir['dy'] - 0.3
                if rw <= 0 or rh <= 0:
                    continue
                ax.add_patch(plt.Rectangle(
                    (rx, ry), rw, rh,
                    facecolor=pnl_color(stk['chg_pct'], _max_abs),
                    edgecolor=SEP_COL, linewidth=0.2, zorder=2))
                self._mkt_rects.append({
                    'rx': rx, 'ry': ry, 'rw': rw, 'rh': rh,
                    **stk, 'industry': grp['name'],
                })
                self._draw_stock_label(ax, stk, rx, ry, rw, rh,
                                       _px_per_ux, _px_per_uy, _pt_per_px)

        period_lbl = {'RT': '即時', '1D': '今日', '5D': '5日', '10D': '10日', '1M': '1月'}.get(
            self._mkt_period, self._mkt_period)
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        sax.text(0.5, 0.5,
                 f'其他業  ·  {len(sub_summary)} 類  ·  {period_lbl}漲跌  ·  '
                 f'點擊空白處返回  ·  {now}',
                 ha='center', va='center', color='#888', fontsize=8,
                 fontfamily=CHART_FONT, transform=sax.transAxes)
        self._mkt_canvas.draw()
        self._mkt_status.set(f'其他業  ·  {len(sub_summary)} 個類股  ·  點擊任意處返回總覽')

    # ── 共用文字標籤 ──────────────────────────────────────────────────────────

    def _draw_stock_label(self, ax, stk, rx, ry, rw, rh,
                          _px_per_ux, _px_per_uy, _pt_per_px):
        min_dim = min(rw, rh)
        if min_dim < 2:
            return
        rw_px      = rw * _px_per_ux
        rh_px      = rh * _px_per_uy
        chars      = max(len(stk['code']), 4)
        name_chars = max(len(stk['name']), 2)
        fs_by_w    = 0.58 * rw_px * _pt_per_px / (chars * 0.60)
        fs_by_h    = 0.65 * rh_px * _pt_per_px / 3.6
        fs_code    = max(min(fs_by_w, fs_by_h, 44), 6)
        fs_name    = max(min(0.80 * rw_px * _pt_per_px / (name_chars * 0.95),
                             fs_code * 0.55, 22), 5)
        fs_sub     = max(fs_code * 0.55, 5)
        pnl_str    = (f"{'+' if stk['chg_pct'] >= 0 else ''}{stk['chg_pct']:.2f}%")
        cx_t  = rx + rw / 2
        mid_y = ry + rh * 0.50
        fcu   = fs_code / _pt_per_px / _px_per_uy
        fnu   = fs_name / _pt_per_px / _px_per_uy
        fsu   = fs_sub  / _pt_per_px / _px_per_uy

        if min_dim >= 8:
            gcn  = (fcu + fnu) * 0.55
            gnp  = (fnu + fsu) * 0.55
            half = (gcn + gnp) / 2
            ax.text(cx_t, mid_y + half, stk['code'],
                    ha='center', va='center', color='white',
                    fontsize=fs_code, fontweight='bold',
                    fontfamily=CHART_FONT, clip_on=True, zorder=4)
            ax.text(cx_t, mid_y + half - gcn, stk['name'],
                    ha='center', va='center', color='#dddddd',
                    fontsize=fs_name, fontfamily=CHART_FONT,
                    clip_on=True, zorder=4)
            ax.text(cx_t, mid_y - half, pnl_str,
                    ha='center', va='center', color='white',
                    fontsize=fs_sub, fontfamily=CHART_FONT,
                    clip_on=True, zorder=4)
        elif min_dim >= 5:
            gap = (fcu + fsu) * 0.55
            ax.text(cx_t, mid_y + gap * 0.5, stk['code'],
                    ha='center', va='center', color='white',
                    fontsize=fs_code, fontweight='bold',
                    fontfamily=CHART_FONT, clip_on=True, zorder=4)
            ax.text(cx_t, mid_y - gap * 0.5, pnl_str,
                    ha='center', va='center', color='white',
                    fontsize=fs_sub, fontfamily=CHART_FONT,
                    clip_on=True, zorder=4)
        else:
            ax.text(cx_t, mid_y, stk['code'],
                    ha='center', va='center', color='white',
                    fontsize=fs_code, fontweight='bold',
                    fontfamily=CHART_FONT, clip_on=True, zorder=4)

    # ── 點擊事件（drill-down / 返回）─────────────────────────────────────────

    def _on_mkt_click(self, event):
        if self._current_page != 4:
            return
        if event.inaxes is None or event.xdata is None:
            return
        x, y = event.xdata, event.ydata

        # 右鍵點個股區塊 → 開啟個股分析
        if event.button == 3:
            for r in self._mkt_rects:
                if (r['rx'] <= x <= r['rx'] + r['rw'] and
                        r['ry'] <= y <= r['ry'] + r['rh']):
                    code = r.get('code', '')
                    name = r.get('name', '')
                    if code:
                        self._mkt_tooltip.withdraw()
                        self._open_stk_analysis(code, name)
                        return
            return   # 右鍵未點到個股，不觸發 drill-down

        self._mkt_tooltip.withdraw()

        # ── 第三層：子類股個股檢視 → 返回其他業/其他概念展開 ───────────────
        if self._mkt_sub_drill:
            self._mkt_sub_drill = None
            self._render_mkt_drill(self._mkt_drill)   # 返回所屬的上層分組
            return

        # ── 第二層：其他業/其他概念展開 → 點擊子類股進第三層，點空白返回總覽 ───
        if self._mkt_drill in ('其他業', '其他概念'):
            for c in self._mkt_cat_rects:
                if (c.get('_is_other_sub') and
                        c['cx'] <= x <= c['cx'] + c['cw'] and
                        c['cy'] <= y <= c['cy'] + c['ch']):
                    sub_name = c['name']
                    self._mkt_sub_drill = sub_name
                    parent_key = self._mkt_drill
                    parent_stocks = self._mkt_groups.get(parent_key, [])
                    sub_stocks = [s for s in parent_stocks if s['industry'] == sub_name]
                    back_lbl = '返回其他業' if parent_key == '其他業' else '返回其他概念'
                    self._render_mkt_drill(sub_name,
                                           stocks=sub_stocks,
                                           back_label=back_lbl)
                    return
            # 沒點到子類股 → 返回總覽
            self._mkt_drill     = None
            self._mkt_sub_drill = None
            self._render_mkt_overview()
            return

        # ── 第二層：一般類股個股檢視 → 返回總覽 ─────────────────────────────
        if self._mkt_drill:
            self._mkt_drill = None
            self._render_mkt_overview()
            return

        # ── 第一層：總覽 → 點擊類股進入 ──────────────────────────────────────
        for c in self._mkt_cat_rects:
            if (c['cx'] <= x <= c['cx'] + c['cw'] and
                    c['cy'] <= y <= c['cy'] + c['ch']):
                self._mkt_drill = c['name']
                self._render_mkt_drill(c['name'])
                return

    # ── Hover tooltip ─────────────────────────────────────────────────────────

    def _on_mkt_motion(self, event):
        if self._current_page != 4:
            self._mkt_tooltip.withdraw()
            return
        if event.inaxes is None or not event.xdata:
            self._mkt_tooltip.withdraw()
            return
        x, y = event.xdata, event.ydata

        # 先找個股（含合併的「其他 N檔」方塊）
        for r in self._mkt_rects:
            if r['rx'] <= x <= r['rx'] + r['rw'] and r['ry'] <= y <= r['ry'] + r['rh']:
                sign = '+' if r['chg_pct'] >= 0 else ''
                tv   = r['trade_val'] / 1e8
                period_lbl = {'1D': '今日', '5D': '5日', '10D': '10日', '1M': '1月'}.get(
                    self._mkt_period, self._mkt_period)
                if r.get('_is_other'):
                    tt = (f"{r['name']}  ({r['industry']})\n"
                          f"{period_lbl}加權平均漲跌：{sign}{r['chg_pct']:.2f}%\n"
                          f"合計成交金額：{tv:.1f} 億\n"
                          f"點擊類股區域可查看明細")
                else:
                    if self._mkt_period in ('1D', 'RT'):
                        chg_line = f"漲跌：{r['change']:+.2f}  ({sign}{r['chg_pct']:.2f}%)"
                    else:
                        s1 = '+' if r['chg_1d'] >= 0 else ''
                        chg_line = (f"今日漲跌：{r['change']:+.2f} ({s1}{r['chg_1d']:.2f}%)  "
                                    f"{period_lbl}漲跌：{sign}{r['chg_pct']:.2f}%")
                    # 三大法人今日（若有資料）
                    inst = self._mkt_inst_today.get(r['code'], {})
                    if inst:
                        def _fmt(v): return f'+{v}' if v > 0 else str(v)
                        inst_line = (f"\n外資 {_fmt(inst.get('foreign',0))}張  "
                                     f"投信 {_fmt(inst.get('trust',0))}張  "
                                     f"自營 {_fmt(inst.get('dealer',0))}張")
                    else:
                        inst_line = ''
                    tt = (f"{r['code']}  {r['name']}\n"
                          f"收盤：{r['close']:.2f}  {chg_line}\n"
                          f"成交金額：{tv:.1f} 億  |  {r['industry']}"
                          f"{inst_line}\n"
                          f"Ctrl+右鍵 → 籌碼分析")
                self._mkt_tt_lbl.config(text=tt)
                self._place_tooltip(self._mkt_tooltip,
                                    event.guiEvent.x_root, event.guiEvent.y_root)
                self._mkt_tooltip.deiconify()
                return

        # 找類股區塊（總覽 + 其他業/其他概念展開層皆檢查）
        for c in self._mkt_cat_rects:
            if c['cx'] <= x <= c['cx'] + c['cw'] and c['cy'] <= y <= c['cy'] + c['ch']:
                sign = '+' if c['avg_chg'] >= 0 else ''
                tv   = c['trade_val'] / 1e8
                s_dir, s_cnt = c.get('streak', (0, 0))
                if s_cnt >= 2:
                    sym = '▲' if s_dir > 0 else '▼'
                    streak_line = f"連續{'上漲' if s_dir > 0 else '下跌'}：{sym}{s_cnt} 天\n"
                else:
                    streak_line = ''

                # 判斷游標是否在標題列（header bar）上
                HDR_H = c.get('hdr_h', 3.2)
                hdr_y = c['cy'] + c['ch'] - HDR_H
                in_header = y >= hdr_y

                if in_header:
                    # 取得此類股的股票清單
                    if c.get('_is_other_sub'):
                        # 其他業/其他概念 子類股：從父類中篩選
                        parent_stks = self._mkt_groups.get(self._mkt_drill, [])
                        stks = [s for s in parent_stks if s.get('industry') == c['name']]
                    else:
                        stks = self._mkt_groups.get(c['name'], [])
                    stks_sorted = sorted(stks, key=lambda s: -s['trade_val'])
                    lines = [f"{c['name']}  {c['count']} 檔  |  加權均漲跌：{sign}{c['avg_chg']:.2f}%  |  成交 {tv:.1f} 億",
                             '─' * 38]
                    for s in stks_sorted[:20]:
                        s_tv = s['trade_val'] / 1e8
                        lines.append(
                            f"  {s['code']:<6}  {s['name']:<8}  "
                            f"{s['chg_pct']:+.2f}%   {s_tv:.1f}億")
                    if len(stks_sorted) > 20:
                        lines.append(f"  … 另有 {len(stks_sorted) - 20} 檔")
                    tt_text = '\n'.join(lines)
                else:
                    drill_hint = '點擊進入子類股詳細' if c.get('_is_other_sub') else '點擊進入類股詳細'
                    tt_text = (
                        f"{c['name']}  共 {c['count']} 檔\n"
                        f"加權平均漲跌：{sign}{c['avg_chg']:.2f}%\n"
                        f"{streak_line}"
                        f"類股成交金額：{tv:.1f} 億\n"
                        f"{drill_hint}")

                self._mkt_tt_lbl.config(text=tt_text)
                self._place_tooltip(self._mkt_tooltip,
                                    event.guiEvent.x_root, event.guiEvent.y_root)
                self._mkt_tooltip.deiconify()
                return

        self._mkt_tooltip.withdraw()

    # ── 另存圖檔 ───────────────────────────────────────────────────────────────

    def _save_market_map(self):
        from tkinter import filedialog
        path = filedialog.asksaveasfilename(
            title='儲存台股總覽圖',
            defaultextension='.png',
            filetypes=[('PNG 圖片', '*.png'), ('JPEG 圖片', '*.jpg'), ('所有檔案', '*.*')],
            initialfile='market_overview.png')
        if path:
            self._mkt_fig.savefig(path, dpi=150, bbox_inches='tight',
                                   facecolor=self._mkt_fig.get_facecolor())
            self._mkt_status.set(f'已儲存：{path}')

    # ═══════════════════════════════════════════════════════════════════════════
    # Tab 7：美股市場總覽
    # ═══════════════════════════════════════════════════════════════════════════

    def _build_tab7(self):
        f = self.tab7

        # ── 控制列 ────────────────────────────────────────────────────────────
        ctrl = ttk.Frame(f)
        ctrl.pack(fill='x', padx=14, pady=8)
        ttk.Label(ctrl, text='美股市場總覽', style='Hdr.TLabel').pack(side='left')
        ttk.Button(ctrl, text='💾  另存圖檔', style='Nav.TButton',
                   command=self._save_us_market_map).pack(side='right', padx=(4, 0))
        ttk.Button(ctrl, text='🔄  更新', style='Nav.TButton',
                   command=self._draw_us_market_map).pack(side='right')

        self._us_mkt_status = tk.StringVar(value='切換至此頁面自動更新')
        ttk.Label(ctrl, textvariable=self._us_mkt_status,
                  foreground=C_FG2, font=('Microsoft JhengHei', 9)).pack(side='left', padx=14)

        # ── 期間選擇 ──────────────────────────────────────────────────────────
        period_bar = tk.Frame(f, bg=C_BG)
        period_bar.pack(fill='x', padx=14, pady=(0, 4))
        tk.Label(period_bar, text='期間：', bg=C_BG, fg=C_FG2,
                 font=('Microsoft JhengHei', 9)).pack(side='left')
        self._us_mkt_period = '1D'
        self._us_mkt_period_btns: dict[str, tk.Label] = {}
        for pid, plbl in [('1D', '1日'), ('5D', '5日'), ('1M', '1月'), ('3M', '3月')]:
            btn = tk.Label(period_bar, text=plbl, bg='#1a1a2e', fg=C_FG2,
                           font=('Microsoft JhengHei', 9), padx=10, pady=3,
                           relief='flat', cursor='hand2')
            btn.pack(side='left', padx=2)
            btn.bind('<Button-1>', lambda e, p=pid: self._set_us_mkt_period(p))
            self._us_mkt_period_btns[pid] = btn
        self._update_us_period_btn_style()

        # ── 摘要列 ────────────────────────────────────────────────────────────
        stats_bar = tk.Frame(f, bg=C_PANEL)
        stats_bar.pack(fill='x', padx=8, pady=(0, 4))

        def _sc(parent, title):
            fr = tk.Frame(parent, bg='#1a1a2e', padx=14, pady=6)
            fr.pack(side='left', padx=3)
            tk.Label(fr, text=title, bg='#1a1a2e', fg='#6a8faf',
                     font=('Microsoft JhengHei', 8)).pack(anchor='w')
            var = tk.StringVar(value='—')
            lbl = tk.Label(fr, textvariable=var, bg='#1a1a2e',
                           fg=C_FG, font=('Microsoft JhengHei', 11, 'bold'))
            lbl.pack(anchor='w')
            return var, lbl

        self._us_sv_sp500,  self._us_sl_sp500  = _sc(stats_bar, 'S&P 500')
        self._us_sv_nasdaq, self._us_sl_nasdaq  = _sc(stats_bar, 'Nasdaq')
        self._us_sv_dow,    self._us_sl_dow     = _sc(stats_bar, 'Dow Jones')
        self._us_sv_up,     self._us_sl_up      = _sc(stats_bar, '上漲')
        self._us_sv_down,   self._us_sl_down    = _sc(stats_bar, '下跌')
        self._us_sv_total,  self._us_sl_total   = _sc(stats_bar, '總計')

        # ── 強勢類股排行（右側）──────────────────────────────────────────────
        self._us_hot_frame = tk.Frame(stats_bar, bg='#1a1a2e')
        self._us_hot_frame.pack(side='left', fill='both', expand=True, padx=(8, 4), pady=3)
        tk.Label(self._us_hot_frame, text='強勢類股', bg='#1a1a2e', fg='#6a8faf',
                 font=('Microsoft JhengHei', 8)).pack(anchor='nw', padx=8, pady=(4, 0))
        self._us_hot_lbl = tk.Label(
            self._us_hot_frame, text='', bg='#1a1a2e', fg='#4ec94e',
            font=('Microsoft JhengHei', 11, 'bold'),
            anchor='w', justify='left', padx=8, pady=0)
        self._us_hot_lbl.pack(fill='x', expand=True)

        # ── Treemap canvas ────────────────────────────────────────────────────
        self._us_mkt_fig = plt.Figure(figsize=(9.5, 5.5), dpi=100, facecolor='#111111')
        self._us_mkt_canvas = FigureCanvasTkAgg(self._us_mkt_fig, master=f)
        w = self._us_mkt_canvas.get_tk_widget()
        w.configure(bg='#111111')
        w.pack(fill='both', expand=True, padx=8, pady=(0, 8))
        w.bind('<Leave>', lambda e: self._us_mkt_tooltip.withdraw()
               if hasattr(self, '_us_mkt_tooltip') else None)

        # ── Tooltip ───────────────────────────────────────────────────────────
        self._us_mkt_tooltip = tk.Toplevel(self)
        self._us_mkt_tooltip.withdraw()
        self._us_mkt_tooltip.overrideredirect(True)
        self._us_mkt_tooltip.configure(bg='#1a1a2e')
        self._us_mkt_tt_lbl = tk.Label(
            self._us_mkt_tooltip, bg='#1a1a2e', fg=C_FG,
            font=('Microsoft JhengHei', 9), justify='left', padx=8, pady=6)
        self._us_mkt_tt_lbl.pack()

        # ── 內部狀態 ──────────────────────────────────────────────────────────
        self._us_mkt_rects:     list = []
        self._us_mkt_cat_rects: list = []
        self._us_mkt_groups:    dict = {}
        self._us_mkt_drill:     str | None = None
        self._us_mkt_hist_cache: dict = {}
        self._us_mkt_up = self._us_mkt_down = 0

        self._us_mkt_canvas.mpl_connect('motion_notify_event', self._on_us_mkt_motion)
        self._us_mkt_canvas.mpl_connect('button_press_event', self._on_us_mkt_click)

    # ── 期間切換 ──────────────────────────────────────────────────────────────

    def _set_us_mkt_period(self, period: str):
        if period == self._us_mkt_period:
            return
        self._us_mkt_period = period
        self._update_us_period_btn_style()
        self._draw_us_market_map()

    def _update_us_period_btn_style(self):
        for pid, btn in self._us_mkt_period_btns.items():
            if pid == self._us_mkt_period:
                btn.config(bg='#2a4a6e', fg='white')
            else:
                btn.config(bg='#1a1a2e', fg=C_FG2)

    # ── 資料抓取 + 繪製（背景執行緒）─────────────────────────────────────────

    def _draw_us_market_map(self):
        self._us_mkt_status.set('🔄 載入美股資料中…')
        self._us_mkt_rects     = []
        self._us_mkt_cat_rects = []
        self._us_mkt_drill     = None
        self._us_mkt_hist_cache.clear()
        threading.Thread(target=self._draw_us_market_map_impl, daemon=True).start()

    def _draw_us_market_map_impl(self):
        try:
            import logging as _log, io, sys as _sys
            _yf_log   = _log.getLogger('yfinance')
            _prev_lvl = _yf_log.level
            _yf_log.setLevel(_log.CRITICAL)

            period    = self._us_mkt_period
            n_map     = {'1D': '2d', '5D': '7d', '1M': '35d', '3M': '100d'}
            yf_period = n_map.get(period, '2d')

            # 收集所有代號
            all_tickers: list[str] = []
            ticker_sector: dict[str, str] = {}
            for sector, tickers in _US_SECTOR_STOCKS.items():
                for t in tickers:
                    if t not in ticker_sector:
                        all_tickers.append(t)
                        ticker_sector[t] = sector

            # 批次下載歷史資料
            _old_stderr, _sys.stderr = _sys.stderr, io.StringIO()
            try:
                raw = yf.download(
                    all_tickers, period=yf_period,
                    progress=False, auto_adjust=True,
                    group_by='ticker', threads=True)
            finally:
                _sys.stderr = _old_stderr
                _yf_log.setLevel(_prev_lvl)

            # 解析每支股票的漲跌幅與成交金額
            base_stocks: dict[str, dict] = {}
            up = down = 0

            def _get_series(ticker):
                try:
                    if len(all_tickers) == 1:
                        return raw
                    return raw[ticker]
                except Exception:
                    return None

            for tkr in all_tickers:
                try:
                    df = _get_series(tkr)
                    if df is None or df.empty:
                        continue
                    closes = df['Close'].dropna()
                    vols   = df['Volume'].dropna()
                    if len(closes) < 2:
                        continue

                    close_now  = float(closes.iloc[-1])
                    close_prev = float(closes.iloc[-2 if period == '1D' else max(-len(closes), -(
                        {'5D': 6, '1M': 22, '3M': 66}.get(period, 2)))])
                    vol_now    = float(vols.iloc[-1]) if not vols.empty else 0.0

                    if close_prev <= 0:
                        continue
                    chg_pct = (close_now - close_prev) / close_prev * 100
                    trade_val = close_now * vol_now   # 近似成交金額（USD）

                    if chg_pct > 0:  up   += 1
                    elif chg_pct < 0: down += 1

                    sector = ticker_sector.get(tkr, '其他')
                    base_stocks[tkr] = {
                        'code':      tkr,
                        'name':      tkr,
                        'sector':    sector,
                        'close':     close_now,
                        'chg_pct':   chg_pct,
                        'trade_val': max(trade_val, 1.0),
                    }
                except Exception:
                    pass

            # 依類股組裝 groups
            MERGE_THRESHOLD = 0.5
            raw_groups: dict[str, list] = {}
            total_tv = sum(s['trade_val'] for s in base_stocks.values()) or 1.0
            for stk in base_stocks.values():
                raw_groups.setdefault(stk['sector'], []).append(stk)

            groups: dict[str, list] = {}
            other_stocks: list = []
            for sec, stks in raw_groups.items():
                sec_tv = sum(s['trade_val'] for s in stks)
                pct = sec_tv / total_tv * 100
                if pct < MERGE_THRESHOLD:
                    other_stocks.extend(stks)
                else:
                    groups[sec] = stks
            if other_stocks:
                groups.setdefault('其他', []).extend(other_stocks)

            # 抓取三大指數
            sp500_price = sp500_chg = None
            nasdaq_price = nasdaq_chg = None
            dow_price = dow_chg = None
            try:
                idx_raw = yf.download(
                    ['^GSPC', '^IXIC', '^DJI'], period='5d',
                    progress=False, auto_adjust=True,
                    group_by='ticker', threads=True)
                def _idx_chg(tkr):
                    try:
                        closes = idx_raw[tkr]['Close'].dropna()
                        if len(closes) >= 2:
                            p = float(closes.iloc[-1])
                            c = (closes.iloc[-1] - closes.iloc[-2]) / closes.iloc[-2] * 100
                            return float(p), float(c)
                    except Exception:
                        pass
                    return None, None
                sp500_price,  sp500_chg  = _idx_chg('^GSPC')
                nasdaq_price, nasdaq_chg = _idx_chg('^IXIC')
                dow_price,    dow_chg    = _idx_chg('^DJI')
            except Exception:
                pass

            _period_lbl = {'1D': '今日', '5D': '5日', '1M': '1月', '3M': '3月'}.get(period, period)
            self._ui_call(lambda: self._render_us_market_map(
                groups, raw_groups, up, down,
                sp500_price, sp500_chg,
                nasdaq_price, nasdaq_chg,
                dow_price, dow_chg,
                _period_lbl))
        except Exception as e:
            _e = e
            self._ui_call(lambda: self._us_mkt_status.set(f'⚠ 載入失敗：{_e}'))

    def _render_us_market_map(self, groups, raw_groups, up, down,
                               sp500_price, sp500_chg,
                               nasdaq_price, nasdaq_chg,
                               dow_price, dow_chg,
                               period_label='今日'):
        self._us_mkt_groups = groups
        self._us_mkt_up     = up
        self._us_mkt_down   = down

        # ── 更新摘要列 ────────────────────────────────────────────────────────
        def _fmt_idx(var, lbl, price, chg):
            if price:
                sign = '+' if (chg or 0) >= 0 else ''
                chg_str = f'  {sign}{chg:.2f}%' if chg is not None else ''
                var.set(f'{price:,.2f}{chg_str}')
                lbl.config(fg='#4ec94e' if (chg or 0) >= 0 else '#f07070')
            else:
                var.set('—')

        _fmt_idx(self._us_sv_sp500,  self._us_sl_sp500,  sp500_price,  sp500_chg)
        _fmt_idx(self._us_sv_nasdaq, self._us_sl_nasdaq, nasdaq_price, nasdaq_chg)
        _fmt_idx(self._us_sv_dow,    self._us_sl_dow,    dow_price,    dow_chg)

        self._us_sv_up   .set(f'{up} 檔');        self._us_sl_up  .config(fg='#4ec94e')
        self._us_sv_down .set(f'{down} 檔');      self._us_sl_down.config(fg='#f07070')
        self._us_sv_total.set(f'{up + down} 檔')

        # 強勢類股
        sector_chg: list[tuple[float, str]] = []
        for sec, stks in raw_groups.items():
            if not stks:
                continue
            total_tv = sum(s['trade_val'] for s in stks)
            if total_tv <= 0:
                continue
            w_chg = sum(s['chg_pct'] * s['trade_val'] for s in stks) / total_tv
            sector_chg.append((w_chg, sec))
        rising = sorted([(c, n) for c, n in sector_chg if c > 0], reverse=True)[:3]
        parts = [f'{n}  +{c:.1f}%' for c, n in rising]
        self._us_hot_lbl.config(text='    '.join(parts) if parts else '—')

        self._us_mkt_period_label = period_label
        if self._us_mkt_drill:
            self._render_us_drill(self._us_mkt_drill)
        else:
            self._render_us_overview()

    # ── 總覽繪製 ──────────────────────────────────────────────────────────────

    def _render_us_overview(self):
        groups = self._us_mkt_groups
        fig    = self._us_mkt_fig
        fig.clf()
        ax  = fig.add_axes([0, 0.03, 1, 0.97])
        ax.set_xlim(0, 100); ax.set_ylim(0, 100); ax.axis('off')
        sax = fig.add_axes([0, 0, 1, 0.03]); sax.axis('off')

        dpi        = fig.dpi
        _px_per_ux = fig.get_size_inches()[0] * dpi / 100
        _px_per_uy = fig.get_size_inches()[1] * dpi * 0.97 / 100
        _pt_per_px = 72 / dpi
        SEP_COL    = '#111111'
        GAP        = 0.3
        HDR_H      = 3.2

        ind_vals    = {sec: sum(s['trade_val'] for s in stks)
                       for sec, stks in groups.items()}
        sorted_inds = sorted(groups, key=lambda x: -ind_vals[x])
        all_pcts    = [s['chg_pct'] for stks in groups.values() for s in stks]
        _max_abs    = max((abs(p) for p in all_pcts), default=5.0) or 5.0
        total_val   = sum(ind_vals.values())

        if not total_val:
            ax.text(50, 50, '無資料', ha='center', va='center',
                    color='#888', fontsize=14, fontfamily=CHART_FONT)
            self._us_mkt_canvas.draw()
            self._us_mkt_status.set('無法取得資料')
            return

        cat_sizes = [ind_vals[i] for i in sorted_inds]
        norm_cats = squarify.normalize_sizes(cat_sizes, 100, 96)
        cats_raw  = squarify.squarify(norm_cats, 0, 4, 100, 96)
        cat_rects = [{'x': r['x'],
                      'y': 4 + 96 - (r['y'] - 4) - r['dy'],
                      'dx': r['dx'], 'dy': r['dy']} for r in cats_raw]

        self._us_mkt_rects     = []
        self._us_mkt_cat_rects = []

        for sec_name, cr in zip(sorted_inds, cat_rects):
            stocks = groups[sec_name]
            cx, cy, cw, ch = cr['x'], cr['y'], cr['dx'], cr['dy']

            hdr_y    = cy + ch - HDR_H
            hdr_h_px = HDR_H * _px_per_uy
            fs_hdr   = min(10, max(6, hdr_h_px * _pt_per_px * 0.62))
            ax.add_patch(plt.Rectangle(
                (cx, hdr_y), cw, HDR_H,
                facecolor='#1a1a2e', edgecolor='none', zorder=6, clip_on=True))

            sec_tv  = ind_vals[sec_name]
            sec_avg = (sum(s['chg_pct'] * s['trade_val'] for s in stocks) / sec_tv
                       if sec_tv else 0)
            sign_hdr = '+' if sec_avg >= 0 else ''
            ax.text(cx + 0.6, hdr_y + HDR_H / 2,
                    f'{sec_name}  {sec_tv/total_val*100:.1f}%  {sign_hdr}{sec_avg:.2f}%',
                    ha='left', va='center', color='#aecde8',
                    fontsize=fs_hdr, fontweight='bold',
                    fontfamily=CHART_FONT, clip_on=True, zorder=7)

            ax.add_patch(plt.Rectangle(
                (cx, cy), cw, ch,
                facecolor='none', edgecolor='#2a2a2a', linewidth=1.5, zorder=8))

            self._us_mkt_cat_rects.append({
                'cx': cx, 'cy': cy, 'cw': cw, 'ch': ch,
                'name': sec_name, 'count': len(stocks),
                'trade_val': sec_tv, 'avg_chg': sec_avg, 'hdr_h': HDR_H,
            })

            inner_h = ch - HDR_H
            if inner_h < 1:
                continue

            STK_MERGE_THR = 2.0
            stks_sorted = sorted(stocks, key=lambda s: -s['trade_val'])
            sec_tv_full = sum(s['trade_val'] for s in stks_sorted) or 1
            main_stks, tiny_stks = [], []
            for s in stks_sorted:
                if s['trade_val'] / sec_tv_full * 100 >= STK_MERGE_THR:
                    main_stks.append(s)
                else:
                    tiny_stks.append(s)
            if tiny_stks:
                tiny_tv  = sum(s['trade_val'] for s in tiny_stks)
                tiny_avg = (sum(s['chg_pct'] * s['trade_val'] for s in tiny_stks) / tiny_tv
                            if tiny_tv else 0)
                main_stks.append({
                    'code': '', 'name': f'其他 {len(tiny_stks)}檔',
                    'close': 0, 'chg_pct': tiny_avg,
                    'trade_val': tiny_tv, 'sector': sec_name,
                    '_is_other': True,
                })
            stks_sorted = main_stks

            svals = [s['trade_val'] for s in stks_sorted]
            if cw >= inner_h:
                norm_s = squarify.normalize_sizes(svals, cw, inner_h)
                sr_raw = squarify.squarify(norm_s, cx, cy, cw, inner_h)
                srects = [{'x': r['x'],
                           'y': cy + inner_h - (r['y'] - cy) - r['dy'],
                           'dx': r['dx'], 'dy': r['dy']} for r in sr_raw]
            else:
                norm_s = squarify.normalize_sizes(svals, inner_h, cw)
                sr_t   = squarify.squarify(norm_s, 0, 0, inner_h, cw)
                srects = [{'x': cx + r['y'],
                           'y': cy + inner_h - r['x'] - r['dx'],
                           'dx': r['dy'], 'dy': r['dx']} for r in sr_t]

            for stk, sr in zip(stks_sorted, srects):
                rx = sr['x'] + GAP / 2;  ry = sr['y'] + GAP / 2
                rw = sr['dx'] - GAP;     rh = sr['dy'] - GAP
                if rw <= 0 or rh <= 0:
                    continue
                ax.add_patch(plt.Rectangle(
                    (rx, ry), rw, rh,
                    facecolor=pnl_color_us(stk['chg_pct'], _max_abs),
                    edgecolor=SEP_COL, linewidth=0.3, zorder=1))
                self._us_mkt_rects.append({
                    'rx': rx, 'ry': ry, 'rw': rw, 'rh': rh,
                    **stk, 'sector': sec_name,
                })
                self._draw_stock_label(ax, stk, rx, ry, rw, rh,
                                       _px_per_ux, _px_per_uy, _pt_per_px)

        period_lbl = getattr(self, '_us_mkt_period_label', '今日')
        sax.text(0.5, 0.5, f'顏色：綠色=上漲  紅色=下跌  | 大小：成交金額（USD）  | {period_lbl}漲跌幅',
                 ha='center', va='center', color='#666', fontsize=7,
                 fontfamily=CHART_FONT, transform=sax.transAxes)

        self._us_mkt_canvas.draw_idle()
        total = self._us_mkt_up + self._us_mkt_down
        self._us_mkt_status.set(
            f'上漲 {self._us_mkt_up} / 下跌 {self._us_mkt_down} / 共 {total} 檔  '
            f'（{datetime.now().strftime("%H:%M:%S")} 更新）')

    # ── Drill-down 類股檢視 ────────────────────────────────────────────────────

    def _render_us_drill(self, sec_name: str):
        stocks  = self._us_mkt_groups.get(sec_name, [])
        fig     = self._us_mkt_fig
        fig.clf()
        ax  = fig.add_axes([0, 0.04, 1, 0.96])
        ax.set_xlim(0, 100); ax.set_ylim(0, 100); ax.axis('off')
        sax = fig.add_axes([0, 0, 1, 0.04]); sax.axis('off')

        dpi        = fig.dpi
        _px_per_ux = fig.get_size_inches()[0] * dpi / 100
        _px_per_uy = fig.get_size_inches()[1] * dpi * 0.96 / 100
        _pt_per_px = 72 / dpi
        SEP_COL    = '#111111'
        GAP        = 0.3

        all_pcts = [s['chg_pct'] for s in stocks]
        _max_abs = max((abs(p) for p in all_pcts), default=5.0) or 5.0
        total_tv = sum(s['trade_val'] for s in stocks) or 1.0
        stks_sorted = sorted(stocks, key=lambda s: -s['trade_val'])

        if not stks_sorted:
            ax.text(50, 50, '無資料', ha='center', va='center',
                    color='#888', fontsize=14, fontfamily=CHART_FONT)
            self._us_mkt_canvas.draw()
            return

        svals  = [s['trade_val'] for s in stks_sorted]
        norm_s = squarify.normalize_sizes(svals, 100, 100)
        sr_raw = squarify.squarify(norm_s, 0, 0, 100, 100)
        srects = [{'x': r['x'], 'y': 100 - r['y'] - r['dy'],
                   'dx': r['dx'], 'dy': r['dy']} for r in sr_raw]

        self._us_mkt_rects     = []
        self._us_mkt_cat_rects = []

        for stk, sr in zip(stks_sorted, srects):
            rx = sr['x'] + GAP / 2;  ry = sr['y'] + GAP / 2
            rw = sr['dx'] - GAP;     rh = sr['dy'] - GAP
            if rw <= 0 or rh <= 0:
                continue
            ax.add_patch(plt.Rectangle(
                (rx, ry), rw, rh,
                facecolor=pnl_color_us(stk['chg_pct'], _max_abs),
                edgecolor=SEP_COL, linewidth=0.3, zorder=1))
            self._us_mkt_rects.append({
                'rx': rx, 'ry': ry, 'rw': rw, 'rh': rh, **stk})
            self._draw_stock_label(ax, stk, rx, ry, rw, rh,
                                   _px_per_ux, _px_per_uy, _pt_per_px)

        sec_avg = sum(s['chg_pct'] * s['trade_val'] for s in stocks) / total_tv
        sign    = '+' if sec_avg >= 0 else ''
        period_lbl = getattr(self, '_us_mkt_period_label', '今日')
        sax.text(0.5, 0.5,
                 f'{sec_name}  共 {len(stks_sorted)} 檔  |  加權平均 {sign}{sec_avg:.2f}%  '
                 f'|  {period_lbl}漲跌  |  點擊空白處返回總覽',
                 ha='center', va='center', color='#888', fontsize=8,
                 fontfamily=CHART_FONT, transform=sax.transAxes)

        self._us_mkt_canvas.draw_idle()

    # ── 點擊事件 ──────────────────────────────────────────────────────────────

    def _on_us_mkt_click(self, event):
        if self._current_page != 6:
            return
        if event.inaxes is None or event.xdata is None:
            return
        x, y = event.xdata, event.ydata
        self._us_mkt_tooltip.withdraw()

        if self._us_mkt_drill:
            self._us_mkt_drill = None
            self._render_us_overview()
            return

        for c in self._us_mkt_cat_rects:
            if (c['cx'] <= x <= c['cx'] + c['cw'] and
                    c['cy'] <= y <= c['cy'] + c['ch']):
                self._us_mkt_drill = c['name']
                self._render_us_drill(c['name'])
                return

    # ── Hover tooltip ─────────────────────────────────────────────────────────

    def _on_us_mkt_motion(self, event):
        if self._current_page != 6:
            self._us_mkt_tooltip.withdraw()
            return
        if event.inaxes is None or not event.xdata:
            self._us_mkt_tooltip.withdraw()
            return
        x, y = event.xdata, event.ydata

        for r in self._us_mkt_rects:
            if (r['rx'] <= x <= r['rx'] + r['rw'] and
                    r['ry'] <= y <= r['ry'] + r['rh']):
                sign = '+' if r['chg_pct'] >= 0 else ''
                tv_b = r['trade_val'] / 1e9
                period_lbl = getattr(self, '_us_mkt_period_label', '今日')
                if r.get('_is_other'):
                    tt = (f"{r['name']}  ({r['sector']})\n"
                          f"{period_lbl}加權平均漲跌：{sign}{r['chg_pct']:.2f}%\n"
                          f"合計成交金額：${tv_b:.2f}B")
                else:
                    tt = (f"{r['code']}\n"
                          f"收盤：${r['close']:.2f}  {period_lbl}漲跌：{sign}{r['chg_pct']:.2f}%\n"
                          f"成交金額：${tv_b:.2f}B  |  {r.get('sector', '')}")
                self._us_mkt_tt_lbl.config(text=tt)
                self._place_tooltip(self._us_mkt_tooltip,
                                    event.guiEvent.x_root, event.guiEvent.y_root)
                self._us_mkt_tooltip.deiconify()
                return

        if not self._us_mkt_drill:
            for c in self._us_mkt_cat_rects:
                if (c['cx'] <= x <= c['cx'] + c['cw'] and
                        c['cy'] <= y <= c['cy'] + c['ch']):
                    sign = '+' if c['avg_chg'] >= 0 else ''
                    tv_b = c['trade_val'] / 1e9
                    self._us_mkt_tt_lbl.config(text=(
                        f"{c['name']}  共 {c['count']} 檔\n"
                        f"加權平均漲跌：{sign}{c['avg_chg']:.2f}%\n"
                        f"類股成交金額：${tv_b:.2f}B\n"
                        f"點擊進入類股詳細"))
                    self._place_tooltip(self._us_mkt_tooltip,
                                        event.guiEvent.x_root, event.guiEvent.y_root)
                    self._us_mkt_tooltip.deiconify()
                    return

        self._us_mkt_tooltip.withdraw()

    # ── 另存圖檔 ───────────────────────────────────────────────────────────────

    def _save_us_market_map(self):
        from tkinter import filedialog
        path = filedialog.asksaveasfilename(
            title='儲存美股總覽圖',
            defaultextension='.png',
            filetypes=[('PNG 圖片', '*.png'), ('JPEG 圖片', '*.jpg'), ('所有檔案', '*.*')],
            initialfile='us_market_overview.png')
        if path:
            self._us_mkt_fig.savefig(path, dpi=150, bbox_inches='tight',
                                      facecolor=self._us_mkt_fig.get_facecolor())
            self._us_mkt_status.set(f'已儲存：{path}')

    # ═══════════════════════════════════════════════════════════════════════════
    # Tab 6 — 籌碼分析
    # ═══════════════════════════════════════════════════════════════════════════

    def _build_tab6(self):
        """個股分析 — K 線圖（同 ETF 分析，但針對任意個股）"""
        f = self.tab6
        CHART_BG = '#111111'

        # ── 標題 + 代號輸入 ────────────────────────────────────────────────
        ctrl = ttk.Frame(f)
        ctrl.pack(fill='x', padx=14, pady=8)
        ttk.Label(ctrl, text='個股分析', style='Hdr.TLabel').pack(side='left')

        ttk.Label(ctrl, text='股票代號：',
                  foreground=C_FG2, font=('Microsoft JhengHei', 9)).pack(side='left', padx=(20, 4))
        self._stk_code = tk.StringVar()
        stk_entry = ttk.Entry(ctrl, textvariable=self._stk_code, width=10)
        stk_entry.pack(side='left', padx=4)
        stk_entry.bind('<Return>', lambda _e: self._load_stk_chart())

        ttk.Button(ctrl, text='📈  查詢', style='Nav.TButton',
                   command=self._load_stk_chart).pack(side='left', padx=(6, 0))

        self._stk_status = tk.StringVar(value='請輸入股票代號後點擊查詢')
        status_bar = tk.Frame(f, bg=C_BG)
        status_bar.pack(fill='x', padx=14, pady=(0, 2))
        ttk.Label(status_bar, textvariable=self._stk_status,
                  foreground='#c8c8d8', font=('Microsoft JhengHei', 13, 'bold')).pack(anchor='w')

        # ── 單一捲動區域：K線 + 財報連續顯示 ───────────────────────────────
        _CTRL_BG = '#1c1c28'

        # 控制列固定在最上方（不隨捲動移動）
        kctrl = tk.Frame(f, bg=_CTRL_BG, pady=4)
        kctrl.pack(fill='x', padx=8, pady=(0, 2))

        _BTN_OFF = dict(bg='#2a2a3e', fg='#8899cc',
                        font=('Microsoft JhengHei', 8, 'bold'),
                        relief='flat', bd=0, padx=9, pady=3,
                        cursor='hand2',
                        highlightbackground='#3a3a5e', highlightthickness=1,
                        activebackground='#3a3a5e', activeforeground='#aaaaee')
        _BTN_ON  = dict(bg='#3a5fcd', fg='#ffffff',
                        font=('Microsoft JhengHei', 8, 'bold'),
                        relief='flat', bd=0, padx=9, pady=3,
                        cursor='hand2',
                        highlightbackground='#2a4fbd', highlightthickness=1,
                        activebackground='#4a70e0', activeforeground='#ffffff')

        def _s6btn(btn, on):
            btn.configure(**(_BTN_ON if on else _BTN_OFF))

        # 區間
        self._stk_period = '3M'
        self._stk_period_btns: dict[str, tk.Button] = {}
        tk.Label(kctrl, text='區間:', bg=_CTRL_BG, fg='#6677aa',
                 font=('Microsoft JhengHei', 8)).pack(side='left', padx=(8, 4))

        def _set_stk_period(lbl):
            self._stk_period = lbl
            for _l, _b in self._stk_period_btns.items():
                _s6btn(_b, _l == lbl)
            self._redraw_stk_kline()

        for _pl in ['1M', '3M', '6M', '1Y', '全部']:
            _b = tk.Button(kctrl, text=_pl, **_BTN_OFF,
                           command=lambda l=_pl: _set_stk_period(l))
            _b.pack(side='left', padx=2)
            self._stk_period_btns[_pl] = _b
        _s6btn(self._stk_period_btns['3M'], True)

        tk.Label(kctrl, text='│', bg=_CTRL_BG, fg='#333355').pack(side='left', padx=6)

        # 指標
        self._stk_ind_state: dict[str, bool] = {
            'MA': True, 'BB': False, 'VOL': True,
            'MACD': True, 'RSI': False, 'KD': True}
        self._stk_ind_btns: dict[str, tk.Button] = {}
        tk.Label(kctrl, text='指標:', bg=_CTRL_BG, fg='#6677aa',
                 font=('Microsoft JhengHei', 8)).pack(side='left', padx=(0, 4))

        def _toggle_stk_ind(lbl):
            self._stk_ind_state[lbl] = not self._stk_ind_state[lbl]
            _s6btn(self._stk_ind_btns[lbl], self._stk_ind_state[lbl])
            self._redraw_stk_kline()

        for _il in ['MA', 'BB', 'VOL', 'MACD', 'RSI', 'KD']:
            _on = self._stk_ind_state[_il]
            _b = tk.Button(kctrl, text=_il,
                           **(_BTN_ON if _on else _BTN_OFF),
                           command=lambda l=_il: _toggle_stk_ind(l))
            _b.pack(side='left', padx=2)
            self._stk_ind_btns[_il] = _b

        # WantGoo 快速連結（右側）
        tk.Label(kctrl, text='│', bg=_CTRL_BG, fg='#333355').pack(side='left', padx=6)
        self._wg_open_btn = tk.Button(
            kctrl, text='🌐 WantGoo', **_BTN_OFF,
            command=self._open_wantgoo_kline)
        self._wg_open_btn.pack(side='left', padx=2)

        # ── 當日資訊列 ───────────────────────────────────────────────────
        info_bar = tk.Frame(f, bg='#0d0f17', pady=4)
        info_bar.pack(fill='x', padx=8, pady=(0, 0))
        _FI  = ('Microsoft JhengHei', 9)
        _FIP = ('Microsoft JhengHei', 15, 'bold')
        self._stk_info_date   = tk.Label(info_bar, text='—', bg='#0d0f17', fg='#777788', font=_FI)
        self._stk_info_date.pack(side='left', padx=(6, 10))
        self._stk_info_close  = tk.Label(info_bar, text='—', bg='#0d0f17', fg='#cccccc', font=_FIP)
        self._stk_info_close.pack(side='left', padx=(0, 4))
        self._stk_info_chg    = tk.Label(info_bar, text='', bg='#0d0f17', fg='#cccccc', font=_FI)
        self._stk_info_chg.pack(side='left', padx=(0, 14))
        for _lbl, _attr in [('開', '_stk_info_o'), ('高', '_stk_info_h'),
                             ('低', '_stk_info_l'), ('量', '_stk_info_v')]:
            tk.Label(info_bar, text=_lbl, bg='#0d0f17', fg='#555566', font=_FI).pack(side='left', padx=(4, 1))
            _w = tk.Label(info_bar, text='—', bg='#0d0f17', fg='#cccccc', font=_FI)
            _w.pack(side='left', padx=(0, 4))
            setattr(self, _attr, _w)

        # ── 繪線工具列 ───────────────────────────────────────────────────
        draw_bar = tk.Frame(f, bg='#1c1c28', pady=3)
        draw_bar.pack(fill='x', padx=8, pady=(2, 0))
        tk.Label(draw_bar, text='繪線：', bg='#1c1c28', fg='#6677aa',
                 font=('Microsoft JhengHei', 8)).pack(side='left', padx=(8, 4))
        self._stk_draw_mode  = 'none'
        self._stk_draw_start = None
        self._stk_drawings   = []
        self._stk_draw_btns  = {}
        _DB_OFF = dict(bg='#2a2a3e', fg='#8899cc', font=('Microsoft JhengHei', 9),
                       relief='flat', padx=8, pady=2, cursor='hand2',
                       highlightthickness=0, activebackground='#3a3a5e',
                       activeforeground='#aaaaee')
        _DB_ON  = dict(bg='#3a5fcd', fg='white', font=('Microsoft JhengHei', 9),
                       relief='flat', padx=8, pady=2, cursor='hand2',
                       highlightthickness=0, activebackground='#4a70e0',
                       activeforeground='white')
        self._stk_db_off = _DB_OFF
        self._stk_db_on  = _DB_ON
        for _mode, _sym in [('line', '╱ 趨勢線'), ('hline', '─ 水平線'),
                             ('vline', '│ 垂直線'), ('erase', '✕ 清除')]:
            _b = tk.Button(draw_bar, text=_sym, **_DB_OFF,
                           command=lambda m=_mode: self._set_draw_mode(m))
            _b.pack(side='left', padx=2)
            self._stk_draw_btns[_mode] = _b

        # 顏色選擇
        tk.Label(draw_bar, text='│', bg='#1c1c28', fg='#333355').pack(side='left', padx=6)
        tk.Label(draw_bar, text='顏色：', bg='#1c1c28', fg='#6677aa',
                 font=('Microsoft JhengHei', 8)).pack(side='left', padx=(0, 4))
        self._stk_draw_color = '#f0c060'   # 預設黃色
        self._stk_color_btns = {}
        _COLORS = [
            ('#f0c060', '黃'),
            ('#ef5350', '紅'),
            ('#26a69a', '綠'),
            ('#4a9cf0', '藍'),
            ('#e060e0', '紫'),
            ('#ffffff', '白'),
        ]
        for _clr, _name in _COLORS:
            _cb = tk.Button(
                draw_bar, text='  ', width=2,
                bg=_clr, activebackground=_clr,
                relief='flat', bd=0, cursor='hand2',
                highlightbackground='#555577', highlightthickness=1,
                command=lambda c=_clr: self._set_draw_color(c))
            _cb.pack(side='left', padx=2)
            self._stk_color_btns[_clr] = _cb
        self._set_draw_color('#f0c060')   # 初始選中黃色

        # ── 共用捲動區域（K線 + 財報連續） ──────────────────────────────
        scroll_outer = tk.Frame(f, bg=CHART_BG)
        scroll_outer.pack(fill='both', expand=True, padx=8, pady=(0, 8))

        vbar = ttk.Scrollbar(scroll_outer, orient='vertical')
        vbar.pack(side='right', fill='y')

        self._stk_sc = tk.Canvas(scroll_outer, bg=CHART_BG,
                                 highlightthickness=0,
                                 yscrollcommand=vbar.set)
        self._stk_sc.pack(side='left', fill='both', expand=True)
        vbar.config(command=self._stk_sc.yview)

        scroll_inner = tk.Frame(self._stk_sc, bg=CHART_BG)
        _win_id = self._stk_sc.create_window((0, 0), window=scroll_inner, anchor='nw')

        def _inner_cfg(_e):
            self._stk_sc.configure(scrollregion=self._stk_sc.bbox('all'))
        scroll_inner.bind('<Configure>', _inner_cfg)

        def _sc_cfg(e):
            self._stk_sc.itemconfig(_win_id, width=e.width)
        self._stk_sc.bind('<Configure>', _sc_cfg)

        def _wheel(e):
            self._stk_sc.yview_scroll(int(-1 * (e.delta / 120)), 'units')
        for _w in [self._stk_sc, scroll_inner]:
            _w.bind('<MouseWheel>', _wheel)

        # ── K線 matplotlib 圖（上方） ────────────────────────────────────
        self._stk_fig = plt.Figure(figsize=(14, 9), dpi=100, facecolor=CHART_BG)
        self._stk_canvas = FigureCanvasTkAgg(self._stk_fig, master=scroll_inner)
        wk = self._stk_canvas.get_tk_widget()
        wk.configure(bg=CHART_BG)
        wk.pack(fill='x', pady=(0, 2))
        wk.bind('<MouseWheel>', _wheel)
        self._stk_canvas.mpl_connect('motion_notify_event', self._on_stk_hover)

        # ── 財報狀態列 + 財報 matplotlib 圖（下方，連續） ──────────────
        fin_status_bar = tk.Frame(scroll_inner, bg='#1c1c28', pady=3)
        fin_status_bar.pack(fill='x')
        self._fin_status = tk.StringVar(value='財報走勢（查詢後顯示）')
        ttk.Label(fin_status_bar, textvariable=self._fin_status,
                  foreground='#6a8faf',
                  font=('Microsoft JhengHei', 8)).pack(side='left', padx=10)

        self._fin_fig = plt.Figure(figsize=(14, 10), dpi=100, facecolor=CHART_BG)
        self._fin_canvas = FigureCanvasTkAgg(self._fin_fig, master=scroll_inner)
        fin_wk = self._fin_canvas.get_tk_widget()
        fin_wk.configure(bg=CHART_BG)
        fin_wk.pack(fill='x')
        fin_wk.bind('<MouseWheel>', _wheel)
        self._fin_canvas.draw()
        self._fin_canvas.mpl_connect('motion_notify_event', self._on_fin_hover)
        self._fin_canvas.mpl_connect('figure_leave_event',  lambda e: self._fin_tooltip_hide())

        # 財報 hover 狀態
        self._fin_tooltip      = None
        self._fin_tt_lbl       = None
        self._fin_ax_rev       = None
        self._fin_ax_rev_twin  = None   # twinx YoY 軸（也會收到 mouse event）
        self._fin_ax_eps       = None
        self._fin_ax_rat       = None
        self._fin_rev_data: list = []   # [(label, rev_億, yoy_or_None), ...]
        self._fin_eps_data: list = []   # [(label, eps), ...]
        self._fin_rat_data: list = []   # [(label, gross, op, net), ...]

        # ── 狀態變數 ────────────────────────────────────────────────────────
        self._stk_ax           = None
        self._stk_ohlcv        = None
        self._stk_n            = 0
        self._stk_header       = None
        self._stk_vline        = None
        self._stk_hline        = None
        self._stk_current_code = None
        self._stk_current_name = ''


    # ── 個股分析（Tab 6）─────────────────────────────────────────────────────

    def _open_stk_analysis(self, code: str, name: str = ''):
        """從樹狀圖右鍵呼叫：切換到個股分析並載入。"""
        self._stk_code.set(code)
        self._show_page(5)
        self._draw_stk_kline(code, name or code)
        self._fin_status.set(f'載入 {code} 財報資料中…')
        import threading
        threading.Thread(target=self._load_fin_data, args=(code,), daemon=True).start()

    def _load_stk_chart(self):
        """查詢按鈕 / Enter：繪製 K 線並載入財報。"""
        code = self._stk_code.get().strip()
        if not code:
            return
        self._stk_status.set(f'載入 {code} 中…')
        self._fin_status.set(f'載入 {code} 財報資料中…')
        self._draw_stk_kline(code, code)
        import threading
        threading.Thread(target=self._load_fin_data, args=(code,), daemon=True).start()

    def _redraw_stk_kline(self):
        """指標 / 期間 toggle 後重繪 K 線。"""
        if self._stk_current_code:
            self._draw_stk_kline(self._stk_current_code, self._stk_current_name)

    def _open_wantgoo_kline(self):
        """在預設瀏覽器開啟 WantGoo 個股K線頁。"""
        import webbrowser
        code = self._stk_current_code
        if code:
            webbrowser.open(f'https://www.wantgoo.com/stock/{code}')

    def _draw_stk_kline(self, code: str, name: str):
        """繪製個股 K 線圖（邏輯與 ETF K 線完全相同）。"""
        import numpy as np

        CHART_BG = '#131722'   # WantGoo 深藍底
        PANEL_BG = '#1e2133'

        self._stk_current_code = code
        self._stk_current_name = name

        fig = self._stk_fig
        fig.clear()
        fig.patch.set_facecolor(CHART_BG)
        self._stk_ax = None
        self._stk_ohlcv = None

        # ── 取得 OHLCV ────────────────────────────────────────────────────
        _period_map = {'1M': '1mo', '3M': '3mo', '6M': '6mo', '1Y': '1y', '全部': '5y'}
        _yf_period = _period_map.get(getattr(self, '_stk_period', '3M'), '3mo')
        ohlcv = None
        for s in ['.TW', '.TWO', '']:
            try:
                t = yf.Ticker(code + s)
                h = t.history(period=_yf_period, auto_adjust=False)
                if not h.empty:
                    ohlcv = h
                    # 若名稱仍是代號，嘗試從 yfinance 取得中文名
                    if name == code:
                        try:
                            raw = t.info.get('longName') or t.info.get('shortName') or code
                            for sfx in ['股份有限公司', ' Inc.', ' Corp.', ' Co.,Ltd.', ', Ltd.']:
                                raw = raw.replace(sfx, '')
                            name = raw.strip()
                            self._stk_current_name = name
                        except Exception:
                            pass
                    break
            except Exception:
                pass

        if ohlcv is None or ohlcv.empty:
            ax = fig.add_subplot(111, facecolor=PANEL_BG)
            ax.text(0.5, 0.5, f'{code} 無法取得 K 線資料',
                    ha='center', va='center', color='#888',
                    fontsize=11, fontfamily=CHART_FONT, transform=ax.transAxes)
            ax.axis('off')
            self._stk_canvas.draw()
            self._stk_status.set(f'{code}  無法取得資料')
            return

        # ── 指標計算（與 ETF K 線相同）────────────────────────────────────
        ind = dict(self._stk_ind_state)
        close = ohlcv['Close']
        ohlcv = ohlcv.copy()

        if ind.get('MA'):
            ohlcv['MA5']   = close.rolling(5,   min_periods=1).mean()
            ohlcv['MA10']  = close.rolling(10,  min_periods=1).mean()
            ohlcv['MA20']  = close.rolling(20,  min_periods=1).mean()
            ohlcv['MA60']  = close.rolling(60,  min_periods=1).mean()
            ohlcv['MA120'] = close.rolling(120, min_periods=1).mean()
            ohlcv['MA240'] = close.rolling(240, min_periods=1).mean()
        if ind.get('BB'):
            _bbm = close.rolling(20, min_periods=5).mean()
            _bbs = close.rolling(20, min_periods=5).std(ddof=0)
            ohlcv['BB_U'] = _bbm + 2 * _bbs
            ohlcv['BB_M'] = _bbm
            ohlcv['BB_L'] = _bbm - 2 * _bbs
        if ind.get('MACD'):
            _e12 = close.ewm(span=12, adjust=False).mean()
            _e26 = close.ewm(span=26, adjust=False).mean()
            ohlcv['MACD_L'] = _e12 - _e26
            ohlcv['MACD_S'] = ohlcv['MACD_L'].ewm(span=9, adjust=False).mean()
            ohlcv['MACD_H'] = ohlcv['MACD_L'] - ohlcv['MACD_S']
        if ind.get('RSI'):
            _d = close.diff()
            _g = _d.clip(lower=0)
            _l = (-_d).clip(lower=0)
            _rs = _g.ewm(com=13, adjust=False).mean() / \
                  _l.ewm(com=13, adjust=False).mean().replace(0, float('nan'))
            ohlcv['RSI'] = 100 - 100 / (1 + _rs)
        if ind.get('KD'):
            _lo = ohlcv.get('Low',  ohlcv['Close']).rolling(9, min_periods=1).min()
            _hi = ohlcv.get('High', ohlcv['Close']).rolling(9, min_periods=1).max()
            _rk = 100 * (close - _lo) / (_hi - _lo + 1e-10)
            ohlcv['KD_K'] = _rk.ewm(com=2, adjust=False).mean()
            ohlcv['KD_D'] = ohlcv['KD_K'].ewm(com=2, adjust=False).mean()

        n  = len(ohlcv)
        xs = np.arange(n)

        # ── 動態佈局（順序：VOL → KD → RSI → MACD，與 WantGoo 一致）────────
        sub_list = []
        if ind.get('VOL'):  sub_list.append('VOL')
        if ind.get('KD'):   sub_list.append('KD')
        if ind.get('RSI'):  sub_list.append('RSI')
        if ind.get('MACD'): sub_list.append('MACD')

        L, W  = 0.07, 0.88
        TOP   = 0.94
        BOT   = 0.07
        GAP   = 0.012
        n_sub = len(sub_list)
        avail = TOP - BOT

        if n_sub == 0:
            sub_h = 0.0
            p_bot = BOT
            p_h   = avail
        else:
            sub_h = min(0.19, (avail * 0.44) / n_sub)
            p_bot = BOT + n_sub * (sub_h + GAP)
            p_h   = TOP - p_bot

        ax_price = fig.add_axes([L, p_bot, W, max(p_h, 0.25)])
        sub_axes: dict = {}
        _y = BOT
        for _name in reversed(sub_list):
            sub_axes[_name] = fig.add_axes([L, _y, W, sub_h - GAP * 0.5])
            _y += sub_h + GAP

        def _style(ax, xticks=False):
            ax.set_facecolor(PANEL_BG)
            ax.tick_params(colors='#888', labelsize=7, length=2)
            ax.yaxis.tick_right()
            ax.set_xlim(-0.5, n - 0.5)
            for sp in ax.spines.values():
                sp.set_color('#333')
            ax.grid(axis='y', color='#2a2a3a', linewidth=0.4, linestyle='-')
            # 每個子圖 Y 軸最多 4 個 tick，並裁掉邊緣避免與相鄰圖重疊
            ax.yaxis.set_major_locator(
                matplotlib.ticker.MaxNLocator(nbins=4, prune='both', integer=False))
            # 自動縮短大數字（千→K，百萬→M）
            ax.yaxis.set_major_formatter(
                matplotlib.ticker.FuncFormatter(
                    lambda v, _: (f'{v/1e6:.1f}M' if abs(v) >= 1e6
                                  else f'{v/1e3:.0f}K' if abs(v) >= 1e3
                                  else f'{v:.1f}')))
            if not xticks:
                ax.set_xticks([])

        # ── 蠟燭圖（寬度依資料量動態調整，接近 WantGoo 比例）─────────────────
        cw = max(0.3, min(0.75, 0.72 - n * 0.001))
        UP_CLR, DN_CLR = '#ef5350', '#26a69a'   # WantGoo 紅漲綠跌
        for i, (_, row) in enumerate(ohlcv.iterrows()):
            o_  = float(row.get('Open',  row['Close']))
            h_  = float(row.get('High',  row['Close']))
            l_  = float(row.get('Low',   row['Close']))
            c_  = float(row['Close'])
            is_up = c_ >= o_
            clr = UP_CLR if is_up else DN_CLR
            # 影線
            ax_price.plot([i, i], [l_, h_], color=clr, linewidth=0.8, zorder=2)
            body_h = max(abs(c_ - o_), 0.005 * c_)
            # 上漲：實心紅；下跌：實心綠
            ax_price.add_patch(plt.Rectangle(
                (i - cw/2, min(o_, c_)), cw, body_h,
                facecolor=clr, edgecolor='none', linewidth=0, zorder=3))

        # ── MA 線（MA5/10/20/60/120/240，對齊 WantGoo 配色）────────────────
        _leg_handles = []
        if ind.get('MA'):
            for col, clr, lbl, lw in [
                    ('MA5',   '#f0e44a', 'MA5',   1.2),
                    ('MA10',  '#4a9cf0', 'MA10',  1.1),
                    ('MA20',  '#f0a04a', 'MA20',  1.1),
                    ('MA60',  '#e060e0', 'MA60',  1.0),
                    ('MA120', '#60c060', 'MA120', 1.0),
                    ('MA240', '#f08040', 'MA240', 1.0)]:
                if col in ohlcv:
                    ln, = ax_price.plot(xs, ohlcv[col].values, color=clr,
                                        linewidth=lw, label=lbl, zorder=4)
                    _leg_handles.append(ln)

        # ── Bollinger Bands ───────────────────────────────────────────────
        if ind.get('BB') and 'BB_U' in ohlcv:
            ax_price.plot(xs, ohlcv['BB_U'].values, color='#88aaff',
                          linewidth=0.8, linestyle='--', label='BB上', zorder=4)
            ax_price.plot(xs, ohlcv['BB_M'].values, color='#ffffff',
                          linewidth=0.6, linestyle='--', label='BB中', zorder=4)
            ax_price.plot(xs, ohlcv['BB_L'].values, color='#88aaff',
                          linewidth=0.8, linestyle='--', label='BB下', zorder=4)
            ax_price.fill_between(xs, ohlcv['BB_U'].values, ohlcv['BB_L'].values,
                                  color='#4a9cf0', alpha=0.06, zorder=1)

        # ── Y 軸縮放 ─────────────────────────────────────────────────────
        _pvals = [ohlcv['High'].dropna().values, ohlcv['Low'].dropna().values]
        if 'BB_U' in ohlcv: _pvals.append(ohlcv['BB_U'].dropna().values)
        if 'BB_L' in ohlcv: _pvals.append(ohlcv['BB_L'].dropna().values)
        _ap = np.concatenate(_pvals)
        _pmin, _pmax = float(np.nanmin(_ap)), float(np.nanmax(_ap))
        _mg = (_pmax - _pmin) * 0.05 or 0.5
        ax_price.set_ylim(_pmin - _mg, _pmax + _mg * 2.0)

        _style(ax_price)
        ax_price.tick_params(labelsize=7.5)
        ax_price.yaxis.set_major_locator(
            matplotlib.ticker.MaxNLocator(nbins=6, prune='upper', integer=False))

        if _leg_handles:
            ax_price.legend(handles=_leg_handles, loc='upper left',
                            fontsize=7.5, facecolor='#1e2133',
                            edgecolor='#2a2a4a', labelcolor='white',
                            handlelength=1.2, framealpha=0.80, borderpad=0.4,
                            ncol=3)

        # ── 副圖 ──────────────────────────────────────────────────────────
        if 'VOL' in sub_axes:
            ax_v = sub_axes['VOL']
            vol = ohlcv.get('Volume', None)
            if vol is not None:
                v_clrs = [UP_CLR if ohlcv['Close'].iloc[i] >= ohlcv['Open'].iloc[i]
                          else DN_CLR for i in range(n)]
                ax_v.bar(xs, vol.values, color=v_clrs, width=0.7, zorder=2)
            ax_v.set_ylabel('VOL', color='#888', fontsize=7, labelpad=2)
            ax_v.yaxis.set_label_position('left')
            _style(ax_v)

        if 'MACD' in sub_axes and 'MACD_L' in ohlcv:
            ax_m = sub_axes['MACD']
            h_clrs = ['#f07070' if v >= 0 else '#4ec94e' for v in ohlcv['MACD_H'].values]
            ax_m.bar(xs, ohlcv['MACD_H'].values, color=h_clrs, width=0.6, zorder=2)
            ax_m.plot(xs, ohlcv['MACD_L'].values, color='#4a9cf0', linewidth=0.9, zorder=3)
            ax_m.plot(xs, ohlcv['MACD_S'].values, color='#f0a04a', linewidth=0.9, zorder=3)
            ax_m.axhline(0, color='#444', linewidth=0.5, linestyle='--')
            ax_m.set_ylabel('MACD', color='#888', fontsize=7, labelpad=2)
            ax_m.yaxis.set_label_position('left')
            _style(ax_m)

        if 'RSI' in sub_axes and 'RSI' in ohlcv:
            ax_r = sub_axes['RSI']
            ax_r.plot(xs, ohlcv['RSI'].values, color='#c46af0', linewidth=0.9)
            ax_r.axhline(70, color='#f07070', linewidth=0.5, linestyle='--')
            ax_r.axhline(30, color='#4ec94e', linewidth=0.5, linestyle='--')
            ax_r.set_ylim(0, 100)
            ax_r.set_ylabel('RSI', color='#888', fontsize=7, labelpad=2)
            ax_r.yaxis.set_label_position('left')
            _style(ax_r)

        if 'KD' in sub_axes and 'KD_K' in ohlcv:
            ax_k = sub_axes['KD']
            ax_k.plot(xs, ohlcv['KD_K'].values, color='#f0e44a', linewidth=0.9, label='K')
            ax_k.plot(xs, ohlcv['KD_D'].values, color='#4a9cf0', linewidth=0.9, label='D')
            ax_k.axhline(80, color='#f07070', linewidth=0.5, linestyle='--')
            ax_k.axhline(20, color='#4ec94e', linewidth=0.5, linestyle='--')
            ax_k.set_ylim(0, 100)
            ax_k.set_ylabel('KD', color='#888', fontsize=7, labelpad=2)
            ax_k.yaxis.set_label_position('left')
            _style(ax_k)

        # ── X 軸日期 ─────────────────────────────────────────────────────
        _bottom_ax = sub_axes[sub_list[-1]] if sub_list else ax_price
        step  = max(1, n // 8)
        xtks  = list(range(0, n, step))
        xlbls = [ohlcv.index[i].strftime('%Y-%m-%d') for i in xtks]
        _bottom_ax.set_xticks(xtks)
        _bottom_ax.set_xticklabels(xlbls, rotation=25, ha='right', fontsize=7, color='#888')

        # ── 游標線（主圖 + 所有副圖共用同一 X）─────────────────────────────
        self._stk_vline = ax_price.axvline(
            -999, color='#888', linewidth=0.8, linestyle='--', zorder=5)
        self._stk_hline = ax_price.axhline(
            -999, color='#aaa', linewidth=0.7, linestyle=':', zorder=5, alpha=0.8)
        self._stk_sub_vlines = [
            ax.axvline(-999, color='#666', linewidth=0.8, linestyle='--', zorder=5)
            for ax in sub_axes.values()]

        # ── 右側 Y 軸股價標籤 ────────────────────────────────────────────
        self._stk_price_ann = ax_price.annotate(
            '', xy=(1.0, 0), xycoords=('axes fraction', 'data'),
            xytext=(4, 0), textcoords='offset points',
            ha='left', va='center', fontsize=8, color='white',
            zorder=10, clip_on=False,
            bbox=dict(boxstyle='square,pad=0.25', facecolor=UP_CLR,
                      edgecolor='none', alpha=0.92))

        # 儲存所有 axes 供 hover 使用
        self._stk_all_axes = [ax_price] + list(sub_axes.values())
        self._stk_ax    = ax_price
        self._stk_ohlcv = ohlcv
        self._stk_n     = n

        # 繪製後連結點擊事件（繪線功能）
        self._stk_canvas.mpl_connect('button_press_event', self._on_stk_click)

        last_close = float(ohlcv['Close'].iloc[-1])
        prev_close = float(ohlcv['Close'].iloc[-2]) if len(ohlcv) >= 2 else last_close
        chg = last_close - prev_close
        is_up = last_close >= prev_close
        clr_txt = '#ef5350' if is_up else '#26a69a'
        arrow = '▲' if is_up else '▼'
        self._stk_canvas.draw()
        self._stk_status.set(f'{code}  {name}  ─  最新：{last_close:.2f}  ─  資料載入完成')
        # 更新資訊列初始值（最後一筆）
        last_row = ohlcv.iloc[-1]
        self._stk_info_date.config(text=ohlcv.index[-1].strftime('%Y-%m-%d'))
        self._stk_info_close.config(text=f'{last_close:.2f}', fg=clr_txt)
        self._stk_info_chg.config(
            text=f'{arrow} {abs(chg):.2f} ({abs(chg/prev_close*100):.2f}%)', fg=clr_txt)
        self._stk_info_o.config(text=f'{float(last_row.get("Open", last_close)):.2f}')
        self._stk_info_h.config(text=f'{float(last_row.get("High", last_close)):.2f}')
        self._stk_info_l.config(text=f'{float(last_row.get("Low",  last_close)):.2f}')
        self._stk_info_v.config(text=f'{int(last_row.get("Volume", 0)):,}')

    def _on_stk_hover(self, event):
        """個股 K 線 hover：更新資訊列、十字游標（延伸至副圖）、右側價格標籤。"""
        if self._stk_ax is None or self._stk_ohlcv is None:
            return
        all_ax = getattr(self, '_stk_all_axes', [self._stk_ax])
        if event.inaxes not in all_ax:
            return
        if event.xdata is None:
            return
        xi = int(round(event.xdata))
        if not (0 <= xi < self._stk_n):
            return

        row      = self._stk_ohlcv.iloc[xi]
        date_str = self._stk_ohlcv.index[xi].strftime('%Y-%m-%d')
        o_  = float(row.get('Open',   row['Close']))
        h_  = float(row.get('High',   row['Close']))
        l_  = float(row.get('Low',    row['Close']))
        c   = float(row['Close'])
        vol = int(row.get('Volume', 0))
        prev_c = float(self._stk_ohlcv['Close'].iloc[xi - 1]) if xi > 0 else c
        chg  = c - prev_c
        pct  = chg / prev_c * 100 if prev_c != 0 else 0
        is_up = c >= prev_c
        clr  = '#ef5350' if is_up else '#26a69a'
        arrow = '▲' if is_up else '▼'

        # 更新 tkinter 資訊列
        self._stk_info_date.config(text=date_str)
        self._stk_info_close.config(text=f'{c:.2f}', fg=clr)
        self._stk_info_chg.config(
            text=f'{arrow} {abs(chg):.2f} ({abs(pct):.2f}%)', fg=clr)
        self._stk_info_o.config(text=f'{o_:.2f}')
        self._stk_info_h.config(text=f'{h_:.2f}')
        self._stk_info_l.config(text=f'{l_:.2f}')
        self._stk_info_v.config(text=f'{vol:,}')

        # 主圖游標線
        if self._stk_vline is not None:
            self._stk_vline.set_xdata([xi, xi])
            self._stk_vline.set_visible(True)
        if self._stk_hline is not None:
            self._stk_hline.set_ydata([c, c])
            self._stk_hline.set_visible(True)

        # 副圖垂直線延伸
        for vl in getattr(self, '_stk_sub_vlines', []):
            vl.set_xdata([xi, xi])
            vl.set_visible(True)

        # 右側 Y 軸股價標籤
        ann = getattr(self, '_stk_price_ann', None)
        if ann is not None:
            ann.xy = (1.0, c)
            ann.set_text(f' {c:.2f} ')
            ann.get_bbox_patch().set_facecolor(clr)
            ann.set_visible(True)

        self._stk_canvas.draw_idle()

    # ── 繪線工具 ──────────────────────────────────────────────────────────

    def _set_draw_mode(self, mode: str):
        if mode == 'erase':
            for obj in self._stk_drawings:
                try:
                    obj.remove()
                except Exception:
                    pass
            self._stk_drawings.clear()
            self._stk_draw_mode = 'none'
            if self._stk_ax:
                self._stk_canvas.draw_idle()
        elif self._stk_draw_mode == mode:
            self._stk_draw_mode = 'none'
        else:
            self._stk_draw_mode = mode
            self._stk_draw_start = None
        for m, b in self._stk_draw_btns.items():
            is_on = (m == self._stk_draw_mode)
            b.config(**(self._stk_db_on if is_on else self._stk_db_off))

    def _set_draw_color(self, color: str):
        self._stk_draw_color = color
        for c, b in self._stk_color_btns.items():
            b.config(highlightbackground='#ffffff' if c == color else '#555577',
                     highlightthickness=2 if c == color else 1)

    def _on_stk_click(self, event):
        """處理 K 線圖上的繪線點擊。"""
        if self._stk_draw_mode == 'none':
            return
        if event.inaxes is not self._stk_ax:
            return
        if event.xdata is None or event.ydata is None:
            return
        xi, yi = event.xdata, event.ydata
        mode = self._stk_draw_mode
        ax = self._stk_ax
        DRAW_CLR = self._stk_draw_color

        if mode == 'hline':
            ln = ax.axhline(yi, color=DRAW_CLR, linewidth=1.1,
                            linestyle='--', zorder=6, alpha=0.85)
            self._stk_drawings.append(ln)
            self._stk_canvas.draw_idle()

        elif mode == 'vline':
            ln = ax.axvline(xi, color=DRAW_CLR, linewidth=1.1,
                            linestyle='--', zorder=6, alpha=0.85)
            self._stk_drawings.append(ln)
            self._stk_canvas.draw_idle()

        elif mode == 'line':
            if self._stk_draw_start is None:
                self._stk_draw_start = (xi, yi)
                # 暫時標示起點
                mk, = ax.plot(xi, yi, 'o', color=DRAW_CLR, markersize=4, zorder=7)
                self._stk_drawings.append(mk)
                self._stk_canvas.draw_idle()
            else:
                x0, y0 = self._stk_draw_start
                # 移除起點標記
                try:
                    self._stk_drawings[-1].remove()
                    self._stk_drawings.pop()
                except Exception:
                    pass
                ln, = ax.plot([x0, xi], [y0, yi], color=DRAW_CLR,
                              linewidth=1.3, zorder=6, alpha=0.9)
                self._stk_drawings.append(ln)
                self._stk_draw_start = None
                self._stk_canvas.draw_idle()

    # ── 舊籌碼資料載入（保留但不再從 UI 呼叫）──────────────────────────────────

    # ── 個股財報載入與繪圖 ─────────────────────────────────────────────────

    def _load_fin_data(self, code: str):
        """背景執行緒：從 yfinance 取得季財報（營收、EPS、三率）。"""
        import pandas as pd

        fin_df = pd.DataFrame()
        eps_df = pd.DataFrame()
        rev_df = pd.DataFrame()
        name   = code
        ticker_obj = None

        for sfx in ['.TW', '.TWO', '']:
            try:
                t = yf.Ticker(code + sfx)
                info = t.fast_info
                price = getattr(info, 'last_price', None)
                if price and price > 0:
                    ticker_obj = t
                    raw = (getattr(info, 'display_name', None)
                           or code)
                    for s in ['股份有限公司', ' Inc.', ' Corp.', ' Co.,Ltd.', ', Ltd.']:
                        raw = raw.replace(s, '')
                    name = raw.strip() or code
                    break
            except Exception:
                continue

        if ticker_obj is None:
            for sfx in ['.TW', '.TWO']:
                try:
                    t = yf.Ticker(code + sfx)
                    h = t.history(period='5d')
                    if not h.empty:
                        ticker_obj = t
                        break
                except Exception:
                    continue

        if ticker_obj is not None:
            try:
                qfin = ticker_obj.quarterly_financials
                if qfin is not None and not qfin.empty:
                    def _row(df, *labels):
                        for lb in labels:
                            if lb in df.index:
                                return df.loc[lb]
                        return None

                    rev = _row(qfin, 'Total Revenue', 'Revenue')
                    gp  = _row(qfin, 'Gross Profit', 'GrossProfit')
                    op  = _row(qfin, 'Operating Income', 'EBIT')
                    ni  = _row(qfin, 'Net Income', 'NetIncome')

                    if rev is not None and rev.abs().max() > 0:
                        rev_df = rev.to_frame('Revenue')
                        rev_df.index = pd.to_datetime(rev_df.index).tz_localize(None)
                        rev_df = rev_df.sort_index().tail(12)
                        cols = {}
                        if gp is not None:
                            cols['gross_m'] = (gp / rev * 100).round(2)
                        if op is not None:
                            cols['op_m'] = (op / rev * 100).round(2)
                        if ni is not None:
                            cols['net_m'] = (ni / rev * 100).round(2)
                        if cols:
                            fin_df = pd.DataFrame(cols)
                            fin_df.index = pd.to_datetime(fin_df.index).tz_localize(None)
                            fin_df = fin_df.sort_index().tail(12)
            except Exception:
                pass

            try:
                qfin2 = ticker_obj.quarterly_financials
                eps_ser = None
                if qfin2 is not None:
                    for lbl in ['Diluted EPS', 'Basic EPS']:
                        if lbl in qfin2.index:
                            eps_ser = qfin2.loc[lbl]
                            break
                if eps_ser is not None:
                    eps_df = eps_ser.to_frame('EPS')
                    eps_df.index = pd.to_datetime(eps_df.index).tz_localize(None)
                    eps_df = eps_df.sort_index().tail(12)
            except Exception:
                pass

        self._stk_current_name = name
        self._ui_call(lambda: self._render_fin_charts(code, name, fin_df, eps_df, rev_df))

    def _render_fin_charts(self, code: str, name: str,
                           fin_df, eps_df, rev_df):
        """在 Tab 6 下半部繪製財報走勢（季營收、EPS、三率）。"""
        import numpy as _np
        import matplotlib.ticker as _mtk

        fig = self._fin_fig
        fig.clf()
        fig.patch.set_facecolor('#111111')

        CHART_BG = '#111111'
        PANEL_BG = '#1a1a2e'
        C_TEXT   = '#aaaacc'
        C_BORDER = '#333355'
        FNT      = CHART_FONT

        def _fmt_q(d):
            q = (d.month - 1) // 3 + 1
            return f"{d.year}Q{q}"

        def _style(ax):
            ax.set_facecolor(PANEL_BG)
            ax.tick_params(colors='#888899', labelsize=6.5)
            for sp in ax.spines.values():
                sp.set_edgecolor(C_BORDER)
            ax.grid(axis='y', color=C_BORDER, linewidth=0.4, alpha=0.5)

        has_rev = not rev_df.empty
        has_eps = not eps_df.empty
        has_fin = not fin_df.empty

        gs = fig.add_gridspec(
            2, 2,
            height_ratios=[1, 1.1],
            hspace=0.55, wspace=0.32,
            left=0.07, right=0.97, top=0.93, bottom=0.10)
        ax_rev = fig.add_subplot(gs[0, 0])
        ax_eps = fig.add_subplot(gs[0, 1])
        ax_rat = fig.add_subplot(gs[1, :])

        for ax in [ax_rev, ax_eps, ax_rat]:
            _style(ax)

        fig.suptitle(f'{code}  {name}  財報走勢',
                     color='#aecde8', fontsize=10, fontfamily=FNT, y=0.97)

        # ── 季營收 長條 + QoQ 折線 + YoY 折線 ─────────────────────────────────
        ax_rev.set_title('季營收（億）', color=C_TEXT, fontsize=9, fontfamily=FNT, pad=4)
        if has_rev:
            rv  = rev_df['Revenue'].values
            xr  = list(range(len(rv)))
            lbl = [_fmt_q(d) for d in rev_df.index]
            ax_rev.bar(xr, [v / 1e8 for v in rv], color='#5b9cf6', alpha=0.85, zorder=2)
            ax_rev.set_xticks(xr)
            ax_rev.set_xticklabels(lbl, rotation=40, ha='right',
                                   fontsize=6, color='#888899', fontfamily=FNT)
            ax_rev.set_ylabel('億', color='#888899', fontsize=7, labelpad=2)

            # QoQ（季增率，只需前一期，2 筆資料就能畫線）
            qoq = [None] + [(((rv[i] / rv[i - 1] - 1) * 100) if rv[i - 1] != 0 else None)
                            for i in range(1, len(rv))]
            qoq_xs = [i for i, v in enumerate(qoq) if v is not None]
            qoq_y  = [v for v in qoq if v is not None]

            # YoY（年增率，需 4 期前資料）
            yoy = [((rv[i] / rv[i - 4] - 1) * 100 if rv[i - 4] != 0 else None)
                   if i >= 4 else None for i in range(len(rv))]
            yoy_xs = [i for i, v in enumerate(yoy) if v is not None]
            yoy_y  = [v for v in yoy if v is not None]

            if qoq_xs or yoy_xs:
                ax2 = ax_rev.twinx()
                self._fin_ax_rev_twin = ax2   # 記錄 twinx 軸，hover 時才能識別
                ax2.set_facecolor(PANEL_BG)
                ax2.axhline(0, color='#333355', linewidth=0.6, linestyle='--', zorder=1)
                if qoq_xs:
                    ax2.plot(qoq_xs, qoq_y, color='#80cfff',
                             linewidth=1.4, linestyle='--', marker='s',
                             markersize=3, zorder=2, label='QoQ%')
                if yoy_xs:
                    ax2.plot(yoy_xs, yoy_y, color='#f0c060',
                             linewidth=1.6, marker='o', markersize=4, zorder=3, label='YoY%')
                ax2.tick_params(axis='y', colors='#aaaaaa', labelsize=6)
                ax2.set_ylabel('增率 %', color='#aaaaaa', fontsize=7, labelpad=2)
                ax2.legend(loc='upper left', fontsize=6,
                           facecolor='#1e2133', edgecolor='none',
                           labelcolor='white', markerscale=0.9)
                for sp in ax2.spines.values():
                    sp.set_edgecolor(C_BORDER)
        else:
            ax_rev.text(0.5, 0.5, '查無資料', ha='center', va='center',
                        color='#445566', fontsize=10, transform=ax_rev.transAxes)

        # ── 季 EPS 長條 + 趨勢線 ─────────────────────────────────────────────
        ax_eps.set_title('季EPS（元）', color=C_TEXT, fontsize=9, fontfamily=FNT, pad=4)
        if has_eps:
            ev   = eps_df['EPS'].values
            xe   = list(range(len(ev)))
            lble = [_fmt_q(d) for d in eps_df.index]
            clrs = ['#f07070' if v >= 0 else '#4ec94e' for v in ev]
            ax_eps.bar(xe, ev, color=clrs, alpha=0.85, zorder=2)
            ax_eps.axhline(0, color='#333355', linewidth=0.6, linestyle='--')
            if len(ev) >= 3:
                z = _np.polyfit(xe, ev, 1)
                ax_eps.plot(xe, _np.polyval(z, xe),
                            color='#f0c060', linewidth=1.5,
                            linestyle='--', alpha=0.8, zorder=3)
            ax_eps.set_xticks(xe)
            ax_eps.set_xticklabels(lble, rotation=40, ha='right',
                                   fontsize=6, color='#888899', fontfamily=FNT)
            ax_eps.set_ylabel('元', color='#888899', fontsize=7, labelpad=2)
        else:
            ax_eps.text(0.5, 0.5, '查無資料', ha='center', va='center',
                        color='#445566', fontsize=10, transform=ax_eps.transAxes)

        # ── 財報三率 折線 ────────────────────────────────────────────────────
        ax_rat.set_title('財報三率 %', color=C_TEXT, fontsize=9, fontfamily=FNT, pad=4)
        if has_fin:
            xf   = list(range(len(fin_df)))
            lblf = [_fmt_q(d) for d in fin_df.index]
            RATE_CFG = [
                ('gross_m', '毛利率',     '#4ec94e', 'o'),
                ('op_m',    '營業利益率', '#5b9cf6', 's'),
                ('net_m',   '淨利率',     '#f07070', '^'),
            ]
            for col, lbl, clr, mk in RATE_CFG:
                if col in fin_df.columns:
                    ax_rat.plot(xf, fin_df[col].values,
                                color=clr, linewidth=1.8,
                                marker=mk, markersize=4, label=lbl, zorder=3)
            ax_rat.axhline(0, color='#333355', linewidth=0.6, linestyle='--')
            ax_rat.set_xticks(xf)
            ax_rat.set_xticklabels(lblf, rotation=40, ha='right',
                                   fontsize=6.5, color='#888899', fontfamily=FNT)
            ax_rat.set_ylabel('%', color='#888899', fontsize=7, labelpad=2)
            ax_rat.legend(fontsize=8, facecolor='#1a1a2e',
                          edgecolor='#333355', labelcolor='white', loc='best')
        else:
            ax_rat.text(0.5, 0.5, '查無資料', ha='center', va='center',
                        color='#445566', fontsize=10, transform=ax_rat.transAxes)

        # ── 儲存 hover 用資料 ────────────────────────────────────────────────
        self._fin_ax_rev      = ax_rev
        self._fin_ax_rev_twin = None    # 若有 twinx 才設定，預先清空
        self._fin_ax_eps      = ax_eps
        self._fin_ax_rat      = ax_rat

        # 季營收資料  (label, rev_億, yoy_or_None, qoq_or_None)
        self._fin_rev_data = []
        if has_rev:
            rv  = rev_df['Revenue'].values
            lbl = [_fmt_q(d) for d in rev_df.index]
            yoy_full = [None] * len(rv)
            qoq_full = [None] * len(rv)
            for i in range(1, len(rv)):
                if rv[i - 1] != 0:
                    qoq_full[i] = (rv[i] / rv[i - 1] - 1) * 100
            if len(rv) >= 5:
                for i in range(4, len(rv)):
                    if rv[i - 4] != 0:
                        yoy_full[i] = (rv[i] / rv[i - 4] - 1) * 100
            self._fin_rev_data = [(lbl[i], float(rv[i]) / 1e8, yoy_full[i], qoq_full[i])
                                  for i in range(len(rv))]

        # 季 EPS 資料
        self._fin_eps_data = []
        if has_eps:
            ev   = eps_df['EPS'].values
            lble = [_fmt_q(d) for d in eps_df.index]
            self._fin_eps_data = [(lble[i], float(ev[i])) for i in range(len(ev))]

        # 財報三率資料
        self._fin_rat_data = []
        if has_fin:
            lblf = [_fmt_q(d) for d in fin_df.index]
            for i in range(len(fin_df)):
                self._fin_rat_data.append((
                    lblf[i],
                    float(fin_df['gross_m'].iloc[i]) if 'gross_m' in fin_df.columns else None,
                    float(fin_df['op_m'].iloc[i])    if 'op_m'    in fin_df.columns else None,
                    float(fin_df['net_m'].iloc[i])   if 'net_m'   in fin_df.columns else None,
                ))

        self._fin_canvas.draw()
        done_txt = f'{code}  {name}  ─  財報資料載入完成'
        self._fin_status.set(done_txt)

    # ── 財報圖 hover tooltip ──────────────────────────────────────────────────

    def _fin_tooltip_hide(self):
        """隱藏財報圖 tooltip。"""
        if self._fin_tooltip is not None:
            self._fin_tooltip.withdraw()

    def _on_fin_hover(self, event):
        """游標移到財報圖時，顯示對應季度資料 tooltip。"""
        ax_rev      = getattr(self, '_fin_ax_rev',      None)
        ax_rev_twin = getattr(self, '_fin_ax_rev_twin', None)
        ax_eps      = getattr(self, '_fin_ax_eps',      None)
        ax_rat      = getattr(self, '_fin_ax_rat',      None)

        # ax_rev_twin 是 twinx YoY 軸，也算在 ax_rev 範圍內
        in_rev = event.inaxes in (ax_rev, ax_rev_twin)

        # 確認游標在三個子圖之一
        if not in_rev and event.inaxes not in (ax_eps, ax_rat):
            self._fin_tooltip_hide()
            return
        if event.xdata is None:
            self._fin_tooltip_hide()
            return

        xi = int(round(event.xdata))

        # ── 建立 / 取得 tooltip 視窗 ─────────────────────────────────────────
        if self._fin_tooltip is None:
            self._fin_tooltip = tk.Toplevel(self)
            self._fin_tooltip.withdraw()
            self._fin_tooltip.overrideredirect(True)
            self._fin_tooltip.configure(bg='#1a1a2e')
            self._fin_tt_lbl = tk.Label(
                self._fin_tooltip, bg='#1a1a2e', fg='#e0e0e0',
                font=('Microsoft JhengHei', 9), justify='left', padx=10, pady=7)
            self._fin_tt_lbl.pack()

        # ── 根據子圖組裝文字 ──────────────────────────────────────────────────
        lines = []

        if in_rev:
            data = getattr(self, '_fin_rev_data', [])
            if not data or xi < 0 or xi >= len(data):
                self._fin_tooltip_hide()
                return
            row = data[xi]
            lbl, rev_e8 = row[0], row[1]
            yoy = row[2] if len(row) > 2 else None
            qoq = row[3] if len(row) > 3 else None
            lines.append(f'📅 {lbl}')
            lines.append(f'季營收：{rev_e8:,.2f} 億')
            if qoq is not None:
                sign = '+' if qoq >= 0 else ''
                arrow = '▲' if qoq >= 0 else '▼'
                lines.append(f'季增率：{arrow} {sign}{qoq:.1f}%')
            else:
                lines.append('季增率：—')
            if yoy is not None:
                sign = '+' if yoy >= 0 else ''
                arrow = '▲' if yoy >= 0 else '▼'
                lines.append(f'年增率：{arrow} {sign}{yoy:.1f}%')
            else:
                lines.append('年增率：—')

        elif event.inaxes is ax_eps:
            data = getattr(self, '_fin_eps_data', [])
            if not data or xi < 0 or xi >= len(data):
                self._fin_tooltip_hide()
                return
            lbl, eps = data[xi]
            sign = '+' if eps >= 0 else ''
            arrow = '▲' if eps > 0 else ('▼' if eps < 0 else '─')
            lines.append(f'📅 {lbl}')
            lines.append(f'季 EPS：{arrow} {sign}{eps:.2f} 元')

        elif event.inaxes is ax_rat:
            data = getattr(self, '_fin_rat_data', [])
            if not data or xi < 0 or xi >= len(data):
                self._fin_tooltip_hide()
                return
            row = data[xi]
            lbl = row[0]
            gross_m, op_m, net_m = row[1], row[2], row[3]
            lines.append(f'📅 {lbl}')
            lines.append(f'毛利率　　：{gross_m:.1f}%' if gross_m is not None else '毛利率　　：—')
            lines.append(f'營業利益率：{op_m:.1f}%'   if op_m    is not None else '營業利益率：—')
            lines.append(f'淨利率　　：{net_m:.1f}%'  if net_m   is not None else '淨利率　　：—')

        if not lines:
            self._fin_tooltip_hide()
            return

        self._fin_tt_lbl.config(text='\n'.join(lines))

        # ── 定位 ─────────────────────────────────────────────────────────────
        try:
            sx = event.guiEvent.x_root
            sy = event.guiEvent.y_root
        except Exception:
            cw = self._fin_canvas.get_tk_widget()
            sx = cw.winfo_rootx() + int(event.x)
            sy = cw.winfo_rooty() + int(cw.winfo_height() - event.y)

        self._place_tooltip(self._fin_tooltip, sx, sy)
        self._fin_tooltip.deiconify()

    # ── 舊籌碼資料載入（保留但不再從 UI 呼叫）──────────────────────────────────

    def _load_chip_data(self, code: str):
        """啟動背景執行緒取得籌碼資料"""
        code = code.strip()
        if not code:
            return
        self._chip_status.set(f'載入 {code} 中…')
        threading.Thread(target=self._load_chip_data_impl,
                         args=(code,), daemon=True).start()

    def _load_chip_data_impl(self, code: str):
        """背景執行緒：股價K線 + 三大法人 + 融資融券 + 財報三率 + EPS + 月營收"""
        from datetime import date as _dt_date

        period = self._chip_period.get()       # 60 / 120 / 180
        today  = _dt_date.today()
        hdr    = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/122.0.0.0'}

        def _roc_to_date(s: str):
            try:
                p = str(s).strip().split('/')
                if len(p) == 3:
                    return _dt_date(int(p[0]) + 1911, int(p[1]), int(p[2]))
            except Exception:
                pass
            return None

        # 需要查詢的月份清單
        needed_months: list[tuple] = []
        ref = today
        while len(needed_months) < (period // 20 + 2):
            ym = (ref.year, ref.month)
            if ym not in needed_months:
                needed_months.append(ym)
            ref = (ref.replace(day=1) - timedelta(days=1))
        needed_months.sort()

        # ── 1. 股價 OHLCV（yfinance） ────────────────────────────────────────
        price_df   = pd.DataFrame()
        stock_name = code
        ticker_obj = None
        is_otc     = False    # True = 上櫃(TPEX) / False = 上市(TWSE)
        try:
            for suffix in ['.TW', '.TWO']:
                t = yf.Ticker(f'{code}{suffix}')
                h = t.history(period=f'{period + 15}d')
                if not h.empty:
                    ticker_obj = t
                    is_otc = (suffix == '.TWO')
                    cols = [c for c in ['Open', 'High', 'Low', 'Close', 'Volume'] if c in h.columns]
                    price_df = h[cols].copy()
                    price_df.index = pd.to_datetime(price_df.index).tz_localize(None)
                    price_df = price_df.tail(period)
                    break
            if ticker_obj is not None:
                info = ticker_obj.fast_info
                lname = getattr(info, 'currency', None)   # dummy access to trigger load
                info2 = ticker_obj.info
                raw_name = info2.get('longName') or info2.get('shortName') or code
                for sfx in ['股份有限公司', ' Inc.', ' Corp.', ' Co.,Ltd.', ', Ltd.', ',Ltd']:
                    raw_name = raw_name.replace(sfx, '')
                stock_name = raw_name.strip()
        except Exception:
            pass

        # ── 2. 三大法人（TWSE TWT38U / TPEX 備用） ───────────────────────────
        def _net_lots(rec: dict, inst_frag: str) -> int:
            """TWSE TWT38U：欄位含機構關鍵字 AND '買賣超' → 張數"""
            for k, v in rec.items():
                if inst_frag in k and '買賣超' in k:
                    try:
                        return int(str(v).replace(',', '').replace('+', '') or '0') // 1000
                    except Exception:
                        pass
            return 0

        def _parse_int(s) -> int:
            try:
                return int(str(s).replace(',', '').replace('+', '').strip() or '0')
            except Exception:
                return 0

        inst_rows: list[dict] = []

        if not is_otc:
            # 上市 → TWSE TWT38U（月查）
            for (yr, mo) in needed_months:
                try:
                    url  = ('https://www.twse.com.tw/rwd/zh/fund/TWT38U'
                            f'?date={yr}{mo:02d}01&stockNo={code}&response=json')
                    data = _cffi_get_json(url, headers=hdr, timeout=15)
                    if not isinstance(data, dict) or data.get('stat') != 'OK':
                        continue
                    fields = data.get('fields', [])
                    for row_vals in data.get('data', []):
                        if len(row_vals) < len(fields):
                            continue
                        rec = dict(zip(fields, row_vals))
                        d   = _roc_to_date(rec.get('日期', ''))
                        if d is None:
                            continue
                        dealer = (_net_lots(rec, '自行買賣') + _net_lots(rec, '避險'))
                        if dealer == 0:
                            dealer = _net_lots(rec, '自營商')
                        inst_rows.append({
                            'date':    d,
                            'foreign': _net_lots(rec, '不含外資自營商'),
                            'trust':   _net_lots(rec, '投信'),
                            'dealer':  dealer,
                        })
                except Exception:
                    continue
        else:
            # 上櫃 → TPEX 三大法人月查（aaData 欄位：日期/外資買賣超/投信買賣超/自營商買賣超）
            # cols: 0=日期,1=外資買進,2=外資賣出,3=外資買賣超,4=投信買進,5=投信賣出,
            #       6=投信買賣超,7=自營商買進,8=自營商賣出,9=自營商買賣超,10=合計
            for (yr, mo) in needed_months:
                try:
                    roc_yr = yr - 1911
                    url = ('https://www.tpex.org.tw/web/stock/3insti/daily_trade/'
                           f'3itrade_hedge_result.php?l=zh-tw&d={roc_yr}/{mo:02d}/15'
                           f'&s={code},0&o=json')
                    data = _cffi_get_json(url, headers=hdr, timeout=15)
                    if not isinstance(data, dict) or not data.get('success'):
                        continue
                    for row in data.get('aaData', []):
                        if len(row) < 10:
                            continue
                        d = _roc_to_date(str(row[0]).strip())
                        if d is None:
                            continue
                        inst_rows.append({
                            'date':    d,
                            'foreign': _parse_int(row[3])  // 1000,
                            'trust':   _parse_int(row[6])  // 1000,
                            'dealer':  _parse_int(row[9])  // 1000,
                        })
                except Exception:
                    continue

        # 去重、排序、截取
        seen_i: dict = {}
        for r in inst_rows:
            seen_i[r['date']] = r
        inst_rows = sorted(seen_i.values(), key=lambda r: r['date'])[-period:]
        print(f'[chip] 三大法人 {len(inst_rows)} 筆  is_otc={is_otc}')

        # ── 3. 融資融券（TWSE MI_MARGN / TPEX 備用） ─────────────────────────
        margin_rows: list[dict] = []

        if not is_otc:
            # 上市 → TWSE MI_MARGN（月查，data1=融資 / data2=融券）
            for (yr, mo) in needed_months:
                try:
                    url  = ('https://www.twse.com.tw/rwd/zh/marginTrading/MI_MARGN'
                            f'?date={yr}{mo:02d}01&stockNo={code}&response=json')
                    data = _cffi_get_json(url, headers=hdr, timeout=15)
                    if not isinstance(data, dict):
                        continue
                    f1 = data.get('fields1', data.get('fields', []))
                    d1 = data.get('data1',   data.get('data',   []))
                    f2 = data.get('fields2', [])
                    d2 = data.get('data2',   [])

                    short_by_date: dict = {}
                    for rv in d2:
                        if not rv:
                            continue
                        d_ = _roc_to_date(str(rv[0]))
                        if d_ is None:
                            continue
                        try:
                            col_idx = next((i for i, h in enumerate(f2) if '餘額' in str(h)), 3)
                            short_by_date[d_] = _parse_int(rv[col_idx])
                        except Exception:
                            short_by_date[d_] = 0

                    for rv in d1:
                        if not rv:
                            continue
                        d_ = _roc_to_date(str(rv[0]))
                        if d_ is None:
                            continue
                        try:
                            col_idx = next((i for i, h in enumerate(f1) if '餘額' in str(h)), 3)
                            m_bal = _parse_int(rv[col_idx])
                        except Exception:
                            m_bal = 0
                        margin_rows.append({
                            'date':       d_,
                            'margin_bal': m_bal,
                            'short_bal':  short_by_date.get(d_, 0),
                        })
                except Exception:
                    continue
        else:
            # 上櫃 → TPEX 融資融券月查
            # aaData cols: 0=日期, 1=融資買進, 2=融資賣出, 3=融資現償, 4=融資餘額,
            #              5=融資限額, 6=融券賣出, 7=融券買進, 8=融券現償, 9=融券餘額,
            #              10=融券限額, 11=資券互抵
            for (yr, mo) in needed_months:
                try:
                    roc_yr = yr - 1911
                    url = ('https://www.tpex.org.tw/web/stock/margin_trading/month/'
                           f'margin_month_result.php?l=zh-tw&d={roc_yr}/{mo:02d}'
                           f'&s={code},0&o=json')
                    data = _cffi_get_json(url, headers=hdr, timeout=15)
                    if not isinstance(data, dict) or not data.get('success'):
                        continue
                    for row in data.get('aaData', []):
                        if len(row) < 10:
                            continue
                        d_ = _roc_to_date(str(row[0]).strip())
                        if d_ is None:
                            continue
                        margin_rows.append({
                            'date':       d_,
                            'margin_bal': _parse_int(row[4]),
                            'short_bal':  _parse_int(row[9]),
                        })
                except Exception:
                    continue

        seen_m: dict = {}
        for r in margin_rows:
            seen_m[r['date']] = r
        margin_rows = sorted(seen_m.values(), key=lambda r: r['date'])[-period:]
        print(f'[chip] 融資融券 {len(margin_rows)} 筆  is_otc={is_otc}')

        # ── 4. 財報三率 & EPS（yfinance） ────────────────────────────────────
        fin_df  = pd.DataFrame()   # columns: gross_m, op_m, net_m  index: date
        eps_df  = pd.DataFrame()   # columns: EPS  index: date
        rev_df  = pd.DataFrame()   # columns: Revenue  index: date

        if ticker_obj is not None:
            try:
                qfin = ticker_obj.quarterly_financials   # rows=metrics, cols=dates
                if qfin is not None and not qfin.empty:
                    def _row(qfin, *labels):
                        for lb in labels:
                            if lb in qfin.index:
                                return qfin.loc[lb]
                        return None

                    rev  = _row(qfin, 'Total Revenue', 'Revenue')
                    gp   = _row(qfin, 'Gross Profit',  'GrossProfit')
                    op   = _row(qfin, 'Operating Income', 'EBIT')
                    ni   = _row(qfin, 'Net Income', 'NetIncome')

                    if rev is not None and rev.abs().max() > 0:
                        cols = {}
                        if gp  is not None: cols['gross_m'] = (gp  / rev * 100).round(2)
                        if op  is not None: cols['op_m']    = (op  / rev * 100).round(2)
                        if ni  is not None: cols['net_m']   = (ni  / rev * 100).round(2)
                        if cols:
                            fin_df = pd.DataFrame(cols)
                            fin_df.index = pd.to_datetime(fin_df.index).tz_localize(None)
                            fin_df = fin_df.sort_index().tail(16)
                    if rev is not None:
                        rev_df = rev.to_frame('Revenue')
                        rev_df.index = pd.to_datetime(rev_df.index).tz_localize(None)
                        rev_df = rev_df.sort_index().tail(16)
            except Exception:
                pass

            try:
                eps_ser = None
                qfin2 = ticker_obj.quarterly_financials
                if qfin2 is not None and 'Diluted EPS' in qfin2.index:
                    eps_ser = qfin2.loc['Diluted EPS']
                elif qfin2 is not None and 'Basic EPS' in qfin2.index:
                    eps_ser = qfin2.loc['Basic EPS']
                else:
                    qe = ticker_obj.quarterly_earnings
                    if qe is not None and not qe.empty and 'Earnings' in qe.columns:
                        eps_ser = qe['Earnings']
                if eps_ser is not None:
                    eps_df = eps_ser.to_frame('EPS')
                    eps_df.index = pd.to_datetime(eps_df.index).tz_localize(None)
                    eps_df = eps_df.sort_index().tail(16)
            except Exception:
                pass

        self._ui_call(lambda: self._render_chip_chart(
            code, stock_name, price_df, inst_rows, margin_rows,
            fin_df, eps_df, rev_df))

    # ── 圖表渲染 ──────────────────────────────────────────────────────────────

    def _render_chip_chart(self, code: str, name: str,
                           price_df: pd.DataFrame,
                           inst_rows: list,
                           margin_rows: list,
                           fin_df: pd.DataFrame,
                           eps_df: pd.DataFrame,
                           rev_df: pd.DataFrame):
        """在 Tab 6 繪製 K線 / 三大法人 / 融資融券 / 財報三率 / EPS / 月營收"""
        fig = self._chip_fig
        ax_candle, ax_vol, ax_inst, ax_mgn, ax_fin, ax_eps, ax_rev = self._chip_axes

        all_axes = list(self._chip_axes)
        # 清除舊圖（含 twinx 子軸）
        for ax in all_axes:
            ax.cla()
            for twin in ax.get_shared_x_axes().get_siblings(ax):
                if twin is not ax:
                    try:
                        twin.cla()
                    except Exception:
                        pass
        fig.clf()

        # 重新建立 axes（clf 會清空 gridspec 子圖）
        gs = fig.add_gridspec(
            5, 2,
            height_ratios=[2.4, 0.65, 1.2, 1.2, 1.2],
            hspace=0.75, wspace=0.40,
            left=0.07, right=0.97, top=0.92, bottom=0.05)
        ax_candle = fig.add_subplot(gs[0, :])
        ax_vol    = fig.add_subplot(gs[1, :])
        ax_inst   = fig.add_subplot(gs[2, :])
        ax_mgn    = fig.add_subplot(gs[3, 0])
        ax_fin    = fig.add_subplot(gs[3, 1])
        ax_eps    = fig.add_subplot(gs[4, 0])
        ax_rev    = fig.add_subplot(gs[4, 1])
        self._chip_axes = (ax_candle, ax_vol, ax_inst, ax_mgn, ax_fin, ax_eps, ax_rev)

        fp  = fm.FontProperties(family=CHART_FONT)
        NDM = '（查無資料）'

        def _style(ax):
            ax.set_facecolor(C_BG)
            ax.tick_params(colors=C_FG2, labelsize=7)
            for sp in ax.spines.values():
                sp.set_edgecolor(C_BORDER)
            ax.grid(axis='y', color=C_BORDER, linewidth=0.4, alpha=0.4)

        def _style2(ax):
            ax.tick_params(axis='y', labelsize=7)
            for sp in ax.spines.values():
                sp.set_edgecolor(C_BORDER)

        for ax in self._chip_axes:
            _style(ax)

        def _fmt_date_axis(ax, dates, n_ticks=8):
            if not dates:
                return
            ax.set_xlim(-0.5, len(dates) - 0.5)
            step = max(1, len(dates) // n_ticks)
            ticks = list(range(0, len(dates), step))
            ax.set_xticks(ticks)
            ax.set_xticklabels([dates[i].strftime('%m/%d') for i in ticks],
                                color=C_FG2, fontsize=6.5)

        period = self._chip_period.get()

        # ══ Panel 1：K線圖（蠟燭圖） ═════════════════════════════════════════
        if not price_df.empty and 'Open' in price_df.columns:
            ohlc = price_df[['Open', 'High', 'Low', 'Close']].dropna()
            xs_c = list(range(len(ohlc)))
            for i, (_, row) in enumerate(ohlc.iterrows()):
                o, h, l, c = row['Open'], row['High'], row['Low'], row['Close']
                col = '#ff4444' if c >= o else '#44cc44'   # 台灣：紅漲綠跌
                ax_candle.plot([i, i], [l, h], color=col, lw=0.8, zorder=2)
                body = abs(c - o) or (h - l) * 0.01
                ax_candle.add_patch(plt.Rectangle(
                    (i - 0.38, min(o, c)), 0.76, body,
                    color=col, alpha=0.88, zorder=3))
            # MA5 / MA20
            closes = ohlc['Close'].values
            if len(closes) >= 5:
                ma5 = pd.Series(closes).rolling(5).mean().values
                ax_candle.plot(xs_c, ma5, color='#ffcc44', lw=0.9,
                               label='MA5', alpha=0.8)
            if len(closes) >= 20:
                ma20 = pd.Series(closes).rolling(20).mean().values
                ax_candle.plot(xs_c, ma20, color='#44aaff', lw=0.9,
                               label='MA20', alpha=0.8)
            ax_candle.legend(loc='upper left', fontsize=7,
                              facecolor=C_PANEL, edgecolor=C_BORDER,
                              labelcolor=C_FG, prop=fp)
            ax_candle.set_ylabel('股價 (元)', color=C_FG2, fontsize=8, fontproperties=fp)
            ax_candle.yaxis.set_tick_params(labelcolor=C_FG2)
            ax_candle.set_title(
                f'K線走勢  最新:{closes[-1]:.2f}',
                color=C_FG2, fontsize=9, loc='left', fontproperties=fp)
            dates_c = [d.date() if hasattr(d, 'date') else d for d in ohlc.index]
            _fmt_date_axis(ax_candle, dates_c)
        elif not price_df.empty and 'Close' in price_df.columns:
            # 只有收盤價（fallback 折線）
            closes = price_df['Close'].dropna().values
            ax_candle.plot(closes, color='#f0c040', lw=1.2)
            ax_candle.fill_between(range(len(closes)), closes, alpha=0.1, color='#f0c040')
            ax_candle.set_title('收盤股價', color=C_FG2, fontsize=9, loc='left', fontproperties=fp)
            dates_c2 = [d.date() if hasattr(d, 'date') else d for d in price_df.index]
            _fmt_date_axis(ax_candle, dates_c2)
        else:
            ax_candle.text(0.5, 0.5, NDM, ha='center', va='center',
                           color=C_FG2, fontsize=11, transform=ax_candle.transAxes,
                           fontproperties=fp)

        # ══ Panel 2：成交量 ═══════════════════════════════════════════════════
        if not price_df.empty and 'Volume' in price_df.columns and 'Close' in price_df.columns:
            vols   = price_df['Volume'].fillna(0).values
            closes = price_df['Close'].values
            opens  = price_df['Open'].values if 'Open' in price_df.columns else closes
            vol_colors = ['#ff5555' if c >= o else '#55cc55'
                          for c, o in zip(closes, opens)]
            ax_vol.bar(range(len(vols)), vols, color=vol_colors, alpha=0.75, width=0.8)
            ax_vol.set_ylabel('成交量', color=C_FG2, fontsize=7, fontproperties=fp)
            ax_vol.yaxis.set_tick_params(labelcolor=C_FG2)
            ax_vol.yaxis.set_major_formatter(
                matplotlib.ticker.FuncFormatter(
                    lambda x, _: f'{x/1e6:.1f}M' if x >= 1e6 else f'{x/1e3:.0f}K'))
            _fmt_date_axis(ax_vol, dates_c if 'dates_c' in dir() else [])
        else:
            ax_vol.text(0.5, 0.5, NDM, ha='center', va='center',
                        color=C_FG2, fontsize=9, transform=ax_vol.transAxes,
                        fontproperties=fp)

        # ══ Panel 3：三大法人淨買超 ═══════════════════════════════════════════
        if inst_rows:
            dates_i = [r['date'] for r in inst_rows]
            xs_i    = list(range(len(inst_rows)))
            f_vals  = [r['foreign'] for r in inst_rows]
            t_vals  = [r['trust']   for r in inst_rows]
            d_vals  = [r['dealer']  for r in inst_rows]

            bw = 0.27
            ax_inst.bar([x - bw for x in xs_i], f_vals, width=bw,
                        color='#4a9eff', label='外資', alpha=0.85)
            ax_inst.bar(xs_i,               t_vals, width=bw,
                        color='#ffaa44', label='投信', alpha=0.85)
            ax_inst.bar([x + bw for x in xs_i], d_vals, width=bw,
                        color='#bb88ff', label='自營商', alpha=0.85)
            ax_inst.axhline(0, color=C_BORDER, lw=0.8)
            ax_inst.set_ylabel('淨買超（張）', color=C_FG2, fontsize=8, fontproperties=fp)
            ax_inst.yaxis.set_tick_params(labelcolor=C_FG2)
            ax_inst.legend(loc='upper left', fontsize=7,
                           facecolor=C_PANEL, edgecolor=C_BORDER,
                           labelcolor=C_FG, prop=fp)
            _fmt_date_axis(ax_inst, dates_i)

            # 外資累計折線（右軸）
            cum_f  = pd.Series(f_vals).cumsum().values
            ax_i2  = ax_inst.twinx()
            ax_i2.plot(xs_i, cum_f, color='#4a9eff', lw=1.1, ls='--', alpha=0.55)
            ax_i2.set_ylabel('外資累計', color='#4a9eff', fontsize=7, fontproperties=fp)
            ax_i2.tick_params(axis='y', colors='#4a9eff', labelsize=6.5)
            _style2(ax_i2)
            ax_inst.set_title('三大法人淨買超', color=C_FG2, fontsize=9,
                               loc='left', fontproperties=fp)
        else:
            ax_inst.text(0.5, 0.5, NDM, ha='center', va='center',
                         color=C_FG2, fontsize=11, transform=ax_inst.transAxes,
                         fontproperties=fp)
            ax_inst.set_title('三大法人淨買超', color=C_FG2, fontsize=9,
                               loc='left', fontproperties=fp)

        # ══ Panel 4：融資融券餘額 ═════════════════════════════════════════════
        if margin_rows:
            dates_m = [r['date']       for r in margin_rows]
            xs_m    = list(range(len(margin_rows)))
            m_bal   = [r['margin_bal'] for r in margin_rows]
            s_bal   = [r['short_bal']  for r in margin_rows]
            ax_mgn.plot(xs_m, m_bal, color='#ff6b6b', lw=1.2, label='融資餘額')
            ax_mgn.fill_between(xs_m, m_bal, alpha=0.10, color='#ff6b6b')
            ax_mgn.set_ylabel('融資（張）', color='#ff6b6b', fontsize=7, fontproperties=fp)
            ax_mgn.yaxis.set_tick_params(labelcolor='#ff6b6b')
            _fmt_date_axis(ax_mgn, dates_m, n_ticks=6)
            ax_m2 = ax_mgn.twinx()
            ax_m2.plot(xs_m, s_bal, color='#5de85d', lw=1.2, label='融券餘額')
            ax_m2.fill_between(xs_m, s_bal, alpha=0.08, color='#5de85d')
            ax_m2.set_ylabel('融券（張）', color='#5de85d', fontsize=7, fontproperties=fp)
            ax_m2.tick_params(axis='y', colors='#5de85d', labelsize=6.5)
            _style2(ax_m2)
            lns = [Line2D([0],[0],color='#ff6b6b',lw=1.2),
                   Line2D([0],[0],color='#5de85d',lw=1.2)]
            ax_mgn.legend(lns, ['融資', '融券'], loc='upper left', fontsize=7,
                          facecolor=C_PANEL, edgecolor=C_BORDER,
                          labelcolor=C_FG, prop=fp)
        else:
            ax_mgn.text(0.5, 0.5, NDM, ha='center', va='center',
                        color=C_FG2, fontsize=10, transform=ax_mgn.transAxes,
                        fontproperties=fp)
        ax_mgn.set_title('融資融券餘額', color=C_FG2, fontsize=8.5,
                          loc='left', fontproperties=fp)

        # ══ Panel 5：財報三率 ═════════════════════════════════════════════════
        if not fin_df.empty:
            xs_f = list(range(len(fin_df)))
            dts_f = [d.strftime(f'%y/Q{((d.month-1)//3)+1}') for d in fin_df.index]
            if 'gross_m' in fin_df.columns:
                ax_fin.plot(xs_f, fin_df['gross_m'], color='#4a9eff', lw=1.1, label='毛利率')
            if 'op_m' in fin_df.columns:
                ax_fin.plot(xs_f, fin_df['op_m'],    color='#ffaa44', lw=1.1, label='營業利益率')
            if 'net_m' in fin_df.columns:
                ax_fin.plot(xs_f, fin_df['net_m'],   color='#ff6b6b', lw=1.1, label='稅後淨利率')
            ax_fin.axhline(0, color=C_BORDER, lw=0.7)
            ax_fin.set_ylabel('%', color=C_FG2, fontsize=7, fontproperties=fp)
            ax_fin.yaxis.set_tick_params(labelcolor=C_FG2)
            step_f = max(1, len(xs_f) // 6)
            ax_fin.set_xticks(xs_f[::step_f])
            ax_fin.set_xticklabels(dts_f[::step_f], color=C_FG2, fontsize=6.5)
            ax_fin.legend(loc='upper left', fontsize=6.5,
                          facecolor=C_PANEL, edgecolor=C_BORDER,
                          labelcolor=C_FG, prop=fp)
        else:
            ax_fin.text(0.5, 0.5, NDM, ha='center', va='center',
                        color=C_FG2, fontsize=10, transform=ax_fin.transAxes,
                        fontproperties=fp)
        ax_fin.set_title('財報三率', color=C_FG2, fontsize=8.5,
                          loc='left', fontproperties=fp)

        # ══ Panel 6：每季 EPS ════════════════════════════════════════════════
        if not eps_df.empty:
            eps_vals = eps_df['EPS'].values
            xs_e     = list(range(len(eps_df)))
            colors_e = ['#ff5555' if v >= 0 else '#55cc55' for v in eps_vals]
            ax_eps.bar(xs_e, eps_vals, color=colors_e, alpha=0.82, width=0.65)
            ax_eps.axhline(0, color=C_BORDER, lw=0.7)
            ax_eps.set_ylabel('EPS (元)', color=C_FG2, fontsize=7, fontproperties=fp)
            ax_eps.yaxis.set_tick_params(labelcolor=C_FG2)
            dts_e = [d.strftime(f'%y/Q{((d.month-1)//3)+1}') for d in eps_df.index]
            step_e = max(1, len(xs_e) // 6)
            ax_eps.set_xticks(xs_e[::step_e])
            ax_eps.set_xticklabels(dts_e[::step_e], color=C_FG2, fontsize=6.5)
        else:
            ax_eps.text(0.5, 0.5, NDM, ha='center', va='center',
                        color=C_FG2, fontsize=10, transform=ax_eps.transAxes,
                        fontproperties=fp)
        ax_eps.set_title('每季 EPS', color=C_FG2, fontsize=8.5,
                          loc='left', fontproperties=fp)

        # ══ Panel 7：每季營收 ════════════════════════════════════════════════
        if not rev_df.empty:
            rev_vals = rev_df['Revenue'].values / 1e8   # 億
            xs_r     = list(range(len(rev_df)))
            ax_rev.bar(xs_r, rev_vals, color='#44aaff', alpha=0.75, width=0.65)
            ax_rev.set_ylabel('營收（億）', color=C_FG2, fontsize=7, fontproperties=fp)
            ax_rev.yaxis.set_tick_params(labelcolor=C_FG2)
            dts_r = [d.strftime(f'%y/Q{((d.month-1)//3)+1}') for d in rev_df.index]
            step_r = max(1, len(xs_r) // 6)
            ax_rev.set_xticks(xs_r[::step_r])
            ax_rev.set_xticklabels(dts_r[::step_r], color=C_FG2, fontsize=6.5)
        else:
            ax_rev.text(0.5, 0.5, NDM, ha='center', va='center',
                        color=C_FG2, fontsize=10, transform=ax_rev.transAxes,
                        fontproperties=fp)
        ax_rev.set_title('每季營收', color=C_FG2, fontsize=8.5,
                          loc='left', fontproperties=fp)

        # ── 主標題 ───────────────────────────────────────────────────────────
        fig.suptitle(f'{code}  {name}  ── 個股分析（近 {period} 日）',
                     color=C_FG, fontsize=12, fontproperties=fp)

        self._chip_canvas.draw()
        self._chip_status.set(f'{code}  {name}  ─  資料載入完成')


# ─── 修改交易對話框 ────────────────────────────────────────────────────────────
class EditDialog(tk.Toplevel):
    """彈出式修改視窗，深色風格，與主視窗一致"""

    def __init__(self, parent: tk.Tk, idx: int, row: pd.Series, on_save):
        super().__init__(parent)
        self._idx     = idx
        self._row     = row
        self._on_save = on_save

        self.title(f'修改交易紀錄  #{idx + 1}')
        self.configure(bg=C_BG)
        self.resizable(False, False)
        self.grab_set()   # modal

        # 置中
        self.update_idletasks()
        pw, ph = parent.winfo_width(), parent.winfo_height()
        px, py = parent.winfo_rootx(), parent.winfo_rooty()
        w, h   = 500, 390
        self.geometry(f'{w}x{h}+{px + (pw - w)//2}+{py + (ph - h)//2}')

        self._vars: dict[str, tk.StringVar] = {}
        self._build()

    def _build(self):
        r = self._row

        pad = dict(padx=14, pady=5)
        lkw = dict(sticky='e', **pad)
        ekw = dict(sticky='ew', **pad)

        f = ttk.Frame(self, style='TFrame')
        f.pack(fill='both', expand=True, padx=14, pady=12)
        f.columnconfigure(1, weight=1)

        today = datetime.today()

        def add_row(row_idx, label, key, widget_fn):
            ttk.Label(f, text=label).grid(row=row_idx, column=0, **lkw)
            w = widget_fn(f)
            w.grid(row=row_idx, column=1, **ekw)
            return w

        # 日期
        ttk.Label(f, text='日期').grid(row=0, column=0, **lkw)
        orig_date = pd.to_datetime(r['日期'])
        self._date_entry = DatePickerEntry(f, initial_date=orig_date)
        self._date_entry.grid(row=0, column=1, **ekw)

        # 股票代號 / 名稱
        for row_idx, (label, key, default) in enumerate([
            ('股票代號', 'code', str(r['股票代號'])),
            ('股票名稱', 'name', str(r['股票名稱'])),
        ], start=1):
            self._vars[key] = tk.StringVar(value=default)
            add_row(row_idx, label, key,
                    lambda p, k=key: ttk.Entry(p, textvariable=self._vars[k], width=22))

        # 分類 / 買賣
        self._vars['category'] = tk.StringVar(value=str(r.get('分類', CATEGORIES[0])))
        ttk.Label(f, text='投資分類').grid(row=3, column=0, **lkw)
        ttk.Combobox(f, textvariable=self._vars['category'],
                     values=CATEGORIES, state='readonly', width=20
                     ).grid(row=3, column=1, **ekw)

        self._vars['side'] = tk.StringVar(value=str(r['買賣']))
        ttk.Label(f, text='買 / 賣').grid(row=4, column=0, **lkw)
        self._side_cb = ttk.Combobox(f, textvariable=self._vars['side'],
                                      values=['買', '賣'], state='readonly', width=20,
                                      style=('Buy.TCombobox' if r['買賣'] == '買'
                                             else 'Sell.TCombobox'))
        self._side_cb.grid(row=4, column=1, **ekw)
        self._vars['side'].trace_add('write', self._on_side_change)

        # 數量 / 價格 / 手續費
        for row_idx, (label, key, default) in enumerate([
            ('數量 (股)',    'qty',   str(int(r['數量(股)']))),
            ('成交價格 (元)', 'price', str(r['價格(元)'])),
            ('手續費 (元)',  'fee',   str(int(r['手續費(元)']))),
        ], start=5):
            self._vars[key] = tk.StringVar(value=default)
            add_row(row_idx, label, key,
                    lambda p, k=key: ttk.Entry(p, textvariable=self._vars[k], width=22))

        # 自動重算淨額提示
        self._note = ttk.Label(f, text='交易稅與淨金額將依買賣方向自動重新計算',
                               foreground=C_FG2, font=('Microsoft JhengHei', 8))
        self._note.grid(row=8, column=0, columnspan=2, pady=(4, 0))

        # 按鈕
        btn_f = ttk.Frame(f)
        btn_f.grid(row=9, column=0, columnspan=2, pady=(10, 0))
        ttk.Button(btn_f, text='💾  儲存修改', command=self._save).pack(side='left', padx=6)
        ttk.Button(btn_f, text='✖  取消',      command=self.destroy).pack(side='left', padx=6)

    def _on_side_change(self, *_):
        s = self._vars['side'].get()
        self._side_cb.configure(style='Buy.TCombobox' if s == '買' else 'Sell.TCombobox')

    def _save(self):
        try:
            date_str = self._date_entry.get_date().strftime('%Y-%m-%d')
            code     = self._vars['code'].get().strip().upper()
            name     = self._vars['name'].get().strip()
            cat      = self._vars['category'].get()
            side     = self._vars['side'].get()
            qty_s    = self._vars['qty'].get().strip()
            price_s  = self._vars['price'].get().strip()
            fee_s    = self._vars['fee'].get().strip()

            if not code or not name:
                raise ValueError('請填寫股票代號和名稱')
            if not qty_s or not price_s or not fee_s:
                raise ValueError('請填寫數量、價格與手續費')

            tx_date = datetime.strptime(date_str, '%Y-%m-%d')
            qty, price, fee = float(qty_s), float(price_s), float(fee_s)
            if qty <= 0 or price <= 0:
                raise ValueError('數量和價格必須大於 0')

            gross = qty * price
            tax   = round(gross * TAX_RATE) if side == '賣' else 0
            net   = round(gross - fee - tax) if side == '賣' else -round(gross + fee)

            new_row = {
                '日期':      tx_date,
                '股票代號':  code,
                '股票名稱':  name,
                '分類':      cat,
                '買賣':      side,
                '數量(股)':  int(qty),
                '價格(元)':  price,
                '手續費(元)': int(fee),
                '交易稅(元)': tax,
                '淨金額(元)': net,
            }
            self._on_save(self._idx, new_row)
            self.destroy()

        except ValueError as e:
            messagebox.showerror('輸入錯誤', str(e), parent=self)
        except Exception as e:
            messagebox.showerror('錯誤', str(e), parent=self)


# ─── 啟動 ────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    app = StockApp()
    app.mainloop()